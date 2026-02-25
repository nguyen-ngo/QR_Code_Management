from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, Response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename 
from functools import wraps
from datetime import datetime, date, time, timedelta
from sqlalchemy import text
from user_agents import parse
import io, os, base64, re, uuid, requests, json, qrcode, math, traceback, googlemaps
import openpyxl.cell.cell
from PIL import Image, ImageDraw
from math import radians, sin, cos, asin, sqrt
from dotenv import load_dotenv
# Import the logging handler
from logger_handler import AppLogger, log_user_activity, log_database_operations

from working_hours_calculator import (WorkingHoursCalculator,
                                      round_time_to_quarter_hour)
from payroll_excel_exporter import PayrollExcelExporter
from enhanced_payroll_excel_exporter import EnhancedPayrollExcelExporter
from time_attendance_import_service import TimeAttendanceImportService
from qr_code_import_service import QRCodeImportService

# Load environment variables in .env
load_dotenv()
from turnstile_utils import turnstile_utils
from db_performance_optimization import initialize_performance_optimizations
from app_performance_middleware import PerformanceMonitor
from address_normalization_fix import normalize_address, addresses_are_similar

# Initialize Flask application
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = os.environ.get('SQLALCHEMY_TRACK_MODIFICATIONS')
app.config['TEMPLATES_AUTO_RELOAD'] = os.environ.get('TEMPLATES_AUTO_RELOAD')

# Session configuration for "Remember Me" functionality
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE')   # Set to True if using HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = os.environ.get('SESSION_COOKIE_HTTPONLY')
app.config['SESSION_COOKIE_SAMESITE'] = os.environ.get('SESSION_COOKIE_SAMESITE')

# Photo Verification Configuration
PHOTO_VERIFICATION_ENABLED = os.environ.get('ENABLE_PHOTO_VERIFICATION', 'true').lower() == 'true'
DISTANCE_THRESHOLD_FOR_VERIFICATION = float(os.environ.get('PHOTO_VERIFICATION_DISTANCE_THRESHOLD', '0.3'))
VERIFICATION_PHOTO_MAX_SIZE = int(os.environ.get('VERIFICATION_PHOTO_MAX_SIZE', str(5 * 1024 * 1024)))

# Initialize database
db = SQLAlchemy(app)

@app.context_processor
def inject_company_name():
    """Make COMPANY_NAME available to all templates"""
    return {
        'COMPANY_NAME': os.environ.get('COMPANY_NAME', 'QR Code Management System')
    }

def create_performance_indexes():
    """Create performance optimization indexes"""
    try:
        print("🔄 Creating performance indexes...")
        
        # Attendance data indexes
        db.session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_attendance_date_employee_location 
            ON attendance_data (check_in_date DESC, employee_id, location_name)
        """))
        
        db.session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_attendance_date_qr_gps 
            ON attendance_data (check_in_date, qr_code_id, latitude, longitude)
        """))
        
        db.session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_attendance_employee_date 
            ON attendance_data (employee_id, check_in_date DESC)
        """))
        
        # QR Code indexes
        db.session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_qrcode_project_active 
            ON qr_codes (project_id, active_status)
        """))
        
        # User indexes
        db.session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_users_username_active 
            ON users (username, active_status)
        """))
        
        # Log event indexes
        db.session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_log_events_timestamp_category 
            ON log_events (created_timestamp DESC, event_category)
        """))
        
        db.session.commit()
        print("✅ Performance indexes created successfully")
        
        # Log the optimization
        logger_handler.log_system_event(
            event_type="database_optimization",
            description="Performance indexes created successfully",
            severity="INFO",
            additional_data={"optimization_type": "indexes"}
        )
        
    except Exception as e:
        print(f"❌ Error creating performance indexes: {e}")
        db.session.rollback()
        logger_handler.log_database_error(
            error_type="index_creation_error",
            error_message=str(e),
            query="CREATE INDEX statements"
        )

def create_audit_triggers():
    """Create audit and integrity triggers"""
    try:
        print("🔄 Creating audit triggers...")
        
        # Create audit table
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS attendance_audit (
                audit_id INT AUTO_INCREMENT PRIMARY KEY,
                record_id INT NOT NULL,
                action_type ENUM('INSERT', 'UPDATE', 'DELETE') NOT NULL,
                old_values JSON,
                new_values JSON,
                changed_by VARCHAR(100),
                change_timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                ip_address VARCHAR(45),
                INDEX idx_audit_record_timestamp (record_id, change_timestamp),
                INDEX idx_audit_action_timestamp (action_type, change_timestamp)
            ) ENGINE=InnoDB
        """))
        
        # Create audit trigger (simplified version)
        db.session.execute(text("""
            DROP TRIGGER IF EXISTS tr_attendance_insert_audit
        """))
        
        db.session.execute(text("""
            CREATE TRIGGER tr_attendance_insert_audit
                AFTER INSERT ON attendance_data
                FOR EACH ROW
                INSERT INTO attendance_audit (
                    record_id, action_type, new_values, changed_by
                ) VALUES (
                    NEW.id, 'INSERT', 
                    JSON_OBJECT(
                        'employee_id', NEW.employee_id,
                        'location_name', NEW.location_name,
                        'check_in_date', NEW.check_in_date
                    ),
                    USER()
                )
        """))
        
        db.session.commit()
        print("✅ Audit triggers created successfully")
        
    except Exception as e:
        print(f"❌ Error creating audit triggers: {e}")
        db.session.rollback()

from location_logging import *

# Initialize Google Maps client
try:
    GOOGLE_MAPS_API_KEY = os.environ.get('GOOGLE_MAPS_API_KEY')
    if GOOGLE_MAPS_API_KEY:
        gmaps_client = googlemaps.Client(key=GOOGLE_MAPS_API_KEY)
        print("✅ Google Maps client initialized successfully")
    else:
        gmaps_client = None
        print("⚠️ Google Maps API key not found, falling back to OpenStreetMap")
except Exception as e:
    gmaps_client = None
    print(f"❌ Error initializing Google Maps client: {e}")

# Geocoding cache to reduce API calls and costs
geocoding_cache = {}
CACHE_MAX_SIZE = 1000
CACHE_EXPIRY_HOURS = 24

def get_cached_coordinates(address):
    """Get coordinates from cache if available and not expired"""
    if address in geocoding_cache:
        cached_data = geocoding_cache[address]
        cache_time = cached_data.get('timestamp', datetime.min)
        
        # Check if cache is still valid (24 hours)
        if datetime.now() - cache_time < timedelta(hours=CACHE_EXPIRY_HOURS):
            print(f"📋 Using cached coordinates for: {address[:50]}...")
            return cached_data.get('lat'), cached_data.get('lng'), cached_data.get('accuracy')
    
    return None, None, None

def cache_coordinates(address, lat, lng, accuracy):
    """Cache coordinates to reduce future API calls"""
    try:
        # Implement simple cache size limit
        if len(geocoding_cache) >= CACHE_MAX_SIZE:
            # Remove oldest entries (simple FIFO)
            oldest_key = min(geocoding_cache.keys(), key=lambda k: geocoding_cache[k]['timestamp'])
            del geocoding_cache[oldest_key]
        
        geocoding_cache[address] = {
            'lat': lat,
            'lng': lng,
            'accuracy': accuracy,
            'timestamp': datetime.now()
        }
        print(f"💾 Cached coordinates for: {address[:50]}...")
    except Exception as e:
        print(f"⚠️ Error caching coordinates: {e}")

# Valid user roles with new additions
VALID_ROLES = ['admin', 'staff', 'payroll', 'project_manager', 'accounting']

# Roles that have staff-level permissions (non-admin roles)
STAFF_LEVEL_ROLES = ['staff', 'payroll', 'project_manager', 'accounting']

# Import and initialize models
from models import set_db
User, QRCode, QRCodeStyle, Project, AttendanceData, Employee, TimeAttendance, UserProjectPermission, UserLocationPermission = set_db(db)

# Initialize the logging system
logger_handler = AppLogger(app, db)

# Utility functions
def is_valid_role(role):
    """Check if role is valid"""
    return role in VALID_ROLES

def has_admin_privileges(role):
    """Check if role has admin privileges"""
    return role == 'admin'

def has_staff_level_access(role):
    """Check if role has staff-level access (includes new roles)"""
    return role in STAFF_LEVEL_ROLES

def get_role_permissions(role):
    """Get permissions description for a role"""
    permissions = {
        'admin': {
            'title': 'Administrator Permissions',
            'permissions': [
                'Full QR code management (create, edit, delete)',
                'Complete user management capabilities',
                'System configuration access',
                'View all system analytics',
                'Bulk operations and data export',
                'Access to all admin features'
            ],
            'restrictions': ['With great power comes great responsibility!']
        },
        'staff': {
            'title': 'Staff User Permissions',
            'permissions': [
                'Create and edit QR codes',
                'View all QR codes in the system',
                'Download QR code images',
                'Update personal profile information',
            ],
            'restrictions': [
                'Cannot delete QR codes',
                'Cannot manage other users',
                'Cannot access admin settings'
            ]
        },
        'payroll': {
            'title': 'Payroll Specialist Permissions',
            'permissions': [
                'Create and edit QR codes',
                'View all QR codes in the system',
                'Download QR code images',
                'Update personal profile information',
                'Access dashboard and reports',
                'Same permissions as Staff (additional features coming soon)'
            ],
            'restrictions': [
                'Cannot delete QR codes',
                'Cannot manage other users',
                'Cannot access admin settings'
            ]
        },
        'project_manager': {
            'title': 'Project Manager Permissions',
            'permissions': [
                'Create and edit QR codes',
                'View all QR codes in the system',
                'Download QR code images',
                'Update personal profile information',
                'Access dashboard and reports',
                'Same permissions as Staff (additional features coming soon)'
            ],
            'restrictions': [
                'Cannot delete QR codes',
                'Cannot manage other users',
                'Cannot access admin settings'
            ]
        },
        'accounting': {
            'title': 'Accounting Specialist Permissions',
            'permissions': [
                'View and modify employee records',
                'Access attendance reports and analytics',
                'View and manage time attendance data',
                'Export payroll and attendance data',
                'Access financial reports and statistics',
                'Update personal profile information',
                'Delete attendance records (same as payroll)'
            ],
            'restrictions': [
                'Cannot create or delete QR codes',
                'Cannot manage other users',
                'Cannot access admin settings',
                'Cannot manage projects'
            ]
        }
    }
    return permissions.get(role, {})

def log_google_maps_usage(operation_type):
    """Log Google Maps API usage for monitoring"""
    try:
        logger_handler.log_user_activity('google_maps_api_usage', f'Google Maps API used: {operation_type}')
    except Exception as e:
        print(f"⚠️ Usage logging error: {e}")

def get_coordinates_from_address(address):
    """
    Get latitude and longitude from address using Google Maps Geocoding API
    Falls back to OpenStreetMap if Google Maps is unavailable
    Returns (lat, lng) tuple or (None, None) if failed
    """
    if not address or address.strip() == '':
        return None, None

    address = address.strip()
    print(f"🌍 Geocoding address: {address}")

    # Log geocoding action
    try:
        logger_handler.log_user_activity('geocoding', f'Geocoding address: {address[:50]}...')
    except Exception as log_error:
        print(f"⚠️ Logging error (non-critical): {log_error}")

    try:
        # Primary: Use Google Maps Geocoding API
        if gmaps_client:
            print(f"🗺️ Using Google Maps Geocoding API")
            
            geocode_result = gmaps_client.geocode(address)
            
            if geocode_result:
                location = geocode_result[0]['geometry']['location']
                lat = location['lat']
                lng = location['lng']
                
                print(f"✅ Google Maps geocoded address '{address[:50]}...' to coordinates: {lat}, {lng}")
                
                # Log successful geocoding
                try:
                    logger_handler.log_user_activity('geocoding_success', f'Successfully geocoded: {address[:50]}... -> {lat}, {lng}')
                except Exception as log_error:
                    print(f"⚠️ Logging error (non-critical): {log_error}")
                
                return lat, lng
            else:
                print(f"⚠️ Google Maps: No results found for address: {address}")
        
        # Fallback: Use OpenStreetMap Nominatim
        print(f"🌐 Falling back to OpenStreetMap Nominatim")
        url = "https://nominatim.openstreetmap.org/search"
        params = {
            'q': address,
            'format': 'json',
            'limit': 1,
            'addressdetails': 1
        }

        headers = {
            'User-Agent': 'QR-Attendance-System/1.0'
        }

        response = requests.get(url, params=params, headers=headers, timeout=10)

        if response.status_code == 200:
            data = response.json()
            if data and len(data) > 0:
                lat = float(data[0]['lat'])
                lng = float(data[0]['lon'])
                print(f"✅ OSM geocoded address '{address[:50]}...' to coordinates: {lat}, {lng}")
                
                # Log fallback geocoding
                try:
                    logger_handler.log_user_activity('geocoding_fallback', f'OSM fallback geocoded: {address[:50]}... -> {lat}, {lng}')
                except Exception as log_error:
                    print(f"⚠️ Logging error (non-critical): {log_error}")
                
                return lat, lng

        print(f"⚠️ Could not geocode address: {address}")
        
        # Log geocoding failure
        try:
            logger_handler.log_user_activity('geocoding_failed', f'Failed to geocode: {address[:50]}...')
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")
        
        return None, None

    except Exception as e:
        print(f"❌ Error geocoding address '{address}': {e}")
        
        # Log geocoding error
        try:
            logger_handler.log_flask_error('geocoding_error', f'Error geocoding {address[:50]}...: {str(e)}')
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")
        
        return None, None

def get_coordinates_from_address_enhanced(address):
    """
    Enhanced geocoding function using Google Maps with caching and better error handling
    Returns (latitude, longitude, accuracy_level)
    """
    if not address or address.strip() == "":
        return None, None, None
        
    address = address.strip()
    print(f"🌍 Enhanced geocoding for: {address}")
        
    # STEP 1: Normalize address before any processing
    normalized_address = normalize_address(address)
        
    # STEP 2: Check cache using normalized address
    cached_lat, cached_lng, cached_accuracy = get_cached_coordinates(normalized_address)
    if cached_lat is not None:
        print(f"✅ Using cached coordinates for normalized address")
        return cached_lat, cached_lng, cached_accuracy

    # Log enhanced geocoding action
    try:
        logger_handler.log_user_activity('enhanced_geocoding', f'Enhanced geocoding: {address[:50]}...')
    except Exception as log_error:
        print(f"⚠️ Logging error (non-critical): {log_error}")

    try:
        # Primary: Use Google Maps Geocoding API
        if gmaps_client:
            print(f"🗺️ Using Google Maps Geocoding API (Enhanced)")
            
            geocode_result = gmaps_client.geocode(address)
            
            if geocode_result:
                result = geocode_result[0]
                location = result['geometry']['location']
                lat = location['lat']
                lng = location['lng']
                
                # Determine accuracy based on Google Maps location type
                location_type = result['geometry'].get('location_type', 'UNKNOWN')
                place_types = result.get('types', [])
                
                # Enhanced accuracy assessment based on Google Maps data
                if location_type == 'ROOFTOP':
                    accuracy = 'excellent'  # Building-level accuracy
                elif location_type == 'RANGE_INTERPOLATED':
                    accuracy = 'good'       # Street-level accuracy
                elif location_type == 'GEOMETRIC_CENTER':
                    if any(ptype in place_types for ptype in ['premise', 'subpremise', 'street_address']):
                        accuracy = 'good'
                    elif any(ptype in place_types for ptype in ['neighborhood', 'sublocality']):
                        accuracy = 'fair'
                    else:
                        accuracy = 'poor'
                elif location_type == 'APPROXIMATE':
                    accuracy = 'poor'      # City/region level
                else:
                    accuracy = 'fair'      # Unknown, assume moderate

                print(f"✅ Google Maps enhanced geocoding successful:")
                print(f"   Coordinates: {lat:.10f}, {lng:.10f}")
                print(f"   Accuracy: {accuracy} (location_type: {location_type})")
                print(f"   Place types: {place_types[:3]}")  # Show first 3 types
                
                # Cache the result
                cache_coordinates(normalized_address, lat, lng, accuracy)
                
                # Log successful enhanced geocoding
                try:
                    logger_handler.log_user_activity('enhanced_geocoding_success', f'Google Maps enhanced: {address[:50]}... -> {lat}, {lng} ({accuracy})')
                except Exception as log_error:
                    print(f"⚠️ Logging error (non-critical): {log_error}")

                return lat, lng, accuracy
            else:
                print(f"⚠️ Google Maps: No results found for enhanced geocoding: {address}")

        # Fallback: Use OpenStreetMap Nominatim with enhanced accuracy
        print(f"🌐 Falling back to OpenStreetMap Nominatim (Enhanced)")
        nominatim_url = "https://nominatim.openstreetmap.org/search"
        params = {
            'q': address,
            'format': 'json',
            'limit': 1,
            'addressdetails': 1,
            'extratags': 1
        }

        headers = {
            'User-Agent': 'QR-Attendance-System/1.0 (Enhanced Location Accuracy)'
        }

        response = requests.get(nominatim_url, params=params, headers=headers, timeout=10)

        if response.status_code == 200:
            results = response.json()

            if results:
                result = results[0]
                lat = float(result['lat'])
                lng = float(result['lon'])

                # Enhanced accuracy assessment for OSM fallback
                place_type = result.get('type', 'unknown')
                osm_type = result.get('osm_type', 'unknown')
                importance = float(result.get('importance', 0))

                if place_type in ['house', 'building', 'shop', 'office'] or osm_type == 'way':
                    accuracy = 'good'  # Building-level (slightly lower than Google's excellent)
                elif place_type in ['neighbourhood', 'suburb', 'quarter', 'residential']:
                    accuracy = 'fair'
                elif place_type in ['city', 'town', 'village'] and importance > 0.5:
                    accuracy = 'poor'
                else:
                    accuracy = 'poor'

                print(f"✅ OSM enhanced geocoding successful:")
                print(f"   Coordinates: {lat:.10f}, {lng:.10f}")
                print(f"   Accuracy: {accuracy} (fallback)")
                
                # Cache the fallback result
                cache_coordinates(normalized_address, lat, lng, accuracy)
                
                # Log fallback enhanced geocoding
                try:
                    logger_handler.log_user_activity('enhanced_geocoding_fallback', f'OSM enhanced fallback: {address[:50]}... -> {lat}, {lng} ({accuracy})')
                except Exception as log_error:
                    print(f"⚠️ Logging error (non-critical): {log_error}")

                return lat, lng, accuracy

        print(f"⚠️ No results from enhanced geocoding for: {address}")
        
        # Log enhanced geocoding failure
        try:
            logger_handler.log_user_activity('enhanced_geocoding_failed', f'Enhanced geocoding failed: {address[:50]}...')
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")
        
        return None, None, None

    except Exception as e:
        print(f"❌ Enhanced geocoding error: {e}")
        
        # Log enhanced geocoding error
        try:
            logger_handler.log_flask_error('enhanced_geocoding_error', f'Enhanced geocoding error {address[:50]}...: {str(e)}')
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")
        
        return None, None, None

def geocode_address_enhanced(address):
    """
    Enhanced geocoding using Nominatim API with better accuracy classification
    Returns: (latitude, longitude, accuracy_level)
    """
    if not address or len(address.strip()) < 5:
        print("❌ Address too short for geocoding")
        return None, None, None

    try:
        # Nominatim API endpoint
        url = "https://nominatim.openstreetmap.org/search"

        params = {
            'q': address.strip(),
            'format': 'json',
            'limit': 1,
            'addressdetails': 1
        }

        headers = {
            'User-Agent': 'QR-Attendance-System/1.0'
        }

        response = requests.get(url, params=params, headers=headers, timeout=10)

        if response.status_code == 200:
            data = response.json()

            if data and len(data) > 0:
                result = data[0]
                lat = float(result['lat'])
                lng = float(result['lon'])

                # Determine accuracy based on result type
                place_type = result.get('type', 'unknown')
                osm_type = result.get('osm_type', 'unknown')

                if place_type in ['house', 'building'] or osm_type == 'way':
                    accuracy = 'high'
                elif place_type in ['neighbourhood', 'suburb', 'quarter']:
                    accuracy = 'medium'
                else:
                    accuracy = 'low'

                print(f"✅ Geocoded address: {address}")
                print(f"   Coordinates: {lat:.10f}, {lng:.10f}")
                print(f"   Accuracy: {accuracy} ({place_type})")

                return lat, lng, accuracy

        print(f"⚠️ No geocoding results for address: {address}")
        return None, None, None

    except Exception as e:
        logger_handler.log_flask_error('geocoding_error', str(e))
        print(f"❌ Geocoding error: {e}")
        return None, None, None

def calculate_distance_miles(lat1, lng1, lat2, lng2):
    """
    Calculate DIRECT straight-line distance between two points using Haversine formula
        
    Args:
        lat1, lng1: First point coordinates (decimal degrees)
        lat2, lng2: Second point coordinates (decimal degrees)
    
    Returns:
        Distance in miles (float) or None if calculation fails
    """
    if any(coord is None for coord in [lat1, lng1, lat2, lng2]):
        print("⚠️ Missing coordinates for distance calculation")
        return None

    try:
        # Convert to float and validate coordinate ranges
        try:
            lat1_val = float(lat1)
            lng1_val = float(lng1)
            lat2_val = float(lat2)
            lng2_val = float(lng2)
        except (ValueError, TypeError) as e:
            print(f"⚠️ Invalid coordinate format: {e}")
            return None
        
        # Validate latitude range [-90, 90]
        if not (-90 <= lat1_val <= 90) or not (-90 <= lat2_val <= 90):
            print(f"⚠️ Invalid latitude values: {lat1_val}, {lat2_val}")
            return None

        # Validate longitude range [-180, 180]
        if not (-180 <= lng1_val <= 180) or not (-180 <= lng2_val <= 180):
            print(f"⚠️ Invalid longitude values: {lng1_val}, {lng2_val}")
            return None

        # Log distance calculation attempt
        try:
            logger_handler.log_user_activity(
            'distance_calculation', 
            f'Calculating direct distance: ({lat1_val:.6f}, {lng1_val:.6f}) to ({lat2_val:.6f}, {lng2_val:.6f})'
        )
        except:
            pass  # Ignore logging errors

        # Use Haversine formula for straight-line distance
        print(f"📐 Calculating direct straight-line distance using Haversine formula")
        
        # Step 1: Convert decimal degrees to radians
        # CRITICAL: Only convert ONCE!
        lat1_rad = radians(lat1_val)
        lng1_rad = radians(lng1_val)
        lat2_rad = radians(lat2_val)
        lng2_rad = radians(lng2_val)
        
        # Step 2: Calculate differences in radians
        dlat = lat2_rad - lat1_rad
        dlng = lng2_rad - lng1_rad
        
        # Step 3: Haversine formula
        # a = sin²(Δφ/2) + cos(φ1) × cos(φ2) × sin²(Δλ/2)
        # where φ is latitude, λ is longitude
        
        sin_dlat_half = sin(dlat / 2.0)
        sin_dlng_half = sin(dlng / 2.0)
        
        a = (sin_dlat_half * sin_dlat_half + 
             cos(lat1_rad) * cos(lat2_rad) * sin_dlng_half * sin_dlng_half)
        
        # Clamp 'a' to valid range [0, 1] to avoid math domain errors
        a = max(0.0, min(1.0, a))
        
        # Step 4: Calculate central angle
        # c = 2 × asin(√a)
        c = 2.0 * asin(sqrt(a))
        
        # Step 5: Calculate distance
        # d = R × c
        # where R is Earth's mean radius in miles
        
        # CRITICAL: Earth's mean radius
        # DO NOT CHANGE THIS VALUE!
        EARTH_RADIUS_MILES = 3959.87433
        
        distance = c * EARTH_RADIUS_MILES
        
        # Round to 4 decimal places
        distance = round(distance, 4)
        
        # Debug output
        print(f"📏 Direct straight-line distance calculation:")
        print(f"   Point 1: ({lat1_val:.10f}, {lng1_val:.10f})")
        print(f"   Point 2: ({lat2_val:.10f}, {lng2_val:.10f})")
        print(f"   Δlat: {abs(lat2_val - lat1_val):.10f}° = {dlat:.12f} radians")
        print(f"   Δlng: {abs(lng2_val - lng1_val):.10f}° = {dlng:.12f} radians")
        print(f"   a value: {a:.15f}")
        print(f"   c value (central angle): {c:.15f} radians")
        print(f"   🎯 Distance: {distance:.4f} miles = {distance * 5280:.2f} feet = {distance * 1609.34:.2f} meters")
        
        # Log fallback calculation
        try:
            logger_handler.log_user_activity(
                'distance_calculation_success', 
                f'Direct distance: {distance:.4f} miles'
            )
        except:
            pass
        
        return distance

    except Exception as e:
        print(f"❌ Error in distance calculation: {e}")
        import traceback
        print(f"   Traceback: {traceback.format_exc()}")
        
        try:
            logger_handler.log_flask_error(
                'distance_calculation_error', 
                f'Distance calculation error: {str(e)}'
            )
        except:
            pass
        
        return None

def calculate_location_accuracy(qr_address, checkin_address, checkin_lat=None, checkin_lng=None):
    """
    Calculate location accuracy by comparing QR code address with check-in location
    Returns distance in miles between the two locations
    """
    print(f"\n📍 CALCULATING LOCATION ACCURACY:")
    print(f"   QR Address: {qr_address}")
    print(f"   Check-in Address: {checkin_address}")
    print(f"   Check-in Coordinates: {checkin_lat}, {checkin_lng}")

    # Get QR code coordinates from address
    qr_lat, qr_lng = get_coordinates_from_address(qr_address)

    if qr_lat is None or qr_lng is None:
        print(f"⚠️ Could not geocode QR address, cannot calculate accuracy")
        return None

    # Use check-in coordinates if available, otherwise geocode check-in address
    if checkin_lat is not None and checkin_lng is not None:
        checkin_coords_lat, checkin_coords_lng = checkin_lat, checkin_lng
        print(f"✅ Using GPS coordinates for check-in location")
    else:
        checkin_coords_lat, checkin_coords_lng = get_coordinates_from_address(checkin_address)
        if checkin_coords_lat is None or checkin_coords_lng is None:
            print(f"⚠️ Could not geocode check-in address, cannot calculate accuracy")
            return None
        print(f"✅ Using geocoded coordinates for check-in address")

    # Calculate distance
    distance = calculate_distance_miles(qr_lat, qr_lng, checkin_coords_lat, checkin_coords_lng)

    if distance is not None:
        print(f"✅ Location accuracy calculated: {distance} miles")

    return distance

def calculate_location_accuracy_enhanced(qr_address, checkin_address, checkin_lat=None, checkin_lng=None):
    """
    ENHANCED location accuracy calculation comparing QR address with check-in location
    This function provides improved precision and better error handling
    Returns:
    - Distance in miles between QR location and check-in location
    """
    print(f"\n🎯 ENHANCED LOCATION ACCURACY CALCULATION:")
    print(f"   QR Address: {qr_address}")
    print(f"   Check-in Address: {checkin_address}")
    print(f"   Check-in GPS: {checkin_lat}, {checkin_lng}")
    print(f"   Timestamp: {datetime.now()}")

    # Validate input parameters
    if not qr_address or qr_address.strip() == "":
        print(f"❌ QR address is empty or invalid")
        return None

    # Step 1: Get coordinates for QR address using enhanced geocoding
    print(f"\n📍 Step 1: Geocoding QR address...")
    try:
        # STEP 0: Check if addresses are essentially the same
        if addresses_are_similar(qr_address, checkin_address, threshold=0.90):
            print(f"🎯 Addresses are essentially identical - returning near-zero distance")
            # Return very small distance (within 50 feet / ~0.01 miles)
            return 0.01
        
        qr_lat, qr_lng, qr_accuracy = get_coordinates_from_address_enhanced(qr_address)
        print(f"   Geocoding result: lat={qr_lat}, lng={qr_lng}, accuracy={qr_accuracy}")

        if qr_lat is None or qr_lng is None:
            print(f"❌ Could not geocode QR address: {qr_address}")
            return None

        print(f"✅ QR location coordinates: {qr_lat:.10f}, {qr_lng:.10f} (accuracy: {qr_accuracy})")
    except Exception as e:
        print(f"❌ Error geocoding QR address: {e}")
        return None

    # Step 2: Determine check-in coordinates
    print(f"\n📱 Step 2: Determining check-in coordinates...")

    checkin_coords_lat = None
    checkin_coords_lng = None
    checkin_source = "unknown"

    # Priority 1: Use GPS coordinates if available and valid
    if checkin_lat is not None and checkin_lng is not None:
        try:
            lat_val = float(checkin_lat)
            lng_val = float(checkin_lng)

            # Validate GPS coordinates
            if -90 <= lat_val <= 90 and -180 <= lng_val <= 180:
                checkin_coords_lat = lat_val
                checkin_coords_lng = lng_val
                checkin_source = "gps"
                print(f"✅ Using GPS coordinates: {lat_val:.10f}, {lng_val:.10f}")
            else:
                print(f"⚠️ Invalid GPS coordinates: {lat_val}, {lng_val}")
        except (ValueError, TypeError) as e:
            print(f"⚠️ Could not parse GPS coordinates: {e}")

    # Priority 2: Fallback to geocoding check-in address
    if checkin_coords_lat is None and checkin_address:
        print(f"🌍 Falling back to geocoding check-in address...")
        try:
            checkin_coords_lat, checkin_coords_lng, checkin_accuracy = get_coordinates_from_address_enhanced(checkin_address)
            print(f"   Checkin geocoding result: lat={checkin_coords_lat}, lng={checkin_coords_lng}, accuracy={checkin_accuracy}")

            if checkin_coords_lat is not None:
                checkin_source = "address"
                print(f"✅ Using geocoded coordinates: {checkin_coords_lat:.10f}, {checkin_coords_lng:.10f} (accuracy: {checkin_accuracy})")
        except Exception as e:
            print(f"❌ Error geocoding check-in address: {e}")

    # Check if we have valid coordinates for both locations
    if checkin_coords_lat is None or checkin_coords_lng is None:
        print(f"❌ Could not determine check-in coordinates")
        print(f"   GPS: {checkin_lat}, {checkin_lng}")
        print(f"   Address: {checkin_address}")
        return None

    # Step 3: Calculate distance
    print(f"\n📏 Step 3: Calculating distance...")
    try:
        print(f"   QR coordinates: {qr_lat:.10f}, {qr_lng:.10f}")
        print(f"   Check-in coordinates: {checkin_coords_lat:.10f}, {checkin_coords_lng:.10f}")
        print(f"   Source: {checkin_source}")

        distance = calculate_distance_miles(qr_lat, qr_lng, checkin_coords_lat, checkin_coords_lng)
        print(f"   Distance calculation result: {distance}")

        if distance is not None:
            accuracy_level = get_location_accuracy_level_enhanced(distance)
            print(f"✅ Enhanced location accuracy calculated successfully!")
            print(f"   Distance: {distance:.4f} miles")
            print(f"   Accuracy Level: {accuracy_level}")
            return distance
        else:
            print(f"❌ Distance calculation returned None")
            return None

    except Exception as e:
        print(f"❌ Error calculating distance: {e}")
        print(f"❌ Distance calculation traceback: {traceback.format_exc()}")
        return None

def generate_qr_url(name, qr_id):
    """Generate a unique URL for QR code destination"""
    # Clean the name for URL use
    clean_name = re.sub(r'[^a-zA-Z0-9\s-]', '', name)
    clean_name = re.sub(r'\s+', '-', clean_name.strip())
    clean_name = clean_name.lower()

    # Create unique URL
    url_slug = f"qr-{qr_id}-{clean_name}"
    return url_slug[:200]  # Limit length

def detect_device_info(user_agent_string):
    """Extract device information from user agent"""
    try:
        user_agent = parse(user_agent_string)
        device_info = f"{user_agent.device.family}"

        if user_agent.os.family:
            device_info += f" - {user_agent.os.family}"
            if user_agent.os.version_string:
                device_info += f" {user_agent.os.version_string}"

        if user_agent.browser.family:
            device_info += f" ({user_agent.browser.family})"

        return device_info[:200]  # Limit length
    except:
        return "Unknown Device"

def get_client_ip():
    """Get client IP address"""
    if request.environ.get('HTTP_X_FORWARDED_FOR') is None:
        return request.environ['REMOTE_ADDR']
    else:
        return request.environ['HTTP_X_FORWARDED_FOR']

def get_location_accuracy_level_enhanced(location_accuracy):
    """
    Enhanced function to categorize location accuracy with more granular levels
    """
    if not location_accuracy or location_accuracy is None:
        return 'unknown'

    # More precise accuracy thresholds
    if location_accuracy <= 0.05:  # Within 264 feet (50 meters)
        return 'excellent'
    elif location_accuracy <= 0.1:  # Within 528 feet (100 meters)
        return 'very_good'
    elif location_accuracy <= 0.25:  # Within 0.25 mile (1320 feet)
        return 'good'
    elif location_accuracy <= 0.5:   # Within 0.5 mile
        return 'fair'
    elif location_accuracy <= 1.0:   # Within 1 mile
        return 'poor'
    else:                            # Greater than 1 mile
        return 'very_poor'

def process_location_data(location_data):
    """
    Process and validate location data from form
    Returns clean location data or None values for invalid data
    """
    processed = {
        'latitude': None,
        'longitude': None,
        'accuracy': None,
        'altitude': None,
        'source': location_data.get('location_source', 'manual'),
        'address': location_data.get('address', '')[:500] if location_data.get('address') else None
    }

    try:
        # Process latitude
        if location_data.get('latitude') and location_data['latitude'] not in ['null', '']:
            lat = float(location_data['latitude'])
            if -90 <= lat <= 90:  # Valid latitude range
                processed['latitude'] = lat
            else:
                print(f"⚠️ Invalid latitude: {lat}")

        # Process longitude
        if location_data.get('longitude') and location_data['longitude'] not in ['null', '']:
            lng = float(location_data['longitude'])
            if -180 <= lng <= 180:  # Valid longitude range
                processed['longitude'] = lng
            else:
                print(f"⚠️ Invalid longitude: {lng}")

        # Process accuracy
        if location_data.get('accuracy') and location_data['accuracy'] not in ['null', '']:
            acc = float(location_data['accuracy'])
            if acc >= 0:  # Accuracy should be positive
                processed['accuracy'] = acc
            else:
                print(f"⚠️ Invalid accuracy: {acc}")

        # Process altitude
        if location_data.get('altitude') and location_data['altitude'] not in ['null', '']:
            alt = float(location_data['altitude'])
            # Altitude can be negative (below sea level)
            processed['altitude'] = alt

    except (ValueError, TypeError) as e:
        print(f"⚠️ Error processing location data: {e}")

    return processed

def reverse_geocode_coordinates(latitude, longitude):
    """
    Convert GPS coordinates to human-readable address using Google Maps Reverse Geocoding
    Falls back to OpenStreetMap if Google Maps is unavailable
    Returns address string or None if failed
    """
    if not latitude or not longitude:
        return None

    try:
        print(f"🌍 Reverse geocoding coordinates: {latitude}, {longitude}")
        
        # Log reverse geocoding action
        try:
            logger_handler.log_user_activity('reverse_geocoding', f'Reverse geocoding: {latitude}, {longitude}')
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")

        # Primary: Use Google Maps Reverse Geocoding API
        if gmaps_client:
            print(f"🗺️ Using Google Maps Reverse Geocoding API")
            
            reverse_geocode_result = gmaps_client.reverse_geocode((latitude, longitude))
            
            if reverse_geocode_result:
                # Get the most detailed address (usually the first result)
                address = reverse_geocode_result[0]['formatted_address']
                print(f"✅ Google Maps reverse geocoded address: {address}")
                
                # Log successful reverse geocoding
                try:
                    logger_handler.log_user_activity('reverse_geocoding_success', f'Google Maps reverse geocoded: {latitude}, {longitude} -> {address[:50]}...')
                except Exception as log_error:
                    print(f"⚠️ Logging error (non-critical): {log_error}")
                
                return address
            else:
                print(f"⚠️ Google Maps: No address found for coordinates")

        # Fallback: Use OpenStreetMap Nominatim reverse geocoding
        print(f"🌐 Falling back to OpenStreetMap Nominatim reverse geocoding")
        url = "https://nominatim.openstreetmap.org/reverse"
        params = {
            'lat': latitude,
            'lon': longitude,
            'format': 'json',
            'addressdetails': 1,
            'zoom': 18  # High detail level
        }

        headers = {
            'User-Agent': 'QR-Attendance-System/1.0'
        }

        response = requests.get(url, params=params, headers=headers, timeout=10)

        if response.status_code == 200:
            data = response.json()

            if data and 'display_name' in data:
                address = data['display_name']
                print(f"✅ OSM reverse geocoded address: {address}")
                
                # Log fallback reverse geocoding
                try:
                    logger_handler.log_user_activity('reverse_geocoding_fallback', f'OSM reverse geocoded: {latitude}, {longitude} -> {address[:50]}...')
                except Exception as log_error:
                    print(f"⚠️ Logging error (non-critical): {log_error}")
                
                return address
            else:
                print(f"⚠️ No address found for coordinates")
                return None
        else:
            print(f"⚠️ Reverse geocoding API returned status: {response.status_code}")
            return None

    except Exception as e:
        print(f"❌ Error in reverse geocoding: {e}")
        
        # Log reverse geocoding error
        try:
            logger_handler.log_flask_error('reverse_geocoding_error', f'Reverse geocoding error {latitude}, {longitude}: {str(e)}')
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")
        
        return None

def process_location_data_enhanced(form_data):
    """
    Enhanced processing of location data from form submission
    Validates and cleans location data for storage, including reverse geocoding
    """
    processed = {
        'latitude': None,
        'longitude': None,
        'accuracy': None,
        'altitude': None,
        'source': form_data.get('location_source', 'manual'),
        'address': None
    }

    try:
        # Process latitude
        if form_data.get('latitude') and form_data['latitude'] not in ['null', '', 'undefined']:
            lat = float(form_data['latitude'])
            if -90 <= lat <= 90:  # Valid latitude range
                processed['latitude'] = lat
            else:
                print(f"⚠️ Invalid latitude: {lat}")

        # Process longitude
        if form_data.get('longitude') and form_data['longitude'] not in ['null', '', 'undefined']:
            lng = float(form_data['longitude'])
            if -180 <= lng <= 180:  # Valid longitude range
                processed['longitude'] = lng
            else:
                print(f"⚠️ Invalid longitude: {lng}")

        # Process GPS accuracy
        if form_data.get('accuracy') and form_data['accuracy'] not in ['null', '', 'undefined']:
            acc = float(form_data['accuracy'])
            if acc >= 0:  # Accuracy should be positive
                processed['accuracy'] = acc
            else:
                print(f"⚠️ Invalid GPS accuracy: {acc}")

        # Process altitude
        if form_data.get('altitude') and form_data['altitude'] not in ['null', '', 'undefined']:
            alt = float(form_data['altitude'])
            processed['altitude'] = alt

        # Process address - First check if address was provided
        if form_data.get('address'):
            address = form_data['address'].strip()
            if address and address not in ['null', '', 'undefined']:
                # Check if the address is just coordinates (like "38.8104192000, -77.1850240000")
                if re.match(r'^-?\d+\.\d+,?\s*-?\d+\.\d+$', address.replace(' ', '')):
                    print(f"🔍 Detected coordinate-format address: {address}")
                    # This is just coordinates, we need to reverse geocode
                    processed['address'] = None  # Reset so reverse geocoding will trigger
                else:
                    # This is a real address
                    processed['address'] = address[:500]  # Limit to 500 characters
                    print(f"✅ Using provided address: {processed['address'][:100]}...")

        # CRITICAL: If we have coordinates but no real address, perform reverse geocoding
        if (processed['latitude'] is not None and processed['longitude'] is not None
            and not processed['address']):
            print(f"🌍 Performing reverse geocoding for coordinates: {processed['latitude']}, {processed['longitude']}")
            reverse_geocoded_address = reverse_geocode_coordinates(processed['latitude'], processed['longitude'])
            if reverse_geocoded_address:
                processed['address'] = reverse_geocoded_address[:500]
                print(f"✅ Reverse geocoded address: {processed['address']}")
            else:
                print(f"⚠️ Could not reverse geocode coordinates, keeping coordinates as fallback")
                processed['address'] = f"{processed['latitude']:.10f}, {processed['longitude']:.10f}"

        print(f"📍 Final processed location data:")
        print(f"   Coordinates: {processed['latitude']}, {processed['longitude']}")
        print(f"   GPS Accuracy: {processed['accuracy']}m")
        print(f"   Source: {processed['source']}")
        print(f"   Address: {processed['address'][:100] if processed['address'] else 'None'}...")

        return processed

    except Exception as e:
        print(f"❌ Error processing location data: {e}")
        return processed

def migrate_to_enhanced_location_accuracy():
    """
    Migration function to recalculate all existing records with enhanced accuracy
    """
    try:
        print("🔄 Starting enhanced location accuracy migration...")

        # Get all records that need recalculation
        records = db.session.execute(text("""
            SELECT ad.id, qc.location_address, ad.address, ad.latitude, ad.longitude, ad.location_accuracy
            FROM attendance_data ad
            LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id
            WHERE qc.location_address IS NOT NULL
        """)).fetchall()

        print(f"📊 Found {len(records)} records to process")

        updated_count = 0
        improved_count = 0

        for record in records:
            try:
                # Calculate enhanced location accuracy
                new_accuracy = calculate_location_accuracy_enhanced(
                    qr_address=record.location_address,
                    checkin_address=record.address,
                    checkin_lat=record.latitude,
                    checkin_lng=record.longitude
                )

                if new_accuracy is not None:
                    # Update the record
                    db.session.execute(text("""
                        UPDATE attendance_data 
                        SET location_accuracy = :accuracy
                        WHERE id = :record_id
                    """), {
                        'accuracy': new_accuracy,
                        'record_id': record.id
                    })

                    updated_count += 1

                    # Check if this is an improvement
                    if record.location_accuracy is None or abs(new_accuracy - (record.location_accuracy or 0)) > 0.001:
                        improved_count += 1
                        print(f"   ✅ Updated record {record.id}: {record.location_accuracy} → {new_accuracy:.4f} miles")

            except Exception as e:
                print(f"   ⚠️ Error processing record {record.id}: {e}")

        # Commit all changes
        db.session.commit()

        print(f"✅ Enhanced migration completed!")
        print(f"   📊 Records processed: {len(records)}")
        print(f"   ✅ Records updated: {updated_count}")
        print(f"   📈 Records improved: {improved_count}")

        return True

    except Exception as e:
        print(f"❌ Enhanced migration failed: {e}")
        db.session.rollback()
        return False

def check_location_accuracy_column_exists():
    """
    Check if location_accuracy column exists in attendance_data table (MySQL compatible)
    """
    try:
        # MySQL-compatible query for checking column existence
        result = db.session.execute(text("""
            SELECT COUNT(*) as count
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'attendance_data' 
            AND COLUMN_NAME = 'location_accuracy'
        """))

        count = result.fetchone().count
        return count > 0

    except Exception as e:
        print(f"Error checking location_accuracy column: {e}")
        return False

def get_employee_checkin_history(employee_id, qr_code_id, date_filter=None):
    """
    Get check-in history for an employee at a specific location
    """
    try:
        if date_filter is None:
            date_filter = date.today()

        checkins = AttendanceData.query.filter_by(
            employee_id=employee_id.upper(),
            qr_code_id=qr_code_id,
            check_in_date=date_filter
        ).order_by(AttendanceData.check_in_time.asc()).all()

        return checkins

    except Exception as e:
        print(f"❌ Error retrieving checkin history: {e}")
        return []

def format_checkin_intervals(checkins):
    """
    Format time intervals between check-ins for display
    """
    if len(checkins) < 2:
        return []

    intervals = []
    for i in range(1, len(checkins)):
        previous_time = datetime.combine(checkins[i-1].check_in_date, checkins[i-1].check_in_time)
        current_time = datetime.combine(checkins[i].check_in_date, checkins[i].check_in_time)

        interval = current_time - previous_time
        interval_minutes = int(interval.total_seconds() / 60)

        intervals.append({
            'from_time': checkins[i-1].check_in_time.strftime('%H:%M'),
            'to_time': checkins[i].check_in_time.strftime('%H:%M'),
            'interval_minutes': interval_minutes,
            'interval_text': format_time_interval(interval_minutes)
        })

    return intervals

def format_time_interval(minutes):
    """
    Format minutes into human-readable time interval
    """
    if minutes < 60:
        return f"{minutes} minutes"
    elif minutes < 1440:  # Less than 24 hours
        hours = minutes // 60
        remaining_minutes = minutes % 60
        if remaining_minutes == 0:
            return f"{hours} hour{'s' if hours != 1 else ''}"
        else:
            return f"{hours}h {remaining_minutes}m"
    else:
        days = minutes // 1440
        remaining_hours = (minutes % 1440) // 60
        if remaining_hours == 0:
            return f"{days} day{'s' if days != 1 else ''}"
        else:
            return f"{days}d {remaining_hours}h"

# Authentication decorator
def login_required(f):
    """Decorator to ensure user is logged in"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator to ensure user has admin privileges"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))

        user_role = session.get('role')
        if not has_admin_privileges(user_role):
            flash('Administrator privileges required for this action.', 'error')
            return redirect(url_for('dashboard'))

        return f(*args, **kwargs)
    return decorated_function

def staff_or_admin_required(f):
    """Decorator to ensure user has staff-level or admin privileges"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))

        user_role = session.get('role')
        if not (has_admin_privileges(user_role) or has_staff_level_access(user_role)):
            flash('Insufficient privileges to access this page.', 'error')
            return redirect(url_for('dashboard'))

        return f(*args, **kwargs)
    return decorated_function

# Add this helper function to check admin requirements more safely
def is_admin_user(user_id):
    """Helper function to safely check if user is admin"""
    try:
        user = User.Query.get(user_id)
        return user and user.active_status and user.role == 'admin'
    except:
        return False

# Utility function to generate QR code
def generate_qr_code(data, fill_color="black", back_color="white", box_size=10, border=4, error_correction='L'):
    # Error correction mapping
    error_correction_map = {
        'L': qrcode.constants.ERROR_CORRECT_L,
        'M': qrcode.constants.ERROR_CORRECT_M,
        'Q': qrcode.constants.ERROR_CORRECT_Q,
        'H': qrcode.constants.ERROR_CORRECT_H
    }

    try:
        qr = qrcode.QRCode(
            version=1,
            error_correction=error_correction_map.get(error_correction, qrcode.constants.ERROR_CORRECT_L),
            box_size=int(box_size),
            border=int(border),
        )
        qr.add_data(data)
        qr.make(fit=True)

        # Generate QR code image with custom colors
        img = qr.make_image(fill_color=fill_color, back_color=back_color)

        # Convert to base64
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        img_str = base64.b64encode(buffer.getvalue()).decode()

        # Log successful generation if logger is available
        try:
            logger_handler.log_qr_code_generated(
                data_length=len(data),
                fill_color=fill_color,
                back_color=back_color,
                box_size=box_size,
                border=border,
                error_correction=error_correction
            )
        except:
            pass  # Ignore logging errors

        return img_str

    except Exception as e:
        logger_handler.log_database_error('qr_code_generation', e)
        # Return default QR code on error
        return generate_default_qr_code(data)

def generate_default_qr_code(data):
    """Fallback function for basic QR code generation"""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    img_str = base64.b64encode(buffer.getvalue()).decode()

    return img_str

def get_qr_styling(qr_code):
    """Extract QR code styling parameters from database record"""
    return {
        'fill_color': getattr(qr_code, 'fill_color', '#000000') or '#000000',
        'back_color': getattr(qr_code, 'back_color', '#FFFFFF') or '#FFFFFF',
        'box_size': getattr(qr_code, 'box_size', 10) or 10,
        'border': getattr(qr_code, 'border', 4) or 4,
        'error_correction': getattr(qr_code, 'error_correction', 'L') or 'L'
    }

@app.template_filter('strftime')
def strftime_filter(value, format='%m/%d/%Y'):
    """Format datetime/date/string as strftime"""
    if isinstance(value, str):
        if value.lower() == 'now':
            return datetime.now().strftime(format)
        try:
            # Try to parse string as datetime
            dt = datetime.fromisoformat(value)
            return dt.strftime(format)
        except (ValueError, TypeError):
            return value

    if hasattr(value, 'strftime'):
        return value.strftime(format)

    return str(value)

# Routes
@app.route('/')
def index():
    """Home page - redirect to login if not authenticated"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
@log_user_activity('user_registration')
def register():
    """User registration endpoint"""
    if request.method == 'POST':
        try:
            full_name = request.form['full_name']
            email = request.form['email']
            username = request.form['username']
            password = request.form['password']

            # Check if user already exists
            if User.query.filter_by(username=username).first():
                flash('Username already exists.', 'error')
                return render_template('register.html')

            if User.query.filter_by(email=email).first():
                flash('Email already registered.', 'error')
                return render_template('register.html')

            # Create new user (default role: staff)
            new_user = User(
                full_name=full_name,
                email=email,
                username=username,
                role='staff'
            )
            new_user.set_password(password)

            db.session.add(new_user)
            db.session.commit()

            # Log successful user registration
            logger_handler.logger.info(f"New user registered: {username} ({email})")

            flash('Registration successful! Please log in.', 'success')
            return redirect(url_for('login'))

        except Exception as e:
            db.session.rollback()
            logger_handler.log_database_error('user_registration', e)
            flash('Registration failed. Please try again.', 'error')

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Enhanced user authentication with Turnstile and comprehensive logging"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        turnstile_response = request.form.get('cf-turnstile-response', '')

        if not username or not password:
            flash('Please enter both username and password.', 'error')
            return render_template('login.html')

        # Verify Turnstile if enabled
        if turnstile_utils.is_enabled():
            if not turnstile_utils.verify_turnstile(turnstile_response):
                # Log failed Turnstile attempt
                logger_handler.log_security_event(
                    event_type="turnstile_verification_failed",
                    description=f"Failed Turnstile verification for username: {username}",
                    severity="HIGH"
                )
                flash('Please complete the security verification.', 'error')
                return render_template('login.html')

        try:
            # Find user (case-insensitive username)
            user = User.query.filter(
                User.username.like(username),
                User.active_status == True
            ).first()

            if user and user.check_password(password):
                # Check if "Remember Me" is checked
                remember_me = request.form.get('remember_me') == 'on'
                
                # Set session as permanent if "Remember Me" is checked
                if remember_me:
                    session.permanent = True
                    session['remember_me'] = True
                else:
                    session.permanent = False
                    session['remember_me'] = False
                
                # Successful login
                session['user_id'] = user.id
                session['username'] = user.username
                session['role'] = user.role
                session['full_name'] = user.full_name
                session['login_time'] = datetime.now().isoformat()

                # Update last login date
                user.last_login_date = datetime.utcnow()
                db.session.commit()

                # Log successful login with Turnstile info
                logger_handler.log_user_login(
                    user_id=user.id,
                    username=user.username,
                    success=True
                )
                
                # Log successful Turnstile verification
                if turnstile_utils.is_enabled():
                    logger_handler.log_security_event(
                        event_type="turnstile_verification_success",
                        description=f"Successful Turnstile verification for user: {user.username}",
                        severity="INFO"
                    )

                flash(f'Welcome back, {user.full_name}!', 'success')
                print(f"User {user.username} logged in successfully")

                # Redirect to intended page or dashboard
                next_page = request.args.get('next')
                return redirect(next_page) if next_page else redirect(url_for('attendance_report'))

            else:
                # Invalid credentials - log failed attempt
                user_id = user.id if user else None
                logger_handler.log_user_login(
                    user_id=user_id,
                    username=username,
                    success=False,
                    failure_reason="Invalid credentials"
                )

                flash('Invalid username or password.', 'error')
                print(f"Failed login attempt for username: {username}")

        except Exception as e:
            logger_handler.log_database_error('user_login', e)
            print(f"Login error: {e}")
            flash('Login error. Please try again.', 'error')

    return render_template('login.html')

@app.route('/logout')
def logout():
    """User logout endpoint with session duration logging"""
    user_id = session.get('user_id')
    username = session.get('username')
    login_time_str = session.get('login_time')

    # Calculate session duration
    session_duration = None
    if login_time_str:
        try:
            login_time = datetime.fromisoformat(login_time_str)
            session_duration = (datetime.now() - login_time).total_seconds() / 60  # minutes
        except:
            pass

    # Log user logout
    if user_id and username:
        logger_handler.log_user_logout(
            user_id=user_id,
            username=username,
            session_duration=session_duration
        )

    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Enhanced project-centric dashboard with search filters"""
    try:
        user = User.query.get(session['user_id'])
        
        # Get search parameters from URL
        search_name = request.args.get('search_name', '').strip()
        search_status = request.args.get('search_status', '').strip()
        
        # Build QR codes query with filters
        qr_query = QRCode.query
        
        # Apply name filter if provided
        if search_name:
            qr_query = qr_query.filter(QRCode.name.ilike(f'%{search_name}%'))
        
        # Apply status filter if provided
        if search_status == 'active':
            qr_query = qr_query.filter(QRCode.active_status == True)
        elif search_status == 'inactive':
            qr_query = qr_query.filter(QRCode.active_status == False)
        
        # Execute query
        qr_codes = qr_query.order_by(QRCode.created_date.desc()).all()
        projects = Project.query.order_by(Project.name.asc()).all()
        
        # Log dashboard access with filter info
        filter_info = []
        if search_name:
            filter_info.append(f"name contains '{search_name}'")
        if search_status:
            filter_info.append(f"status is {search_status}")
        
        log_message = f"User {session['username']} accessed dashboard: {len(qr_codes)} QR codes"
        if filter_info:
            log_message += f" (filtered: {', '.join(filter_info)})"
        
        logger_handler.logger.info(log_message)
        
        return render_template('dashboard.html',
                             user=user,
                             qr_codes=qr_codes,
                             projects=projects,
                             search_name=search_name,
                             search_status=search_status)
    
    except Exception as e:
        logger_handler.log_database_error('dashboard_load', e)
        print(f"Error loading dashboard: {e}")
        flash('Error loading dashboard. Please try again.', 'error')
        return redirect(url_for('login'))

@app.route('/project/<int:project_id>/qr-codes')
@login_required
def project_qr_codes(project_id):
    """
    View all QR codes for a specific project with search filters
    Allows filtering by name and status within the project
    """
    try:
        # Get the project
        project = Project.query.get_or_404(project_id)
        
        # Get search parameters from URL
        search_name = request.args.get('search_name', '').strip()
        search_status = request.args.get('search_status', '').strip()
        
        # Build QR codes query with filters for this project only
        qr_query = QRCode.query.filter_by(project_id=project_id)
        
        # Apply name filter if provided
        if search_name:
            qr_query = qr_query.filter(QRCode.name.ilike(f'%{search_name}%'))
        
        # Apply status filter if provided
        if search_status == 'active':
            qr_query = qr_query.filter(QRCode.active_status == True)
        elif search_status == 'inactive':
            qr_query = qr_query.filter(QRCode.active_status == False)
        
        # Execute query
        qr_codes = qr_query.order_by(QRCode.created_date.desc()).all()
        
        # Log access with filter info
        filter_info = []
        if search_name:
            filter_info.append(f"name contains '{search_name}'")
        if search_status:
            filter_info.append(f"status is {search_status}")
        
        log_message = f"User {session['username']} viewed project '{project.name}' QR codes: {len(qr_codes)} QR codes"
        if filter_info:
            log_message += f" (filtered: {', '.join(filter_info)})"
        
        logger_handler.logger.info(log_message)
        
        return render_template('project_qr_codes.html',
                             project=project,
                             qr_codes=qr_codes,
                             search_name=search_name,
                             search_status=search_status)
    
    except Exception as e:
        logger_handler.log_database_error('project_qr_codes_view', e)
        print(f"Error loading project QR codes: {e}")
        flash('Error loading project QR codes. Please try again.', 'error')
        return redirect(url_for('dashboard'))
    
@app.route('/dashboard/search', methods=['GET'])
@login_required
def search_qr_codes():
    """Search QR codes - redirect to dashboard with filters"""
    search_name = request.args.get('search_name', '').strip()
    search_status = request.args.get('search_status', '').strip()
    
    # Log search activity
    logger_handler.logger.info(
        f"User {session['username']} searched QR codes: "
        f"name='{search_name}', status='{search_status}'"
    )
    
    # Redirect to dashboard with search parameters
    return redirect(url_for('dashboard', search_name=search_name, search_status=search_status))

@app.route('/api/dashboard/stats')
@login_required
def dashboard_stats_api():
    """API endpoint for dashboard statistics"""
    try:
        # Get current stats
        total_qr_codes = QRCode.query.filter_by(active_status=True).count()
        
        # Today's check-ins
        today = datetime.utcnow().date()
        today_checkins = AttendanceData.query.filter(
            AttendanceData.check_in_date == today
        ).count()
        
        # Active projects
        active_projects = Project.query.filter_by(active_status=True).count()
        
        # Unique locations
        unique_locations = db.session.query(
            AttendanceData.location_name
        ).distinct().count()
        
        # Calculate trends (compared to last month)
        last_month = datetime.utcnow() - timedelta(days=30)
        
        # QR codes trend
        old_qr_count = QRCode.query.filter(
            QRCode.created_date <= last_month,
            QRCode.active_status == True
        ).count()
        qr_change = ((total_qr_codes - old_qr_count) / max(old_qr_count, 1)) * 100
        
        # Check-ins trend (yesterday)
        yesterday = today - timedelta(days=1)
        yesterday_checkins = AttendanceData.query.filter(
            AttendanceData.check_in_date == yesterday
        ).count()
        checkin_change = ((today_checkins - yesterday_checkins) / max(yesterday_checkins, 1)) * 100
        
        return jsonify({
            'success': True,
            'total_qr_codes': total_qr_codes,
            'today_checkins': today_checkins,
            'active_projects': active_projects,
            'unique_locations': unique_locations,
            'qr_change': round(qr_change, 1),
            'checkin_change': round(checkin_change, 1),
            'project_change': 0,  # You can calculate this based on your needs
            'location_change': 0   # You can calculate this based on your needs
        })
        
    except Exception as e:
        logger_handler.log_database_error('dashboard_stats_api', e)
        return jsonify({
            'success': False,
            'error': 'Failed to fetch dashboard statistics'
        }), 500

@app.route('/api/dashboard/realtime')
@login_required
def dashboard_realtime_api():
    """API endpoint for real-time dashboard data"""
    try:
        # Get recent activity (last 10 check-ins)
        recent_activity = db.session.query(
            AttendanceData.employee_id,
            AttendanceData.location_name,
            AttendanceData.check_in_time,
            AttendanceData.check_in_date
        ).order_by(
            AttendanceData.check_in_date.desc(),
            AttendanceData.check_in_time.desc()
        ).limit(10).all()
        
        activity_data = [
            {
                'employee_id': activity.employee_id,
                'location': activity.location_name,
                'time': activity.check_in_time.strftime('%H:%M'),
                'date': activity.check_in_date.strftime('%Y-%m-%d')
            }
            for activity in recent_activity
        ]
        
        return jsonify({
            'success': True,
            'recent_activity': activity_data
        })
        
    except Exception as e:
        logger_handler.log_database_error('dashboard_realtime_api', e)
        return jsonify({
            'success': False,
            'error': 'Failed to fetch real-time data'
        }), 500
    
# USER MANAGEMENT ROUTES
@app.route('/profile', methods=['GET', 'POST'])
@login_required
@log_user_activity('profile_update')
def profile():
    """User profile management with logging"""
    try:
        user = User.query.get(session['user_id'])

        if request.method == 'POST':
            form_type = request.form.get('form_type')

            if form_type == 'profile':
                # Track changes for logging
                old_name = user.full_name
                old_email = user.email

                # Update profile information
                user.full_name = request.form['full_name']
                user.email = request.form['email']

                # Check for changes
                changes = {}
                if old_name != user.full_name:
                    changes['full_name'] = {'old': old_name, 'new': user.full_name}
                if old_email != user.email:
                    changes['email'] = {'old': old_email, 'new': user.email}

                db.session.commit()

                # Log profile update if there were changes
                if changes:
                    logger_handler.logger.info(f"User profile updated: {user.username} - Changes: {json.dumps(changes)}")

                flash('Profile updated successfully!', 'success')

            elif form_type == 'password':
                # Update password
                current_password = request.form['current_password']
                new_password = request.form['new_password']

                if user.check_password(current_password):
                    user.set_password(new_password)
                    db.session.commit()

                    # Log password change
                    logger_handler.log_security_event(
                        event_type="password_change",
                        description=f"User {user.username} changed password",
                        severity="MEDIUM"
                    )

                    flash('Password updated successfully!', 'success')
                else:
                    # Log failed password change attempt
                    logger_handler.log_security_event(
                        event_type="password_change_failed",
                        description=f"Failed password change attempt for user {user.username}",
                        severity="HIGH"
                    )
                    flash('Current password is incorrect.', 'error')

        return render_template('profile.html', user=user)

    except Exception as e:
        logger_handler.log_database_error('profile_update', e)
        flash('Profile update failed. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/users')
@admin_required
def users():
    """Display all users (Admin only)"""
    try:
        users = User.query.order_by(User.created_date.desc()).all()
        return render_template('users.html', users=users)
    except Exception as e:
        logger_handler.log_database_error('users_list', e)
        flash('Error loading users list.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/users/create', methods=['GET', 'POST'])
@admin_required
@log_database_operations('user_creation')
def create_user():
    """Create new user (Admin only) with Project Manager permissions support"""
    if request.method == 'POST':
        try:
            # Get basic form data
            full_name = request.form.get('full_name', '').strip()
            email = request.form.get('email', '').strip()
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            role = request.form.get('role', '')

            # Validate required fields
            if not all([full_name, email, username, password, role]):
                flash('All fields are required.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                locations = get_all_locations_from_qr_codes()
                return render_template('create_user.html', projects=projects, locations=locations)

            # Validate role
            if role not in VALID_ROLES:
                flash(f'Invalid role selected. Valid roles: {", ".join(VALID_ROLES)}', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                locations = get_all_locations_from_qr_codes()
                return render_template('create_user.html', projects=projects, locations=locations)

            # Check if user already exists
            if User.query.filter_by(username=username).first():
                flash('Username already exists.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                locations = get_all_locations_from_qr_codes()
                return render_template('create_user.html', projects=projects, locations=locations)

            if User.query.filter_by(email=email).first():
                flash('Email already registered.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                locations = get_all_locations_from_qr_codes()
                return render_template('create_user.html', projects=projects, locations=locations)

            # Create new user
            new_user = User(
                full_name=full_name,
                email=email,
                username=username,
                role=role,
                created_by=session['user_id']
            )
            new_user.set_password(password)

            db.session.add(new_user)
            db.session.flush()  # Get the user ID without committing

            # Handle Project Manager permissions
            if role == 'project_manager':
                # Get selected projects - getlist returns empty list if field doesn't exist
                selected_projects = request.form.getlist('assigned_projects')
                
                # Validate and filter project IDs
                valid_project_ids = []
                if selected_projects:
                    for pid in selected_projects:
                        try:
                            project_id = int(pid)
                            # Verify project exists
                            if Project.query.get(project_id):
                                valid_project_ids.append(project_id)
                        except (ValueError, TypeError):
                            logger_handler.logger.warning(f"Invalid project ID received: {pid}")
                
                # Add project permissions
                if valid_project_ids:
                    for project_id in valid_project_ids:
                        try:
                            permission = UserProjectPermission(
                                user_id=new_user.id,
                                project_id=project_id
                            )
                            db.session.add(permission)
                        except Exception as e:
                            logger_handler.logger.error(f"Error adding project permission: {e}")
                    
                    logger_handler.logger.info(
                        f"Admin {session['username']} assigned {len(valid_project_ids)} projects to new Project Manager {username}"
                    )

                # Get selected locations
                selected_locations = request.form.getlist('assigned_locations')
                
                # Filter and clean location names
                valid_locations = []
                if selected_locations:
                    for location in selected_locations:
                        location_clean = location.strip()
                        if location_clean:
                            valid_locations.append(location_clean)
                
                # Add location permissions
                if valid_locations:
                    for location_name in valid_locations:
                        try:
                            permission = UserLocationPermission(
                                user_id=new_user.id,
                                location_name=location_name
                            )
                            db.session.add(permission)
                        except Exception as e:
                            logger_handler.logger.error(f"Error adding location permission: {e}")
                    
                    logger_handler.logger.info(
                        f"Admin {session['username']} assigned {len(valid_locations)} locations to new Project Manager {username}"
                    )

            # Commit all changes
            db.session.commit()

            # Log user creation
            logger_handler.logger.info(f"Admin user {session['username']} created new user: {username} with role {role}")

            flash(f'User "{full_name}" created successfully with role "{role}".', 'success')
            return redirect(url_for('users'))

        except KeyError as e:
            db.session.rollback()
            logger_handler.logger.error(f"Missing form field: {e}")
            flash(f'Missing required field: {e}. Please fill in all fields.', 'error')
            projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
            locations = get_all_locations_from_qr_codes()
            return render_template('create_user.html', projects=projects, locations=locations)
        except Exception as e:
            db.session.rollback()
            logger_handler.log_database_error('user_creation', e)
            logger_handler.logger.error(f"User creation error details: {str(e)}")
            flash('User creation failed. Please try again.', 'error')
            projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
            locations = get_all_locations_from_qr_codes()
            return render_template('create_user.html', projects=projects, locations=locations)

    # GET request - load form with projects and locations
    try:
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        locations = get_all_locations_from_qr_codes()
        return render_template('create_user.html', projects=projects, locations=locations)
    except Exception as e:
        logger_handler.logger.error(f"Error loading create user form: {e}")
        flash('Error loading form. Please try again.', 'error')
        return redirect(url_for('users'))

def get_all_locations_from_qr_codes():
    """Helper function to get all unique locations from QR codes"""
    try:
        result = db.session.execute(text("""
            SELECT DISTINCT location 
            FROM qr_codes 
            WHERE location IS NOT NULL 
            AND active_status = 1
            ORDER BY location
        """))
        return [row[0] for row in result.fetchall()]
    except Exception as e:
        logger_handler.logger.error(f"Error loading locations: {e}")
        return []
    
@app.route('/users/<int:user_id>/delete', methods=['GET', 'POST'])
@admin_required
def delete_user(user_id):
    """Deactivate user (Admin only) - Fixed with proper validation"""
    try:
        user_to_delete = User.query.get(user_id)
        current_user = User.query.get(session['user_id'])

        if not user_to_delete:
            flash('User not found.', 'error')
            return redirect(url_for('users'))

        # Prevent self-deletion
        if user_to_delete.id == current_user.id:
            flash('You cannot deactivate your own account. Ask another admin to do this.', 'error')
            return redirect(url_for('users'))

        # Check if trying to delete the last admin
        if user_to_delete.role == 'admin':
            active_admin_count = User.query.filter_by(role='admin', active_status=True).count()
            if active_admin_count <= 1:
                flash('Cannot deactivate the last admin user. Promote another user to admin first.', 'error')
                return redirect(url_for('users'))

        # Deactivate the user instead of deleting
        user_to_delete.active_status = False
        db.session.commit()

        flash(f'User "{user_to_delete.full_name}" has been deactivated successfully.', 'success')
        print(f"Admin {current_user.username} deactivated user: {user_to_delete.username}")

        return redirect(url_for('users'))

    except Exception as e:
        db.session.rollback()
        print(f"Error deactivating user: {e}")
        flash('Error deactivating user. Please try again.', 'error')
        return redirect(url_for('users'))

@app.route('/users/<int:user_id>/reactivate', methods=['GET', 'POST'])
@admin_required
def reactivate_user(user_id):
    """Reactivate a deactivated user (Admin only)"""
    try:
        user_to_reactivate = User.query.get(user_id)
        current_user = User.query.get(session['user_id'])

        if not user_to_reactivate:
            flash('User not found.', 'error')
            return redirect(url_for('users'))

        if user_to_reactivate.active_status:
            flash('User is already active.', 'info')
        else:
            user_to_reactivate.active_status = True
            db.session.commit()
            flash(f'User "{user_to_reactivate.full_name}" has been reactivated successfully.', 'success')
            print(f"Admin {current_user.username} reactivated user: {user_to_reactivate.username}")

        return redirect(url_for('users'))

    except Exception as e:
        db.session.rollback()
        print(f"Error reactivating user: {e}")
        flash('Error reactivating user. Please try again.', 'error')
        return redirect(url_for('users'))

@app.route('/users/<int:user_id>/promote', methods=['GET', 'POST'])
@admin_required
def promote_user(user_id):
    """Promote a staff user to admin (Admin only)"""
    try:
        user_to_promote = User.query.get(user_id)
        current_user = User.query.get(session['user_id'])

        if not user_to_promote:
            flash('User not found.', 'error')
            return redirect(url_for('users'))

        if user_to_promote.role == 'admin':
            flash('User is already an admin.', 'info')
        else:
            user_to_promote.role = 'admin'
            db.session.commit()
            flash(f'"{user_to_promote.full_name}" has been promoted to admin.', 'success')
            print(f"Admin {current_user.username} promoted user {user_to_promote.username} to admin")

        return redirect(url_for('users'))

    except Exception as e:
        db.session.rollback()
        print(f"Error promoting user: {e}")
        flash('Error promoting user. Please try again.', 'error')
        return redirect(url_for('users'))

@app.route('/users/<int:user_id>/demote', methods=['GET', 'POST'])
@admin_required
def demote_user(user_id):
    """Demote an admin user to staff (Admin only)"""
    try:
        user_to_demote = User.query.get(user_id)
        current_user = User.query.get(session['user_id'])

        if not user_to_demote:
            flash('User not found.', 'error')
            return redirect(url_for('users'))

        # Prevent self-demotion
        if user_to_demote.id == current_user.id:
            flash('You cannot demote yourself. Have another admin do this.', 'error')
            return redirect(url_for('users'))

        # Check if this is the last admin
        active_admin_count = User.query.filter_by(role='admin', active_status=True).count()
        if active_admin_count <= 1 and user_to_demote.role == 'admin':
            flash('Cannot demote the last admin user. Promote another user to admin first.', 'error')
            return redirect(url_for('users'))

        if has_staff_level_access(user_to_demote.role):
            flash('User already has staff-level permissions.', 'info')
        else:
            user_to_demote.role = 'staff'
            db.session.commit()
            flash(f'"{user_to_demote.full_name}" has been demoted to staff.', 'success')
            print(f"Admin {current_user.username} demoted user {user_to_demote.username} to staff")

        return redirect(url_for('users'))

    except Exception as e:
        db.session.rollback()
        print(f"Error demoting user: {e}")
        flash('Error demoting user. Please try again.', 'error')
        return redirect(url_for('users'))

@app.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
@log_database_operations('user_edit')
def edit_user(user_id):
    """Edit existing user with Project Manager permissions support"""
    try:
        user_to_edit = User.query.get_or_404(user_id)
        
        # Track old role for permission cleanup
        old_role = user_to_edit.role

        if request.method == 'POST':
            # Store old values for change tracking
            old_values = {
                'full_name': user_to_edit.full_name,
                'email': user_to_edit.email,
                'username': user_to_edit.username,
                'role': user_to_edit.role,
                'active_status': user_to_edit.active_status
            }
            changes = {}

            # Update basic info with validation
            full_name = request.form.get('full_name', '').strip()
            email = request.form.get('email', '').strip()
            username = request.form.get('username', '').strip()
            
            if not all([full_name, email, username]):
                flash('Name, email, and username are required.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                locations = get_all_locations_from_qr_codes()
                assigned_project_ids = []
                assigned_location_names = []
                if user_to_edit.role == 'project_manager':
                    assigned_project_ids = [p.project_id for p in UserProjectPermission.query.filter_by(user_id=user_id).all()]
                    assigned_location_names = [l.location_name for l in UserLocationPermission.query.filter_by(user_id=user_id).all()]
                return render_template('edit_user.html', user=user_to_edit, valid_roles=VALID_ROLES,
                                     projects=projects, locations=locations,
                                     assigned_project_ids=assigned_project_ids,
                                     assigned_location_names=assigned_location_names)

            user_to_edit.full_name = full_name
            user_to_edit.email = email
            user_to_edit.username = username
            
            # Update role with validation
            new_role = request.form.get('role', '')
            if new_role not in VALID_ROLES:
                flash(f'Invalid role selected. Valid roles: {", ".join(VALID_ROLES)}', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                locations = get_all_locations_from_qr_codes()
                assigned_project_ids = []
                assigned_location_names = []
                if user_to_edit.role == 'project_manager':
                    assigned_project_ids = [p.project_id for p in UserProjectPermission.query.filter_by(user_id=user_id).all()]
                    assigned_location_names = [l.location_name for l in UserLocationPermission.query.filter_by(user_id=user_id).all()]
                return render_template('edit_user.html', user=user_to_edit, valid_roles=VALID_ROLES,
                                     projects=projects, locations=locations,
                                     assigned_project_ids=assigned_project_ids,
                                     assigned_location_names=assigned_location_names)

            user_to_edit.role = new_role

            # Handle password update if provided
            new_password = request.form.get('new_password', '')
            if new_password and new_password.strip():
                user_to_edit.set_password(new_password)
                changes['password'] = 'Password updated'
                # Log password change
                logger_handler.log_security_event(
                    event_type="admin_password_change",
                    description=f"Admin {session['username']} changed password for user {user_to_edit.username}",
                    severity="MEDIUM"
                )

            # Handle Project Manager permissions
            if new_role == 'project_manager':
                # Update project permissions
                # First, remove existing project permissions
                try:
                    UserProjectPermission.query.filter_by(user_id=user_id).delete()
                except Exception as e:
                    logger_handler.logger.error(f"Error deleting old project permissions: {e}")
                
                # Add new project permissions
                selected_projects = request.form.getlist('assigned_projects')
                
                # Validate project IDs
                valid_project_ids = []
                if selected_projects:
                    for pid in selected_projects:
                        try:
                            project_id = int(pid)
                            # Verify project exists
                            if Project.query.get(project_id):
                                valid_project_ids.append(project_id)
                        except (ValueError, TypeError):
                            logger_handler.logger.warning(f"Invalid project ID received: {pid}")
                
                # Add validated project permissions
                if valid_project_ids:
                    for project_id in valid_project_ids:
                        try:
                            permission = UserProjectPermission(
                                user_id=user_id,
                                project_id=project_id
                            )
                            db.session.add(permission)
                        except Exception as e:
                            logger_handler.logger.error(f"Error adding project permission: {e}")
                    
                    changes['assigned_projects'] = f'{len(valid_project_ids)} projects assigned'
                    logger_handler.logger.info(
                        f"Admin {session['username']} updated project permissions for Project Manager {user_to_edit.username}: {len(valid_project_ids)} projects"
                    )

                # Update location permissions
                # First, remove existing location permissions
                try:
                    UserLocationPermission.query.filter_by(user_id=user_id).delete()
                except Exception as e:
                    logger_handler.logger.error(f"Error deleting old location permissions: {e}")
                
                # Add new location permissions
                selected_locations = request.form.getlist('assigned_locations')
                
                # Validate and clean locations
                valid_locations = []
                if selected_locations:
                    for location in selected_locations:
                        location_clean = location.strip()
                        if location_clean:
                            valid_locations.append(location_clean)
                
                # Add validated location permissions
                if valid_locations:
                    for location_name in valid_locations:
                        try:
                            permission = UserLocationPermission(
                                user_id=user_id,
                                location_name=location_name
                            )
                            db.session.add(permission)
                        except Exception as e:
                            logger_handler.logger.error(f"Error adding location permission: {e}")
                    
                    changes['assigned_locations'] = f'{len(valid_locations)} locations assigned'
                    logger_handler.logger.info(
                        f"Admin {session['username']} updated location permissions for Project Manager {user_to_edit.username}: {len(valid_locations)} locations"
                    )
            
            # If role changed from project_manager to something else, remove permissions
            elif old_role == 'project_manager' and new_role != 'project_manager':
                try:
                    UserProjectPermission.query.filter_by(user_id=user_id).delete()
                    UserLocationPermission.query.filter_by(user_id=user_id).delete()
                    logger_handler.logger.info(
                        f"Admin {session['username']} removed Project Manager permissions from user {user_to_edit.username} (role changed to {new_role})"
                    )
                except Exception as e:
                    logger_handler.logger.error(f"Error removing permissions: {e}")

            # Track changes
            for field, old_value in old_values.items():
                new_value = getattr(user_to_edit, field)
                if old_value != new_value:
                    changes[field] = {'old': old_value, 'new': new_value}

            # Commit all changes
            db.session.commit()

            # Log user update
            if changes:
                logger_handler.logger.info(f"Admin user {session['username']} updated user {user_to_edit.username}: {json.dumps(changes, default=str)}")

            flash(f'User "{user_to_edit.full_name}" updated successfully.', 'success')
            return redirect(url_for('users'))

        # GET request - load form with current assignments
        try:
            projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
            locations = get_all_locations_from_qr_codes()
            
            # Get current assignments if user is a project manager
            assigned_project_ids = []
            assigned_location_names = []
            
            if user_to_edit.role == 'project_manager':
                try:
                    assigned_project_ids = [p.project_id for p in UserProjectPermission.query.filter_by(user_id=user_id).all()]
                    assigned_location_names = [l.location_name for l in UserLocationPermission.query.filter_by(user_id=user_id).all()]
                except Exception as e:
                    logger_handler.logger.error(f"Error loading current permissions: {e}")

            return render_template('edit_user.html', 
                                 user=user_to_edit, 
                                 valid_roles=VALID_ROLES,
                                 projects=projects,
                                 locations=locations,
                                 assigned_project_ids=assigned_project_ids,
                                 assigned_location_names=assigned_location_names)
        except Exception as e:
            logger_handler.logger.error(f"Error loading edit user form: {e}")
            flash('Error loading edit form. Please try again.', 'error')
            return redirect(url_for('users'))

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('user_update', e)
        logger_handler.logger.error(f"User update error details: {str(e)}")
        flash('Error updating user. Please try again.', 'error')
        return redirect(url_for('users'))

@app.route('/users/<int:user_id>/toggle-status', methods=['POST'])
@admin_required
def toggle_user_status(user_id):
    """Toggle user active status via AJAX (Admin only)"""
    try:
        user_to_toggle = User.query.get(user_id)
        current_user = User.query.get(session['user_id'])

        if not user_to_toggle:
            return jsonify({
                'success': False,
                'message': 'User not found.'
            }), 404

        # Prevent self-deactivation
        if user_to_toggle.id == current_user.id:
            return jsonify({
                'success': False,
                'message': 'You cannot deactivate yourself.'
            }), 400

        # Check if trying to deactivate the last admin
        if (user_to_toggle.role == 'admin' and
            user_to_toggle.active_status and
            User.query.filter_by(role='admin', active_status=True).count() <= 1):
            return jsonify({
                'success': False,
                'message': 'Cannot deactivate the last admin user.'
            }), 400

        # Toggle the status
        new_status = not user_to_toggle.active_status
        user_to_toggle.active_status = new_status
        db.session.commit()

        action = 'activated' if new_status else 'deactivated'
        message = f'"{user_to_toggle.full_name}" has been {action} successfully.'

        # Log status change
        logger_handler.logger.info(f"Admin {current_user.username} {action} user {user_to_toggle.username}")

        print(f"Admin {current_user.username} {action} user {user_to_toggle.username}")

        return jsonify({
            'success': True,
            'message': message,
            'new_status': new_status,
            'user_id': user_id
        })

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('user_status_toggle', e)
        print(f"Error toggling user status: {e}")
        return jsonify({
            'success': False,
            'message': 'Error updating user status. Please try again.'
        }), 500

@app.route('/users/<int:user_id>/activate', methods=['GET', 'POST'])
@admin_required
def activate_user(user_id):
    """Activate a user (Admin only) - Alternative route"""
    try:
        user_to_activate = User.query.get(user_id)
        current_user = User.query.get(session['user_id'])

        if not user_to_activate:
            flash('User not found.', 'error')
            return redirect(url_for('users'))

        if user_to_activate.active_status:
            flash('User is already active.', 'info')
        else:
            user_to_activate.active_status = True
            db.session.commit()

            # Log activation
            logger_handler.logger.info(f"Admin {current_user.username} activated user {user_to_activate.username}")

            flash(f'"{user_to_activate.full_name}" has been activated.', 'success')
            print(f"Admin {current_user.username} activated user {user_to_activate.username}")

        return redirect(url_for('users'))

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('user_activation', e)
        print(f"Error activating user: {e}")
        flash('Error activating user. Please try again.', 'error')
        return redirect(url_for('users'))

@app.route('/users/<int:user_id>/deactivate', methods=['GET', 'POST'])
@admin_required
def deactivate_user(user_id):
    """Deactivate a user (Admin only) - Alternative route"""
    try:
        user_to_deactivate = User.query.get(user_id)
        current_user = User.query.get(session['user_id'])

        if not user_to_deactivate:
            flash('User not found.', 'error')
            return redirect(url_for('users'))

        # Prevent self-deactivation
        if user_to_deactivate.id == current_user.id:
            flash('You cannot deactivate yourself.', 'error')
            return redirect(url_for('users'))

        # Check if this is the last admin
        if user_to_deactivate.role == 'admin' and user_to_deactivate.active_status:
            active_admin_count = User.query.filter_by(role='admin', active_status=True).count()
            if active_admin_count <= 1:
                flash('Cannot deactivate the last admin user.', 'error')
                return redirect(url_for('users'))

        if not user_to_deactivate.active_status:
            flash('User is already inactive.', 'info')
        else:
            user_to_deactivate.active_status = False
            db.session.commit()

            # Log deactivation
            logger_handler.logger.info(f"Admin {current_user.username} deactivated user {user_to_deactivate.username}")

            flash(f'"{user_to_deactivate.full_name}" has been deactivated.', 'success')
            print(f"Admin {current_user.username} deactivated user {user_to_deactivate.username}")

        return redirect(url_for('users'))

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('user_deactivation', e)
        print(f"Error deactivating user: {e}")
        flash('Error deactivating user. Please try again.', 'error')
        return redirect(url_for('users'))

# ENHANCED USER STATISTICS API
@app.route('/api/users/stats')
@admin_required
def user_stats_api():
    """API endpoint to get user statistics for dashboard"""
    try:
        # Get current date for recent activity calculations
        one_week_ago = datetime.now() - timedelta(days=7)

        total_users = User.query.count()
        active_users = User.query.filter_by(active_status=True).count()
        admin_users = User.query.filter_by(role='admin', active_status=True).count()
        staff_users = User.query.filter_by(role='staff', active_status=True).count()
        payroll_users = User.query.filter_by(role='payroll', active_status=True).count()
        project_manager_users = User.query.filter_by(role='project_manager', active_status=True).count()
        accounting_users = User.query.filter_by(role='accounting', active_status=True).count()
        inactive_users = User.query.filter_by(active_status=False).count()

        recent_registrations = User.query.filter(
            User.created_date >= one_week_ago
        ).count()

        recent_logins = User.query.filter(
            User.last_login_date >= one_week_ago
        ).count()

        return jsonify({
            'total_users': total_users,
            'active_users': active_users,
            'admin_users': admin_users,
            'staff_users': staff_users,
            'payroll_users': payroll_users,
            'project_manager_users': project_manager_users,
            'accounting_users': accounting_users,
            'inactive_users': inactive_users,
            'recent_registrations': recent_registrations,
            'recent_logins': recent_logins
        })

    except Exception as e:
        logger_handler.log_database_error('user_stats_api', e)
        print(f"Error fetching user stats: {e}")
        return jsonify({'error': 'Failed to fetch user statistics'}), 500
    
@app.route('/api/locations-by-projects', methods=['POST'])
@admin_required
def get_locations_by_projects():
    """Get locations that belong to selected projects"""
    try:
        data = request.get_json()
        project_ids = data.get('project_ids', [])
        
        if not project_ids:
            # No projects selected, return empty list
            return jsonify({
                'success': True,
                'locations': [],
                'message': 'No projects selected'
            })
        
        # Get unique locations from QR codes that belong to selected projects
        result = db.session.execute(text("""
            SELECT DISTINCT location 
            FROM qr_codes 
            WHERE project_id IN :project_ids
            AND location IS NOT NULL 
            AND active_status = 1
            ORDER BY location
        """), {'project_ids': tuple(project_ids)})
        
        locations = [row[0] for row in result.fetchall()]
        
        return jsonify({
            'success': True,
            'locations': locations,
            'count': len(locations)
        })
        
    except Exception as e:
        logger_handler.logger.error(f"Error fetching locations by projects: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/roles/permissions')
@admin_required
def role_permissions_api():
    """API endpoint to get role permissions data"""
    try:
        permissions_data = {}
        for role in VALID_ROLES:
            permissions_data[role] = get_role_permissions(role)

        return jsonify({
            'success': True,
            'roles': permissions_data,
            'valid_roles': VALID_ROLES,
            'staff_level_roles': STAFF_LEVEL_ROLES
        })

    except Exception as e:
        print(f"Error fetching role permissions: {e}")
        return jsonify({'error': 'Failed to fetch role permissions'}), 500

@app.route('/api/geocode', methods=['POST'])
@login_required
def geocode_address_api():
    """API endpoint to geocode an address and return coordinates using Google Maps"""
    try:
        data = request.get_json()
        address = data.get('address', '').strip()

        if not address:
            return jsonify({
                'success': False,
                'message': 'Address is required'
            }), 400

        # Log API geocoding request
        try:
            logger_handler.log_user_activity('api_geocoding_request', f'API geocoding request: {address[:50]}...')
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")

        # Use the enhanced function that returns 3 values
        lat, lng, accuracy = get_coordinates_from_address_enhanced(address)

        if lat is not None and lng is not None:
            # Log successful API geocoding
            try:
                logger_handler.log_user_activity('api_geocoding_success', f'API geocoding success: {address[:50]}... -> {lat}, {lng} ({accuracy})')
            except Exception as log_error:
                print(f"⚠️ Logging error (non-critical): {log_error}")

            return jsonify({
                'success': True,
                'data': {
                    'latitude': lat,
                    'longitude': lng,
                    'accuracy': accuracy,
                    'coordinates_display': f"{lat:.10f}, {lng:.10f}",
                    'service_used': 'Google Maps' if gmaps_client else 'OpenStreetMap'
                },
                'message': f'Address geocoded successfully with {accuracy} accuracy using {"Google Maps" if gmaps_client else "OpenStreetMap"}'
            })
        else:
            # Log failed API geocoding
            try:
                logger_handler.log_user_activity('api_geocoding_failed', f'API geocoding failed: {address[:50]}...')
            except Exception as log_error:
                print(f"⚠️ Logging error (non-critical): {log_error}")

            return jsonify({
                'success': False,
                'message': 'Unable to geocode the provided address. Please verify the address is complete and accurate.'
            }), 404

    except Exception as e:
        print(f"❌ Geocoding API error: {e}")
        
        # Log API geocoding error
        try:
            logger_handler.log_flask_error('api_geocoding_error', f'API geocoding error: {str(e)}')
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")

        return jsonify({
            'success': False,
            'message': 'Internal server error during geocoding. Please try again.'
        }), 500

@app.route('/api/reverse-geocode', methods=['POST'])
@login_required
def reverse_geocode_api():
    """API endpoint for reverse geocoding coordinates to address using Google Maps"""
    try:
        data = request.get_json()
        latitude = data.get('latitude')
        longitude = data.get('longitude')

        if not latitude or not longitude:
            return jsonify({
                'success': False,
                'message': 'Latitude and longitude are required'
            }), 400

        # Log API reverse geocoding request
        try:
            logger_handler.log_user_activity('api_reverse_geocoding_request', f'API reverse geocoding: {latitude}, {longitude}')
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")

        # Use the reverse geocoding function
        address = reverse_geocode_coordinates(latitude, longitude)

        if address:
            return jsonify({
                'success': True,
                'data': {
                    'address': address,
                    'coordinates': f"{latitude}, {longitude}",
                    'service_used': 'Google Maps' if gmaps_client else 'OpenStreetMap'
                },
                'message': f'Coordinates reverse geocoded successfully using {"Google Maps" if gmaps_client else "OpenStreetMap"}'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Unable to reverse geocode the provided coordinates.'
            }), 404

    except Exception as e:
        print(f"❌ Reverse geocoding API error: {e}")
        
        # Log API reverse geocoding error
        try:
            logger_handler.log_flask_error('api_reverse_geocoding_error', f'API reverse geocoding error: {str(e)}')
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")

        return jsonify({
            'success': False,
            'message': 'Internal server error during reverse geocoding. Please try again.'
        }), 500

@app.route('/users/<int:user_id>/permanently-delete', methods=['GET', 'POST'])
@admin_required
def permanently_delete_user(user_id):
    """Permanently delete user but preserve associated QR codes (Admin only)"""
    try:
        user_to_delete = User.query.get_or_404(user_id)
        current_user = User.query.get(session['user_id'])

        # Security checks
        if user_to_delete.id == current_user.id:
            flash('You cannot delete your own account.', 'error')
            return redirect(url_for('users'))

        # Only allow deletion of inactive users for safety
        if user_to_delete.active_status:
            flash('User must be deactivated before permanent deletion.', 'error')
            return redirect(url_for('users'))

        # If deleting an admin, ensure at least one admin remains
        if user_to_delete.role == 'admin':
            active_admin_count = User.query.filter_by(role='admin', active_status=True).count()
            if active_admin_count <= 1:
                flash('Cannot delete the last admin user in the system.', 'error')
                return redirect(url_for('users'))

        user_name = user_to_delete.full_name
        user_qr_count = user_to_delete.created_qr_codes.count()
        username = user_to_delete.username

        # MODIFIED: Preserve QR codes by setting created_by to NULL instead of deleting them
        orphaned_qr_codes = QRCode.query.filter_by(created_by=user_id).all()
        for qr_code in orphaned_qr_codes:
            qr_code.created_by = None

        # Update any users that were created by this user (set created_by to None)
        created_users = User.query.filter_by(created_by=user_id).all()
        for created_user in created_users:
            created_user.created_by = None

        # Log user deletion before actual deletion
        logger_handler.log_security_event(
            event_type="user_permanent_deletion",
            description=f"Admin {current_user.username} permanently deleted user {username}",
            severity="HIGH",
            additional_data={'deleted_user': username, 'qr_codes_orphaned': user_qr_count}
        )

        # Delete the user
        db.session.delete(user_to_delete)
        db.session.commit()

        # Updated flash message to reflect QR codes are preserved
        flash(f'User "{user_name}" has been permanently deleted. {user_qr_count} QR codes created by this user are now orphaned but preserved.', 'success')
        print(f"Admin {current_user.username} permanently deleted user: {username}, preserved {user_qr_count} QR codes")

        return redirect(url_for('users'))

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('user_permanent_deletion', e)
        print(f"Error permanently deleting user: {e}")
        flash('Error deleting user. Please try again.', 'error')
        return redirect(url_for('users'))

# Admin logging routes
@app.route('/admin/logs')
@admin_required
def admin_logs():
    """Admin logging dashboard"""
    try:
        # Get log statistics for the last 7 days
        stats = logger_handler.get_log_statistics(days=7)
        return render_template('admin_logs.html', log_stats=stats)
    except Exception as e:
        logger_handler.log_database_error('admin_logs_load', e)
        flash('Error loading log statistics.', 'error')
        return redirect(url_for('dashboard'))

def check_google_maps_health():
    """Check if Google Maps services are working properly"""
    try:
        if not gmaps_client:
            return False, "Google Maps client not initialized"
        
        # Test with a known address
        test_result = gmaps_client.geocode("1600 Amphitheatre Parkway, Mountain View, CA")
        
        if test_result:
            return True, "Google Maps services are operational"
        else:
            return False, "Google Maps API not returning results"
            
    except Exception as e:
        return False, f"Google Maps health check failed: {str(e)}"

# Optional: Add health check route
@app.route('/admin/health/google-maps')
@admin_required
def google_maps_health():
    """Admin route to check Google Maps service health"""
    is_healthy, message = check_google_maps_health()
    
    return jsonify({
        'healthy': is_healthy,
        'message': message,
        'service': 'Google Maps',
        'fallback_available': True,
        'timestamp': datetime.now().isoformat()
    })

# API endpoints for logging data (admin only)
@app.route('/api/logs/recent')
@admin_required
def api_recent_logs():
    """API endpoint to get recent log entries with full details and pagination support"""
    try:
        days = request.args.get('days', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        page = request.args.get('page', 1, type=int)
        category = request.args.get('category', '')
        severity = request.args.get('severity', '')
        search = request.args.get('search', '')

        print(f"📊 API request - Days: {days}, Limit: {limit}, Page: {page}")
        print(f"📊 Filters - Category: {category}, Severity: {severity}, Search: {search}")

        cutoff_date = datetime.now() - timedelta(days=days)

        # Calculate offset for pagination
        offset = (page - 1) * limit

        # Build the base SQL query with filters
        base_sql = """
        SELECT 
            event_id,
            event_type, 
            event_category, 
            event_description, 
            event_data,
            severity_level, 
            created_timestamp, 
            username, 
            user_id,
            ip_address
        FROM log_events 
        WHERE created_timestamp >= :cutoff_date
        """

        count_sql = """
        SELECT COUNT(*) as total_count
        FROM log_events 
        WHERE created_timestamp >= :cutoff_date
        """

        params = {'cutoff_date': cutoff_date}

        # Add category filter
        if category:
            base_sql += " AND event_category = :category"
            count_sql += " AND event_category = :category"
            params['category'] = category

        # Add severity filter
        if severity:
            base_sql += " AND severity_level = :severity"
            count_sql += " AND severity_level = :severity"
            params['severity'] = severity

        # Add search filter
        if search:
            search_condition = " AND (event_type LIKE :search OR event_description LIKE :search OR username LIKE :search)"
            base_sql += search_condition
            count_sql += search_condition
            params['search'] = f'%{search}%'

        # Get total count first
        count_result = db.session.execute(text(count_sql), params).fetchone()
        total_count = count_result.total_count if count_result else 0

        # Add ordering, limit and offset to main query
        base_sql += " ORDER BY created_timestamp DESC LIMIT :limit OFFSET :offset"
        params['limit'] = limit
        params['offset'] = offset

        # Execute main query
        result = db.session.execute(text(base_sql), params).fetchall()

        logs = []
        for row in result:
            # Parse event_data if it's JSON
            event_data = None
            if row.event_data:
                try:
                    event_data = json.loads(row.event_data) if isinstance(row.event_data, str) else row.event_data
                except (json.JSONDecodeError, TypeError):
                    event_data = row.event_data

            logs.append({
                'event_id': row.event_id,
                'event_type': row.event_type,
                'event_category': row.event_category,
                'description': row.event_description,
                'event_data': event_data,
                'severity': row.severity_level,
                'timestamp': row.created_timestamp.isoformat(),
                'username': row.username or 'System',
                'user_id': row.user_id,
                'ip_address': row.ip_address or '-'
            })

        print(f"📊 Returning {len(logs)} logs out of {total_count} total")

        return jsonify({
            'success': True,
            'logs': logs,
            'total': total_count,
            'page': page,
            'limit': limit,
            'total_pages': math.ceil(total_count / limit) if total_count > 0 else 0,
            'has_next': offset + limit < total_count,
            'has_prev': page > 1
        })

    except Exception as e:
        logger_handler.log_database_error('api_recent_logs', e)
        print(f"Error in api_recent_logs: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to fetch recent logs: {str(e)}'
        }), 500

@app.route('/api/logs/stats')
@admin_required
def api_log_stats():
    """API endpoint to get logging statistics"""
    try:
        days = request.args.get('days', 7, type=int)
        print(f"📊 Getting log statistics for last {days} days")

        # Get statistics from logger handler
        stats = logger_handler.get_log_statistics(days=days)
        print(f"📈 Retrieved stats: {stats}")

        # Ensure all expected keys exist with updated categories
        expected_stats = {
            'total_events': stats.get('total_events', 0),
            'security_events': stats.get('security_events', 0),
            'authentication_events': stats.get('authentication_events', 0),
            'qr_management_events': stats.get('qr_management_events', 0),
            'database_errors': stats.get('database_errors', 0),
            'application_events': stats.get('application_events', 0),
            'system_events': stats.get('system_events', 0)
        }

        return jsonify({
            'success': True,
            'stats': expected_stats,
            'days': days,
            'timestamp': datetime.now().isoformat()
        })

    except Exception as e:
        logger_handler.log_database_error('api_log_stats', e)
        print(f"❌ Error in api_log_stats: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to fetch log statistics: {str(e)}',
            'stats': {
                'total_events': 0,
                'security_events': 0,
                'authentication_events': 0,
                'qr_management_events': 0,
                'database_errors': 0,
                'application_events': 0,
                'system_events': 0
            }
        }), 500

@app.route('/api/logs/cleanup', methods=['POST'])
@admin_required
def api_cleanup_logs():
    """API endpoint to cleanup old log entries"""
    try:
        # Get JSON data
        data = request.get_json()
        if not data:
            print("❌ No JSON data provided")
            return jsonify({
                'success': False,
                'error': 'No JSON data provided'
            }), 400

        days_to_keep = data.get('days_to_keep', 90)
        print(f"🧹 Cleanup request: keep last {days_to_keep} days")

        # Validate input
        if not isinstance(days_to_keep, int) or days_to_keep < 7:
            print(f"❌ Invalid days_to_keep: {days_to_keep}")
            return jsonify({
                'success': False,
                'error': 'days_to_keep must be an integer >= 7'
            }), 400

        if days_to_keep > 365:
            print(f"❌ days_to_keep too large: {days_to_keep}")
            return jsonify({
                'success': False,
                'error': 'days_to_keep cannot exceed 365 days'
            }), 400

        # Perform cleanup using logger handler
        deleted_count = logger_handler.cleanup_old_logs(days_to_keep=days_to_keep)

        admin_username = session.get('username', 'unknown')
        print(f"✅ Cleanup completed by {admin_username}: {deleted_count} records deleted")

        # Log the admin action
        logger_handler.log_security_event(
            event_type="admin_log_cleanup",
            description=f"Admin {admin_username} performed log cleanup: {deleted_count} entries removed (keeping last {days_to_keep} days)",
            severity="HIGH",
            additional_data={
                'admin_user': admin_username,
                'days_to_keep': days_to_keep,
                'deleted_count': deleted_count,
                'ip_address': request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)
            }
        )

        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'days_to_keep': days_to_keep,
            'message': f'Successfully cleaned up {deleted_count} old log entries (keeping last {days_to_keep} days)',
            'performed_by': admin_username,
            'performed_at': datetime.now().isoformat()
        })

    except Exception as e:
        logger_handler.log_database_error('api_cleanup_logs', e)
        print(f"❌ Error in api_cleanup_logs: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to cleanup old logs: {str(e)}'
        }), 500

@app.route('/api/logs/clear', methods=['POST'])
@admin_required
def api_clear_logs():
    """API endpoint to clear ALL log entries"""
    try:
        admin_username = session.get('username', 'unknown')
        print(f"🧹 Clear logs request by admin: {admin_username}")

        # Count existing logs before deletion
        try:
            count_sql = "SELECT COUNT(*) as total_logs FROM log_events"
            count_result = db.session.execute(text(count_sql)).fetchone()
            total_logs = count_result.total_logs if count_result else 0

            print(f"📊 Total logs to be cleared: {total_logs}")

            if total_logs == 0:
                print("✅ No logs found to clear")
                return jsonify({
                    'success': True,
                    'deleted_count': 0,
                    'message': 'No logs found to clear'
                })

        except Exception as count_error:
            print(f"⚠️ Error counting logs: {count_error}")
            total_logs = 0

        # Perform the clear operation
        try:
            clear_sql = "DELETE FROM log_events"
            result = db.session.execute(text(clear_sql))
            deleted_count = result.rowcount
            db.session.commit()

            print(f"🗑️ Successfully cleared {deleted_count} log entries")

            # Log the clear operation (this will be the first entry in the new log)
            logger_handler.log_security_event(
                event_type="admin_log_clear",
                description=f"Admin {admin_username} cleared all log entries: {deleted_count} records deleted",
                severity="HIGH",
                additional_data={
                    'admin_user': admin_username,
                    'deleted_count': deleted_count,
                    'ip_address': request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)
                }
            )

            return jsonify({
                'success': True,
                'deleted_count': deleted_count,
                'message': f'Successfully cleared {deleted_count} log entries',
                'performed_by': admin_username,
                'performed_at': datetime.now().isoformat()
            })

        except Exception as delete_error:
            print(f"❌ Error during log clearing: {delete_error}")
            db.session.rollback()
            return jsonify({
                'success': False,
                'error': f'Failed to clear logs: {str(delete_error)}'
            }), 500

    except Exception as e:
        logger_handler.log_database_error('api_clear_logs', e)
        print(f"❌ Error in api_clear_logs: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to clear logs: {str(e)}'
        }), 500

@app.route('/api/logs/clear-old', methods=['POST'])
@admin_required
def api_clear_old_logs():
    """API endpoint to clear log entries older than specified days"""
    try:
        # Get JSON data
        data = request.get_json()
        if not data:
            print("❌ No JSON data provided")
            return jsonify({
                'success': False,
                'error': 'No JSON data provided'
            }), 400

        days_threshold = data.get('days_threshold', 90)
        admin_username = session.get('username', 'unknown')
        print(f"🧹 Clear old logs request by admin: {admin_username}, threshold: {days_threshold} days")

        # Validate input
        if not isinstance(days_threshold, int) or days_threshold not in [30, 60, 90]:
            print(f"❌ Invalid days_threshold: {days_threshold}")
            return jsonify({
                'success': False,
                'error': 'days_threshold must be 30, 60, or 90'
            }), 400

        # Calculate cutoff date
        cutoff_date = datetime.now() - timedelta(days=days_threshold)

        # Count existing logs before deletion
        try:
            count_sql = "SELECT COUNT(*) as total_logs FROM log_events WHERE created_timestamp < :cutoff_date"
            count_result = db.session.execute(text(count_sql), {'cutoff_date': cutoff_date}).fetchone()
            total_logs = count_result.total_logs if count_result else 0

            print(f"📊 Total logs older than {days_threshold} days to be cleared: {total_logs}")

            if total_logs == 0:
                print("✅ No old logs found to clear")
                return jsonify({
                    'success': True,
                    'deleted_count': 0,
                    'message': f'No logs older than {days_threshold} days found to clear'
                })

        except Exception as count_error:
            print(f"⚠️ Error counting old logs: {count_error}")
            total_logs = 0

        # Perform the clear operation
        try:
            clear_sql = "DELETE FROM log_events WHERE created_timestamp < :cutoff_date"
            result = db.session.execute(text(clear_sql), {'cutoff_date': cutoff_date})
            deleted_count = result.rowcount
            db.session.commit()

            print(f"🗑️ Successfully cleared {deleted_count} log entries older than {days_threshold} days")

            # Log the clear operation
            logger_handler.log_security_event(
                event_type="admin_clear_old_logs",
                description=f"Admin {admin_username} cleared {deleted_count} log entries older than {days_threshold} days",
                severity="HIGH",
                additional_data={
                    'admin_user': admin_username,
                    'days_threshold': days_threshold,
                    'deleted_count': deleted_count,
                    'cutoff_date': cutoff_date.isoformat(),
                    'ip_address': request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)
                }
            )

            return jsonify({
                'success': True,
                'deleted_count': deleted_count,
                'days_threshold': days_threshold,
                'message': f'Successfully cleared {deleted_count} log entries older than {days_threshold} days',
                'performed_by': admin_username,
                'performed_at': datetime.now().isoformat()
            })

        except Exception as delete_error:
            print(f"❌ Error during old log clearing: {delete_error}")
            db.session.rollback()
            return jsonify({
                'success': False,
                'error': f'Failed to clear old logs: {str(delete_error)}'
            }), 500

    except Exception as e:
        logger_handler.log_database_error('api_clear_old_logs', e)
        print(f"❌ Error in api_clear_old_logs: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to clear old logs: {str(e)}'
        }), 500

@app.route('/api/logs/export')
@admin_required
def api_export_logs():
    """API endpoint to export log entries"""
    try:
        days = request.args.get('days', 7, type=int)
        category = request.args.get('category', '')
        severity = request.args.get('severity', '')
        search = request.args.get('search', '')

        admin_username = session.get('username', 'unknown')
        print(f"📊 Export logs request by admin: {admin_username}")

        cutoff_date = datetime.now() - timedelta(days=days)

        # Build the SQL query with filters
        base_sql = """
        SELECT 
            event_id,
            event_type, 
            event_category, 
            event_description, 
            event_data,
            severity_level, 
            created_timestamp, 
            username, 
            user_id,
            ip_address
        FROM log_events 
        WHERE created_timestamp >= :cutoff_date
        """

        params = {'cutoff_date': cutoff_date}

        # Add category filter
        if category:
            base_sql += " AND event_category = :category"
            params['category'] = category

        # Add severity filter
        if severity:
            base_sql += " AND severity_level = :severity"
            params['severity'] = severity

        # Add search filter
        if search:
            base_sql += " AND (event_type LIKE :search OR event_description LIKE :search OR username LIKE :search)"
            params['search'] = f'%{search}%'

        base_sql += " ORDER BY created_timestamp DESC"

        result = db.session.execute(text(base_sql), params).fetchall()

        logs = []
        for row in result:
            # Parse event_data if it's JSON
            event_data = None
            if row.event_data:
                try:
                    event_data = json.loads(row.event_data) if isinstance(row.event_data, str) else row.event_data
                except (json.JSONDecodeError, TypeError):
                    event_data = row.event_data

            logs.append({
                'event_id': row.event_id,
                'event_type': row.event_type,
                'event_category': row.event_category,
                'description': row.event_description,
                'event_data': event_data,
                'severity': row.severity_level,
                'timestamp': row.created_timestamp.isoformat(),
                'username': row.username or 'System',
                'user_id': row.user_id,
                'ip_address': row.ip_address or '-'
            })

        # Log the export operation
        logger_handler.log_security_event(
            event_type="admin_log_export",
            description=f"Admin {admin_username} exported {len(logs)} log entries (last {days} days)",
            severity="MEDIUM",
            additional_data={
                'admin_user': admin_username,
                'exported_count': len(logs),
                'days_exported': days,
                'filters': {
                    'category': category,
                    'severity': severity,
                    'search': search
                },
                'ip_address': request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)
            }
        )

        return jsonify({
            'success': True,
            'logs': logs,
            'total': len(logs),
            'filters_applied': {
                'days': days,
                'category': category,
                'severity': severity,
                'search': search
            }
        })

    except Exception as e:
        logger_handler.log_database_error('api_export_logs', e)
        print(f"❌ Error in api_export_logs: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to export logs: {str(e)}'
        }), 500

# PROJECT MANAGEMENT ROUTES
@app.route('/projects')
@admin_required
def projects():
    """Display all projects"""
    try:
        projects = Project.query.order_by(Project.created_date.desc()).all()
        return render_template('projects.html', projects=projects)
    except Exception as e:
        logger_handler.log_database_error('projects_list', e)
        flash('Error loading projects list.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/projects/create', methods=['GET', 'POST'])
@admin_required
@log_database_operations('project_creation')
def create_project():
    """Create new project"""
    if request.method == 'POST':
        try:
            name = request.form['name']
            description = request.form.get('description', '')

            # Check if project name already exists
            if Project.query.filter_by(name=name).first():
                flash('Project name already exists.', 'error')
                return render_template('create_project.html')

            # Create new project
            new_project = Project(
                name=name,
                description=description,
                created_by=session['user_id']
            )

            db.session.add(new_project)
            db.session.commit()

            # Log project creation
            logger_handler.logger.info(f"User {session['username']} created new project: {name}")

            flash(f'Project "{name}" created successfully.', 'success')
            return redirect(url_for('projects'))

        except Exception as e:
            db.session.rollback()
            logger_handler.log_database_error('project_creation', e)
            flash('Project creation failed. Please try again.', 'error')

    return render_template('create_project.html')

@app.route('/projects/<int:project_id>/edit', methods=['GET', 'POST'])
@admin_required
@log_database_operations('project_edit')
def edit_project(project_id):
    """Edit existing project"""
    try:
        project = Project.query.get_or_404(project_id)

        if request.method == 'POST':
            old_name = project.name
            old_description = project.description

            project.name = request.form['name']
            project.description = request.form.get('description', '')

            db.session.commit()

            # Log project update
            changes = {}
            if old_name != project.name:
                changes['name'] = {'old': old_name, 'new': project.name}
            if old_description != project.description:
                changes['description'] = {'old': old_description, 'new': project.description}

            if changes:
                logger_handler.logger.info(f"User {session['username']} updated project {project_id}: {json.dumps(changes)}")

            flash(f'Project "{project.name}" updated successfully.', 'success')
            return redirect(url_for('projects'))

        return render_template('edit_project.html', project=project)

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('project_edit', e)
        flash('Project update failed. Please try again.', 'error')
        return redirect(url_for('projects'))

@app.route('/projects/<int:project_id>/toggle', methods=['POST'])
@admin_required
@log_database_operations('project_toggle')
def toggle_project(project_id):
    """Toggle project active status"""
    try:
        project = Project.query.get_or_404(project_id)
        old_status = project.active_status
        project.active_status = not project.active_status

        db.session.commit()

        # Log status change
        status = "activated" if project.active_status else "deactivated"
        logger_handler.logger.info(f"User {session['username']} {status} project: {project.name}")

        flash(f'Project "{project.name}" {status} successfully.', 'success')

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('project_toggle', e)
        flash('Failed to update project status.', 'error')

    return redirect(url_for('projects'))

# API ENDPOINTS FOR DROPDOWN FUNCTIONALITY
@app.route('/api/projects/active')
@login_required
def api_active_projects():
    """Get active projects for dropdown"""
    try:
        projects = Project.query.filter_by(active_status=True).order_by(Project.name.asc()).all()

        projects_data = [
            {
                'id': project.id,
                'name': project.name,
                'description': project.description,
                'qr_count': project.qr_count
            }
            for project in projects
        ]

        return jsonify({
            'success': True,
            'projects': projects_data
        })

    except Exception as e:
        logger_handler.log_database_error('api_active_projects', e)
        return jsonify({
            'success': False,
            'error': 'Failed to fetch projects'
        }), 500

# QR CODE MANAGEMENT ROUTES
@app.route('/qr-codes/create', methods=['GET', 'POST'])
@login_required
@log_database_operations('qr_code_creation')
def create_qr_code():
    """Enhanced create QR code with customization options"""
    if request.method == 'POST':
        try:
            # Existing form data
            name = request.form['name']
            location = request.form['location']
            location_address = request.form['location_address']
            location_event = request.form.get('location_event', '')
            project_id = request.form.get('project_id')

            # Extract coordinates data from form
            latitude = request.form.get('latitude')
            longitude = request.form.get('longitude')
            coordinate_accuracy = request.form.get('coordinate_accuracy', 'geocoded')

            # NEW: QR Code customization data
            fill_color = request.form.get('fill_color', '#000000')
            back_color = request.form.get('back_color', '#FFFFFF')
            box_size = int(request.form.get('box_size', 10))
            border = int(request.form.get('border', 4))
            error_correction = request.form.get('error_correction', 'L')
            style_id = request.form.get('style_id')  # Pre-defined style

            # Validate colors (basic hex validation)
            if not (fill_color.startswith('#') and len(fill_color) == 7):
                fill_color = '#000000'
            if not (back_color.startswith('#') and len(back_color) == 7):
                back_color = '#FFFFFF'

            # Convert coordinates to float if they exist
            address_latitude = None
            address_longitude = None
            has_coordinates = False

            if latitude and longitude:
                try:
                    address_latitude = float(latitude)
                    address_longitude = float(longitude)
                    has_coordinates = True
                    print(f"✓ Coordinates received: {address_latitude}, {address_longitude}")
                except (ValueError, TypeError) as e:
                    print(f"⚠️ Invalid coordinates format: {e}")
                    address_latitude = None
                    address_longitude = None
                    has_coordinates = False

            # Validate project_id if provided
            project = None
            if project_id:
                try:
                    project_id = int(project_id)
                    project = Project.query.get(project_id)
                    if not project or not project.active_status:
                        flash('Selected project is not valid or inactive.', 'error')
                        return render_template('create_qr_code.html',
                                             projects=Project.query.filter_by(active_status=True).all(),
                                             styles=QRCodeStyle.query.all() if 'QRCodeStyle' in globals() else [])
                except (ValueError, TypeError):
                    flash('Invalid project selection.', 'error')
                    return render_template('create_qr_code.html',
                                         projects=Project.query.filter_by(active_status=True).all(),
                                         styles=QRCodeStyle.query.all() if 'QRCodeStyle' in globals() else [])

            # Create new QR code record first (without URL and image)
            new_qr_code = QRCode(
                name=name,
                location=location,
                location_address=location_address,
                location_event=location_event,
                qr_code_image="",  # Will be updated after URL generation
                qr_url="",  # Will be updated after ID is assigned
                created_by=session['user_id'],
                project_id=project_id,
                address_latitude=address_latitude,
                address_longitude=address_longitude,
                coordinate_accuracy=coordinate_accuracy if has_coordinates else None,
                coordinates_updated_date=datetime.utcnow() if has_coordinates else None,
                # NEW: Customization fields (only if columns exist)
                **({
                    'fill_color': fill_color,
                    'back_color': back_color,
                    'box_size': box_size,
                    'border': border,
                    'error_correction': error_correction,
                    'style_id': int(style_id) if style_id and style_id.isdigit() else None
                } if hasattr(QRCode, 'fill_color') else {})
            )

            # Add to session and flush to get the ID
            db.session.add(new_qr_code)
            db.session.flush()  # This assigns the ID without committing

            # Now generate the readable URL using the ID
            qr_url = generate_qr_url(name, new_qr_code.id)

            # Generate QR code data with the destination URL and custom styling
            qr_data = f"{request.url_root}qr/{qr_url}"
            qr_image = generate_qr_code(
                data=qr_data,
                fill_color=fill_color,
                back_color=back_color,
                box_size=box_size,
                border=border,
                error_correction=error_correction
            )

            # Update the QR code with the URL and image
            new_qr_code.qr_url = qr_url
            new_qr_code.qr_code_image = qr_image

            # Now commit all changes
            db.session.commit()

            # Enhanced logging with customization information
            logger_handler.log_qr_code_created(
                qr_code_id=new_qr_code.id,
                qr_code_name=name,
                created_by_user_id=session['user_id'],
                qr_data={
                    'location': location,
                    'location_address': location_address,
                    'location_event': location_event,
                    'has_coordinates': has_coordinates,
                    'customization': {
                        'fill_color': fill_color,
                        'back_color': back_color,
                        'box_size': box_size,
                        'border': border,
                        'error_correction': error_correction
                    }
                }
            )

            # Success message with customization info
            project_info = f" in project '{project.name}'" if project else ""
            coord_info = f" with coordinates ({new_qr_code.coordinates_display})" if has_coordinates else ""
            style_info = f" with custom styling (Fill: {fill_color}, Background: {back_color})"

            flash(f'QR Code "{name}" created successfully{project_info}{coord_info}{style_info}! URL: {qr_url}', 'success')
            return redirect(url_for('dashboard'))

        except Exception as e:
            db.session.rollback()
            logger_handler.log_database_error('qr_code_creation', e)
            flash('QR Code creation failed. Please try again.', 'error')
            print(f"❌ QR Code creation error: {e}")

    # Get active projects and styles for dropdown
    projects = Project.query.filter_by(active_status=True).order_by(Project.name.asc()).all()
    styles = QRCodeStyle.query.order_by(QRCodeStyle.name.asc()).all() if 'QRCodeStyle' in globals() else []

    return render_template('create_qr_code.html', projects=projects, styles=styles)

@app.route('/qr-codes/bulk-import', methods=['GET', 'POST'])
@login_required
@log_database_operations('qr_code_bulk_import')
def import_bulk_qr_codes():
    """Bulk import QR codes from Excel file"""
    
    if request.method == 'GET':
        return render_template('bulk_qr_import.html')
    
    try:
        proceed_import = request.form.get('proceed_import') == 'true'
        
        if proceed_import:
            if 'pending_qr_import_file' not in session or 'pending_qr_import_filename' not in session:
                flash('Import session expired. Please upload the file again.', 'error')
                return redirect(url_for('import_bulk_qr_codes'))
            
            temp_path = session['pending_qr_import_file']
            filename = session['pending_qr_import_filename']
            
            if not os.path.exists(temp_path):
                flash('Temporary file not found. Please upload the file again.', 'error')
                session.pop('pending_qr_import_file', None)
                session.pop('pending_qr_import_filename', None)
                return redirect(url_for('import_bulk_qr_codes'))
        else:
            if 'file' not in request.files:
                flash('No file uploaded.', 'error')
                return redirect(request.url)
            
            file = request.files['file']
            if file.filename == '':
                flash('No file selected.', 'error')
                return redirect(request.url)
            
            if not file.filename.lower().endswith(('.xlsx', '.xls')):
                flash('Please upload an Excel file (.xlsx or .xls).', 'error')
                return redirect(request.url)
            
            filename = secure_filename(file.filename)
            temp_path = os.path.join(app.config.get('UPLOAD_FOLDER', '/tmp'), 
                                   f"temp_qr_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
            
            os.makedirs(os.path.dirname(temp_path), exist_ok=True)
            file.save(temp_path)
            
            session['pending_qr_import_file'] = temp_path
            session['pending_qr_import_filename'] = filename
        
        validate_only = request.form.get('validate_only') == 'true' and not proceed_import
        
        import_service = QRCodeImportService(db, logger_handler)
        
        if validate_only:
            validation_result = import_service.validate_excel_file(temp_path)
            
            if validation_result['success']:
                flash(f"Validation successful! Found {validation_result['valid_rows']} valid records.", 'success')
            else:
                flash(f"Validation found errors. Please fix them before importing.", 'error')
            
            return render_template('bulk_qr_import.html', validation_result=validation_result)
        
        projects = Project.query.filter_by(active_status=True).all()
        project_lookup = {p.name: p.id for p in projects}
        
        import_result = import_service.import_from_excel(
            file_path=temp_path,
            created_by=session['user_id'],
            generate_qr_code_func=generate_qr_code,
            generate_qr_url_func=generate_qr_url,
            request_url_root=request.url_root,
            project_lookup=project_lookup,
            QRCode=QRCode,
            Project=Project,
            geocode_func=get_coordinates_from_address_enhanced
        )
        
        if import_result['success']:
            logger_handler.logger.info(
                f"User {session['username']} successfully imported {import_result['imported_records']} QR codes via bulk import "
                f"({import_result.get('geocoded_records', 0)} addresses auto-geocoded)"
            )
            
            flash(f"Import successful! Imported {import_result['imported_records']} QR codes "
                f"out of {import_result['total_rows']} total records.", 'success')
            
            # Show geocoding info
            if import_result.get('geocoded_records', 0) > 0:
                flash(f"✓ {import_result['geocoded_records']} addresses were automatically geocoded using Google Maps.", 'info')
            
            if import_result['failed_records'] > 0:
                flash(f"Note: {import_result['failed_records']} records failed to import. "
                    f"Check the error details below.", 'warning')
        else:
            flash(f"Import failed: {import_result.get('error', 'Unknown error')}", 'error')
        
        session.pop('pending_qr_import_file', None)
        session.pop('pending_qr_import_filename', None)
        
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as cleanup_error:
            logger_handler.logger.warning(f"Failed to cleanup temp file: {cleanup_error}")
        
        return render_template('bulk_qr_import.html', import_result=import_result)
    
    except Exception as e:
        logger_handler.log_database_error('qr_code_bulk_import', e)
        flash(f'Import failed: {str(e)}', 'error')
        return redirect(url_for('import_bulk_qr_codes'))


@app.route('/qr-codes/bulk-import/template')
@login_required
def download_qr_import_template():
    """Download Excel template for bulk QR code import"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "QR Code Import Template"
        
        headers = [
            'QR Code Name',
            'QR Code Location',
            'Project',
            'Location Address',
            'Event',
            'Latitude',
            'Longitude'
        ]
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        example_data = [
            ['HQ-Entrance', 'Main Building', 'Corporate HQ', '123 Main St, Springfield, IL 62701', 'Check IN', 39.781721, -89.650148],
            ['HQ-Exit', 'Main Building', 'Corporate HQ', '123 Main St, Springfield, IL 62701', 'Check OUT', 39.781721, -89.650148],
            ['Site-A-Gate1', 'Construction Site A', 'Construction Projects', '456 Oak Ave, Chicago, IL 60601', 'Check IN', '', '']
        ]
        
        for row_num, row_data in enumerate(example_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        column_widths = [20, 20, 20, 40, 15, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger_handler.logger.info(f"User {session.get('username', 'unknown')} downloaded QR import template")
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='QR_Code_Import_Template.xlsx'
        )
    
    except Exception as e:
        logger_handler.log_flask_error('qr_import_template_download', str(e))
        flash('Error generating template. Please try again.', 'error')
        return redirect(url_for('import_bulk_qr_codes'))
    
@app.route('/qr-codes/<int:qr_id>/edit', methods=['GET', 'POST'])
@login_required
@log_database_operations('qr_code_edit')
def edit_qr_code(qr_id):
    """Enhanced edit QR code with customization support"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)

        if request.method == 'POST':
            # Track changes for logging
            old_data = {
                'name': qr_code.name,
                'location': qr_code.location,
                'location_address': qr_code.location_address,
                'location_event': qr_code.location_event,
                'project_id': qr_code.project_id,
                'qr_url': qr_code.qr_url,
                'address_latitude': qr_code.address_latitude,
                'address_longitude': qr_code.address_longitude,
                'coordinate_accuracy': qr_code.coordinate_accuracy,
                # Track old styling
                'fill_color': getattr(qr_code, 'fill_color', '#000000'),
                'back_color': getattr(qr_code, 'back_color', '#FFFFFF')
            }

            # Update QR code fields
            new_name = request.form['name']
            qr_code.name = new_name
            qr_code.location = request.form['location']
            qr_code.location_address = request.form['location_address']
            qr_code.location_event = request.form.get('location_event', '')

            # Handle coordinates
            latitude = request.form.get('address_latitude')
            longitude = request.form.get('address_longitude')
            coordinate_accuracy = request.form.get('coordinate_accuracy', 'geocoded')

            if latitude and longitude:
                try:
                    qr_code.address_latitude = float(latitude)
                    qr_code.address_longitude = float(longitude)
                    qr_code.coordinate_accuracy = coordinate_accuracy
                    qr_code.coordinates_updated_date = datetime.utcnow()
                except (ValueError, TypeError):
                    pass
            elif latitude == '' and longitude == '':
                qr_code.address_latitude = None
                qr_code.address_longitude = None
                qr_code.coordinate_accuracy = None
                qr_code.coordinates_updated_date = None

            # Handle project association
            new_project_id = request.form.get('project_id')
            if new_project_id and new_project_id.strip():
                try:
                    new_project_id = int(new_project_id)
                    project = Project.query.get(new_project_id)
                    if project and project.active_status:
                        qr_code.project_id = new_project_id
                    else:
                        flash('Selected project is not valid or inactive.', 'error')
                        return render_template('edit_qr_code.html', qr_code=qr_code,
                                             projects=Project.query.filter_by(active_status=True).all(),
                                             styles=QRCodeStyle.query.all() if 'QRCodeStyle' in globals() else [])
                except (ValueError, TypeError):
                    flash('Invalid project selection.', 'error')
                    return render_template('edit_qr_code.html', qr_code=qr_code,
                                         projects=Project.query.filter_by(active_status=True).all(),
                                         styles=QRCodeStyle.query.all() if 'QRCodeStyle' in globals() else [])
            else:
                qr_code.project_id = None

            # Handle QR code customization (only if columns exist)
            fill_color = request.form.get('fill_color', '#000000')
            back_color = request.form.get('back_color', '#FFFFFF')
            box_size = int(request.form.get('box_size', 10))
            border = int(request.form.get('border', 4))
            error_correction = request.form.get('error_correction', 'L')
            style_id = request.form.get('style_id')

            # Update styling fields if they exist
            if hasattr(qr_code, 'fill_color'):
                qr_code.fill_color = fill_color
                qr_code.back_color = back_color
                qr_code.box_size = box_size
                qr_code.border = border
                qr_code.error_correction = error_correction
                qr_code.style_id = int(style_id) if style_id and style_id.isdigit() else None

            # Check if QR code needs regeneration
            name_changed = old_data['name'] != new_name
            styling_changed = (hasattr(qr_code, 'fill_color') and
                             (old_data['fill_color'] != fill_color or
                              old_data['back_color'] != back_color))

            if name_changed:
                new_qr_url = generate_qr_url(new_name, qr_code.id)
                qr_code.qr_url = new_qr_url

            # Regenerate QR code if name or styling changed
            if name_changed or styling_changed:
                qr_data = f"{request.url_root}qr/{qr_code.qr_url}"

                # Use new styling if available, otherwise use defaults
                styling = get_qr_styling(qr_code)
                qr_code.qr_code_image = generate_qr_code(
                    data=qr_data,
                    fill_color=styling['fill_color'],
                    back_color=styling['back_color'],
                    box_size=styling['box_size'],
                    border=styling['border'],
                    error_correction=styling['error_correction']
                )

            db.session.commit()

            # Success message
            flash(f'QR Code "{qr_code.name}" updated successfully!', 'success')
            return redirect(url_for('dashboard'))

        # GET request - render edit form
        projects = Project.query.filter_by(active_status=True).order_by(Project.name.asc()).all()
        styles = QRCodeStyle.query.order_by(QRCodeStyle.name.asc()).all() if 'QRCodeStyle' in globals() else []

        return render_template('edit_qr_code.html', qr_code=qr_code, projects=projects, styles=styles)

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('qr_code_edit', e)
        flash('QR Code update failed. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/qr-codes/<int:qr_id>/delete', methods=['GET', 'POST'])
@admin_required
@log_database_operations('qr_code_deletion')
def delete_qr_code(qr_id):
    """Permanently delete QR code (Admin only) - Hard delete - PRESERVING EXACT ROUTE"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        print(f"✅ Found QR Code: {qr_code.name}")

        if request.method == 'POST':
            qr_name = qr_code.name
            qr_code_id = qr_code.id
            print(f"🗑️ ATTEMPTING TO DELETE: {qr_name}")

            # Check if QR exists before delete
            before_count = QRCode.query.count()
            print(f"📊 QR count before delete: {before_count}")

            # Log QR code deletion before actual deletion
            logger_handler.log_qr_code_deleted(
                qr_code_id=qr_code_id,
                qr_code_name=qr_name,
                deleted_by_user_id=session['user_id']
            )

            # Delete the QR code
            db.session.delete(qr_code)
            print("💾 Called db.session.delete()")

            db.session.commit()
            print("💾 Called db.session.commit()")

            # Check count after delete
            after_count = QRCode.query.count()
            print(f"📊 QR count after delete: {after_count}")
            print(f"✅ DELETE SUCCESS! Removed {before_count - after_count} records")

            flash(f'QR code "{qr_name}" has been permanently deleted!', 'success')
            return redirect(url_for('dashboard'))

        # GET request - show confirmation page
        print("📄 Showing confirmation page")
        return render_template('confirm_delete_qr.html', qr_code=qr_code)

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('qr_code_deletion', e)
        print(f"❌ ERROR in delete route: {e}")
        print(f"❌ Exception type: {type(e)}")
        print(f"❌ Traceback: {traceback.format_exc()}")
        flash('Error deleting QR code. Please try again.', 'error')
        return redirect(url_for('dashboard'))
    
@app.route('/qr/<string:qr_url>')
def qr_destination(qr_url):
    """QR code destination page where staff check in - PRESERVING EXACT ROUTE"""
    try:
        # Find QR code by URL
        qr_code = QRCode.query.filter_by(qr_url=qr_url, active_status=True).first()

        if not qr_code:
            # Log invalid QR code access attempt
            logger_handler.log_security_event(
                event_type="invalid_qr_access",
                description=f"Attempt to access invalid QR code URL: {qr_url}",
                severity="MEDIUM"
            )
            flash('QR code not found or inactive.', 'error')
            return redirect(url_for('index'))

        # Log QR code access
        logger_handler.log_qr_code_accessed(
            qr_code_id=qr_code.id,
            qr_code_name=qr_code.name,
            access_method='scan'
        )

        return render_template('qr_destination.html', qr_code=qr_code)

    except Exception as e:
        logger_handler.log_database_error('qr_code_scan', e)
        flash('Error processing QR code scan.', 'error')
        return redirect(url_for('index'))

@app.route('/qr/<string:qr_url>/checkin', methods=['POST'])
def qr_checkin(qr_url):
    """
    Enhanced staff check-in with location accuracy calculation
    Allows multiple check-ins with minimum interval between them
    PRESERVES coordinate-to-address conversion functionality
    """
    try:
        print(f"\n🚀 STARTING ENHANCED CHECK-IN PROCESS")
        print(f"   QR URL: {qr_url}")
        print(f"   Timestamp: {datetime.now()}")

        # Find QR code by URL
        qr_code = QRCode.query.filter_by(qr_url=qr_url, active_status=True).first()

        if not qr_code:
            print(f"❌ QR code not found or inactive: {qr_url}")
            return jsonify({
                'success': False,
                'message': 'QR code not found or inactive.'
            }), 404

        print(f"✅ Found QR code: {qr_code.name} (ID: {qr_code.id})")
        print(f"   Location: {qr_code.location}")
        print(f"   QR Address: {qr_code.location_address}")

        # Get and validate employee ID
        employee_id = request.form.get('employee_id', '').strip()

        if not employee_id:
            return jsonify({
                'success': False,
                'message': 'Employee ID is required.'
            }), 400

        # Check for recent check-ins with 30-minute interval validation
        today = date.today()
        current_time = datetime.now()
        time_interval = int(os.environ.get('TIME_INTERVAL'))
        the_last_checkin_time = current_time - timedelta(minutes=time_interval)

        # Find the most recent check-in for this employee at this location today
        recent_checkin = AttendanceData.query.filter_by(
            qr_code_id=qr_code.id,
            employee_id=employee_id.upper(),
            check_in_date=today
        ).order_by(AttendanceData.check_in_time.desc()).first()

        if recent_checkin:
            # Convert check_in_time (time) to datetime for comparison
            recent_checkin_datetime = datetime.combine(today, recent_checkin.check_in_time)

            # Check if 30 minutes have passed since the last check-in
            if recent_checkin_datetime > the_last_checkin_time:
                minutes_remaining = time_interval - int((current_time - recent_checkin_datetime).total_seconds() / 60)
                print(f"⚠️ Too soon for another {qr_code.location_event} for {employee_id}")
                print(f"   Last {qr_code.location_event}: {recent_checkin.check_in_time.strftime('%H:%M')}")
                print(f"   Minutes remaining: {minutes_remaining}")

                return jsonify({
                    'success': False,
                    'message': f"You can {qr_code.location_event} again in {minutes_remaining} minutes. Last {qr_code.location_event} was at {recent_checkin.check_in_time.strftime("%H:%M")}. \n"
                               f"Puedes volver a registrarte en {minutes_remaining} minutos. El último registro fue a las {recent_checkin.check_in_time.strftime("%H:%M")}."
                }), 400
            else:
                print(f"✅ {time_interval}-minute interval satisfied. Allowing new {qr_code.location_event} for {employee_id}")
        else:
            print(f"✅ First {qr_code.location_event} today for {employee_id}")

        # Process location data with coordinate-to-address conversion
        location_data = process_location_data_enhanced(request.form)

        # Get device and network info
        user_agent_string = request.headers.get('User-Agent', '')
        device_info = detect_device_info(user_agent_string)
        client_ip = get_client_ip()

        print(f"📱 Device Info: {device_info}")
        print(f"🌐 IP Address: {client_ip}")
        print(f"📍 Location Data: {location_data}")

        # Create attendance record
        print(f"\n💾 CREATING ATTENDANCE RECORD:")

        attendance = AttendanceData(
            qr_code_id=qr_code.id,
            employee_id=employee_id.upper(),
            check_in_date=today,
            check_in_time=datetime.now().time(),
            device_info=device_info,
            user_agent=user_agent_string,
            ip_address=client_ip,
            location_name=qr_code.location,
            latitude=location_data['latitude'],
            longitude=location_data['longitude'],
            accuracy=location_data['accuracy'],
            altitude=location_data['altitude'],
            location_source=location_data['source'],
            address=location_data['address'],
            status='present',
            verification_required=False,  # Will be set below if needed
            verification_status=None
        )

        print(f"✅ Created base attendance record")

        # ENHANCED DEBUG: Calculate location accuracy with detailed logging
        print(f"\n🎯 CALCULATING LOCATION ACCURACY WITH ENHANCED DEBUG...")
        print(f"   📊 QR Code Details:")
        print(f"      ID: {qr_code.id}")
        print(f"      Name: {qr_code.name}")
        print(f"      Location: {qr_code.location}")
        print(f"      Location Address: {qr_code.location_address}")
        print(f"      Has location_address: {qr_code.location_address is not None}")
        print(f"      Location Address Length: {len(qr_code.location_address) if qr_code.location_address else 0}")

        print(f"   📍 Check-in Data:")
        print(f"      Latitude: {location_data['latitude']}")
        print(f"      Longitude: {location_data['longitude']}")
        print(f"      GPS Accuracy: {location_data['accuracy']}")
        print(f"      Address: {location_data['address']}")
        print(f"      Address Length: {len(location_data['address']) if location_data['address'] else 0}")
        print(f"      Source: {location_data['source']}")

        location_accuracy = None

        try:
            # Check if we have the required data
            if not qr_code.location_address:
                print(f"❌ QR code location_address is empty or None")
                print(f"   QR Code location_address value: '{qr_code.location_address}'")
            elif not location_data['address'] and not (location_data['latitude'] and location_data['longitude']):
                print(f"❌ No check-in address or coordinates available")
                print(f"   Check-in address: '{location_data['address']}'")
                print(f"   Check-in coords: {location_data['latitude']}, {location_data['longitude']}")
            else:
                print(f"✅ Required data available, proceeding with calculation...")

                location_accuracy = calculate_location_accuracy_enhanced(
                    qr_address=qr_code.location_address,
                    checkin_address=location_data['address'],
                    checkin_lat=location_data['latitude'],
                    checkin_lng=location_data['longitude']
                )

                print(f"📐 Location accuracy calculation result: {location_accuracy}")

                if location_accuracy is not None:
                    attendance.location_accuracy = location_accuracy
                    accuracy_level = get_location_accuracy_level_enhanced(location_accuracy)
                    print(f"✅ Location accuracy set successfully: {location_accuracy:.4f} miles ({accuracy_level})")
                    print(f"📊 Final attendance.location_accuracy value: {attendance.location_accuracy}")
                else:
                    print(f"⚠️ Could not calculate location accuracy - calculation returned None")

                # CHECK DISTANCE THRESHOLD FOR PHOTO VERIFICATION
                print(f"\n📸 CHECKING PHOTO VERIFICATION REQUIREMENT:")
                print(f"   Photo Verification Enabled: {PHOTO_VERIFICATION_ENABLED}")
                requires_verification = False
                verification_photo_data = None
                
                if PHOTO_VERIFICATION_ENABLED and location_accuracy is not None and location_accuracy > DISTANCE_THRESHOLD_FOR_VERIFICATION:
                    print(f"⚠️ Distance ({location_accuracy:.3f} mi) exceeds threshold ({DISTANCE_THRESHOLD_FOR_VERIFICATION} mi)")
                    
                    # Check if photo was provided
                    verification_photo_data = request.form.get('verification_photo', None)
                    
                    if verification_photo_data:
                        print(f"✅ Verification photo provided (size: {len(verification_photo_data)} chars)")
                        
                        # Validate photo data (basic validation)
                        if verification_photo_data.startswith('data:image/'):
                            attendance.verification_photo = verification_photo_data
                            attendance.verification_required = True
                            attendance.verification_status = 'pending'
                            attendance.verification_timestamp = datetime.now()
                            print(f"✅ Photo verification set to PENDING status")
                        else:
                            print(f"⚠️ Invalid photo format provided")
                            return jsonify({
                                'success': False,
                                'message': 'Invalid photo format. Please try again.',
                                'requires_verification': True
                            }), 400
                    else:
                        print(f"❌ Photo verification REQUIRED but not provided")
                        return jsonify({
                            'success': False,
                            'message': 'Photo verification required. Distance from location is too far.',
                            'requires_verification': True,
                            'distance': round(location_accuracy, 3),
                            'threshold': DISTANCE_THRESHOLD_FOR_VERIFICATION
                        }), 400
                else:
                    print(f"✅ Distance within threshold - no verification needed")

        except Exception as e:
            print(f"❌ Error in location accuracy calculation: {e}")
            print(f"❌ Full traceback: {traceback.format_exc()}")

        # ENHANCED DEBUG: Save to database with verification
        try:
            print(f"\n💾 SAVING TO DATABASE...")
            print(f"   Attendance object before save:")
            print(f"      Employee ID: {attendance.employee_id}")
            print(f"      Location: {attendance.location_name}")
            print(f"      Latitude: {attendance.latitude}")
            print(f"      Longitude: {attendance.longitude}")
            print(f"      Address: {attendance.address}")
            print(f"      Location Accuracy: {attendance.location_accuracy}")

            db.session.add(attendance)
            db.session.commit()

            # Log verification if required
            if attendance.verification_required:
                logger_handler.log_photo_verification(
                    employee_id=attendance.employee_id,
                    qr_code_id=qr_code.id,
                    distance=location_accuracy,
                    status='pending'
                )

            # VERIFICATION: Read back from database
            saved_record = AttendanceData.query.get(attendance.id)
            print(f"✅ Successfully saved attendance record with ID: {attendance.id}")
            print(f"📊 Verification - location accuracy in database: {saved_record.location_accuracy}")

            if saved_record.location_accuracy != attendance.location_accuracy:
                print(f"⚠️ WARNING: Database value differs from object value!")
                print(f"   Object value: {attendance.location_accuracy}")
                print(f"   Database value: {saved_record.location_accuracy}")

            # Add enhanced logging for location accuracy save
            if attendance.location_accuracy is not None:
                logger_handler.logger.info(f"Location accuracy calculated and saved: {attendance.location_accuracy:.4f} miles for employee {attendance.employee_id}")
            else:
                logger_handler.logger.warning(f"Location accuracy could not be calculated for employee {attendance.employee_id} at QR {qr_code.name}")

            # Count total check-ins for today for this employee at this location
            today_checkin_count = AttendanceData.query.filter_by(
                qr_code_id=qr_code.id,
                employee_id=employee_id.upper(),
                check_in_date=today
            ).count()

            checkin_sequence_text = f"{qr_code.location_event} details"

        except Exception as e:
            print(f"❌ Database error: {e}")
            print(f"❌ Full traceback: {traceback.format_exc()}")
            db.session.rollback()
            logger_handler.log_database_error('checkin_save', e)
            return jsonify({
                'success': False,
                'message': 'Database error occurred.'
            }), 500

        # Return success response with sequence information
        response_data = {
            'success': True,
            'message': f'Check-in successful! {checkin_sequence_text} for today.',
            'data': {
                'employee_id': attendance.employee_id,
                'location': qr_code.location_address,
                'location_event': qr_code.location_event,
                'event': qr_code.location_event,  # Add both for compatibility
                'check_in_time': attendance.check_in_time.strftime('%I:%M %p'),  # 12-hour format
                'check_in_date': attendance.check_in_date.strftime('%B %d, %Y'),  # Full date format
                'device_info': attendance.device_info,
                'ip_address': attendance.ip_address,
                'location_accuracy': location_accuracy,
                'checkin_count_today': today_checkin_count,
                'checkin_sequence': checkin_sequence_text
            }
        }

        if location_data['address']:
            response_data['data']['address'] = location_data['address']

        if location_data['latitude'] and location_data['longitude']:
            response_data['data']['coordinates'] = f"{location_data['latitude']:.10f}, {location_data['longitude']:.10f}"

        # Enhanced logging for successful check-in with all details
        print(f"✅ Check-in completed successfully")
        print(f"   Employee ID: {attendance.employee_id}")
        print(f"   Time: {attendance.check_in_time.strftime('%I:%M %p')}")
        print(f"   Date: {attendance.check_in_date.strftime('%B %d, %Y')}")
        print(f"   Location: {attendance.location_name}")
        print(f"   Action: {qr_code.location_event}")
        print(f"   Address: {attendance.address}")
        print(f"   Today's count: {today_checkin_count}")
        
        # Log to database for audit trail
        logger_handler.logger.info(f"Check-in success - Employee: {attendance.employee_id}, Location: {attendance.location_name}, Time: {attendance.check_in_time}, Action: {qr_code.location_event}")

        return jsonify(response_data), 200

    except Exception as e:
        print(f"❌ Unexpected error in check-in process: {e}")
        print(f"❌ Traceback: {traceback.format_exc()}")

        return jsonify({
            'success': False,
            'message': 'An unexpected error occurred during check-in.'
        }), 500

@app.route('/qr-codes/<int:qr_id>/toggle-status', methods=['POST'])
@login_required
def toggle_qr_status(qr_id):
    """Toggle QR code active/inactive status"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)

        # Toggle the status
        qr_code.active_status = not qr_code.active_status
        db.session.commit()

        status_text = "activated" if qr_code.active_status else "deactivated"
        flash(f'QR code "{qr_code.name}" has been {status_text} successfully!', 'success')

        return jsonify({
            'success': True,
            'new_status': qr_code.active_status,
            'status_text': 'Active' if qr_code.active_status else 'Inactive',
            'message': f'QR code {status_text} successfully!'
        })

    except Exception as e:
        db.session.rollback()
        print(f"Error toggling QR status: {e}")
        return jsonify({
            'success': False,
            'message': 'Error updating QR code status. Please try again.'
        }), 500

@app.route('/qr-codes/<int:qr_id>/copy-url', methods=['POST'])
@login_required
def copy_qr_url(qr_id):
    """Log QR code URL copy action"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        
        # Log URL copy action
        logger_handler.logger.info(f"User {session.get('username', 'unknown')} copied URL for QR code {qr_code.name} (ID: {qr_id})")
        
        return jsonify({
            'success': True,
            'message': f'QR code URL copied to clipboard!',
            'url': f"{request.url_root}qr/{qr_code.qr_url}"
        })

    except Exception as e:
        logger_handler.logger.error(f"Error copying QR URL for ID {qr_id}: {e}")
        return jsonify({
            'success': False,
            'message': 'Error copying QR code URL.'
        }), 500

@app.route('/qr-codes/<int:qr_id>/open-link', methods=['POST'])
@login_required
def open_qr_link(qr_id):
    """Log QR code link open action"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        
        # Log link open action
        logger_handler.logger.info(f"User {session.get('username', 'unknown')} opened link for QR code {qr_code.name} (ID: {qr_id})")
        
        return jsonify({
            'success': True,
            'message': f'Opening QR code link...',
            'url': f"{request.url_root}qr/{qr_code.qr_url}"
        })

    except Exception as e:
        logger_handler.logger.error(f"Error opening QR link for ID {qr_id}: {e}")
        return jsonify({
            'success': False,
            'message': 'Error opening QR code link.'
        }), 500

@app.route('/qr-codes/<int:qr_id>/activate', methods=['POST'])
@login_required
def activate_qr_code(qr_id):
    """Activate a QR code"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        qr_code.active_status = True
        db.session.commit()

        flash(f'QR code "{qr_code.name}" has been activated successfully!', 'success')
        return jsonify({
            'success': True,
            'new_status': True,
            'status_text': 'Active',
            'message': 'QR code activated successfully!'
        })

    except Exception as e:
        db.session.rollback()
        print(f"Error activating QR code: {e}")
        return jsonify({
            'success': False,
            'message': 'Error activating QR code. Please try again.'
        }), 500

@app.route('/qr-codes/<int:qr_id>/deactivate', methods=['POST'])
@login_required
def deactivate_qr_code(qr_id):
    """Deactivate a QR code"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        qr_code.active_status = False
        db.session.commit()

        flash(f'QR code "{qr_code.name}" has been deactivated successfully!', 'success')
        return jsonify({
            'success': True,
            'new_status': False,
            'status_text': 'Inactive',
            'message': 'QR code deactivated successfully!'
        })

    except Exception as e:
        db.session.rollback()
        print(f"Error deactivating QR code: {e}")
        return jsonify({
            'success': False,
            'message': 'Error deactivating QR code. Please try again.'
        }), 500

@app.route('/qr-codes/<int:qr_id>/toggle-status', methods=['POST'])
@admin_required
def toggle_qr_status_api(qr_id):
    """Toggle QR code active/inactive status - Enhanced JSON API"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        qr_code.active_status = not qr_code.active_status
        db.session.commit()

        status_text = "activated" if qr_code.active_status else "deactivated"

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'new_status': qr_code.active_status,
                'status_text': 'Active' if qr_code.active_status else 'Inactive',
                'message': f'QR code "{qr_code.name}" has been {status_text} successfully!'
            })
        else:
            flash(f'QR code "{qr_code.name}" has been {status_text} successfully!', 'success')
            return redirect(url_for('dashboard'))

    except Exception as e:
        db.session.rollback()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': 'Error updating QR code status. Please try again.'
            }), 500
        else:
            flash('Error updating QR code status. Please try again.', 'error')
            return redirect(url_for('dashboard'))

@app.route('/attendance')
@login_required
def attendance_report():
    """Safe attendance report with backward compatibility for location_accuracy and fixed datetime handling"""
    try:
        print("📊 Loading attendance report...")

        # Log attendance report access
        try:
            user_role = session.get('role', 'unknown')
            logger_handler.logger.info(f"User {session.get('username', 'unknown')} accessed attendance report")
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")

        # Check if location_accuracy column exists
        has_location_accuracy = check_location_accuracy_column_exists()
        print(f"🔍 Location accuracy column exists: {has_location_accuracy}")

        # Get filter parameters
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        location_filter = request.args.get('location', '')
        employee_filter = request.args.get('employee', '')
        project_filter = request.args.get('project', '')

        # Get employee display name for filter if employee ID is provided
        employee_display_name = ''
        if employee_filter:
            try:
                employee = Employee.query.filter_by(id=int(employee_filter)).first()
                if employee:
                    employee_display_name = f"{employee.lastName}, {employee.firstName}"
                else:
                    employee_display_name = f"ID: {employee_filter}"
            except (ValueError, TypeError):
                employee_display_name = employee_filter

        # ============================================================
        # PROJECT MANAGER ACCESS CONTROL
        # ============================================================
        user_role = session.get('role')
        user_id = session.get('user_id')
        
        # Initialize permission filters
        allowed_project_ids = []
        allowed_location_names = []
        
        # Check if user is Project Manager and get their permissions
        if user_role == 'project_manager':
            print(f"🔒 Project Manager access control enabled for user {session.get('username')}")
            
            try:
                # Get assigned projects
                assigned_projects = UserProjectPermission.query.filter_by(user_id=user_id).all()
                allowed_project_ids = [p.project_id for p in assigned_projects]
                
                # Get assigned locations
                assigned_locations = UserLocationPermission.query.filter_by(user_id=user_id).all()
                allowed_location_names = [l.location_name for l in assigned_locations]
                
                # Log the permissions
                logger_handler.logger.info(
                    f"🔒 Project Manager {session.get('username')} restricted to: "
                    f"Projects: {allowed_project_ids}, Locations: {allowed_location_names}"
                )
                
                print(f"🔒 Allowed projects: {allowed_project_ids}")
                print(f"🔒 Allowed locations: {allowed_location_names}")
            except Exception as perm_error:
                print(f"⚠️ Error loading permissions: {perm_error}")
                logger_handler.logger.error(f"Error loading Project Manager permissions: {perm_error}")
            
            # If no permissions assigned, user cannot view anything
            if not allowed_project_ids and not allowed_location_names:
                logger_handler.logger.warning(
                    f"Project Manager {session.get('username')} has no assigned projects or locations"
                )
                flash('You do not have access to any projects or locations. Please contact an administrator.', 'warning')
                
                # Create empty stats object using named tuple style
                from collections import namedtuple
                Stats = namedtuple('Stats', ['total_checkins', 'unique_employees', 'active_locations', 
                                             'today_checkins', 'records_with_gps', 'records_with_accuracy', 
                                             'avg_location_accuracy'])
                empty_stats = Stats(0, 0, 0, 0, 0, 0, 0)
                
                # Return empty template
                return render_template('attendance_report.html',
                                    attendance_records=[],
                                    locations=[],
                                    projects=[],
                                    stats=empty_stats,
                                    date_from=date_from,
                                    date_to=date_to,
                                    location_filter=location_filter,
                                    employee_filter=employee_filter,
                                    employee_display_name=employee_display_name,
                                    project_filter=project_filter,
                                    today_date=datetime.now().strftime('%Y-%m-%d'),
                                    current_date_formatted=datetime.now().strftime('%B %d'),
                                    has_location_accuracy_feature=has_location_accuracy,
                                    user_role=user_role)
            
        # ============================================================
        # END: PROJECT MANAGER ACCESS CONTROL
        # ============================================================

        # Build base query - conditional based on column existence
        if has_location_accuracy:
            # New query with location accuracy
            base_query = """
                SELECT 
                    ad.id,
                    ad.employee_id,
                    ad.check_in_date,
                    ad.check_in_time,
                    ad.location_name,
                    qc.location_event,
                    qc.location_address as qr_address,
                    ad.address as checked_in_address,
                    ad.latitude,
                    ad.longitude,
                    ad.location_accuracy,
                    ad.accuracy as gps_accuracy,
                    ad.device_info,
                    ad.created_timestamp,
                    ad.updated_timestamp,
                    CONCAT(e.firstName, ' ', e.lastName) as employee_name,
                    ad.verification_required,
                    ad.verification_status,
                    ad.verification_photo
                FROM attendance_data ad
                LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id
                LEFT JOIN employee e ON CAST(ad.employee_id AS UNSIGNED) = e.id
                WHERE 1=1
            """
        else:
            # Fallback query without location accuracy
            base_query = """
                SELECT 
                    ad.id,
                    ad.employee_id,
                    ad.check_in_date,
                    ad.check_in_time,
                    ad.location_name,
                    qc.location_event,
                    qc.location_address as qr_address,
                    ad.address as checked_in_address,
                    ad.latitude,
                    ad.longitude,
                    NULL as location_accuracy,
                    ad.accuracy as gps_accuracy,
                    ad.device_info,
                    ad.created_timestamp,
                    ad.updated_timestamp,
                    CONCAT(e.firstName, ' ', e.lastName) as employee_name,
                    ad.verification_required,
                    ad.verification_status,
                    ad.verification_photo
                FROM attendance_data ad
                LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id
                LEFT JOIN employee e ON CAST(ad.employee_id AS UNSIGNED) = e.id
                WHERE 1=1
            """

        # Prepare filter conditions and parameters
        filter_conditions = []
        query_params = {}

        # ============================================================
        # APPLY PROJECT MANAGER FILTERS TO SQL QUERY
        # ============================================================
        if user_role == 'project_manager':
            # Filter by allowed projects
            if allowed_project_ids:
                project_placeholders = ','.join([f':project_{i}' for i in range(len(allowed_project_ids))])
                filter_conditions.append(f"qc.project_id IN ({project_placeholders})")
                for i, pid in enumerate(allowed_project_ids):
                    query_params[f'project_{i}'] = pid
            
            # Filter by allowed locations
            if allowed_location_names:
                location_placeholders = ','.join([f':location_{i}' for i in range(len(allowed_location_names))])
                filter_conditions.append(f"ad.location_name IN ({location_placeholders})")
                for i, loc in enumerate(allowed_location_names):
                    query_params[f'location_{i}'] = loc
        # ============================================================
        # END: APPLY PROJECT MANAGER FILTERS
        # ============================================================

        # Apply user-selected filters
        if date_from:
            filter_conditions.append("ad.check_in_date >= :date_from")
            query_params['date_from'] = date_from

        if date_to:
            filter_conditions.append("ad.check_in_date <= :date_to")
            query_params['date_to'] = date_to

        if location_filter:
            filter_conditions.append("ad.location_name = :location")
            query_params['location'] = location_filter

        if employee_filter:
            filter_conditions.append("ad.employee_id = :employee")
            query_params['employee'] = employee_filter

        if project_filter:
            filter_conditions.append("qc.project_id = :project")
            query_params['project'] = project_filter

        # Combine query with filters
        if filter_conditions:
            base_query += " AND " + " AND ".join(filter_conditions)

        base_query += " ORDER BY ad.check_in_date DESC, ad.check_in_time DESC LIMIT 1000"

        print(f"🔍 Executing attendance query with filters: {list(query_params.keys())}")

        # Execute query
        result = db.session.execute(text(base_query), query_params)
        records = result.fetchall()
        print(f"✅ Loaded {len(records)} attendance records")

        # Process records
        processed_records = []
        for record in records:
            try:
                record_dict = {
                    'id': record[0],
                    'employee_id': record[1],
                    'check_in_date': record[2],
                    'check_in_time': record[3],
                    'location_name': record[4],
                    'location_event': record[5],
                    'qr_address': record[6],
                    'checked_in_address': record[7],
                    'latitude': record[8],
                    'longitude': record[9],
                    'location_accuracy': record[10] if has_location_accuracy else None,
                    'gps_accuracy': record[11],
                    'device_info': record[12],
                    'created_timestamp': record[13],
                    'updated_timestamp': record[14],
                    'employee_name': record[15] or 'Unknown Employee',
                    'verification_required': record[16] if len(record) > 16 else False,
                    'verification_status': record[17] if len(record) > 17 else None,
                    'verification_photo': record[18] if len(record) > 18 else None
                }
                
                # Calculate accuracy_level for template display
                if record_dict['location_accuracy'] is not None:
                    accuracy_value = float(record_dict['location_accuracy'])
                    if accuracy_value <= 0.3:
                        record_dict['accuracy_level'] = 'accurate'
                    else:
                        record_dict['accuracy_level'] = 'inaccurate'
                else:
                    record_dict['accuracy_level'] = 'unknown'
                processed_records.append(record_dict)
            except Exception as rec_error:
                print(f"⚠️ Error processing record: {rec_error}")
                continue

        # Get unique locations for filter dropdown
        try:
            # ============================================================
            # FILTER LOCATIONS FOR PROJECT MANAGER
            # ============================================================
            if user_role == 'project_manager' and allowed_location_names:
                # Only show locations the PM has access to
                locations = sorted(allowed_location_names)
                print(f"✅ Filtered to {len(locations)} locations for Project Manager")
            else:
                # Show all locations for Admin/Staff/Payroll
                locations_query = db.session.execute(text("""
                    SELECT DISTINCT location_name 
                    FROM attendance_data 
                    WHERE location_name IS NOT NULL 
                    ORDER BY location_name
                """))
                locations = [row[0] for row in locations_query.fetchall()]
                print(f"✅ Found {len(locations)} unique locations")
            # ============================================================
            # END: FILTER LOCATIONS FOR PROJECT MANAGER
            # ============================================================
        except Exception as e:
            print(f"⚠️ Error loading locations: {e}")
            locations = []

        # Get projects for filter dropdown
        try:
            # ============================================================
            # FILTER PROJECTS FOR PROJECT MANAGER
            # ============================================================
            if user_role == 'project_manager' and allowed_project_ids:
                # Only show projects the PM has access to
                project_placeholders = ','.join([str(pid) for pid in allowed_project_ids])
                projects_query = db.session.execute(text(f"""
                    SELECT p.id, p.name, COUNT(DISTINCT ad.id) as attendance_count
                    FROM projects p
                    LEFT JOIN qr_codes qc ON qc.project_id = p.id
                    LEFT JOIN attendance_data ad ON ad.qr_code_id = qc.id
                    WHERE p.active_status = true AND p.id IN ({project_placeholders})
                    GROUP BY p.id, p.name
                    ORDER BY p.name
                """))
                projects = projects_query.fetchall()
                print(f"✅ Filtered to {len(projects)} projects for Project Manager")
            else:
                # Show all projects for Admin/Staff/Payroll
                projects = db.session.execute(text("""
                    SELECT p.id, p.name, COUNT(DISTINCT ad.id) as attendance_count
                    FROM projects p
                    LEFT JOIN qr_codes qc ON qc.project_id = p.id
                    LEFT JOIN attendance_data ad ON ad.qr_code_id = qc.id
                    WHERE p.active_status = true
                    GROUP BY p.id, p.name
                    HAVING COUNT(DISTINCT ad.id) > 0
                    ORDER BY p.name
                """)).fetchall()
                print(f"✅ Loaded {len(projects)} projects with attendance data")
            # ============================================================
            # END: FILTER PROJECTS FOR PROJECT MANAGER
            # ============================================================
        except Exception as e:
            print(f"⚠️ Error loading projects: {e}")
            projects = []

        # ============================================================
        # STATISTICS - COMPLETELY REWRITTEN FOR SAFETY
        # ============================================================
        print("📊 Loading statistics...")
        
        # Create simple dict for stats (most compatible approach)
        stats_dict = {
            'total_checkins': 0,
            'unique_employees': 0,
            'active_locations': 0,
            'today_checkins': 0,
            'records_with_gps': 0,
            'records_with_accuracy': 0,
            'avg_location_accuracy': 0.0
        }
        
        try:
            # Build stats query
            if has_location_accuracy:
                stats_select = """
                    SELECT 
                        COALESCE(COUNT(*), 0) as total_checkins,
                        COALESCE(COUNT(DISTINCT employee_id), 0) as unique_employees,
                        COALESCE(COUNT(DISTINCT qr_code_id), 0) as active_locations,
                        COALESCE(COUNT(CASE WHEN check_in_date = CURRENT_DATE THEN 1 END), 0) as today_checkins,
                        COALESCE(COUNT(CASE WHEN latitude IS NOT NULL AND longitude IS NOT NULL THEN 1 END), 0) as records_with_gps,
                        COALESCE(COUNT(CASE WHEN location_accuracy IS NOT NULL THEN 1 END), 0) as records_with_accuracy,
                        COALESCE(AVG(location_accuracy), 0) as avg_location_accuracy
                """
            else:
                stats_select = """
                    SELECT 
                        COALESCE(COUNT(*), 0) as total_checkins,
                        COALESCE(COUNT(DISTINCT employee_id), 0) as unique_employees,
                        COALESCE(COUNT(DISTINCT qr_code_id), 0) as active_locations,
                        COALESCE(COUNT(CASE WHEN check_in_date = CURRENT_DATE THEN 1 END), 0) as today_checkins,
                        COALESCE(COUNT(CASE WHEN latitude IS NOT NULL AND longitude IS NOT NULL THEN 1 END), 0) as records_with_gps,
                        0 as records_with_accuracy,
                        0 as avg_location_accuracy
                """
            
            stats_query_text = stats_select + " FROM attendance_data ad"
            stats_params = {}
            
            # Add filters for Project Manager
            if user_role == 'project_manager':
                stats_query_text += " LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id WHERE 1=1"
                
                stats_conditions = []
                
                if allowed_project_ids:
                    project_placeholders = ','.join([f':stat_project_{i}' for i in range(len(allowed_project_ids))])
                    stats_conditions.append(f"qc.project_id IN ({project_placeholders})")
                    for i, pid in enumerate(allowed_project_ids):
                        stats_params[f'stat_project_{i}'] = pid
                
                if allowed_location_names:
                    location_placeholders = ','.join([f':stat_location_{i}' for i in range(len(allowed_location_names))])
                    stats_conditions.append(f"ad.location_name IN ({location_placeholders})")
                    for i, loc in enumerate(allowed_location_names):
                        stats_params[f'stat_location_{i}'] = loc
                
                if stats_conditions:
                    stats_query_text += " AND " + " AND ".join(stats_conditions)
            
            print(f"📊 Executing stats query...")
            print(f"📊 Stats params: {list(stats_params.keys())}")
            
            # Execute stats query
            stats_result = db.session.execute(text(stats_query_text), stats_params)
            stats_row = stats_result.fetchone()
            
            print(f"📊 Stats row type: {type(stats_row)}")
            print(f"📊 Stats row value: {stats_row}")
            
            # Safely extract stats from row
            if stats_row is not None and len(stats_row) >= 7:
                try:
                    stats_dict['total_checkins'] = int(stats_row[0]) if stats_row[0] is not None else 0
                    stats_dict['unique_employees'] = int(stats_row[1]) if stats_row[1] is not None else 0
                    stats_dict['active_locations'] = int(stats_row[2]) if stats_row[2] is not None else 0
                    stats_dict['today_checkins'] = int(stats_row[3]) if stats_row[3] is not None else 0
                    stats_dict['records_with_gps'] = int(stats_row[4]) if stats_row[4] is not None else 0
                    stats_dict['records_with_accuracy'] = int(stats_row[5]) if stats_row[5] is not None else 0
                    stats_dict['avg_location_accuracy'] = float(stats_row[6]) if stats_row[6] is not None else 0.0
                    print(f"✅ Loaded statistics: {stats_dict['total_checkins']} total check-ins")
                except (IndexError, TypeError, ValueError) as extract_error:
                    print(f"⚠️ Error extracting stats values: {extract_error}")
                    # stats_dict already has default values
            else:
                print("⚠️ Stats query returned None or insufficient columns, using default stats")
                
        except Exception as stats_error:
            print(f"❌ Error loading statistics: {stats_error}")
            import traceback
            print(f"❌ Stats error traceback: {traceback.format_exc()}")
            # stats_dict already has default values
        
        # Convert dict to object-like for template compatibility
        class StatsObject:
            def __init__(self, stats_dict):
                for key, value in stats_dict.items():
                    setattr(self, key, value)
        
        stats = StatsObject(stats_dict)
        print(f"✅ Stats object created: total_checkins={stats.total_checkins}")
        
        # ============================================================
        # END: STATISTICS
        # ============================================================
        
        # Add today's date for template
        today_date = datetime.now().strftime('%Y-%m-%d')
        current_date_formatted = datetime.now().strftime('%B %d')

        print("✅ Rendering attendance report template")
        print(f"✅ Stats object: {stats}")

        return render_template('attendance_report.html',
                     attendance_records=processed_records,
                     locations=locations,
                     projects=projects,
                     stats=stats,
                     date_from=date_from,
                     date_to=date_to,
                     location_filter=location_filter,
                     employee_filter=employee_filter,
                     employee_display_name=employee_display_name,
                     project_filter=project_filter,
                     today_date=datetime.now().strftime('%Y-%m-%d'),
                     current_date_formatted=datetime.now().strftime('%B %d'),
                     has_location_accuracy_feature=has_location_accuracy,
                     user_role=user_role)

    except Exception as e:
        print(f"❌ Error loading attendance report: {e}")
        print(f"❌ Exception type: {type(e)}")
        
        import traceback
        error_traceback = traceback.format_exc()
        print(f"❌ Traceback: {error_traceback}")

        # Log the error
        try:
            logger_handler.log_database_error('attendance_report', e)
        except Exception as log_error:
            print(f"⚠️ Additional logging error: {log_error}")

        flash('Error loading attendance report. Please check the server logs for details.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/attendance/<int:record_id>/edit', methods=['GET', 'POST'])
@login_required
@log_database_operations('attendance_update')
def edit_attendance(record_id):
    """Edit attendance record (Admin and Payroll only)"""
    # Check if user has permission to edit attendance records
    if session.get('role') not in ['admin', 'payroll', 'accounting']:
        flash('Access denied. Only administrators and accounting staff can edit attendance records.', 'error')
        return redirect(url_for('attendance_report'))

    try:
        attendance_record = AttendanceData.query.get_or_404(record_id)

        if request.method == 'POST':
            # Get the audit note from form - REQUIRED
            edit_note = request.form.get('edit_note', '').strip()
            if not edit_note:
                flash('Edit reason is required for audit purposes.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('edit_attendance.html',
                                     attendance_record=attendance_record,
                                     projects=projects,
                                     qr_codes=QRCode.query.filter_by(active_status=True).all())

            # Track changes for logging
            changes = {}
            old_values = {
                'employee_id': attendance_record.employee_id,
                'check_in_date': attendance_record.check_in_date,
                'check_in_time': attendance_record.check_in_time,
                'location_name': attendance_record.location_name,
                'qr_code_id': attendance_record.qr_code_id,
                'location_event': attendance_record.qr_code.location_event if attendance_record.qr_code else None
            }

            # Update attendance record fields
            new_employee_id = request.form['employee_id'].strip().upper()
            new_check_in_date = datetime.strptime(request.form['check_in_date'], '%Y-%m-%d').date()
            new_check_in_time = datetime.strptime(request.form['check_in_time'], '%H:%M').time()
            new_location_name = request.form['location_name'].strip()
            
            # Get the new QR code ID from the form (this determines the location event)
            new_qr_code_id = request.form.get('qr_code_id', '').strip()
            if not new_qr_code_id:
                flash('Location event selection is required.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('edit_attendance.html',
                                     attendance_record=attendance_record,
                                     projects=projects,
                                     qr_codes=QRCode.query.filter_by(active_status=True).all())
            
            # Validate the QR code exists
            new_qr_code = QRCode.query.get(int(new_qr_code_id))
            if not new_qr_code:
                flash('Selected location event not found.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('edit_attendance.html',
                                     attendance_record=attendance_record,
                                     projects=projects,
                                     qr_codes=QRCode.query.filter_by(active_status=True).all())

            # Track what changed
            if attendance_record.employee_id != new_employee_id:
                changes['employee_id'] = f"{attendance_record.employee_id} → {new_employee_id}"
            if attendance_record.check_in_date != new_check_in_date:
                changes['check_in_date'] = f"{attendance_record.check_in_date} → {new_check_in_date}"
            if attendance_record.check_in_time != new_check_in_time:
                changes['check_in_time'] = f"{attendance_record.check_in_time} → {new_check_in_time}"
            if attendance_record.location_name != new_location_name:
                changes['location_name'] = f"{attendance_record.location_name} → {new_location_name}"
            if attendance_record.qr_code_id != int(new_qr_code_id):
                old_event = attendance_record.qr_code.location_event if attendance_record.qr_code else 'Unknown'
                new_event = new_qr_code.location_event
                changes['location_event'] = f"{old_event} → {new_event}"
                changes['qr_code_id'] = f"{attendance_record.qr_code_id} → {new_qr_code_id}"

            # Apply changes
            attendance_record.employee_id = new_employee_id
            attendance_record.check_in_date = new_check_in_date
            attendance_record.check_in_time = new_check_in_time
            attendance_record.location_name = new_location_name
            attendance_record.qr_code_id = int(new_qr_code_id)
            attendance_record.updated_timestamp = datetime.utcnow()
            
            # Store the audit note with timestamp and user info
            timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
            username = session.get('username', 'Unknown')
            role = session.get('role', 'unknown')
            
            new_note_entry = f"[{timestamp}] {role.title()} '{username}': {edit_note}"
            
            if attendance_record.edit_note:
                # Append to existing notes
                attendance_record.edit_note = f"{attendance_record.edit_note}\n\n{new_note_entry}"
            else:
                # First edit note
                attendance_record.edit_note = new_note_entry

            db.session.commit()

            # Enhanced logging with audit note
            if changes:
                logger_handler.log_security_event(
                    event_type="attendance_record_update",
                    description=f"{session.get('role', 'unknown').title()} {session.get('username')} updated attendance record {record_id}",
                    severity="MEDIUM",
                    additional_data={
                        'record_id': record_id, 
                        'changes': changes, 
                        'user_role': session.get('role'),
                        'edit_reason': edit_note,
                        'editor_username': session.get('username')
                    }
                )
                print(f"[LOG] {session.get('role', 'unknown').title()} {session.get('username')} updated attendance record {record_id}: {changes}")
                print(f"[LOG] Edit reason: {edit_note}")
            else:
                # Log even if no changes were made (for audit purposes)
                logger_handler.log_security_event(
                    event_type="attendance_record_edit_no_changes",
                    description=f"{session.get('role', 'unknown').title()} {session.get('username')} accessed edit form for record {record_id} but made no changes",
                    severity="LOW",
                    additional_data={
                        'record_id': record_id,
                        'user_role': session.get('role'),
                        'edit_reason': edit_note,
                        'editor_username': session.get('username')
                    }
                )
                print(f"[LOG] {session.get('role', 'unknown').title()} {session.get('username')} edited record {record_id} with no changes")
                print(f"[LOG] Edit reason: {edit_note}")

            flash(f'Attendance record for {new_employee_id} updated successfully! Edit reason logged for audit.', 'success')
            return redirect(url_for('attendance_report'))

        # GET request - show edit form
        # Get available projects for the dropdown
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        
        # Get available QR codes for location dropdown (for backward compatibility)
        qr_codes = QRCode.query.filter_by(active_status=True).all()

        return render_template('edit_attendance.html',
                             attendance_record=attendance_record,
                             projects=projects,
                             qr_codes=qr_codes)

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('attendance_update', e)
        print(f"[LOG] Error updating attendance record {record_id}: {e}")
        flash('Error updating attendance record. Please try again.', 'error')
        return redirect(url_for('attendance_report'))

@app.route('/attendance/add', methods=['GET'])
@login_required
@log_user_activity('manual_attendance_access')
def add_manual_attendance():
    """
    Display form to manually add attendance record
    Only accessible by admin and accounting roles
    """
    try:
        user_role = session.get('role')
        
        # Check authorization
        if user_role not in ['admin', 'accounting']:
            flash('You do not have permission to manually add attendance records.', 'error')
            return redirect(url_for('attendance_report'))
        
        # Get all active projects
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        
        # Get today's date for form
        today_date = datetime.now().strftime('%Y-%m-%d')
        
        logger_handler.logger.info(
            f"User {session.get('username')} ({user_role}) accessed manual attendance entry form"
        )
        
        return render_template('add_manual_attendance.html',
                             projects=projects,
                             today_date=today_date)
    
    except Exception as e:
        logger_handler.logger.error(f"Error loading manual attendance form: {e}")
        flash('Error loading form. Please try again.', 'error')
        return redirect(url_for('attendance_report'))


@app.route('/attendance/save_manual', methods=['POST'])
@login_required
@log_user_activity('manual_attendance_creation')
@log_database_operations('manual_attendance_insert')
def save_manual_attendance():
    """
    Save manually created attendance record
    Only accessible by admin and accounting roles
    """
    try:
        user_role = session.get('role')
        
        # Check authorization
        if user_role not in ['admin', 'accounting']:
            return jsonify({
                'success': False,
                'message': 'You do not have permission to manually add attendance records.'
            }), 403
        
        # Get form data
        employee_id = request.form.get('employee_id', '').strip()
        location_id = request.form.get('location_id', '').strip()
        check_date = request.form.get('check_date', '').strip()
        check_time = request.form.get('check_time', '').strip()
        
        # Validate required fields
        if not all([employee_id, location_id, check_date, check_time]):
            flash('All fields are required.', 'error')
            return redirect(url_for('add_manual_attendance'))
        
        # Validate employee exists
        employee = Employee.query.filter_by(id=int(employee_id)).first()
        if not employee:
            flash(f'Employee with ID {employee_id} not found.', 'error')
            return redirect(url_for('add_manual_attendance'))
        
        # Get QR code (location)
        qr_code = QRCode.query.get(int(location_id))
        if not qr_code:
            flash('Selected location not found.', 'error')
            return redirect(url_for('add_manual_attendance'))
        
        # Parse date and time
        try:
            check_date_obj = datetime.strptime(check_date, '%Y-%m-%d').date()
            check_time_obj = datetime.strptime(check_time, '%H:%M').time()
        except ValueError as e:
            flash('Invalid date or time format.', 'error')
            logger_handler.logger.error(f"Date/time parsing error: {e}")
            return redirect(url_for('add_manual_attendance'))
        
        # Check if record already exists for this employee, location, date, and time
        existing_record = AttendanceData.query.filter_by(
            employee_id=str(employee_id),
            qr_code_id=qr_code.id,
            check_in_date=check_date_obj,
            check_in_time=check_time_obj
        ).first()
        
        if existing_record:
            flash('An attendance record already exists for this employee at this location, date, and time.', 'warning')
            return redirect(url_for('add_manual_attendance'))
        
        # Create new attendance record
        # Use QR code's location address for both QR address and check-in address
        # Set fixed distance of 0.010 miles
        new_attendance = AttendanceData(
            qr_code_id=qr_code.id,
            employee_id=str(employee_id),
            check_in_date=check_date_obj,
            check_in_time=check_time_obj,
            location_name=qr_code.location,
            # Use QR code's coordinates
            latitude=qr_code.address_latitude,
            longitude=qr_code.address_longitude,
            # Use QR code's address for both
            address=qr_code.location_address,
            # Set fixed distance
            location_accuracy=0.010,
            accuracy=0.010,
            # Mark as manual entry
            location_source='manual_entry',
            device_info='Manual Entry by Admin/Accounting',
            user_agent=f'Manual Entry - User: {session.get("username")}',
            ip_address=get_client_ip(),
            status='present',
            verification_required=False,
            verification_status='approved',
            created_timestamp=datetime.utcnow(),
            updated_timestamp=datetime.utcnow()
        )
        
        db.session.add(new_attendance)
        db.session.commit()
        
        # Log the manual entry
        logger_handler.logger.info(
            f"Manual attendance record created by {session.get('username')} ({user_role}): "
            f"Employee {employee.firstName} {employee.lastName} (ID: {employee_id}), "
            f"Location: {qr_code.location}, Event: {qr_code.location_event}, "
            f"Date: {check_date}, Time: {check_time}"
        )
        
        flash(f'Attendance record successfully created for {employee.firstName} {employee.lastName}.', 'success')
        return redirect(url_for('attendance_report'))
    
    except Exception as e:
        db.session.rollback()
        logger_handler.logger.error(f"Error saving manual attendance record: {e}")
        logger_handler.logger.error(f"Traceback: {traceback.format_exc()}")
        flash('Error saving attendance record. Please try again.', 'error')
        return redirect(url_for('add_manual_attendance'))


@app.route('/api/search_employees')
@login_required
def search_employees_api():
    """
    API endpoint to search employees by name or ID
    Returns JSON with employee list
    """
    try:
        search_query = request.args.get('q', '').strip()
        
        if not search_query or len(search_query) < 2:
            return jsonify({'employees': []})
        
        # Search by ID or name
        search_pattern = f"%{search_query}%"
        
        employees = Employee.query.filter(
            db.or_(
                Employee.id.like(search_pattern),
                Employee.firstName.like(search_pattern),
                Employee.lastName.like(search_pattern),
                db.func.concat(Employee.firstName, ' ', Employee.lastName).like(search_pattern)
            )
        ).limit(10).all()
        
        employee_list = [{
            'id': emp.id,
            'firstName': emp.firstName,
            'lastName': emp.lastName,
            'full_name': f"{emp.firstName} {emp.lastName}"
        } for emp in employees]
        
        return jsonify({'employees': employee_list})
    
    except Exception as e:
        logger_handler.logger.error(f"Error searching employees: {e}")
        return jsonify({'employees': [], 'error': str(e)}), 500


@app.route('/api/get_project_locations')
@login_required
def get_project_locations_api():
    """
    API endpoint to get locations for a specific project
    Returns JSON with location list
    """
    try:
        project_id = request.args.get('project_id', '').strip()
        
        if not project_id:
            return jsonify({'success': False, 'locations': [], 'error': 'Project ID required'})
        
        # Get active QR codes for this project
        qr_codes = QRCode.query.filter_by(
            project_id=int(project_id),
            active_status=True
        ).order_by(QRCode.location).all()
        
        # Group QR codes by location to get unique locations
        locations_dict = {}
        for qr in qr_codes:
            location_key = f"{qr.location}||{qr.location_address}"
            
            if location_key not in locations_dict:
                locations_dict[location_key] = {
                    'location': qr.location,
                    'location_address': qr.location_address,
                    'qr_codes': {}
                }
            
            # Store QR code ID for each event type
            locations_dict[location_key]['qr_codes'][qr.location_event] = qr.id

        # Convert to list format
        location_list = [{
            'location': loc_data['location'],
            'location_address': loc_data['location_address'],
            'qr_codes': loc_data['qr_codes']
        } for loc_data in locations_dict.values()]
        
        return jsonify({'success': True, 'locations': location_list})
    
    except Exception as e:
        logger_handler.logger.error(f"Error getting project locations: {e}")
        return jsonify({'success': False, 'locations': [], 'error': str(e)}), 500
    
@app.route('/attendance/<int:record_id>/delete', methods=['POST'])
@login_required
@log_database_operations('attendance_delete')
def delete_attendance(record_id):
    """Delete attendance record (Admin and Payroll only)"""
    # Check if user has permission to delete attendance records
    if session.get('role') not in ['admin', 'payroll', 'accounting']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': 'Access denied. Only administrators and payroll staff can delete attendance records.'
            }), 403
        else:
            flash('Access denied. Only administrators and payroll staff can delete attendance records.', 'error')
            return redirect(url_for('attendance_report'))

    try:
        attendance_record = AttendanceData.query.get_or_404(record_id)

        # Store record info for logging before deletion
        employee_id = attendance_record.employee_id
        location_name = attendance_record.location_name
        check_in_date = attendance_record.check_in_date

        # Log the deletion
        logger_handler.log_security_event(
            event_type="attendance_record_deletion",
            description=f"{session.get('role', 'unknown').title()} {session.get('username')} deleted attendance record {record_id}",
            severity="HIGH",
            additional_data={
                'record_id': record_id,
                'employee_id': employee_id,
                'location_name': location_name,
                'check_in_date': str(check_in_date),
                'user_role': session.get('role')
            }
        )

        # Delete the record
        db.session.delete(attendance_record)
        db.session.commit()

        print(f"[LOG] {session.get('role', 'unknown').title()} {session.get('username')} deleted attendance record {record_id} for employee {employee_id}")

        # Return JSON response for AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': f'Attendance record for {employee_id} deleted successfully!'
            })
        else:
            flash(f'Attendance record for {employee_id} deleted successfully!', 'success')
            return redirect(url_for('attendance_report'))

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('attendance_delete', e)
        print(f"[LOG] Error deleting attendance record {record_id}: {e}")

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': 'Error deleting attendance record. Please try again.'
            }), 500
        else:
            flash('Error deleting attendance record. Please try again.', 'error')
            return redirect(url_for('attendance_report'))
        
@app.route('/verification-review')
@login_required
def verification_review():
    """Admin page to review pending photo verifications"""
    try:
        # Only admins can access
        if session.get('role') not in ['admin', 'payroll', 'accounting']:
            flash('Unauthorized access.', 'error')
            return redirect(url_for('dashboard'))
        
        # Get filter parameters
        status_filter = request.args.get('status', 'pending')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        project_filter = request.args.get('project', '')
        location_filter = request.args.get('location', '')
        employee_filter = request.args.get('employee', '')
        
        # Build query - join with QRCode to access project_id
        query = AttendanceData.query.join(QRCode).filter(
            AttendanceData.verification_required == True
        )
        
        if status_filter and status_filter != 'all':
            query = query.filter(AttendanceData.verification_status == status_filter)
        
        if date_from:
            query = query.filter(AttendanceData.check_in_date >= date_from)
        
        if date_to:
            query = query.filter(AttendanceData.check_in_date <= date_to)
        
        # Apply project filter
        if project_filter:
            try:
                query = query.filter(QRCode.project_id == int(project_filter))
            except (ValueError, TypeError):
                pass
        
        # Apply location filter
        if location_filter:
            query = query.filter(AttendanceData.location_name.ilike(f'%{location_filter}%'))
        
        # Apply employee ID filter
        if employee_filter:
            query = query.filter(AttendanceData.employee_id.ilike(f'%{employee_filter}%'))
        
        # Get records with QR code information
        verifications = query.order_by(
            AttendanceData.verification_timestamp.desc()
        ).all()
        
        # Build a dictionary for employee names lookup
        employee_names = {}
        for record in verifications:
            if record.employee_id and record.employee_id not in employee_names:
                try:
                    employee = Employee.query.filter_by(id=int(record.employee_id)).first()
                    if employee:
                        employee_names[record.employee_id] = f"{employee.lastName}, {employee.firstName}"
                    else:
                        employee_names[record.employee_id] = None
                except (ValueError, TypeError):
                    employee_names[record.employee_id] = None
        
        # Build a dictionary for project names lookup
        project_names = {}
        for record in verifications:
            if record.qr_code and record.qr_code.project_id:
                project_id = record.qr_code.project_id
                if project_id not in project_names:
                    try:
                        project = Project.query.get(project_id)
                        if project:
                            project_names[project_id] = project.name
                        else:
                            project_names[project_id] = None
                    except Exception:
                        project_names[project_id] = None
        
        # Get counts for status badges
        pending_count = AttendanceData.query.filter(
            AttendanceData.verification_status == 'pending'
        ).count()
        
        approved_count = AttendanceData.query.filter(
            AttendanceData.verification_status == 'approved'
        ).count()
        
        rejected_count = AttendanceData.query.filter(
            AttendanceData.verification_status == 'rejected'
        ).count()
        
        # Get all projects for filter dropdown
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        
        # Get unique locations for filter dropdown
        locations = db.session.query(AttendanceData.location_name).filter(
            AttendanceData.verification_required == True
        ).distinct().order_by(AttendanceData.location_name).all()
        location_list = [loc[0] for loc in locations if loc[0]]
        
        # Log access
        logger_handler.logger.info(
            f"User {session.get('username')} ({session.get('role')}) accessed verification review page"
        )
        
        return render_template('verification_review.html',
                             verifications=verifications,
                             pending_count=pending_count,
                             approved_count=approved_count,
                             rejected_count=rejected_count,
                             status_filter=status_filter,
                             date_from=date_from,
                             date_to=date_to,
                             project_filter=project_filter,
                             location_filter=location_filter,
                             employee_filter=employee_filter,
                             projects=projects,
                             locations=location_list,
                             employee_names=employee_names,
                             project_names=project_names)
    
    except Exception as e:
        logger_handler.logger.error(f"Error in verification review: {e}")
        flash('Error loading verification review.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/verification-review/<int:record_id>/update', methods=['POST'])
@login_required
@log_database_operations('verification_update')
def update_verification_status(record_id):
    """Update verification status (approve/reject)"""
    try:
        # Only admins can update
        if session.get('role') not in ['admin', 'payroll', 'accounting']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403
        
        record = AttendanceData.query.get_or_404(record_id)
        
        new_status = request.json.get('status')
        admin_note = request.json.get('note', '')
        
        if new_status not in ['approved', 'rejected']:
            return jsonify({
                'success': False,
                'message': 'Invalid status'
            }), 400
        
        # Update record
        record.verification_status = new_status
        record.edit_note = f"Verification {new_status} by {session.get('username')}. {admin_note}"
        
        db.session.commit()
        
        # Log the action
        logger_handler.log_photo_verification(
            employee_id=record.employee_id,
            qr_code_id=record.qr_code_id,
            distance=record.location_accuracy or 0,
            status=new_status
        )
        
        return jsonify({
            'success': True,
            'message': f'Verification {new_status} successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        logger_handler.logger.error(f"Error updating verification: {e}")
        return jsonify({
            'success': False,
            'message': 'Error updating verification status'
        }), 500
    
@app.route('/api/attendance/<int:record_id>/verification-details')
@login_required
def get_verification_details(record_id):
    """API endpoint to get verification details for a specific record"""
    try:
        # Get the attendance record with verification data
        record = AttendanceData.query.get_or_404(record_id)
        
        # DEBUG: Log record details
        print(f"=== VERIFICATION DETAILS DEBUG ===")
        print(f"Record ID: {record.id}")
        print(f"Employee: {record.employee_id}")
        print(f"check_in_date type: {type(record.check_in_date)}")
        print(f"check_in_date value: {record.check_in_date}")
        print(f"check_in_time type: {type(record.check_in_time)}")
        print(f"check_in_time value: {record.check_in_time}")
        print(f"verification_photo exists: {record.verification_photo is not None}")
        print(f"verification_status: {record.verification_status}")
        print(f"==================================")
        
        # Check if user has permission to view
        # Allow admin and payroll staff to view verification details
        if session.get('role') not in ['admin', 'payroll', 'accounting']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403
        
        # Log the access for security audit
        logger_handler.logger.info(f"User {session.get('username')} ({session.get('role')}) accessed verification details for record {record_id}")
        
        # Safely format dates/times with error handling
        try:
            check_in_date_str = record.check_in_date.strftime('%Y-%m-%d') if record.check_in_date else 'N/A'
        except Exception as e:
            print(f"Error formatting check_in_date: {e}")
            check_in_date_str = str(record.check_in_date) if record.check_in_date else 'N/A'
        
        try:
            check_in_time_str = record.check_in_time.strftime('%I:%M %p') if record.check_in_time else 'N/A'
        except Exception as e:
            print(f"Error formatting check_in_time: {e}")
            check_in_time_str = str(record.check_in_time) if record.check_in_time else 'N/A'
        
        # Prepare record data with safe formatting
        try:
            check_in_date_str = record.check_in_date.strftime('%Y-%m-%d') if record.check_in_date else 'N/A'
        except:
            check_in_date_str = str(record.check_in_date) if record.check_in_date else 'N/A'
        
        try:
            check_in_time_str = record.check_in_time.strftime('%I:%M %p') if record.check_in_time else 'N/A'
        except:
            check_in_time_str = str(record.check_in_time) if record.check_in_time else 'N/A'
        
        record_data = {
            'id': record.id,
            'employee_id': record.employee_id,
            'location_name': record.location_name or 'Unknown',
            'check_in_date': check_in_date_str,
            'check_in_time': check_in_time_str,
            'location_accuracy': float(record.location_accuracy) if record.location_accuracy else None,
            'checked_in_address': record.address or 'No address',
            'verification_photo': record.verification_photo,
            'verification_status': record.verification_status,
            'verification_required': record.verification_required,
            'device_info': record.device_info or 'Unknown'
        }
        
        return jsonify({
            'success': True,
            'record': record_data
        })
    
    except Exception as e:
        logger_handler.logger.error(f"Error getting verification details for record {record_id}: {e}")
        print(f"❌ Error in get_verification_details for record {record_id}: {e}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")
        
        return jsonify({
            'success': False,
            'message': 'Error loading verification details'
        }), 500

@app.route('/verification-review/<int:record_id>')
@login_required
def verification_review_detail(record_id):
    """Review a single verification photo on a dedicated page"""
    try:
        # Check permissions
        if session.get('role') not in ['admin', 'payroll', 'accounting']:
            flash('Access denied. Only administrators, payroll, and accounting staff can review verification photos.', 'error')
            return redirect(url_for('attendance_report'))
        
        # Get the attendance record
        record = AttendanceData.query.get_or_404(record_id)
        
        # Check if this record has verification
        if not record.verification_required:
            flash('This record does not require verification.', 'warning')
            return redirect(url_for('attendance_report'))
        
        # Get the QR code information for additional context
        qr_code = QRCode.query.get(record.qr_code_id) if record.qr_code_id else None
        
        # Get employee name from Employee table
        employee_name = None
        try:
            if record.employee_id:
                employee = Employee.query.filter_by(id=int(record.employee_id)).first()
                if employee:
                    employee_name = f"{employee.lastName}, {employee.firstName}"
                else:
                    employee_name = f"Unknown (ID: {record.employee_id})"
        except (ValueError, TypeError) as e:
            logger_handler.logger.warning(f"Could not lookup employee name for ID {record.employee_id}: {e}")
            employee_name = f"Unknown (ID: {record.employee_id})"
        
        # Get event type from QR code (Check In/Check Out)
        location_event = qr_code.location_event if qr_code and qr_code.location_event else 'N/A'
        
        # Log the access for audit trail
        logger_handler.logger.info(
            f"User {session.get('username')} ({session.get('role')}) "
            f"accessed verification review for record {record_id}"
        )
        
        # Format date and time for display
        try:
            check_in_date = record.check_in_date.strftime('%m/%d/%Y') if record.check_in_date else 'N/A'
        except:
            check_in_date = str(record.check_in_date) if record.check_in_date else 'N/A'
        
        try:
            check_in_time = record.check_in_time.strftime('%I:%M %p') if record.check_in_time else 'N/A'
        except:
            check_in_time = str(record.check_in_time) if record.check_in_time else 'N/A'
        
        return render_template('verification_review_detail.html',
                             record=record,
                             qr_code=qr_code,
                             check_in_date=check_in_date,
                             check_in_time=check_in_time,
                             employee_name=employee_name,
                             location_event=location_event)
    
    except Exception as e:
        logger_handler.logger.error(f"Error loading verification review detail: {e}")
        flash('Error loading verification details.', 'error')
        return redirect(url_for('attendance_report'))

@app.route('/api/attendance/stats')
@admin_required
def attendance_stats_api():
    """API endpoint for attendance statistics"""
    try:
        # Daily stats for the last 7 days
        daily_stats = db.session.execute(text("""
            SELECT 
                check_in_date,
                COUNT(*) as checkins,
                COUNT(DISTINCT employee_id) as unique_employees
            FROM attendance_data 
            WHERE check_in_date >= CURRENT_DATE - INTERVAL '7 days'
            GROUP BY check_in_date
            ORDER BY check_in_date DESC
        """)).fetchall()

        # Location stats
        location_stats = db.session.execute(text("""
            SELECT 
                location_name,
                COUNT(*) as total_checkins,
                COUNT(DISTINCT employee_id) as unique_employees
            FROM attendance_data
            GROUP BY location_name
            ORDER BY total_checkins DESC
            LIMIT 10
        """)).fetchall()

        # Peak hours
        hourly_stats = db.session.execute(text("""
            SELECT 
                EXTRACT(hour FROM check_in_time) as hour,
                COUNT(*) as checkins
            FROM attendance_data
            WHERE check_in_date >= CURRENT_DATE - INTERVAL '30 days'
            GROUP BY EXTRACT(hour FROM check_in_time)
            ORDER BY hour
        """)).fetchall()

        return jsonify({
            'daily_stats': [{'date': str(row[0]), 'checkins': row[1], 'employees': row[2]} for row in daily_stats],
            'location_stats': [{'location': row[0], 'checkins': row[1], 'employees': row[2]} for row in location_stats],
            'hourly_stats': [{'hour': int(row[0]), 'checkins': row[1]} for row in hourly_stats]
        })

    except Exception as e:
        print(f"Error fetching attendance stats: {e}")
        return jsonify({'error': 'Failed to fetch attendance statistics'}), 500

@app.route('/export-configuration')
@login_required
def export_configuration():
    """Display export configuration page for customizing Excel exports"""
    try:
        user_role = session.get('role')
        if user_role not in ['admin', 'payroll', 'accounting']:
            logger_handler.logger.warning(f"User {session.get('username', 'unknown')} (role: {user_role}) attempted unauthorized access to export configuration")
            flash('Access denied. Only administrators and payroll staff can access export configuration.', 'error')
            return redirect(url_for('attendance_report'))

        # Log export configuration access using your existing logger
        try:
            logger_handler.logger.info(f"User {session.get('username', 'unknown')} (role: {user_role}) accessed export configuration")
            logger_handler.logger.info(f"User {session.get('username', 'unknown')} accessed export configuration page")
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")

        # Get current filters from session or request args
        filters = {
            'date_from': request.args.get('date_from', ''),
            'date_to': request.args.get('date_to', ''),
            'location_filter': request.args.get('location', ''),
            'employee_filter': request.args.get('employee', ''),
            'project_filter': request.args.get('project', '')
        }

        print(f"📊 Filters: {filters}")
        
        # Get project name if project filter is applied
        project_name = None
        if filters.get('project_filter'):
            try:
                from models.project import Project
                project = Project.query.get(int(filters['project_filter']))
                if project:
                    project_name = project.name
                    print(f"📊 Project filter: ID={filters['project_filter']}, Name={project_name}")
            except Exception as e:
                print(f"⚠️ Error fetching project name: {e}")

        # Check if location accuracy feature exists
        try:
            has_location_accuracy = check_location_accuracy_column_exists()
        except Exception as e:
            print(f"⚠️ Error checking location accuracy column: {e}")
            has_location_accuracy = False

        # Define all available columns with their default settings
        available_columns = [
            {'key': 'employee_id', 'label': 'Employee ID', 'default_name': 'ID', 'enabled': True},
            {'key': 'employee_name', 'label': 'Employee Name', 'default_name': 'Employee Name', 'enabled': False},
            {'key': 'location_name', 'label': 'Location', 'default_name': 'Location Name', 'enabled': True},
            {'key': 'status', 'label': 'Event', 'default_name': 'Action Description', 'enabled': True},
            {'key': 'check_in_date', 'label': 'Date', 'default_name': 'Date', 'enabled': True},
            {'key': 'check_in_time', 'label': 'Time', 'default_name': 'Time', 'enabled': True},
            {'key': 'qr_address', 'label': 'QR Address', 'default_name': 'Event Description', 'enabled': True},
            {'key': 'address', 'label': 'Check-in Address', 'default_name': 'Recorded Address', 'enabled': True},
            {'key': 'device_info', 'label': 'Device', 'default_name': 'Platform', 'enabled': True},
            {'key': 'ip_address', 'label': 'IP Address', 'default_name': 'IP Address', 'enabled': False},
            {'key': 'user_agent', 'label': 'User Agent', 'default_name': 'Browser/User Agent', 'enabled': False},
            {'key': 'latitude', 'label': 'Latitude', 'default_name': 'GPS Latitude', 'enabled': False},
            {'key': 'longitude', 'label': 'Longitude', 'default_name': 'GPS Longitude', 'enabled': False},
            {'key': 'accuracy', 'label': 'GPS Accuracy', 'default_name': 'GPS Accuracy (meters)', 'enabled': False},
        ]

        # Add location accuracy column if feature exists
        if has_location_accuracy:
            available_columns.append({
                'key': 'location_accuracy',
                'label': 'Location Accuracy',
                'default_name': 'Distance',
                'enabled': True  # Changed from False to True
            })

        print(f"📊 Rendering export configuration with {len(available_columns)} columns")

        return render_template('export_configuration.html',
                             available_columns=available_columns,
                             filters=filters,
                             project_name=project_name,
                             has_location_accuracy_feature=has_location_accuracy)

    except Exception as e:
        print(f"❌ Error in export_configuration route: {e}")
        print(f"❌ Traceback: {traceback.format_exc()}")

        # Use your existing logger error method with correct parameters
        try:
            logger_handler.log_flask_error(
                error_type="export_configuration_error",
                error_message=str(e),
                stack_trace=traceback.format_exc()
            )
        except Exception as log_error:
            print(f"⚠️ Could not log error: {log_error}")

        flash('Error loading export configuration page.', 'error')
        return redirect(url_for('attendance_report'))

@app.route('/generate-excel-export', methods=['POST'])
@login_required
def generate_excel_export():
    """Generate and download Excel file with selected columns in specified order"""
    try:
        user_role = session.get('role')
        if user_role not in ['admin', 'payroll', 'accounting']:
            logger_handler.logger.warning(f"User {session.get('username', 'unknown')} (role: {user_role}) attempted unauthorized Excel export")
            flash('Access denied. Only administrators and payroll staff can export data.', 'error')
            return redirect(url_for('attendance_report'))

        print("📊 Excel export generation started")

        # Log export action using your existing logger
        try:
            logger_handler.logger.info(f"User {session.get('username', 'unknown')} generated Excel export")
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")

        # Get selected columns and custom names from form
        selected_columns_raw = request.form.getlist('selected_columns')
        print(f"📊 Selected columns (raw): {selected_columns_raw}")

        # Get column order from form
        column_order_json = request.form.get('column_order', '[]')
        try:
            column_order = json.loads(column_order_json) if column_order_json else []
        except (json.JSONDecodeError, TypeError):
            column_order = []

        print(f"📊 Column order from form: {column_order}")

        # Determine final column order
        if column_order:
            # Use the specified order, but only include actually selected columns
            selected_columns = [col for col in column_order if col in selected_columns_raw]
            # Add any selected columns that weren't in the order (shouldn't happen, but safety check)
            for col in selected_columns_raw:
                if col not in selected_columns:
                    selected_columns.append(col)
        else:
            # Fallback to raw selection order
            selected_columns = selected_columns_raw

        print(f"📊 Final column order: {selected_columns}")

        if not selected_columns:
            flash('Please select at least one column to export.', 'error')
            return redirect(url_for('export_configuration'))

        column_names = {}
        for column in selected_columns:
            column_names[column] = request.form.get(f'name_{column}', column)

        # Get filters
        filters = {
            'date_from': request.form.get('date_from'),
            'date_to': request.form.get('date_to'),
            'location_filter': request.form.get('location_filter'),
            'employee_filter': request.form.get('employee_filter'),
            'project_filter': request.form.get('project_filter')
        }

        print(f"📊 Export filters: {filters}")
        print(f"📊 Column names: {column_names}")

        # Save user preferences in session for next time
        session['export_preferences'] = {
            'selected_columns': selected_columns,
            'column_names': column_names,
            'column_order': selected_columns  # This is now the ordered list
        }

        # Generate Excel file with ordered columns
        excel_file = create_excel_export_ordered(selected_columns, column_names, filters)

        if excel_file:
            # Get project name if project filter exists
            project_name_for_filename = ''
            if filters.get('project_filter'):
                try:
                    from models.project import Project
                    project = Project.query.get(int(filters['project_filter']))
                    if project:
                        # Replace spaces and special characters with underscores
                        project_name_safe = project.name.replace(' ', '_').replace('/', '_').replace('\\', '_')
                        project_name_for_filename = f"{project_name_safe}_"
                except Exception as e:
                    print(f"⚠️ Error getting project name for filename: {e}")
            
            # Format dates for filename (MMDDYYYY format)
            date_from_formatted = ''
            date_to_formatted = ''
            if filters.get('date_from'):
                try:
                    date_obj = datetime.strptime(filters['date_from'], '%Y-%m-%d')
                    date_from_formatted = date_obj.strftime('%m%d%Y')
                except ValueError:
                    pass
            
            if filters.get('date_to'):
                try:
                    date_obj = datetime.strptime(filters['date_to'], '%Y-%m-%d')
                    date_to_formatted = date_obj.strftime('%m%d%Y')
                except ValueError:
                    pass
            
            # Build filename components
            # Format: [project_name_]attendance_report_[fromdate_todate].xlsx
            date_range_str = ''
            if date_from_formatted and date_to_formatted:
                date_range_str = f"{date_from_formatted}_{date_to_formatted}"
            elif date_from_formatted:
                date_range_str = f"{date_from_formatted}"
            elif date_to_formatted:
                date_range_str = f"{date_to_formatted}"
            
            filename = f'{project_name_for_filename}attendance_report_{date_range_str}.xlsx'

            print(f"📊 Excel file generated successfully: {filename}")
            print(f"📊 Column order in export: {selected_columns}")

            # Log successful export using your existing logger
            try:
                logger_handler.logger.info(f"Excel export generated successfully with {len(selected_columns)} columns in custom order by user {session.get('username', 'unknown')}: {filename}")
            except Exception as log_error:
                print(f"⚠️ Logging error (non-critical): {log_error}")

            return send_file(
                excel_file,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('Error generating Excel file.', 'error')
            return redirect(url_for('export_configuration'))

    except Exception as e:
        print(f"❌ Error in generate_excel_export route: {e}")
        print(f"❌ Traceback: {traceback.format_exc()}")

        # Use your existing logger error method with correct parameters
        try:
            logger_handler.log_flask_error(
                error_type="excel_export_error",
                error_message=str(e),
                stack_trace=traceback.format_exc()
            )
        except Exception as log_error:
            print(f"⚠️ Could not log error: {log_error}")

        flash('Error generating Excel export.', 'error')
        return redirect(url_for('export_configuration'))

def create_excel_export(selected_columns, column_names, filters):
    """Create Excel file with selected attendance data - Updated to include employee names"""
    try:
        print(f"📊 Creating Excel export with {len(selected_columns)} columns")

        # Import openpyxl modules
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            print(f"❌ openpyxl import error: {e}")
            print("💡 Install openpyxl: pip install openpyxl")
            return None

        # Build query based on filters - JOIN with QRCode to get location_event and location_address
        # Now also JOIN with Employee table to get employee names
        query = db.session.query(AttendanceData, QRCode, Employee).join(
            QRCode, AttendanceData.qr_code_id == QRCode.id
        ).outerjoin(
            Employee, text("CAST(attendance_data.employee_id AS UNSIGNED) = employee.id")
        )

        # Apply date filters
        if filters.get('date_from'):
            try:
                date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
                query = query.filter(AttendanceData.check_in_date >= date_from)
                print(f"📊 Applied date_from filter: {date_from}")
            except ValueError as e:
                print(f"⚠️ Invalid date_from format: {e}")

        if filters.get('date_to'):
            try:
                date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()
                query = query.filter(AttendanceData.check_in_date <= date_to)
                print(f"📊 Applied date_to filter: {date_to}")
            except ValueError as e:
                print(f"⚠️ Invalid date_to format: {e}")

        # Apply location filter
        if filters.get('location_filter'):
            query = query.filter(AttendanceData.location_name.like(f"%{filters['location_filter']}%"))
            print(f"📊 Applied location filter: {filters['location_filter']}")

        # Apply employee filter
        if filters.get('employee_filter'):
            query = query.filter(AttendanceData.employee_id.like(f"%{filters['employee_filter']}%"))
            print(f"📊 Applied employee filter: {filters['employee_filter']}")

        # Apply project filter
        if filters.get('project_filter'):
            try:
                project_id = int(filters['project_filter'])
                query = query.filter(QRCode.project_id == project_id)
                print(f"📊 Applied project filter: {project_id}")
            except (ValueError, TypeError) as e:
                print(f"⚠️ Invalid project filter: {e}")

        # Order by date and time
        query = query.order_by(AttendanceData.check_in_date.desc(), AttendanceData.check_in_time.desc())

        # Execute query
        results = query.all()
        print(f"📊 Query returned {len(results)} records")

        if not results:
            print("⚠️ No records found for export")
            return None

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Set headers based on selected columns
        headers = []
        for column_key in selected_columns:
            header_name = column_names.get(column_key, column_key)
            headers.append(header_name)

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, (attendance_record, qr_record, employee_record) in enumerate(results, 2):
            for col_idx, column_key in enumerate(selected_columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                try:
                    # Handle each column type
                    if column_key == 'employee_id':
                        cell.value = format_employee_id_for_excel(attendance_record.employee_id)
                    elif column_key == 'employee_name':
                        # NEW: Handle employee name from joined Employee table
                        if employee_record:
                            cell.value = f"{employee_record.lastName}, {employee_record.firstName}"
                        else:
                            cell.value = f"Unknown (ID: {attendance_record.employee_id})"
                    elif column_key == 'location_name':
                        cell.value = attendance_record.location_name or ''
                    elif column_key == 'status':
                        cell.value = qr_record.location_event if qr_record.location_event else 'Check In'
                    elif column_key == 'check_in_date':
                        cell.value = attendance_record.check_in_date.strftime('%Y-%m-%d') if attendance_record.check_in_date else ''
                    elif column_key == 'check_in_time':
                        cell.value = attendance_record.check_in_time.strftime('%H:%M:%S') if attendance_record.check_in_time else ''
                    elif column_key == 'qr_address':
                        cell.value = qr_record.location_address if qr_record else ''
                    elif column_key == 'address':
                        # Check-in address logic based on location accuracy WITH HYPERLINKS
                        # If location accuracy < 0.3 miles, use QR address; otherwise use actual check-in address
                        if hasattr(attendance_record, 'location_accuracy') and attendance_record.location_accuracy is not None:
                            try:
                                accuracy_value = float(attendance_record.location_accuracy)
                                if accuracy_value < 0.3:
                                    # High accuracy - use QR code ADDRESS (not location) with hyperlink
                                    address_text = qr_record.location_address if qr_record and qr_record.location_address else ''
                                    if address_text and hasattr(qr_record, 'address_latitude') and hasattr(qr_record, 'address_longitude') and qr_record.address_latitude and qr_record.address_longitude:
                                        # Format coordinates with 10 decimal places
                                        lat_formatted = f"{float(qr_record.address_latitude):.10f}"
                                        lng_formatted = f"{float(qr_record.address_longitude):.10f}"
                                        hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                        cell.value = hyperlink_formula
                                        print(f"📍 Added QR address hyperlink for employee {attendance_record.employee_id}")
                                    else:
                                        cell.value = address_text
                                    print(f"📍 Using QR address for employee {attendance_record.employee_id} (accuracy: {accuracy_value:.3f} miles)")
                                else:
                                    # Lower accuracy - use actual check-in address with hyperlink
                                    address_text = attendance_record.address or ''
                                    if address_text and attendance_record.latitude and attendance_record.longitude:
                                        # Format coordinates with 10 decimal places
                                        lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                        lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                        hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                        cell.value = hyperlink_formula
                                        print(f"📍 Added check-in address hyperlink for employee {attendance_record.employee_id}")
                                    else:
                                        cell.value = address_text
                                    print(f"📍 Using check-in address for employee {attendance_record.employee_id} (accuracy: {accuracy_value:.3f} miles)")
                            except (ValueError, TypeError):
                                # If accuracy can't be converted to float, use check-in address with hyperlink
                                address_text = attendance_record.address or ''
                                if address_text and attendance_record.latitude and attendance_record.longitude:
                                    # Format coordinates with 10 decimal places
                                    lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                    lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                    hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                    cell.value = hyperlink_formula
                                    print(f"📍 Added check-in address hyperlink for employee {attendance_record.employee_id} (fallback)")
                                else:
                                    cell.value = address_text
                        else:
                            # No location accuracy data - use actual check-in address with hyperlink
                            address_text = attendance_record.address or ''
                            if address_text and attendance_record.latitude and attendance_record.longitude:
                                # Format coordinates with 10 decimal places
                                lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                cell.value = hyperlink_formula
                                print(f"📍 Added check-in address hyperlink for employee {attendance_record.employee_id} (no accuracy data)")
                            else:
                                cell.value = address_text
                    elif column_key == 'device_info':
                        cell.value = attendance_record.device_info or ''
                    elif column_key == 'ip_address':
                        cell.value = attendance_record.ip_address or ''
                    elif column_key == 'user_agent':
                        cell.value = attendance_record.user_agent or ''
                    elif column_key == 'latitude':
                       cell.value = attendance_record.latitude or ''
                    elif column_key == 'longitude':
                        cell.value = attendance_record.longitude or ''
                    elif column_key == 'accuracy':
                        cell.value = attendance_record.accuracy or ''
                    elif column_key == 'location_accuracy':
                        cell.value = attendance_record.location_accuracy or ''
                    else:
                        cell.value = ''
                except Exception as cell_error:
                    print(f"⚠️ Error setting cell value for {column_key}: {cell_error}")
                    cell.value = ''

        # Auto-adjust column widths based on content and header
        for col_idx, column_key in enumerate(selected_columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Get header name length
            header_name = column_names.get(column_key, column_key)
            max_length = len(str(header_name))
            
            # Check content in all rows (sample first 100 rows for performance)
            for row_idx in range(2, min(102, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    cell_value = str(cell.value) if cell.value else ''
                    # For HYPERLINK formulas, extract the display text
                    if cell_value.startswith('=HYPERLINK'):
                        # Extract text between last quotes: HYPERLINK("url","display_text")
                        import re
                        match = re.search(r',"([^"]+)"\)$', cell_value)
                        if match:
                            cell_value = match.group(1)
                    
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            # Set width based on column type with reasonable limits
            # Define optimal widths for specific column types
            column_width_rules = {
                'employee_id': {'min': 8, 'max': 15},
                'employee_name': {'min': 20, 'max': 30},
                'location_name': {'min': 15, 'max': 35},
                'status': {'min': 12, 'max': 20},
                'check_in_date': {'min': 12, 'max': 15},
                'check_in_time': {'min': 10, 'max': 12},
                'qr_address': {'min': 20, 'max': 40},
                'address': {'min': 20, 'max': 45},
                'device_info': {'min': 12, 'max': 20},
                'ip_address': {'min': 14, 'max': 18},
                'user_agent': {'min': 15, 'max': 30},
                'latitude': {'min': 12, 'max': 15},
                'longitude': {'min': 12, 'max': 15},
                'accuracy': {'min': 10, 'max': 15},
                'location_accuracy': {'min': 10, 'max': 15}
            }
            
            # Get rules for this column or use defaults
            rules = column_width_rules.get(column_key, {'min': 10, 'max': 40})
            
            # Calculate adjusted width: add 2 for padding, respect min/max
            adjusted_width = max_length + 2
            adjusted_width = max(rules['min'], min(adjusted_width, rules['max']))
            
            ws.column_dimensions[column_letter].width = adjusted_width
            
            print(f"📏 Column {column_letter} ({column_key}): set width to {adjusted_width} (content: {max_length} chars)")

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        print("📊 Excel file created successfully with employee names")
        
        # Log export action with employee name column
        try:
            logger_handler.logger.info(f"Excel export with employee names generated by user {session.get('username', 'unknown')}")
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")
            
        return excel_buffer

    except Exception as e:
        print(f"❌ Error creating Excel export: {e}")
        print(f"❌ Traceback: {traceback.format_exc()}")
        
        # Log error
        try:
            logger_handler.log_flask_error(
                error_type="excel_export_error",
                error_message=str(e),
                stack_trace=traceback.format_exc()
            )
        except Exception as log_error:
            print(f"⚠️ Could not log error: {log_error}")
            
        return None

def format_employee_id_for_excel(employee_id):
    if not employee_id:
        return ''
    emp_id_str = str(employee_id).strip()
    if emp_id_str.isdigit():
        return int(emp_id_str)
    else:
        return emp_id_str
    
def create_excel_export_ordered(selected_columns, column_names, filters):
    """Create Excel file with selected attendance data in specified column order"""
    try:
        print(f"📊 Creating Excel export with {len(selected_columns)} columns in order: {selected_columns}")

        # Import openpyxl modules
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            print(f"❌ openpyxl import error: {e}")
            print("💡 Install openpyxl: pip install openpyxl")
            return None

        # Build query based on filters - JOIN with QRCode to get location_event and location_address
        # Now also JOIN with Employee table to get employee names
        query = db.session.query(AttendanceData, QRCode, Employee).join(
            QRCode, AttendanceData.qr_code_id == QRCode.id
        ).outerjoin(
            Employee, text("CAST(attendance_data.employee_id AS UNSIGNED) = employee.id")
        )

        # Apply date filters
        if filters.get('date_from'):
            try:
                date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
                query = query.filter(AttendanceData.check_in_date >= date_from)
                print(f"📊 Applied date_from filter: {date_from}")
            except ValueError as e:
                print(f"⚠️ Invalid date_from format: {e}")

        if filters.get('date_to'):
            try:
                date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()
                query = query.filter(AttendanceData.check_in_date <= date_to)
                print(f"📊 Applied date_to filter: {date_to}")
            except ValueError as e:
                print(f"⚠️ Invalid date_to format: {e}")

        # Apply location filter
        if filters.get('location_filter'):
            query = query.filter(AttendanceData.location_name.like(f"%{filters['location_filter']}%"))
            print(f"📊 Applied location filter: {filters['location_filter']}")

        # Apply employee filter
        if filters.get('employee_filter'):
            query = query.filter(AttendanceData.employee_id.like(f"%{filters['employee_filter']}%"))
            print(f"📊 Applied employee filter: {filters['employee_filter']}")

        # Apply project filter
        if filters.get('project_filter'):
            try:
                project_id = int(filters['project_filter'])
                query = query.filter(QRCode.project_id == project_id)
                print(f"📊 Applied project filter: {project_id}")
            except (ValueError, TypeError) as e:
                print(f"⚠️ Invalid project filter: {e}")

        # Order by date and time
        query = query.order_by(AttendanceData.check_in_date.desc(), AttendanceData.check_in_time.desc())

        # Execute query
        results = query.all()
        print(f"📊 Query returned {len(results)} records")

        if not results:
            print("⚠️ No records found for export")
            return None

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Verification status color fills for location_accuracy column
        # Yellow for pending, Green for approved, Red for rejected
        verification_fill_pending = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        verification_fill_approved = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
        verification_fill_rejected = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Light Red

        # Set headers based on selected columns in the specified order
        headers = []
        for column_key in selected_columns:
            header_name = column_names.get(column_key, column_key)
            headers.append(header_name)

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, (attendance_record, qr_record, employee_record) in enumerate(results, 2):
            for col_idx, column_key in enumerate(selected_columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                try:
                    # Handle each column type
                    if column_key == 'employee_id':
                        cell.value = format_employee_id_for_excel(attendance_record.employee_id)
                    elif column_key == 'employee_name':
                        # NEW: Handle employee name from joined Employee table
                        if employee_record:
                            cell.value = f"{employee_record.lastName}, {employee_record.firstName}"
                        else:
                            cell.value = f"Unknown (ID: {attendance_record.employee_id})"
                    elif column_key == 'location_name':
                        cell.value = attendance_record.location_name or ''
                    elif column_key == 'status':
                        cell.value = qr_record.location_event if qr_record.location_event else 'Check In'
                    elif column_key == 'check_in_date':
                        cell.value = attendance_record.check_in_date.strftime('%Y-%m-%d') if attendance_record.check_in_date else ''
                    elif column_key == 'check_in_time':
                        cell.value = attendance_record.check_in_time.strftime('%H:%M:%S') if attendance_record.check_in_time else ''
                    elif column_key == 'qr_address':
                        cell.value = qr_record.location_address if qr_record else ''
                    elif column_key == 'address':
                        # Check-in address logic based on location accuracy WITH HYPERLINKS
                        # If location accuracy < 0.3 miles, use QR address; otherwise use actual check-in address
                        if hasattr(attendance_record, 'location_accuracy') and attendance_record.location_accuracy is not None:
                            try:
                                accuracy_value = float(attendance_record.location_accuracy)
                                if accuracy_value < 0.3:
                                    # High accuracy - use QR code ADDRESS (not location) with hyperlink
                                    address_text = qr_record.location_address if qr_record and qr_record.location_address else ''
                                    if address_text and hasattr(qr_record, 'address_latitude') and hasattr(qr_record, 'address_longitude') and qr_record.address_latitude and qr_record.address_longitude:
                                        # Format coordinates with 10 decimal places
                                        lat_formatted = f"{float(qr_record.address_latitude):.10f}"
                                        lng_formatted = f"{float(qr_record.address_longitude):.10f}"
                                        hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                        cell.value = hyperlink_formula
                                        print(f"📍 Added QR address hyperlink for employee {attendance_record.employee_id}")
                                    else:
                                        cell.value = address_text
                                    print(f"📍 Using QR address for employee {attendance_record.employee_id} (accuracy: {accuracy_value:.3f} miles)")
                                else:
                                    # Lower accuracy - use actual check-in address with hyperlink
                                    address_text = attendance_record.address or ''
                                    if address_text and attendance_record.latitude and attendance_record.longitude:
                                        # Format coordinates with 10 decimal places
                                        lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                        lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                        hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                        cell.value = hyperlink_formula
                                        print(f"📍 Added check-in address hyperlink for employee {attendance_record.employee_id}")
                                    else:
                                        cell.value = address_text
                                    print(f"📍 Using check-in address for employee {attendance_record.employee_id} (accuracy: {accuracy_value:.3f} miles)")
                            except (ValueError, TypeError):
                                # If accuracy can't be converted to float, use check-in address with hyperlink
                                address_text = attendance_record.address or ''
                                if address_text and attendance_record.latitude and attendance_record.longitude:
                                    # Format coordinates with 10 decimal places
                                    lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                    lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                    hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                    cell.value = hyperlink_formula
                                    print(f"📍 Added check-in address hyperlink for employee {attendance_record.employee_id} (fallback)")
                                else:
                                    cell.value = address_text
                        else:
                            # No location accuracy data - use actual check-in address with hyperlink
                            address_text = attendance_record.address or ''
                            if address_text and attendance_record.latitude and attendance_record.longitude:
                                # Format coordinates with 10 decimal places
                                lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                cell.value = hyperlink_formula
                                print(f"📍 Added check-in address hyperlink for employee {attendance_record.employee_id} (no accuracy data)")
                            else:
                                cell.value = address_text
                    elif column_key == 'device_info':
                        cell.value = attendance_record.device_info or ''
                    elif column_key == 'ip_address':
                        cell.value = attendance_record.ip_address or ''
                    elif column_key == 'user_agent':
                        cell.value = attendance_record.user_agent or ''
                    elif column_key == 'latitude':
                       cell.value = attendance_record.latitude or ''
                    elif column_key == 'longitude':
                        cell.value = attendance_record.longitude or ''
                    elif column_key == 'accuracy':
                        cell.value = attendance_record.accuracy or ''
                    elif column_key == 'location_accuracy':
                        cell.value = attendance_record.location_accuracy or ''
                        # Apply color fill based on verification_status
                        # Only apply color if verification_status is not NULL
                        if hasattr(attendance_record, 'verification_status') and attendance_record.verification_status:
                            if attendance_record.verification_status == 'pending':
                                cell.fill = verification_fill_pending  # Yellow
                            elif attendance_record.verification_status == 'approved':
                                cell.fill = verification_fill_approved  # Green
                            elif attendance_record.verification_status == 'rejected':
                                cell.fill = verification_fill_rejected  # Red
                    else:
                        cell.value = ''
                except Exception as cell_error:
                    print(f"⚠️ Error setting cell value for {column_key}: {cell_error}")
                    cell.value = ''

        # Auto-adjust column widths based on content and header
        for col_idx, column_key in enumerate(selected_columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Get header name length
            header_name = column_names.get(column_key, column_key)
            max_length = len(str(header_name))
            
            # Check content in all rows (sample first 100 rows for performance)
            for row_idx in range(2, min(102, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    cell_value = str(cell.value) if cell.value else ''
                    # For HYPERLINK formulas, extract the display text
                    if cell_value.startswith('=HYPERLINK'):
                        # Extract text between last quotes: HYPERLINK("url","display_text")
                        import re
                        match = re.search(r',"([^"]+)"\)$', cell_value)
                        if match:
                            cell_value = match.group(1)
                    
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            # Set width based on column type with reasonable limits
            # Define optimal widths for specific column types
            column_width_rules = {
                'employee_id': {'min': 8, 'max': 15},
                'employee_name': {'min': 20, 'max': 30},
                'location_name': {'min': 15, 'max': 35},
                'status': {'min': 12, 'max': 20},
                'check_in_date': {'min': 12, 'max': 15},
                'check_in_time': {'min': 10, 'max': 12},
                'qr_address': {'min': 20, 'max': 40},
                'address': {'min': 20, 'max': 45},
                'device_info': {'min': 12, 'max': 20},
                'ip_address': {'min': 14, 'max': 18},
                'user_agent': {'min': 15, 'max': 30},
                'latitude': {'min': 12, 'max': 15},
                'longitude': {'min': 12, 'max': 15},
                'accuracy': {'min': 10, 'max': 15},
                'location_accuracy': {'min': 10, 'max': 15}
            }
            
            # Get rules for this column or use defaults
            rules = column_width_rules.get(column_key, {'min': 10, 'max': 40})
            
            # Calculate adjusted width: add 2 for padding, respect min/max
            adjusted_width = max_length + 2
            adjusted_width = max(rules['min'], min(adjusted_width, rules['max']))
            
            ws.column_dimensions[column_letter].width = adjusted_width
            
            print(f"📏 Column {column_letter} ({column_key}): set width to {adjusted_width} (content: {max_length} chars)")

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        print("📊 Excel file created successfully with employee names and verification status coloring")
        
        # Log export action with employee name column and verification status coloring
        try:
            logger_handler.logger.info(f"Excel export with employee names and verification status coloring generated by user {session.get('username', 'unknown')}")
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")
            
        return excel_buffer

    except Exception as e:
        print(f"❌ Error creating Excel export: {e}")
        print(f"❌ Traceback: {traceback.format_exc()}")
        
        # Log error
        try:
            logger_handler.log_flask_error(
                error_type="excel_export_ordered_error",
                error_message=str(e),
                stack_trace=traceback.format_exc()
            )
        except Exception as log_error:
            print(f"⚠️ Could not log error: {log_error}")
            
        return None

@app.route('/payroll')
@login_required
def payroll_dashboard():
    """Payroll dashboard for calculating and exporting working hours"""
    try:
        # Check if user has payroll access
        user_role = session.get('role')
        if user_role not in ['admin', 'payroll', 'accounting']:
            logger_handler.logger.warning(f"User {session.get('username', 'unknown')} (role: {user_role}) attempted to access payroll dashboard without permissions")
            flash('Access denied. Only administrators and payroll staff can access payroll features.', 'error')
            return redirect(url_for('dashboard'))

        print("📊 Loading payroll dashboard")

        # Log payroll dashboard access
        logger_handler.logger.info(f"User {session.get('username', 'unknown')} accessed payroll dashboard")

        # Get filter parameters with defaults
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        project_filter = request.args.get('project_filter', '')
        
        # Set default date range if not provided (last 2 weeks)
        if not date_from or not date_to:
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=13)  # 2 weeks (14 days)
            date_from = start_date.strftime('%Y-%m-%d')
            date_to = end_date.strftime('%Y-%m-%d')

        # Get list of projects for dropdown
        projects = []
        try:
            projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
            print(f"📊 Found {len(projects)} active projects for filter")
        except Exception as e:
            print(f"⚠️ Error loading projects: {e}")

        # Get attendance records for the period
        attendance_records = []
        working_hours_data = None

        if date_from and date_to:
            try:
                start_date = datetime.strptime(date_from, '%Y-%m-%d')
                end_date = datetime.strptime(date_to, '%Y-%m-%d')

                # Query attendance records with optional project filter
                query = db.session.query(AttendanceData).join(QRCode, AttendanceData.qr_code_id == QRCode.id)

                # Apply date filter
                query = query.filter(
                    AttendanceData.check_in_date >= start_date.date(),
                    AttendanceData.check_in_date <= end_date.date()
                )

                # Apply project filter if selected
                if project_filter and project_filter != '':
                    query = query.filter(QRCode.project_id == int(project_filter))
                    print(f"📊 Applied project filter: {project_filter}")

                query = query.order_by(AttendanceData.employee_id, AttendanceData.check_in_date, AttendanceData.check_in_time)

                attendance_records = query.all()
                print(f"📊 Found {len(attendance_records)} attendance records for payroll calculation")

                # Calculate working hours if we have records
                if attendance_records:
                    calculator = WorkingHoursCalculator()
                    working_hours_data = calculator.calculate_all_employees_hours(
                        start_date, end_date, attendance_records
                    )
                    print(f"📊 Calculated hours for {working_hours_data['employee_count']} employees")

            except ValueError as e:
                print(f"⚠️ Invalid date format: {e}")
                flash('Invalid date format. Please use YYYY-MM-DD format.', 'error')
            except Exception as e:
                print(f"❌ Error calculating working hours: {e}")
                logger_handler.log_database_error('payroll_calculation', e)
                flash('Error calculating working hours. Please check the server logs.', 'error')

        # Get employee names for display
        employee_names = {}
        if working_hours_data:
            try:
                # Use the same SQL approach as attendance report - JOIN with CAST
                employee_ids = list(working_hours_data['employees'].keys())
                if employee_ids:
                    # Build a query similar to attendance report
                    placeholders = ','.join([f"'{emp_id}'" for emp_id in employee_ids])
                    employee_query = db.session.execute(text(f"""
                        SELECT 
                            ad.employee_id,
                            CONCAT(e.lastName, ',', e.firstName) as full_name 
                        FROM attendance_data ad
                        LEFT JOIN employee e ON CAST(ad.employee_id AS UNSIGNED) = e.id
                        WHERE ad.employee_id IN ({placeholders})
                        GROUP BY ad.employee_id, e.firstName, e.lastName
                    """))

                    for row in employee_query:
                        if row[1]:  # Only add if we got a name
                            employee_names[str(row[0])] = row[1]

                print(f"📊 Retrieved names for {len(employee_names)} employees using CAST method")

            except Exception as e:
                print(f"⚠️ Could not load employee names: {e}")
                import traceback
                print(f"⚠️ Traceback: {traceback.format_exc()}")
                # Continue without names - will use employee IDs

        # Get selected project name for display
        selected_project_name = ''
        if project_filter:
            try:
                selected_project = Project.query.get(int(project_filter))
                if selected_project:
                    selected_project_name = selected_project.name
            except Exception as e:
                print(f"⚠️ Error getting selected project name: {e}")

        return render_template('payroll_dashboard.html',
                             working_hours_data=working_hours_data,
                             employee_names=employee_names,
                             projects=projects,
                             date_from=date_from,
                             date_to=date_to,
                             project_filter=project_filter,
                             selected_project_name=selected_project_name,
                             user_role=user_role)

    except Exception as e:
        print(f"❌ Error loading payroll dashboard: {e}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")

        logger_handler.log_flask_error(
            error_type="payroll_dashboard_error",
            error_message=str(e),
            stack_trace=traceback.format_exc()
        )

        flash('Error loading payroll dashboard. Please check the server logs.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/payroll/export-excel', methods=['POST'])
@login_required
@log_database_operations('payroll_excel_export')
def export_payroll_excel():
    """Export payroll report to Excel with working hours calculations including SP/PW support"""
    try:
        # Check permissions
        user_role = session.get('role')
        if user_role not in ['admin', 'payroll', 'accounting']:
            logger_handler.logger.warning(f"User {session.get('username', 'unknown')} (role: {user_role}) attempted unauthorized payroll Excel export")
            flash('Access denied. Only administrators and payroll staff can export payroll data.', 'error')
            return redirect(url_for('payroll_dashboard'))

        print("📊 Payroll Excel export started")

        # Get parameters from form
        date_from = request.form.get('date_from')
        date_to = request.form.get('date_to')
        project_filter = request.form.get('project_filter', '')
        report_type = request.form.get('report_type', 'payroll')  # 'payroll', 'detailed', 'template', 'enhanced', 'detailed_sp_pw'

        if not date_from or not date_to:
            flash('Please provide both start and end dates for the export.', 'error')
            return redirect(url_for('payroll_dashboard'))

        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
        except ValueError:
            flash('Invalid date format. Please use YYYY-MM-DD format.', 'error')
            return redirect(url_for('payroll_dashboard'))

        # Get attendance records with project filter and QR code data
        query = db.session.query(AttendanceData, QRCode).join(QRCode, AttendanceData.qr_code_id == QRCode.id)

        # Apply date filter
        query = query.filter(
            AttendanceData.check_in_date >= start_date.date(),
            AttendanceData.check_in_date <= end_date.date()
        )

        # Apply project filter if selected
        if project_filter and project_filter != '':
            query = query.filter(QRCode.project_id == int(project_filter))
            print(f"📊 Applied project filter to export: {project_filter}")

        query = query.order_by(AttendanceData.employee_id, AttendanceData.check_in_date, AttendanceData.check_in_time)

        # Get the results and attach QR code data to attendance records
        query_results = query.all()
        attendance_records = []

        for attendance_data, qr_code in query_results:
            # Attach the QR code object to the attendance record
            attendance_data.qr_code = qr_code
            attendance_records.append(attendance_data)

        print(f"📊 Export: Found {len(attendance_records)} records with QR data")

        if not attendance_records:
            flash('No attendance records found for the selected date range and project.', 'warning')
            return redirect(url_for('payroll_dashboard'))

        print(f"📊 Exporting {len(attendance_records)} attendance records to Excel")

        # Get employee names using the same method as dashboard
        employee_names = {}
        try:
            employee_ids = list(set(str(record.employee_id) for record in attendance_records))
            if employee_ids:
                # Use the same SQL approach as attendance report - JOIN with CAST
                placeholders = ','.join([f"'{emp_id}'" for emp_id in employee_ids])
                employee_query = db.session.execute(text(f"""
                    SELECT 
                        ad.employee_id,
                        CONCAT(e.firstName, ' ', e.lastName) as full_name 
                    FROM attendance_data ad
                    LEFT JOIN employee e ON CAST(ad.employee_id AS UNSIGNED) = e.id
                    WHERE ad.employee_id IN ({placeholders})
                    GROUP BY ad.employee_id, e.firstName, e.lastName
                """))

                for row in employee_query:
                    if row[1]:  # Only add if we got a name
                        employee_names[str(row[0])] = row[1]

            print(f"📊 Retrieved names for {len(employee_names)} employees for export using CAST method")

        except Exception as e:
            print(f"⚠️ Could not load employee names for export: {e}")
            import traceback
            print(f"⚠️ Traceback: {traceback.format_exc()}")

        # Get project name for enhanced reports and filename
        project_name = None
        project_name_for_filename = ''
        if project_filter:
            try:
                project = Project.query.get(int(project_filter))
                if project:
                    project_name = project.name
                    project_name_for_filename = f"_{project.name.replace(' ', '_')}"
            except Exception as e:
                print(f"⚠️ Error getting project name: {e}")

        # Generate Excel file based on report type
        excel_file = None
        filename_prefix = 'payroll_report'

        if report_type == 'enhanced':
            # Use enhanced exporter for SP/PW reports
            print("📊 Creating enhanced payroll report with SP/PW support")
            try:
                from enhanced_payroll_excel_exporter import EnhancedPayrollExcelExporter
                exporter = EnhancedPayrollExcelExporter(company_name=os.environ.get('COMPANY_NAME', 'Your Company'))
                excel_file = exporter.create_enhanced_payroll_report(
                    start_date, end_date, attendance_records, employee_names, project_name
                )
                filename_prefix = 'enhanced_payroll_report'
                print("✅ Enhanced payroll report created successfully")
            except ImportError:
                print("⚠️ Enhanced exporter not available, falling back to standard exporter")
                # Fall back to standard exporter
                exporter = PayrollExcelExporter(
                    company_name=os.environ.get('COMPANY_NAME', 'Your Company'),
                    contract_name=os.environ.get('CONTRACT_NAME', 'Default Contract')
                )
                excel_file = exporter.create_payroll_report(
                    start_date, end_date, attendance_records, employee_names
                )
                filename_prefix = 'payroll_report'
            except Exception as e:
                print(f"⚠️ Error with enhanced exporter: {e}, falling back to standard exporter")
                # Fall back to standard exporter
                exporter = PayrollExcelExporter(
                    company_name=os.environ.get('COMPANY_NAME', 'Your Company'),
                    contract_name=os.environ.get('CONTRACT_NAME', 'Default Contract')
                )
                excel_file = exporter.create_payroll_report(
                    start_date, end_date, attendance_records, employee_names
                )
                filename_prefix = 'payroll_report'

        elif report_type == 'detailed_sp_pw':
            # Detailed daily SP/PW breakdown
            print("📊 Creating detailed SP/PW daily breakdown report")
            try:
                from enhanced_payroll_excel_exporter import EnhancedPayrollExcelExporter
                exporter = EnhancedPayrollExcelExporter(company_name=os.environ.get('COMPANY_NAME', 'Your Company'))
                excel_file = exporter.create_detailed_sp_pw_report(
                    start_date, end_date, attendance_records, employee_names
                )
                filename_prefix = 'detailed_sp_pw_report'
                print("✅ Detailed SP/PW report created successfully")
            except ImportError:
                print("⚠️ Enhanced exporter not available, falling back to detailed hours report")
                # Fall back to standard detailed report
                exporter = PayrollExcelExporter(
                    company_name=os.environ.get('COMPANY_NAME', 'Your Company'),
                    contract_name=os.environ.get('CONTRACT_NAME', 'Default Contract')
                )
                excel_file = exporter.create_detailed_hours_report(
                    start_date, end_date, attendance_records, employee_names
                )
                filename_prefix = 'detailed_hours_report'
            except Exception as e:
                print(f"⚠️ Error with enhanced exporter: {e}, falling back to detailed hours report")
                # Fall back to standard detailed report
                exporter = PayrollExcelExporter(
                    company_name=os.environ.get('COMPANY_NAME', 'Your Company'),
                    contract_name=os.environ.get('CONTRACT_NAME', 'Default Contract')
                )
                excel_file = exporter.create_detailed_hours_report(
                    start_date, end_date, attendance_records, employee_names
                )
                filename_prefix = 'detailed_hours_report'

        else:
            # Use standard exporter for existing report types
            exporter = PayrollExcelExporter(
                company_name=os.environ.get('COMPANY_NAME', 'Your Company'),
                contract_name=os.environ.get('CONTRACT_NAME', 'Default Contract')
            )

            if report_type == 'detailed':
                excel_file = exporter.create_detailed_hours_report(
                    start_date, end_date, attendance_records, employee_names
                )
                filename_prefix = 'detailed_hours_report'
            elif report_type == 'template':
                excel_file = exporter.create_template_format_report(
                    start_date, end_date, attendance_records, employee_names, project_name
                )
                filename_prefix = 'time_attendance_report'
            else:
                # Default payroll report
                excel_file = exporter.create_payroll_report(
                    start_date, end_date, attendance_records, employee_names
                )
                filename_prefix = 'payroll_report'

        if excel_file:
            # Generate filename with timestamp and project name
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{filename_prefix}_{date_from}_to_{date_to}{project_name_for_filename}_{timestamp}.xlsx'

            print(f"📊 Payroll Excel file generated successfully: {filename}")

            # Log successful export
            logger_handler.logger.info(f"Payroll Excel export generated by user {session.get('username', 'unknown')}: {filename}")
            if report_type == 'template':
                logger_handler.logger.info(f"Template format hours export generated by user {session.get('username', 'unknown')}: {filename}")
            elif report_type == 'enhanced':
                logger_handler.logger.info(f"Enhanced payroll export with SP/PW generated by user {session.get('username', 'unknown')}: {filename}")
            elif report_type == 'detailed_sp_pw':
                logger_handler.logger.info(f"Detailed SP/PW breakdown export generated by user {session.get('username', 'unknown')}: {filename}")

            return send_file(
                excel_file,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('Error generating payroll Excel file.', 'error')
            return redirect(url_for('payroll_dashboard'))

    except Exception as e:
        print(f"❌ Error in export_payroll_excel route: {e}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")

        logger_handler.log_flask_error(
            error_type="payroll_excel_export_error",
            error_message=str(e),
            stack_trace=traceback.format_exc()
        )

        flash('Error generating payroll Excel export. Please check the server logs.', 'error')
        return redirect(url_for('payroll_dashboard'))

@app.route('/api/working-hours/calculate', methods=['POST'])
@login_required
@log_database_operations('working_hours_api_calculation')
def calculate_working_hours_api():
    """API endpoint for calculating working hours"""
    try:
        # Check permissions
        user_role = session.get('role')
        if user_role not in ['admin', 'payroll', 'accounting']:
            return jsonify({
                'success': False,
                'message': 'Access denied. Insufficient permissions.'
            }), 403

        # Get parameters from JSON request
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'message': 'No data provided'
            }), 400

        employee_id = data.get('employee_id')
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        if not all([employee_id, date_from, date_to]):
            return jsonify({
                'success': False,
                'message': 'Missing required parameters: employee_id, date_from, date_to'
            }), 400

        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
        except ValueError:
            return jsonify({
                'success': False,
                'message': 'Invalid date format. Use YYYY-MM-DD.'
            }), 400

        # Get attendance records for the employee
        query = db.session.query(AttendanceData).filter(
            AttendanceData.employee_id == str(employee_id),
            AttendanceData.check_in_date >= start_date.date(),
            AttendanceData.check_in_date <= end_date.date()
        ).order_by(AttendanceData.check_in_date, AttendanceData.check_in_time)

        attendance_records = query.all()

        # Calculate working hours using single check-in calculator
        calculator = SingleCheckInCalculator()
        hours_data = calculator.calculate_employee_hours(
            str(employee_id), start_date, end_date, attendance_records
        )

        # Log API usage
        logger_handler.logger.info(f"Working hours API used by {session.get('username', 'unknown')} for employee {employee_id}")

        return jsonify({
            'success': True,
            'data': hours_data
        })

    except Exception as e:
        print(f"❌ Error in calculate_working_hours_api: {e}")
        logger_handler.log_flask_error(
            error_type="working_hours_api_error",
            error_message=str(e),
            stack_trace=traceback.format_exc()
        )

        return jsonify({
            'success': False,
            'message': 'Internal server error. Please check the server logs.'
        }), 500

@app.route('/api/employee/<employee_id>/miss-punch-details', methods=['GET'])
@login_required
@log_database_operations('miss_punch_details_api')
def get_miss_punch_details(employee_id):
    """API endpoint to get detailed miss punch information for an employee"""
    try:
        # Check permissions
        user_role = session.get('role')
        if user_role not in ['admin', 'payroll', 'accounting']:
            return jsonify({
                'success': False,
                'message': 'Access denied. Insufficient permissions.'
            }), 403

        # Get date parameters from query string (from the current payroll filters)
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        project_filter = request.args.get('project_filter', '')
        
        if not all([date_from, date_to]):
            return jsonify({
                'success': False,
                'message': 'Missing required parameters: date_from, date_to'
            }), 400

        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
        except ValueError:
            return jsonify({
                'success': False,
                'message': 'Invalid date format. Use YYYY-MM-DD.'
            }), 400

        # Get employee name using proper firstName and lastName fields
        try:
            employee_query = db.session.execute(text("""
                                                     SELECT e.id,
                                                            CONCAT(e.firstName, ' ', e.lastName) as full_name
                                                     FROM employee e
                                                     WHERE e.id = :emp_id
                                                     """), {'emp_id': int(employee_id)})

            employee_row = employee_query.fetchone()
            employee_name = employee_row.full_name if employee_row and employee_row.full_name else f"Employee {employee_id}"
            print(f"📋 Retrieved employee name: {employee_name} for ID: {employee_id}")
        except Exception as e:
            print(f"⚠️ Could not load employee name for ID {employee_id}: {e}")
            import traceback
            print(f"⚠️ Traceback: {traceback.format_exc()}")
            employee_name = f"Employee {employee_id}"

        # Get attendance records for the employee within the period
        query = db.session.query(AttendanceData).filter(
            AttendanceData.employee_id == str(employee_id),
            AttendanceData.check_in_date >= start_date.date(),
            AttendanceData.check_in_date <= end_date.date()
        )

        # Apply project filter if provided
        if project_filter:
            try:
                project_id = int(project_filter)
                query = query.join(QRCode, AttendanceData.qr_code_id == QRCode.id) \
                    .filter(QRCode.project_id == project_id)
            except ValueError:
                pass  # Invalid project_id, ignore filter

        attendance_records = query.order_by(
            AttendanceData.check_in_date,
            AttendanceData.check_in_time
        ).all()

        # Convert to the format expected by the calculator
        converted_records = []
        for record in records:
            # Get distance from the TimeAttendance record
            distance_value = getattr(record, 'distance', None)
            
            converted_record = type('Record', (), {
                'id': record.id,
                'employee_id': str(record.employee_id),
                'check_in_date': record.attendance_date,
                'check_in_time': record.attendance_time,
                'location_name': record.location_name,
                'latitude': None,
                'longitude': None,
                'distance': distance_value,  # ADD THIS LINE
                'qr_code': type('QRCode', (), {
                    'location': record.location_name,
                    'location_address': record.recorded_address or '',
                    'project': None
                })()
            })()
            converted_records.append(converted_record)

        # Calculate working hours using the same calculator as the dashboard
        #calculator = SingleCheckInCalculator()

        # Calculate hours for this employee
        hours_data = calculator.calculate_employee_hours(
            str(employee_id), start_date, end_date, converted_records
        )

        # Extract miss punch details
        miss_punch_days = []
        if 'daily_hours' in hours_data:
            for date_str, day_data in hours_data['daily_hours'].items():
                if day_data.get('is_miss_punch', False):
                    # Get the actual records for this day
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                    day_records = [r for r in converted_records if r.check_in_date == date_obj]

                    # Format the records information with event types
                    record_details = []
                    for i, record in enumerate(day_records):
                        # Determine event type based on position (alternating check-in/check-out)
                        # First record is always check-in, then alternates
                        event_type = "Check In" if i % 2 == 0 else "Check Out"

                        record_details.append({
                            'time': record.check_in_time.strftime('%H:%M:%S'),
                            'event_type': event_type,
                            'location': record.location_name or 'Unknown Location',
                            'has_gps': record.latitude is not None and record.longitude is not None
                        })

                    miss_punch_days.append({
                        'date': date_str,
                        'date_formatted': datetime.strptime(date_str, '%Y-%m-%d').strftime('%B %d, %Y (%A)'),
                        'records_count': day_data.get('records_count', 0),
                        'records': record_details,
                        'reason': 'Incomplete punch pairs - missing check-in or check-out' if len(
                            day_records) % 2 != 0 else 'Invalid work period duration'
                    })

        # Log the API access
        logger_handler.logger.info(
            f"Miss punch details API accessed by {session.get('username', 'unknown')} for employee {employee_id}")

        return jsonify({
            'success': True,
            'data': {
                'employee_id': employee_id,
                'employee_name': employee_name,
                'period': f"{date_from} to {date_to}",
                'miss_punch_count': len(miss_punch_days),
                'miss_punch_days': miss_punch_days
            }
        })

    except Exception as e:
        print(f"❌ Error in get_miss_punch_details: {e}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")

        logger_handler.log_flask_error(
            error_type="miss_punch_details_api_error",
            error_message=str(e),
            stack_trace=traceback.format_exc()
        )

        return jsonify({
            'success': False,
            'message': 'Internal server error. Please check the server logs.'
        }), 500

def get_employee_name(employee_id):
    """Helper function to get employee full name by ID"""
    try:
        result = db.session.execute(text("""
            SELECT CONCAT(firstName, ' ', lastName) as full_name 
            FROM employee 
            WHERE id = :employee_id
        """), {'employee_id': employee_id})

        row = result.fetchone()
        return row[0] if row else f"Employee {employee_id}"

    except Exception as e:
        print(f"⚠️ Error getting employee name for ID {employee_id}: {e}")
        return f"Employee {employee_id}"

def get_qr_code_checkin_count(qr_code_id):
    """Helper function to get total check-ins count for a QR code"""
    try:
        count = AttendanceData.query.filter_by(qr_code_id=qr_code_id).count()
        logger_handler.logger.info(f"QR Code {qr_code_id} total check-ins: {count}")
        return count
    except Exception as e:
        logger_handler.logger.error(f"Error getting check-ins count for QR {qr_code_id}: {e}")
        return 0
    
@app.context_processor
def inject_payroll_utils():
    """Inject payroll utility functions into templates"""
    return {
        'get_employee_name': get_employee_name,
        'format_hours': lambda hours: f"{hours:.2f}" if hours else "0.00"
    }

@app.context_processor
def inject_dashboard_utils():
    """Inject dashboard utility functions into templates"""
    return {
        'get_qr_code_checkin_count': get_qr_code_checkin_count
    }

@app.route('/statistics')
@login_required
def qr_statistics():
    """QR Code Statistics Dashboard with comprehensive analytics"""
    try:
        # Log statistics page access
        logger_handler.logger.info(f"User {session.get('username', 'unknown')} accessed QR code statistics dashboard")
        
        # Get filter parameters
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        qr_code_filter = request.args.get('qr_code', '')
        project_filter = request.args.get('project', '')
        
        # Build date filter
        date_filter = ""
        if date_from:
            date_filter += f" AND ad.check_in_date >= '{date_from}'"
        if date_to:
            date_filter += f" AND ad.check_in_date <= '{date_to}'"
        
        # QR Code filter
        qr_filter = ""
        if qr_code_filter:
            qr_filter = f" AND ad.qr_code_id = {qr_code_filter}"
            
        # Project filter
        project_filter_clause = ""
        if project_filter:
            project_filter_clause = f" AND qc.project_id = {project_filter}"

        # 1. General Statistics
        general_stats = db.session.execute(text(f"""
            SELECT 
                COUNT(*) as total_scans,
                COUNT(DISTINCT ad.employee_id) as unique_users,
                COUNT(DISTINCT ad.qr_code_id) as active_qr_codes,
                COUNT(DISTINCT DATE(ad.check_in_date)) as active_days,
                COUNT(CASE WHEN ad.check_in_date = CURRENT_DATE THEN 1 END) as today_scans,
                COUNT(CASE WHEN ad.check_in_date >= DATE_SUB(CURRENT_DATE, INTERVAL 7 DAY) THEN 1 END) as week_scans,
                COUNT(CASE WHEN ad.latitude IS NOT NULL AND ad.longitude IS NOT NULL THEN 1 END) as gps_enabled_scans
            FROM attendance_data ad
            LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id
            WHERE 1=1 {date_filter} {qr_filter} {project_filter_clause}
        """)).fetchone()

        # 2. Device Statistics
        device_stats = db.session.execute(text(f"""
            SELECT 
                CASE 
                    WHEN device_info LIKE '%iPhone%' OR device_info LIKE '%iOS%' THEN 'iOS'
                    WHEN device_info LIKE '%Android%' THEN 'Android'
                    WHEN device_info LIKE '%Windows%' THEN 'Windows'
                    WHEN device_info LIKE '%Mac%' OR device_info LIKE '%macOS%' THEN 'macOS'
                    WHEN device_info LIKE '%Linux%' THEN 'Linux'
                    ELSE 'Other'
                END as device_type,
                COUNT(*) as scan_count,
                COUNT(DISTINCT employee_id) as unique_users
            FROM attendance_data ad
            LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id
            WHERE device_info IS NOT NULL {date_filter} {qr_filter} {project_filter_clause}
            GROUP BY device_type
            ORDER BY scan_count DESC
        """)).fetchall()

        # 3. Browser Statistics (from User Agent)
        browser_stats = db.session.execute(text(f"""
            SELECT 
                CASE 
                    WHEN user_agent LIKE '%Chrome%' AND user_agent NOT LIKE '%Edge%' THEN 'Chrome'
                    WHEN user_agent LIKE '%Safari%' AND user_agent NOT LIKE '%Chrome%' THEN 'Safari'
                    WHEN user_agent LIKE '%Firefox%' THEN 'Firefox'
                    WHEN user_agent LIKE '%Edge%' THEN 'Edge'
                    WHEN user_agent LIKE '%Opera%' THEN 'Opera'
                    ELSE 'Other'
                END as browser_type,
                COUNT(*) as scan_count,
                COUNT(DISTINCT employee_id) as unique_users
            FROM attendance_data ad
            LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id
            WHERE user_agent IS NOT NULL {date_filter} {qr_filter} {project_filter_clause}
            GROUP BY browser_type
            ORDER BY scan_count DESC
        """)).fetchall()

        # 4. Location Statistics  
        location_stats = db.session.execute(text(f"""
            SELECT 
                qc.name as qr_name,
                qc.location as qr_location,
                qc.location_event,
                COUNT(*) as total_scans,
                COUNT(DISTINCT ad.employee_id) as unique_users,
                COUNT(CASE WHEN ad.latitude IS NOT NULL THEN 1 END) as gps_scans,
                MIN(ad.check_in_date) as first_scan,
                MAX(ad.check_in_date) as last_scan
            FROM attendance_data ad
            JOIN qr_codes qc ON ad.qr_code_id = qc.id
            WHERE 1=1 {date_filter} {qr_filter} {project_filter_clause}
            GROUP BY qc.id, qc.name, qc.location, qc.location_event
            ORDER BY total_scans DESC
        """)).fetchall()

        # 5. IP Address Analysis (Top 3 Most Active)
        ip_stats = db.session.execute(text(f"""
            SELECT 
                ip_address,
                COUNT(*) as scan_count,
                COUNT(DISTINCT employee_id) as unique_users,
                COUNT(DISTINCT qr_code_id) as qr_codes_used,
                MIN(check_in_date) as first_scan,
                MAX(check_in_date) as last_scan
            FROM attendance_data ad
            LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id
            WHERE ip_address IS NOT NULL {date_filter} {qr_filter} {project_filter_clause}
            GROUP BY ip_address
            ORDER BY scan_count DESC
            LIMIT 3
        """)).fetchall()

        # 6. Project Statistics (if projects exist)
        project_stats = db.session.execute(text(f"""
            SELECT 
                p.id,
                p.name as project_name,
                COUNT(*) as total_scans,
                COUNT(DISTINCT ad.employee_id) as unique_users,
                COUNT(DISTINCT ad.qr_code_id) as qr_codes_in_project,
                AVG(CASE WHEN ad.latitude IS NOT NULL THEN 1.0 ELSE 0.0 END) * 100 as gps_usage_percentage
            FROM attendance_data ad
            JOIN qr_codes qc ON ad.qr_code_id = qc.id
            LEFT JOIN projects p ON qc.project_id = p.id
            WHERE p.id IS NOT NULL {date_filter} {qr_filter} {project_filter_clause}
            GROUP BY p.id, p.name
            ORDER BY total_scans DESC
        """)).fetchall()

        # Get dropdown options for filters
        qr_codes_list = db.session.execute(text("""
            SELECT DISTINCT qc.id, qc.name, qc.location
            FROM qr_codes qc
            JOIN attendance_data ad ON qc.id = ad.qr_code_id
            WHERE qc.active_status = true
            ORDER BY qc.name
        """)).fetchall()

        projects_list = db.session.execute(text("""
            SELECT DISTINCT p.id, p.name
            FROM projects p
            JOIN qr_codes qc ON p.id = qc.project_id
            JOIN attendance_data ad ON qc.id = ad.qr_code_id
            WHERE p.active_status = true
            ORDER BY p.name
        """)).fetchall()

        # Log successful statistics generation
        logger_handler.logger.info(
            f"Generated statistics report for user {session.get('username', 'unknown')} "
            f"with {general_stats.total_scans} total scans. Filters applied: "
            f"date_from={date_from}, date_to={date_to}, qr_code={qr_code_filter}, project={project_filter}"
        )

        return render_template('statistics.html',
                             general_stats=general_stats,
                             device_stats=device_stats,
                             browser_stats=browser_stats,
                             location_stats=location_stats,
                             ip_stats=ip_stats,
                             project_stats=project_stats,
                             qr_codes_list=qr_codes_list,
                             projects_list=projects_list,
                             date_from=date_from,
                             date_to=date_to,
                             qr_code_filter=qr_code_filter,
                             project_filter=project_filter,
                             today_date=datetime.now().strftime('%Y-%m-%d'))

    except Exception as e:
        # Log the error using the correct method
        logger_handler.log_database_error('statistics_page_error', e)
        print(f"❌ Error loading statistics: {e}")
        print(f"❌ Traceback: {traceback.format_exc()}")
        
        flash('Error loading statistics. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/api/statistics/export')
@login_required
def export_statistics():
    """Export statistics data to CSV/Excel"""
    try:
        # Check permissions
        if session.get('role') not in ['admin', 'payroll', 'accounting']:
            return jsonify({'error': 'Access denied'}), 403
            
        # Log export attempt
        logger_handler.logger.info(
            f"User {session.get('username', 'unknown')} (role: {session.get('role')}) "
            f"attempted to export statistics data in {request.args.get('format', 'csv')} format"
        )
        
        # Get comprehensive statistics for export
        export_data = db.session.execute(text("""
            SELECT 
                ad.id,
                ad.employee_id,
                COALESCE(CONCAT(e.firstName, ' ', e.lastName), ad.employee_id) as employee_name,
                ad.check_in_date,
                ad.check_in_time,
                qc.name as qr_code_name,
                qc.location as qr_location,
                qc.location_event,
                p.name as project_name,
                ad.device_info,
                ad.user_agent,
                ad.ip_address,
                ad.latitude,
                ad.longitude,
                ad.address,
                ad.location_name,
                ad.created_timestamp
            FROM attendance_data ad
            JOIN qr_codes qc ON ad.qr_code_id = qc.id
            LEFT JOIN projects p ON qc.project_id = p.id
            LEFT JOIN employee e ON CAST(ad.employee_id AS UNSIGNED) = e.id
            ORDER BY ad.created_timestamp DESC
        """)).fetchall()
        
        # Create CSV content
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow([
            'ID', 'Employee ID', 'Employee Name', 'Date', 'Time', 
            'QR Code', 'QR Location', 'Event', 'Project', 'Device', 
            'Browser Info', 'IP Address', 'Latitude', 'Longitude', 
            'Address', 'Location Name', 'Timestamp'
        ])
        
        # Write data
        for row in export_data:
            writer.writerow([
                row.id, row.employee_id, row.employee_name, 
                str(row.check_in_date), str(row.check_in_time),
                row.qr_code_name, row.qr_location, row.location_event,
                row.project_name or 'No Project', row.device_info or 'Unknown',
                row.user_agent or 'Unknown', row.ip_address or 'Unknown',
                row.latitude or '', row.longitude or '', 
                row.address or '', row.location_name or '',
                str(row.created_timestamp)
            ])
        
        output.seek(0)
        
        # Create response with proper file handling
        csv_data = output.getvalue()
        
        # Log successful export
        logger_handler.logger.info(
            f"User {session.get('username', 'unknown')} successfully exported "
            f"{len(export_data)} statistics records"
        )
        
        # Create response
        response = app.make_response(csv_data)
        response.headers["Content-Disposition"] = f"attachment; filename=qr_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response.headers["Content-type"] = "text/csv"
        
        return response
        
    except Exception as e:
        logger_handler.log_database_error('statistics_export_error', e)
        print(f"❌ Error exporting statistics: {e}")
        return jsonify({'error': 'Export failed'}), 500

    except Exception as e:
        # Log the error
        logger_handler.log_database_error('statistics_page_error', e)
        print(f"❌ Error loading statistics: {e}")
        print(f"❌ Traceback: {traceback.format_exc()}")
        
        flash('Error loading statistics. Please try again.', 'error')
        return redirect(url_for('dashboard'))

# EMPLOYEE MANAGEMENT ROUTES
@app.route('/employees')
@login_required
def employees():
    """Display employee management page with search and pagination"""
    try:
        # Log user accessing employee management
        try:
            logger_handler.logger.info(f"User {session['username']} accessed employee management list")
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")
        
        # Get search parameters
        search = request.args.get('search', '').strip()
        page = request.args.get('page', 1, type=int)
        per_page = 20  # Number of employees per page
        
        # Build query based on search
        query = Employee.query.outerjoin(Project, Employee.contractId == Project.id)

        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                db.or_(
                    Employee.firstName.like(search_pattern),
                    Employee.lastName.like(search_pattern),
                    Employee.title.like(search_pattern),
                    Employee.id.like(search_pattern)
                )
            )

        # Order by first name, then last name
        query = query.order_by(Employee.firstName, Employee.lastName)
        
        # Paginate results
        employees = query.paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        # Get summary statistics
        total_employees = Employee.query.count()
        employees_with_title = Employee.query.filter(Employee.title.isnot(None)).filter(Employee.title != '').count()
        unique_titles = db.session.query(Employee.title).filter(Employee.title.isnot(None)).filter(Employee.title != '').distinct().count()
        
        stats = {
            'total_employees': total_employees,
            'employees_with_title': employees_with_title,
            'unique_titles': unique_titles,
            'search_results': employees.total if search else total_employees
        }
        
        return render_template('employees.html', 
                             employees=employees, 
                             search=search,
                             stats=stats)
        
    except Exception as e:
        logger_handler.log_database_error('employee_list', e)
        flash('Error loading employee list. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/employees/create', methods=['GET', 'POST'])
@login_required
@log_database_operations('employee_creation')
def create_employee():
    """Create new employee (Admin only)"""
    if request.method == 'POST':
        try:
            # Get form data
            employee_id = request.form['employee_id'].strip()
            first_name = request.form['first_name'].strip()
            last_name = request.form['last_name'].strip()
            title = request.form.get('title', '').strip()
            contract_id = request.form.get('contract_id', '1').strip()
            
            # Validate required fields
            if not all([employee_id, first_name, last_name, contract_id]):
                flash('Employee ID, First Name, Last Name, and Project are required.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('create_employee.html', projects=projects)
            
            # Validate employee ID is numeric
            try:
                employee_id_int = int(employee_id)
                contract_id_int = int(contract_id)
            except ValueError:
                flash('Employee ID must be numeric and Project must be selected.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('create_employee.html', projects=projects)
            
            # Check if employee ID already exists
            existing_employee = Employee.query.filter_by(id=employee_id_int).first()
            if existing_employee:
                flash(f'Employee with ID {employee_id} already exists.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('create_employee.html', projects=projects)
            
            # Create new employee
            new_employee = Employee(
                id=employee_id_int,
                firstName=first_name,
                lastName=last_name,
                title=title if title else None,
                contractId=contract_id_int
            )
            
            db.session.add(new_employee)
            db.session.commit()
            
            # Log employee creation with project info
            try:
                project = Project.query.get(contract_id_int)
                project_name = project.name if project else f"Project {contract_id_int}"
                logger_handler.logger.info(f"Admin user {session['username']} created new employee: {employee_id_int} - {first_name} {last_name} assigned to {project_name}")
            except Exception as log_error:
                print(f"⚠️ Logging error (non-critical): {log_error}")
            
            flash(f'Employee "{first_name} {last_name}" (ID: {employee_id}) created successfully.', 'success')
            return redirect(url_for('employees'))
            
        except Exception as e:
            db.session.rollback()
            logger_handler.log_database_error('employee_creation', e)
            flash('Failed to create employee. Please try again.', 'error')
            projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
            return render_template('create_employee.html', projects=projects)
    
    # GET request - load the form with projects
    projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
    return render_template('create_employee.html', projects=projects)

@app.route('/employees/<int:employee_index>/edit', methods=['GET', 'POST'])
@login_required  
@log_database_operations('employee_update')
def edit_employee(employee_index):
    """Edit existing employee (Admin only)"""
    try:
        # Get employee by index (primary key)
        employee = Employee.query.get_or_404(employee_index)
        
        if request.method == 'POST':
            # Get form data
            employee_id = request.form['employee_id'].strip()
            first_name = request.form['first_name'].strip()
            last_name = request.form['last_name'].strip()
            title = request.form.get('title', '').strip()
            contract_id = request.form.get('contract_id', '1').strip()

            # Validate required fields
            if not all([employee_id, first_name, last_name, contract_id]):
                flash('Employee ID, First Name, Last Name, and Project are required.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('edit_employee.html', employee=employee, projects=projects)
            
            # Validate numeric fields
            try:
                employee_id_int = int(employee_id)
                contract_id_int = int(contract_id)
            except ValueError:
                flash('Employee ID must be numeric and Project must be selected.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('edit_employee.html', employee=employee, projects=projects)
            
            # Check if employee ID already exists (but not for this employee)
            existing_employee = Employee.query.filter_by(id=employee_id_int).first()
            if existing_employee and existing_employee.index != employee.index:
                flash(f'Employee with ID {employee_id} already exists.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('edit_employee.html', employee=employee, projects=projects)
            
            # Store original values for logging
            original_data = {
                'id': employee.id,
                'firstName': employee.firstName,
                'lastName': employee.lastName,
                'title': employee.title,
                'contractId': employee.contractId
            }
            
            # Update employee data
            employee.id = employee_id_int
            employee.firstName = first_name
            employee.lastName = last_name
            employee.title = title if title else None
            employee.contractId = contract_id_int
            
            db.session.commit()
            
            # Log employee update with project info
            try:
                project = Project.query.get(contract_id_int)
                project_name = project.name if project else f"Project {contract_id_int}"
                logger_handler.logger.info(f"Admin user {session['username']} updated employee: {employee_index} - {first_name} {last_name} assigned to {project_name}")
            except Exception as log_error:
                print(f"⚠️ Logging error (non-critical): {log_error}")
            
            flash(f'Employee "{first_name} {last_name}" updated successfully.', 'success')
            return redirect(url_for('employees'))
        
        # GET request - load the form with projects
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        return render_template('edit_employee.html', employee=employee, projects=projects)
        
    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('employee_update', e)
        flash('Error updating employee. Please try again.', 'error')
        return redirect(url_for('employees'))

@app.route('/employees/<int:employee_index>/delete', methods=['POST'])
@login_required
@log_database_operations('employee_deletion')
def delete_employee(employee_index):
    """Delete employee (Admin only) - Enhanced with better logging"""
    try:
        print(f"🗑️ DELETE REQUEST: Employee index {employee_index}")
        print(f"📋 Request method: {request.method}")
        print(f"👤 User: {session.get('username', 'Unknown')}")
        
        # Get employee by index (primary key)
        employee = Employee.query.get_or_404(employee_index)
        print(f"✅ Found employee: {employee.firstName} {employee.lastName} (ID: {employee.id})")
        
        # Store employee data for logging before deletion
        employee_data = {
            'index': employee.index,
            'id': employee.id,
            'firstName': employee.firstName,
            'lastName': employee.lastName,
            'title': employee.title,
            'contractId': employee.contractId
        }
        
        # Check if employee has attendance records
        from models.attendance import AttendanceData
        attendance_count = AttendanceData.query.filter_by(employee_id=str(employee.id)).count()
        print(f"📊 Attendance records found: {attendance_count}")
        
        if attendance_count > 0:
            error_msg = f'Cannot delete employee "{employee.full_name}". Employee has {attendance_count} attendance records. Please contact system administrator.'
            print(f"❌ DELETION BLOCKED: {error_msg}")
            flash(error_msg, 'error')
            return redirect(url_for('employees'))
        
        # Proceed with deletion
        print(f"🗑️ Proceeding with deletion of employee: {employee_data['firstName']} {employee_data['lastName']}")
        
        db.session.delete(employee)
        db.session.commit()
        print("✅ Employee successfully deleted from database")
        
        # Log employee deletion
        try:
            logger_handler.logger.info(f"Admin user {session['username']} deleted employee: {employee_data['firstName']} {employee_data['lastName']} (ID: {employee_data['id']})")
            print(f"📋 Deletion logged successfully")
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")
        
        success_msg = f'Employee "{employee_data["firstName"]} {employee_data["lastName"]}" deleted successfully.'
        flash(success_msg, 'success')
        print(f"✅ SUCCESS: {success_msg}")
        
        return redirect(url_for('employees'))
        
    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('employee_deletion', e)
        error_msg = f'Error deleting employee. Please try again.'
        print(f"❌ ERROR in delete_employee: {e}")
        print(f"❌ Exception type: {type(e)}")
        flash(error_msg, 'error')
        return redirect(url_for('employees'))

@app.route('/api/employees/search')
@login_required
def api_employees_search():
    """API endpoint for employee search (for AJAX)"""
    try:
        search = request.args.get('q', '').strip()
        limit = request.args.get('limit', 10, type=int)
        
        if not search:
            return jsonify({'employees': []})
        
        employees = Employee.search_employees(search)[:limit]
        
        result = {
            'employees': [emp.to_dict() for emp in employees]
        }
        
        return jsonify(result)
        
    except Exception as e:
        logger_handler.log_database_error('employee_search_api', e)
        return jsonify({'error': 'Search failed'}), 500

@app.route('/employees/<int:employee_index>')
@login_required
def employee_detail(employee_index):
    """View employee details with attendance summary"""
    try:
        # Get employee by index (primary key)
        employee = Employee.query.outerjoin(Project, Employee.contractId == Project.id).filter(Employee.index == employee_index).first_or_404()
        
        # Get attendance statistics for this employee
        from models.attendance import AttendanceData
        
        # Total attendance records
        total_attendance = AttendanceData.query.filter_by(employee_id=str(employee.id)).count()
        
        # Recent attendance (last 30 days)
        from datetime import datetime, timedelta
        thirty_days_ago = datetime.now() - timedelta(days=30)
        recent_attendance = AttendanceData.query.filter(
            AttendanceData.employee_id == str(employee.id),
            AttendanceData.check_in_date >= thirty_days_ago.date()
        ).count()
        
        # Most recent attendance record
        latest_attendance = AttendanceData.query.filter_by(employee_id=str(employee.id)).order_by(
            AttendanceData.check_in_date.desc(),
            AttendanceData.check_in_time.desc()
        ).first()
        
        # Get unique projects this employee has attended
        unique_projects = db.session.query(Project).join(
            QRCode, Project.id == QRCode.project_id
        ).join(
            AttendanceData, QRCode.id == AttendanceData.qr_code_id
        ).filter(
            AttendanceData.employee_id == str(employee.id)
        ).distinct().all()
        
        attendance_stats = {
            'total_attendance': total_attendance,
            'recent_attendance': recent_attendance,
            'latest_attendance': latest_attendance,
            'unique_projects': len(unique_projects),
            'projects': unique_projects
        }
        
        # Log employee detail view
        try:
            logger_handler.logger.info(f"User {session['username']} viewed employee detail: {employee.full_name} (ID: {employee.id})")
        except Exception as log_error:
            print(f"⚠️ Logging error (non-critical): {log_error}")
        
        return render_template('employee_detail.html', 
                             employee=employee, 
                             attendance_stats=attendance_stats)
        
    except Exception as e:
        logger_handler.log_database_error('employee_detail', e)
        flash('Error loading employee details. Please try again.', 'error')
        return redirect(url_for('employees'))

@app.route('/time-attendance')
@login_required
@log_user_activity('time_attendance_view')
def time_attendance_dashboard():
    """Display time attendance dashboard with table layout"""
    try:
        # Initialize default values
        total_records = 0
        unique_employees = 0
        unique_locations = 0
        recent_imports = []
        recent_records = []
        employees = []
        locations = []
        
        # Try to get data from TimeAttendance model if it exists
        try:
            from models.time_attendance import TimeAttendance
            
            # Get summary statistics
            total_records = TimeAttendance.query.count()
            
            if total_records > 0:
                unique_employees = db.session.query(TimeAttendance.employee_id).distinct().count()
                unique_locations = db.session.query(TimeAttendance.location_name).distinct().count()
                
                # Get recent records (last 20 records for table display)
                recent_records = TimeAttendance.query.order_by(
                    TimeAttendance.attendance_date.desc(),
                    TimeAttendance.attendance_time.desc()
                ).limit(20).all()
                
                # Get recent imports (last 10 import batches)
                recent_imports = db.session.query(
                    TimeAttendance.import_batch_id,
                    TimeAttendance.import_date,
                    TimeAttendance.import_source,
                    db.func.count(TimeAttendance.id).label('record_count')
                ).filter(
                    TimeAttendance.import_batch_id.isnot(None)
                ).group_by(
                    TimeAttendance.import_batch_id,
                    TimeAttendance.import_date,
                    TimeAttendance.import_source
                ).order_by(
                    TimeAttendance.import_date.desc()
                ).limit(10).all()
                
                # Get filter options
                employees = TimeAttendance.get_unique_employees()
                locations = TimeAttendance.get_unique_locations()
                
        except ImportError:
            # TimeAttendance model doesn't exist yet - use defaults
            pass
        except Exception as e:
            # Database table doesn't exist yet or other error - use defaults
            print(f"TimeAttendance query error: {e}")
            pass
        
        return render_template('time_attendance_dashboard.html',
                             total_records=total_records,
                             unique_employees=unique_employees,
                             unique_locations=unique_locations,
                             recent_imports=recent_imports,
                             recent_records=recent_records,
                             employees=employees,
                             locations=locations)
                             
    except Exception as e:
        logger_handler.logger.error(f"Error in time attendance dashboard: {e}")
        flash('Error loading time attendance dashboard.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/time-attendance/import', methods=['GET', 'POST'])
@login_required
@log_database_operations('time_attendance_import')
def import_time_attendance():
    """Enhanced import with duplicate review"""
    if request.method == 'GET':
        # Load active projects for dropdown
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        return render_template('time_attendance_import.html', projects=projects)

    if request.method == 'POST':
        try:
            # Check if this is coming from invalid review (file is already in session)
            coming_from_invalid_review = request.form.get('from_invalid_review', 'false').lower() == 'true'
            coming_from_duplicate_review = request.form.get('from_duplicate_review', 'false').lower() == 'true'

            print(f"\n🔍 IMPORT FLOW DEBUG:")
            print(f"   Coming from invalid review: {coming_from_invalid_review}")
            print(f"   Coming from duplicate review: {coming_from_duplicate_review}")

            if coming_from_invalid_review or coming_from_duplicate_review:
                # Retrieve file from session
                if 'pending_import_file' not in session or 'pending_import_filename' not in session:
                    flash('Session expired. Please upload the file again.', 'error')
                    return redirect(url_for('import_time_attendance'))
                
                temp_path = session['pending_import_file']
                filename = session['pending_import_filename']
                
                # Verify file still exists
                if not os.path.exists(temp_path):
                    flash('Temporary file not found. Please upload the file again.', 'error')
                    session.pop('pending_import_file', None)
                    session.pop('pending_import_filename', None)
                    return redirect(url_for('import_time_attendance'))
                
                print(f"✅ Retrieved file from session: {filename}")
                print(f"✅ Temp path exists: {os.path.exists(temp_path)}")
                
            else:
                # Normal file upload flow - now supports multiple files
                if 'files' not in request.files:
                    flash('No files uploaded.', 'error')
                    return redirect(request.url)
                
                files = request.files.getlist('files')
                if not files or len(files) == 0:
                    flash('No files selected.', 'error')
                    return redirect(request.url)
                
                # Validate all files and save them temporarily
                temp_paths = []
                filenames = []
                
                for file in files:
                    if file.filename == '':
                        continue
                    
                    # Validate file extension
                    if not file.filename.lower().endswith(('.xlsx', '.xls')):
                        flash(f'Invalid file format: {file.filename}. Please upload only Excel files (.xlsx or .xls).', 'error')
                        # Clean up already saved files
                        for saved_path in temp_paths:
                            if os.path.exists(saved_path):
                                os.remove(saved_path)
                        return redirect(request.url)
                    
                    # Save uploaded file temporarily
                    filename = secure_filename(file.filename)
                    temp_path = os.path.join(app.config.get('UPLOAD_FOLDER', '/tmp'), 
                                           f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
                    
                    os.makedirs(os.path.dirname(temp_path), exist_ok=True)
                    file.save(temp_path)
                    
                    temp_paths.append(temp_path)
                    filenames.append(filename)
                    
                    print(f"✅ Uploaded file {len(temp_paths)}: {filename}")
                    print(f"✅ Saved to: {temp_path}")
                
                if len(temp_paths) == 0:
                    flash('No valid files selected.', 'error')
                    return redirect(request.url)
                
                # Store file paths in session for duplicate/invalid review
                session['pending_import_file'] = temp_paths[0] if len(temp_paths) == 1 else temp_paths
                session['pending_import_filename'] = filenames[0] if len(filenames) == 1 else filenames
                session['pending_import_files_multiple'] = len(temp_paths) > 1
                
                temp_path = temp_paths[0] if len(temp_paths) == 1 else temp_paths
                filename = filenames[0] if len(filenames) == 1 else ', '.join(filenames)
                
                print(f"✅ Total files uploaded: {len(temp_paths)}")

            # Determine if we're processing multiple files
            is_multiple_files = session.get('pending_import_files_multiple', False)
            files_to_process = []
            
            if is_multiple_files:
                # Multiple files mode
                if isinstance(temp_path, list):
                    files_to_process = list(zip(temp_path, filename.split(', ') if isinstance(filename, str) else filename))
                else:
                    files_to_process = [(temp_path, filename)]
            else:
                # Single file mode (existing behavior)
                files_to_process = [(temp_path, filename)]
            
            print(f"📁 Processing {len(files_to_process)} file(s)")
            
            try:
                import_service = TimeAttendanceImportService(db, logger_handler)
                
                # Get import options
                skip_duplicates = request.form.get('skip_duplicates', 'true').lower() == 'true'
                validate_only = request.form.get('validate_only', 'false').lower() == 'true'
                analyze_duplicates = request.form.get('analyze_duplicates', 'false').lower() == 'true'
                analyze_invalid = request.form.get('analyze_invalid', 'false').lower() == 'true'
                
                print(f"📋 Import Options:")
                print(f"   Skip duplicates: {skip_duplicates}")
                print(f"   Validate only: {validate_only}")
                print(f"   Analyze duplicates: {analyze_duplicates}")
                print(f"   Analyze invalid: {analyze_invalid}")
                print(f"   Coming from invalid review: {coming_from_invalid_review}")
                
                # Store combined results for multiple files
                all_results = {
                    'total_files': len(files_to_process),
                    'successful_files': 0,
                    'failed_files': 0,
                    'total_imported': 0,
                    'total_duplicates': 0,
                    'total_failed': 0,
                    'file_results': [],
                    'errors': [],
                    'warnings': []
                }
                
                # Process each file
                for file_index, (current_temp_path, current_filename) in enumerate(files_to_process, 1):
                    print(f"\n📄 Processing file {file_index}/{len(files_to_process)}: {current_filename}")
                    
                    import_result = None  # Initialize to prevent reference errors
                    
                    try:
                        # For multiple files, skip review screens and import directly
                        if is_multiple_files:
                            print(f"   📦 Batch mode: processing directly without review screens")
                            
                            # Validate the file first
                            validation_result = import_service.validate_excel_file(current_temp_path)
                            
                            if not validation_result['valid']:
                                raise Exception(f"Validation failed: {'; '.join(validation_result['errors'])}")
                            
                            # Get import settings
                            project_id = request.form.get('project_id')
                            project_id = int(project_id) if project_id and project_id != '' else None
                            import_source = request.form.get('import_source', f"Batch Import - {current_filename}")
                            
                            # Import the file (always skip duplicates in batch mode)
                            import_result = import_service.import_from_excel(
                                current_temp_path,
                                created_by=session['user_id'],
                                import_source=import_source,
                                skip_duplicates=True,  # Always skip duplicates in batch mode
                                force_import_hashes=set(),
                                project_id=project_id
                            )
                            
                        else:
                            # Single file - use existing review workflow logic below
                            # This continues to the existing code after the loop
                            pass
                        
                        # Accumulate results if import was performed
                        if import_result and import_result.get('success'):
                            all_results['successful_files'] += 1
                            all_results['total_imported'] += import_result.get('imported_records', 0)
                            all_results['total_duplicates'] += import_result.get('duplicate_records', 0)
                            all_results['file_results'].append({
                                'filename': current_filename,
                                'status': 'success',
                                'imported': import_result.get('imported_records', 0),
                                'batch_id': import_result.get('batch_id', '')
                            })
                            print(f"   ✅ Imported {import_result.get('imported_records', 0)} records")
                        elif import_result:
                            # Import ran but failed
                            all_results['failed_files'] += 1
                            all_results['total_failed'] += import_result.get('failed_records', 0)
                            all_results['errors'].append(f"{current_filename}: Import failed")
                            all_results['file_results'].append({
                                'filename': current_filename,
                                'status': 'failed',
                                'error': 'Import returned unsuccessful status'
                            })
                        
                    except Exception as file_error:
                        print(f"❌ Error processing file {current_filename}: {file_error}")
                        logger_handler.logger.error(f"Error processing file {current_filename}: {file_error}")
                        all_results['failed_files'] += 1
                        all_results['errors'].append(f"{current_filename}: {str(file_error)}")
                        all_results['file_results'].append({
                            'filename': current_filename,
                            'status': 'failed',
                            'error': str(file_error)
                        })
                        continue
                    
                    finally:
                        # Cleanup individual file (only for multiple file mode, single file cleanup happens later)
                        if is_multiple_files and os.path.exists(current_temp_path):
                            try:
                                os.remove(current_temp_path)
                                print(f"   🗑️ Cleaned up temp file")
                            except Exception as cleanup_error:
                                print(f"   ⚠️ Failed to cleanup temp file: {cleanup_error}")
                
                # After processing all files
                if is_multiple_files:
                    # Log the batch import activity
                    logger_handler.logger.info(
                        f"Batch Import: User {session.get('username', 'unknown')} imported time attendance data from {len(files_to_process)} files - "
                        f"Successful: {all_results['successful_files']}/{all_results['total_files']}, "
                        f"Total imported: {all_results['total_imported']}, "
                        f"Duplicates: {all_results['total_duplicates']}"
                    )
                    
                    # Show combined results
                    if all_results['successful_files'] > 0:
                        flash(f"✅ Successfully imported {all_results['total_imported']} records from {all_results['successful_files']}/{all_results['total_files']} files.", 'success')
                        
                        if all_results['total_duplicates'] > 0:
                            flash(f"ℹ️ Skipped {all_results['total_duplicates']} duplicate records across all files.", 'info')
                    
                    if all_results['failed_files'] > 0:
                        flash(f"❌ {all_results['failed_files']} file(s) failed to import.", 'error')
                        
                        # Show first few error details
                        for error in all_results['errors'][:3]:
                            flash(f"Error: {error}", 'error')
                        
                        if len(all_results['errors']) > 3:
                            flash(f"...and {len(all_results['errors']) - 3} more errors", 'error')
                    
                    # Clear session
                    session.pop('pending_import_file', None)
                    session.pop('pending_import_filename', None)
                    session.pop('pending_import_files_multiple', None)
                    
                    print(f"\n📊 Batch Import Summary:")
                    print(f"   Total files: {all_results['total_files']}")
                    print(f"   Successful: {all_results['successful_files']}")
                    print(f"   Failed: {all_results['failed_files']}")
                    print(f"   Total imported: {all_results['total_imported']}")
                    print(f"   Total duplicates: {all_results['total_duplicates']}")
                    
                    return redirect(url_for('time_attendance_dashboard'))

                # Check if this is coming from duplicate review
                force_import_hashes = request.form.getlist('force_import_hashes[]')
                
                # If analyzing for duplicates, show review page (but not if coming from invalid/duplicate review)
                if analyze_duplicates and not force_import_hashes and not coming_from_invalid_review and not coming_from_duplicate_review:
                    print("🔍 Analyzing for duplicates...")
                    duplicate_analysis = import_service.analyze_for_duplicates(temp_path)
                    
                    if duplicate_analysis['duplicate_records'] > 0:
                        print(f"⚠️ Found {duplicate_analysis['duplicate_records']} duplicates")
                        # Get project_id from form
                        project_id = request.form.get('project_id')
                        # Show duplicate review page
                        return render_template('time_attendance_duplicate_review.html',
                                            analysis=duplicate_analysis,
                                            filename=filename,
                                            project_id=project_id)
                    else:
                        print("✅ No duplicates found")
                        flash('No duplicates found. Proceeding with import.', 'info')
                
                # Check for invalid rows and show review if any (but not if coming from invalid review)
                if analyze_invalid and not coming_from_invalid_review:
                    print("🔍 Analyzing for invalid rows...")
                    invalid_analysis = import_service.analyze_for_invalid_rows(temp_path)
                    
                    if invalid_analysis['invalid_rows'] > 0:
                        print(f"⚠️ Found {invalid_analysis['invalid_rows']} invalid rows")
                        # Get project_id from form
                        project_id = request.form.get('project_id')
                        # Show invalid row review page
                        return render_template('time_attendance_invalid_review.html',
                                            analysis=invalid_analysis,
                                            filename=filename,
                                            project_id=project_id)
                    else:
                        print("✅ All rows are valid")
                        flash('All rows are valid. Proceeding with import.', 'info')
                
                # If coming from invalid review, skip validation (already done)
                if not coming_from_invalid_review:
                    print("🔍 Validating file...")
                    # Validate file
                    validation_result = import_service.validate_excel_file(temp_path)
                    
                    if not validation_result['valid']:
                        print(f"❌ Validation failed: {validation_result['errors']}")
                        flash(f"File validation failed: {'; '.join(validation_result['errors'])}", 'error')
                        return render_template('time_attendance_import.html', 
                                             validation_result=validation_result)
                    
                    if validation_result['warnings']:
                        for warning in validation_result['warnings']:
                            flash(warning, 'warning')
                    
                    if validate_only:
                        print(f"✅ Validation successful: {validation_result['valid_rows']} valid records")
                        flash(f"File validation successful! Found {validation_result['valid_rows']} valid records.", 'success')
                        return render_template('time_attendance_import.html', 
                                             validation_result=validation_result)
                else:
                    print("⏭️ Skipping validation (already validated)")
                
                # Proceed with import
                print("🚀 Starting import process...")
                import_source = request.form.get('import_source', f"Manual Import - {filename}")
                project_id = request.form.get('project_id')
                project_id = int(project_id) if project_id and project_id != '' else None

                import_result = import_service.import_from_excel(
                    temp_path,
                    created_by=session['user_id'],
                    import_source=import_source,
                    skip_duplicates=skip_duplicates,
                    force_import_hashes=force_import_hashes,
                    project_id=project_id
                )
                
                if import_result['success']:
                    print(f"✅ Import successful!")
                    print(f"   Batch ID: {import_result['batch_id']}")
                    print(f"   Imported: {import_result['imported_records']}/{import_result['total_records']}")
                    print(f"   Duplicates: {import_result['duplicate_records']}")
                    print(f"   Failed: {import_result['failed_records']}")
                    
                    logger_handler.logger.info(
                        f"User {session['username']} successfully imported time attendance data - "
                        f"Batch: {import_result['batch_id']}, "
                        f"Records: {import_result['imported_records']}/{import_result['total_records']}, "
                        f"Duplicates: {import_result['duplicate_records']}, "
                        f"Forced: {import_result['forced_duplicates']}, "
                        f"Failed: {import_result['failed_records']}"
                    )
                    
                    flash(f"Import successful! Imported {import_result['imported_records']} records "
                          f"out of {import_result['total_records']} total records.", 'success')
                    
                    if import_result['duplicate_records'] > 0:
                        flash(f"Skipped {import_result['duplicate_records']} duplicate records.", 'info')
                    
                    if import_result['forced_duplicates'] > 0:
                        flash(f"Imported {import_result['forced_duplicates']} duplicate records as requested.", 'info')
                    
                    if import_result['failed_records'] > 0:
                        flash(f"Note: {import_result['failed_records']} records failed to import. "
                              f"Check the error details below.", 'warning')
                    
                    # Clean up temp file after successful import
                    if os.path.exists(temp_path):
                        try:
                            os.remove(temp_path)
                            session.pop('pending_import_file', None)
                            session.pop('pending_import_filename', None)
                            print("🗑️ Cleaned up temp file")
                        except Exception as cleanup_error:
                            print(f"⚠️ Failed to cleanup temp file: {cleanup_error}")
                    
                    return render_template('time_attendance_import_result.html', 
                                         import_result=import_result)
                else:
                    print(f"❌ Import failed: {import_result['errors']}")
                    flash(f"Import failed: {'; '.join(import_result['errors'][:3])}", 'error')
                    if len(import_result['errors']) > 3:
                        flash(f"...and {len(import_result['errors']) - 3} more errors", 'warning')
                    return render_template('time_attendance_import.html', 
                                         import_result=import_result)
                
            except Exception as import_error:
                print(f"❌ Import exception: {import_error}")
                import traceback
                print(f"❌ Traceback: {traceback.format_exc()}")
                raise
                
        except Exception as e:
            logger_handler.log_database_error('time_attendance_import', e)
            print(f"❌ Top-level exception: {e}")
            import traceback
            print(f"❌ Traceback: {traceback.format_exc()}")
            flash('Import failed due to an unexpected error.', 'error')
            return render_template('time_attendance_import.html')
    
    # GET request
    return render_template('time_attendance_import.html')


@app.route('/time-attendance/import/analyze-duplicates', methods=['POST'])
@login_required
def analyze_import_duplicates():
    """AJAX endpoint to analyze file for duplicates"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Invalid file format'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(app.config.get('UPLOAD_FOLDER', '/tmp'), 
                               f"analyze_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
        
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)
        
        # Store in session
        session['pending_import_file'] = temp_path
        session['pending_import_filename'] = filename
        
        try:
            import_service = TimeAttendanceImportService(db, logger_handler)
            analysis = import_service.analyze_for_duplicates(temp_path)
            
            # Convert datetime objects to strings for JSON
            for duplicate in analysis.get('duplicates', []):
                if 'new_record' in duplicate:
                    if duplicate['new_record'].get('attendance_date'):
                        duplicate['new_record']['attendance_date'] = str(duplicate['new_record']['attendance_date'])
                    if duplicate['new_record'].get('attendance_time'):
                        duplicate['new_record']['attendance_time'] = str(duplicate['new_record']['attendance_time'])
                
                if 'existing_record' in duplicate:
                    if duplicate['existing_record'].get('attendance_date'):
                        duplicate['existing_record']['attendance_date'] = str(duplicate['existing_record']['attendance_date'])
                    if duplicate['existing_record'].get('attendance_time'):
                        duplicate['existing_record']['attendance_time'] = str(duplicate['existing_record']['attendance_time'])
                    if duplicate['existing_record'].get('import_date'):
                        duplicate['existing_record']['import_date'] = str(duplicate['existing_record']['import_date'])
            
            return jsonify({
                'success': True,
                'analysis': analysis
            })
        
        except Exception as e:
            # Cleanup on error
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise e
    
    except Exception as e:
        logger_handler.logger.error(f"Duplicate analysis error: {e}")
        return jsonify({
            'success': False,
            'message': f'Analysis failed: {str(e)}'
        }), 500
    
@app.route('/time-attendance/import/analyze-invalid', methods=['POST'])
@login_required
def analyze_import_invalid():
    """AJAX endpoint to analyze file for invalid rows"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Invalid file format'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(app.config.get('UPLOAD_FOLDER', '/tmp'), 
                               f"analyze_invalid_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
        
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)
        
        # Store in session
        session['pending_import_file'] = temp_path
        session['pending_import_filename'] = filename
        
        try:
            import_service = TimeAttendanceImportService(db, logger_handler)
            analysis = import_service.analyze_for_invalid_rows(temp_path)
            
            # Convert datetime objects to strings for JSON
            for invalid in analysis.get('invalid_details', []):
                if 'row_data' in invalid:
                    if invalid['row_data'].get('attendance_date'):
                        invalid['row_data']['attendance_date'] = str(invalid['row_data']['attendance_date'])
                    if invalid['row_data'].get('attendance_time'):
                        invalid['row_data']['attendance_time'] = str(invalid['row_data']['attendance_time'])
            
            return jsonify({
                'success': True,
                'analysis': analysis
            })
        
        except Exception as e:
            logger_handler.logger.error(f"Invalid row analysis error: {e}")
            return jsonify({
                'success': False,
                'message': f'Analysis failed: {str(e)}'
            }), 500
        
    except Exception as e:
        logger_handler.logger.error(f"Invalid row analysis error: {e}")
        return jsonify({
            'success': False,
            'message': f'Analysis failed: {str(e)}'
        }), 500


# ---------------------------------------------------------------------------
# Time Attendance Import — SSE progress streaming (disk-based, multi-worker safe)
#
# Design: progress state is written to a small JSON file on disk so that any
# gunicorn worker process can read it.  No shared in-memory state is required.
# The /stream endpoint runs the import itself (synchronously inside the SSE
# generator) while writing progress to the file and yielding events to the
# browser — compatible with gunicorn gevent workers.
# ---------------------------------------------------------------------------

def _progress_file_path(job_id: str) -> str:
    """Return the path for the on-disk progress file for a given job_id."""
    upload_dir = app.config.get('UPLOAD_FOLDER', '/tmp')
    os.makedirs(upload_dir, exist_ok=True)
    return os.path.join(upload_dir, f"import_progress_{job_id}.json")


def _write_progress(job_id: str, event: dict) -> None:
    """Atomically write the latest progress event to disk."""
    path = _progress_file_path(job_id)
    try:
        tmp = path + '.tmp'
        with open(tmp, 'w') as f:
            json.dump(event, f)
        os.replace(tmp, path)
    except Exception:
        pass  # Best-effort; import will continue regardless


@app.route('/time-attendance/import/start', methods=['POST'])
@login_required
def start_import_job():
    """
    Validates the uploaded file, saves it to disk, stores import options in a
    progress file, then returns a job_id.  The actual import runs inside the
    SSE stream endpoint so no background thread or shared memory is needed.
    """
    try:
        if 'files' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded.'}), 400

        files = request.files.getlist('files')
        if not files or files[0].filename == '':
            return jsonify({'success': False, 'error': 'No file selected.'}), 400

        file = files[0]
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Invalid file format.'}), 400

        filename = secure_filename(file.filename)
        upload_dir = app.config.get('UPLOAD_FOLDER', '/tmp')
        os.makedirs(upload_dir, exist_ok=True)
        job_id = str(uuid.uuid4())
        temp_path = os.path.join(upload_dir,
                                 f"stream_{job_id}_{filename}")
        file.save(temp_path)

        # Store import options alongside the file so the stream endpoint can
        # read them without depending on session or shared memory.
        job_meta = {
            'type': 'pending',
            'temp_path': temp_path,
            'filename': filename,
            'skip_duplicates': request.form.get('skip_duplicates', 'true').lower() == 'true',
            'project_id': int(request.form.get('project_id')) if request.form.get('project_id') else None,
            'import_source': request.form.get('import_source', f"Manual Import - {filename}"),
            'created_by': session['user_id'],
            'username': session.get('username', 'unknown'),
        }
        _write_progress(job_id, job_meta)

        logger_handler.logger.info(
            f"User {job_meta['username']} queued time attendance import job {job_id} for file {filename}"
        )
        return jsonify({'success': True, 'job_id': job_id})

    except Exception as e:
        logger_handler.logger.error(f"Error queuing import job: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/time-attendance/import/stream/<job_id>')
@login_required
def stream_import_progress(job_id):
    """
    SSE endpoint — runs the import synchronously while streaming progress to
    the browser.  Works across multiple gunicorn workers because all state is
    stored on disk (no in-memory job store).
    """
    progress_path = _progress_file_path(job_id)

    def generate():
        import time as _time

        # ── Read the job metadata written by /start ────────────────────────
        deadline = _time.time() + 15  # Wait up to 15 s for the file to appear
        meta = None
        while _time.time() < deadline:
            if os.path.exists(progress_path):
                try:
                    with open(progress_path) as f:
                        meta = json.load(f)
                    break
                except Exception:
                    pass
            yield "data: " + json.dumps({'type': 'heartbeat'}) + "\n\n"
            _time.sleep(0.3)

        if not meta or meta.get('type') != 'pending':
            yield "data: " + json.dumps({
                'type': 'error',
                'message': 'Job metadata not found. Please try importing again.'
            }) + "\n\n"
            return

        temp_path    = meta['temp_path']
        skip_dupes   = meta['skip_duplicates']
        project_id   = meta['project_id']
        import_source= meta['import_source']
        created_by   = meta['created_by']
        username     = meta['username']

        if not os.path.exists(temp_path):
            yield "data: " + json.dumps({
                'type': 'error',
                'message': 'Uploaded file not found. Please try importing again.'
            }) + "\n\n"
            return

        yield "data: " + json.dumps({'type': 'status', 'message': 'Reading and validating file...'}) + "\n\n"

        # ── Run the import with a progress callback ────────────────────────
        try:
            svc = TimeAttendanceImportService(db, logger_handler)

            # progress_callback writes to disk AND yields an SSE event.
            # We collect events in a list so the generator can yield them.
            _pending_events = []

            def on_progress(current, total, message):
                pct = int(current / total * 100) if total else 0
                event = {
                    'type': 'progress',
                    'current': current,
                    'total': total,
                    'percent': pct,
                    'message': message,
                }
                _write_progress(job_id, event)
                _pending_events.append(event)

            # We need to interleave yielding with the synchronous import loop.
            # Strategy: run import_from_excel; the callback appends to
            # _pending_events; after every DB commit batch (50 records) we
            # flush pending events to the SSE stream.
            import threading as _threading
            result_holder = [None]
            error_holder  = [None]
            done_event    = _threading.Event()

            def _run():
                # Push an application context so the thread can access
                # Flask-SQLAlchemy, Employee.query, etc.
                with app.app_context():
                    try:
                        result_holder[0] = svc.import_from_excel(
                            temp_path,
                            created_by=created_by,
                            import_source=import_source,
                            skip_duplicates=skip_dupes,
                            force_import_hashes=[],
                            project_id=project_id,
                            progress_callback=on_progress,
                        )
                    except Exception as exc:
                        error_holder[0] = exc
                    finally:
                        done_event.set()

            t = _threading.Thread(target=_run, daemon=True)
            t.start()

            # Yield progress events as they arrive while the import thread runs
            while not done_event.is_set():
                while _pending_events:
                    yield "data: " + json.dumps(_pending_events.pop(0)) + "\n\n"
                yield "data: " + json.dumps({'type': 'heartbeat'}) + "\n\n"
                _time.sleep(0.4)

            # Drain any remaining events after the thread finishes
            while _pending_events:
                yield "data: " + json.dumps(_pending_events.pop(0)) + "\n\n"

            if error_holder[0]:
                raise error_holder[0]

            result = result_holder[0]

            if result and result['success']:
                logger_handler.logger.info(
                    f"User {username} imported {result['imported_records']} time attendance records "
                    f"via stream (batch: {result['batch_id']})"
                )

            # Sanitize result dict for JSON serialization — convert any
            # datetime objects (e.g. import_date) to ISO-format strings.
            if result and isinstance(result.get('import_date'), datetime):
                result['import_date'] = result['import_date'].isoformat()
            done_event_data = {'type': 'done', 'result': result}
            _write_progress(job_id, done_event_data)
            yield "data: " + json.dumps(done_event_data) + "\n\n"

        except Exception as e:
            logger_handler.logger.error(f"Import stream error for job {job_id}: {e}")
            error_event = {'type': 'error', 'message': str(e)}
            _write_progress(job_id, error_event)
            yield "data: " + json.dumps(error_event) + "\n\n"

        finally:
            # Clean up temp files
            for path in (temp_path, progress_path):
                try:
                    if os.path.exists(path):
                        os.remove(path)
                except Exception:
                    pass

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',   # Disable nginx buffering for SSE
        }
    )


@app.route('/time-attendance/import/cancel-pending')
@login_required
def cancel_pending_import():
    """Cancel pending import and cleanup temp file"""
    try:
        if 'pending_import_file' in session:
            temp_path = session['pending_import_file']
            if os.path.exists(temp_path):
                os.remove(temp_path)
            session.pop('pending_import_file')
        
        if 'pending_import_filename' in session:
            session.pop('pending_import_filename')
        
        flash('Import cancelled.', 'info')
    except Exception as e:
        logger_handler.logger.error(f"Error cancelling import: {e}")
    
    return redirect(url_for('import_time_attendance'))



@app.route('/time-attendance/import/validate', methods=['POST'])
@login_required
def validate_import_file():
    """AJAX endpoint to validate Excel file before import"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Invalid file format'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(app.config.get('UPLOAD_FOLDER', '/tmp'), 
                               f"validate_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
        
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)
        
        try:
            # Validate file
            import_service = TimeAttendanceImportService(db, logger_handler)
            validation_result = import_service.validate_excel_file(temp_path)
            
            return jsonify({
                'success': True,
                'validation': validation_result
            })
        
        finally:
            # Cleanup
            if os.path.exists(temp_path):
                os.remove(temp_path)
    
    except Exception as e:
        logger_handler.logger.error(f"Validation error: {e}")
        return jsonify({
            'success': False,
            'message': f'Validation failed: {str(e)}'
        }), 500
    
@app.route('/time-attendance/import/batch/<batch_id>')
@login_required
@log_user_activity('view_import_batch')
def view_import_batch(batch_id):
    """View details of a specific import batch"""
    try:
        import_service = TimeAttendanceImportService(db, logger_handler)
        batch_summary = import_service.get_import_summary(batch_id)
        
        if not batch_summary:
            flash('Import batch not found.', 'error')
            return redirect(url_for('time_attendance_dashboard'))
        
        return render_template('time_attendance_batch_detail.html',
                             batch_summary=batch_summary)
    
    except Exception as e:
        logger_handler.logger.error(f"Error viewing batch {batch_id}: {e}")
        flash('Error loading batch details.', 'error')
        return redirect(url_for('time_attendance_dashboard'))


@app.route('/time-attendance/import/batch/<batch_id>/delete', methods=['POST'])
@admin_required
@log_database_operations('delete_import_batch')
def delete_import_batch(batch_id):
    """Delete an entire import batch"""
    try:
        import_service = TimeAttendanceImportService(db, logger_handler)
        result = import_service.delete_import_batch(batch_id, deleted_by=session['user_id'])
        
        if result['success']:
            flash(result['message'], 'success')
            logger_handler.logger.info(
                f"User {session['username']} deleted import batch {batch_id} - "
                f"{result['deleted_count']} records removed"
            )
        else:
            flash(result['message'], 'error')
        
        return redirect(url_for('time_attendance_dashboard'))
    
    except Exception as e:
        logger_handler.logger.error(f"Error deleting batch {batch_id}: {e}")
        flash('Error deleting import batch.', 'error')
        return redirect(url_for('time_attendance_dashboard'))


@app.route('/time-attendance/import/download-template')
@login_required
def download_import_template():
    """Download Excel template for time attendance import"""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Time Attendance Template"
        
        # Define headers
        headers = ['ID', 'Name', 'Platform', 'Date', 'Time', 'Location Name', 
                   'Action Description', 'Event Description', 'Recorded Address', 'Distance']
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add sample data rows
        sample_data = [
            ['12345', 'John Doe', 'iPhone - iOS', '2025-10-06', '09:00:00', 
             'HQ Suite 210', 'Check In', 'Main Office', '123 Main St', '0.125'],
            ['67890', 'Jane Smith', 'Android', '2025-10-06', '08:45:00', 
             'Branch Office', 'Check In', 'Morning Entry', '456 Oak Avenue', '0.250'],
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Time Attendance Import Template - Instructions"],
            [""],
            ["Required Columns:"],
            ["- ID: Employee ID (required)"],
            ["- Name: Employee full name (required)"],
            ["- Date: Attendance date in YYYY-MM-DD format (required)"],
            ["- Time: Attendance time in HH:MM:SS format (required)"],
            ["- Location Name: Location where attendance was recorded (required)"],
            ["- Action Description: Type of action (e.g., Check In, Check Out) (required)"],
            [""],
            ["Optional Columns:"],
            ["- Platform: Device platform (e.g., iPhone - iOS, Android)"],
            ["- Event Description: Additional event details"],
            ["- Recorded Address: Physical address where attendance was recorded"],
            ["- Distance: Distance in miles between Building and Recorded Address (optional)"],
            [""],
            ["Important Notes:"],
            ["- Do not modify the header row"],
            ["- Ensure all required fields have values"],
            ["- Date format must be YYYY-MM-DD (e.g., 2025-10-06)"],
            ["- Time format must be HH:MM:SS (e.g., 09:00:00)"],
            ["- Remove the sample data rows before importing your actual data"],
            ["- Duplicate records will be automatically detected and skipped"],
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row_num, column=1, value=instruction[0])
        
        ws_instructions.column_dimensions['A'].width = 80
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Log download
        logger_handler.logger.info(f"User {session['username']} downloaded import template")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'time_attendance_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    except Exception as e:
        logger_handler.logger.error(f"Error generating template: {e}")
        flash('Error generating template file.', 'error')
        return redirect(url_for('import_time_attendance'))
    
@app.route('/time-attendance/export')
@login_required
@log_user_activity('time_attendance_export')
def export_time_attendance():
    """Export time attendance records to CSV or Excel"""
    try:
        export_format = request.args.get('format', 'excel').lower()
        
        # Get filter parameters (same as records page)
        employee_filter = request.args.get('employee_id')
        location_filter = request.args.get('location_name')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        import_batch = request.args.get('import_batch')
        project_filter = request.args.get('project_id')
        
        # Build query with same filters as the view
        from models.time_attendance import TimeAttendance
        query = TimeAttendance.query
        
        # Apply filters
        if employee_filter:
            query = query.filter(TimeAttendance.employee_id == employee_filter)
        
        if location_filter:
            query = query.filter(TimeAttendance.location_name == location_filter)
        
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(TimeAttendance.attendance_date >= start_date_obj)
            except ValueError:
                flash('Invalid start date format.', 'error')
                return redirect(url_for('time_attendance_records'))
        
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(TimeAttendance.attendance_date <= end_date_obj)
            except ValueError:
                flash('Invalid end date format.', 'error')
                return redirect(url_for('time_attendance_records'))
        
        if import_batch:
            query = query.filter(TimeAttendance.import_batch_id == import_batch)
        
        if project_filter:
            query = query.filter(TimeAttendance.project_id == project_filter)
        
        # Order by date and time (most recent first)
        records = query.order_by(
            TimeAttendance.attendance_date.desc(),
            TimeAttendance.attendance_time.desc()
        ).all()
        
        if not records:
            flash('No records found to export.', 'warning')
            return redirect(url_for('time_attendance_records'))
        
        # Get project name if project filter exists
        project_name_for_filename = ''
        if project_filter:
            try:
                from models.project import Project
                project = Project.query.get(int(project_filter))
                if project:
                    # Replace spaces and special characters with underscores
                    project_name_safe = project.name.replace(' ', '_').replace('/', '_').replace('\\', '_')
                    project_name_for_filename = f"{project_name_safe}_"
            except Exception as e:
                print(f"⚠️ Error getting project name for filename: {e}")
        
        # Log export
        logger_handler.logger.info(
            f"User {session['username']} exported {len(records)} time attendance records "
            f"in {export_format.upper()} format"
        )
        
        # Format dates for filename (MMDDYYYY format)
        date_from_formatted = ''
        date_to_formatted = ''
        if start_date:
            try:
                date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                date_from_formatted = date_obj.strftime('%m%d%Y')
            except ValueError:
                pass
        
        if end_date:
            try:
                date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                date_to_formatted = date_obj.strftime('%m%d%Y')
            except ValueError:
                pass
        
        # Build filename with date range
        # Format: [project_name_]time_attendance_[fromdate_todate].xlsx/csv
        date_range_str = ''
        if date_from_formatted and date_to_formatted:
            date_range_str = f"{date_from_formatted}_{date_to_formatted}"
        elif date_from_formatted:
            date_range_str = f"from_{date_from_formatted}"
        elif date_to_formatted:
            date_range_str = f"to_{date_to_formatted}"
        
        # Keep the filter_str for backward compatibility (but not in filename anymore)
        filter_desc = []
        if employee_filter:
            filter_desc.append(f"emp_{employee_filter}")
        if location_filter:
            filter_desc.append(f"loc_{location_filter[:10]}")
        
        filter_str = "_".join(filter_desc) if filter_desc else "all"
        
        return export_time_attendance_excel(records, project_name_for_filename, date_range_str, filter_str, start_date, end_date)
    
    except Exception as e:
        logger_handler.logger.error(f"Error exporting time attendance records: {e}")
        flash('Error generating export file. Please try again.', 'error')
        return redirect(url_for('time_attendance_records'))


def calculate_possible_violation(distance_value):
    """
    Calculate possible violation status based on distance
    
    Args:
        distance_value: Distance in miles (float or None)
    
    Returns:
        'Yes' if distance > 0.3, 'No' otherwise
    """
    if distance_value is None:
        return 'No'
    
    try:
        distance_float = float(distance_value)
        return 'Yes' if distance_float > 0.3 else 'No'
    except (ValueError, TypeError):
        return 'No'
    
def export_time_attendance_excel(records, project_name_for_filename, date_range_str, filter_str, start_date_filter=None, end_date_filter=None):
    """Generate Excel export with template format matching the provided template"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    
    # Get date range for calculations
    if start_date_filter and end_date_filter:
        # Convert string dates to date objects if needed
        if isinstance(start_date_filter, str):
            start_date = datetime.strptime(start_date_filter, '%Y-%m-%d').date()
        else:
            start_date = start_date_filter
        
        if isinstance(end_date_filter, str):
            end_date = datetime.strptime(end_date_filter, '%Y-%m-%d').date()
        else:
            end_date = end_date_filter
    elif records:
        # Fallback to calculating from records if no filter dates provided
        start_date = min(r.attendance_date for r in records)
        end_date = max(r.attendance_date for r in records)
    else:
        return None
    
    # Import parse function at the beginning for work type detection
    from working_hours_calculator import parse_employee_id_for_work_type
    
    # Convert TimeAttendance records to format expected by calculator
    converted_records = []
    for record in records:
        # Get distance value from the record
        distance_value = getattr(record, 'distance', None)
        
        # CRITICAL: Determine record_type from action_description
        record_type = 'check_in'  # Default
        if hasattr(record, 'action_description') and record.action_description:
            action_lower = record.action_description.lower()
            if 'out' in action_lower or 'checkout' in action_lower:
                record_type = 'check_out'
        
        # Extract work type (PT, SP, PW) from employee_id for location display in Excel
        _, work_type = parse_employee_id_for_work_type(str(record.employee_id))
        
        # Create display location name with work type suffix if applicable
        base_location_name = record.location_name
        if work_type and work_type in ('PT', 'SP', 'PW'):
            display_location_name = f"{base_location_name} ({work_type})"
        else:
            display_location_name = base_location_name
        
        converted_record = type('Record', (), {
            'id': record.id,
            'employee_id': str(record.employee_id),
            'check_in_date': record.attendance_date,
            'check_in_time': record.attendance_time,
            'location_name': display_location_name,  # Use display name with work type for Excel export
            'original_location_name': base_location_name,  # Keep original for internal grouping
            'work_type': work_type,  # Store work type for reference
            'latitude': None,
            'longitude': None,
            'distance': distance_value,
            'record_type': record_type,
            'action_description': record.action_description,
            'event_description': record.event_description or '',
            'recorded_address': record.recorded_address or '',
            'qr_code': type('QRCode', (), {
                'location': base_location_name,  # Keep original for QR code matching
                'location_address': record.recorded_address or '',
                'project': None
            })()
        })()
        converted_records.append(converted_record)
    
    # Log count of records with work types for audit trail
    work_type_counts = {'PT': 0, 'SP': 0, 'PW': 0, 'Regular': 0}
    for r in converted_records:
        wt = getattr(r, 'work_type', None)
        if wt in work_type_counts:
            work_type_counts[wt] += 1
        else:
            work_type_counts['Regular'] += 1
    
    if any(work_type_counts[wt] > 0 for wt in ['PT', 'SP', 'PW']):
        logger_handler.logger.info(
            f"Excel Export: Processing records with work types - "
            f"Regular: {work_type_counts['Regular']}, PT: {work_type_counts['PT']}, "
            f"SP: {work_type_counts['SP']}, PW: {work_type_counts['PW']}"
        )
    
    # Calculate working hours using SingleCheckInCalculator
    calculator = WorkingHoursCalculator()
    hours_data = calculator.calculate_all_employees_hours(
        datetime.combine(start_date, datetime.min.time()),
        datetime.combine(end_date, datetime.max.time()),
        converted_records
    )
    
    # Get employee names - map BASE employee IDs to names for consolidated display
    # Look up from Employee table using the numeric base_id to get the correct name,
    # regardless of what is stored in the employee_name column (which may contain
    # work type characters such as 'Employee 3937SP' if imported with a decorated ID).
    from working_hours_calculator import parse_employee_id_for_work_type
    employee_names = {}
    for record in records:
        base_id, _ = parse_employee_id_for_work_type(str(record.employee_id))
        if base_id not in employee_names:
            try:
                emp = Employee.query.filter_by(id=int(base_id)).first()
                if emp:
                    employee_names[base_id] = f"{emp.lastName}, {emp.firstName}"
                else:
                    # Fallback: use stored name if Employee table lookup fails
                    employee_names[base_id] = record.employee_name
                    logger_handler.logger.warning(f"Employee ID {base_id} not found in employee table during export; using stored name.")
            except Exception as e:
                employee_names[base_id] = record.employee_name
                logger_handler.logger.warning(f"Could not lookup employee name for ID {base_id} during export: {e}")
    
    # Setup styles
    # White bold text on black background for column header row (no border)
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # CHANGED: Sample format uses ONLY a bottom border on the last row of each day group.
    # Intermediate rows and first rows have no borders at all (no left/right/top).
    border_day_middle = Border()  # No borders on intermediate rows

    border_day_last = Border(
        bottom=Side(style='thin')  # Only bottom border on the last row of a day group
    )

    border_day_single = Border(
        bottom=Side(style='thin')  # Single-row days also get only bottom border
    )

    # border_day_first is same as middle (no borders) — kept for compatibility
    border_day_first = Border()

    def get_day_border(row_position, total_rows):
        """
        Get appropriate border style based on row position within a day.
        Matches sample.xlsx: only the LAST row of each day group has a bottom border.

        Args:
            row_position: Current row number (0-indexed) within the day
            total_rows: Total number of rows for this day

        Returns:
            Border object
        """
        if total_rows == 1:
            return border_day_single
        elif row_position == total_rows - 1:
            return border_day_last
        else:
            return border_day_middle

    # Orange background for Missed Punch
    missed_punch_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    # Write main headers
    current_row = 1
    
    # Row 1: Company name
    ws.merge_cells(f'A{current_row}:N{current_row}')
    title_cell = ws.cell(row=current_row, column=1, value=os.environ.get('COMPANY_NAME', 'Your Company'))
    title_cell.font = Font(name='Arial', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 2: Summary title
    ws.merge_cells(f'A{current_row}:N{current_row}')
    summary_cell = ws.cell(row=current_row, column=1, value='Summary report of Hours worked')
    summary_cell.font = Font(name='Arial', size=12, bold=True)
    summary_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 3: Project name
    project_display = project_name_for_filename.replace('_', ' ').strip() if project_name_for_filename else "[Project Name]"
    project_cell = ws.cell(row=current_row, column=1, value=project_display)
    project_cell.font = Font(name='Arial', size=11, bold=True)
    project_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 4: Date range
    date_range_text = f"Date range: {start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"
    ws.merge_cells(f'A{current_row}:N{current_row}')
    date_cell = ws.cell(row=current_row, column=1, value=date_range_text)
    date_cell.font = Font(name='Arial', size=11)
    date_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 5: Empty row
    current_row += 1
    
    # Empty row before first employee
    current_row += 1
    
    # Sort employees by name for organized output
    sorted_employees = sorted(
        hours_data['employees'].items(),
        key=lambda x: employee_names.get(x[0], f'Employee {x[0]}').lower()
    )

    # Write data for each employee (sorted by name)
    for employee_id, emp_data in sorted_employees:
        employee_name = employee_names.get(employee_id, f'Employee {employee_id}')
        
        # Employee header row (merged A to O)
        ws.merge_cells(f'A{current_row}:O{current_row}')
        emp_header = ws.cell(row=current_row, column=1, 
                            value=f'Employee ID {employee_id}: {employee_name}')
        emp_header.font = Font(name='Arial', size=11, bold=True)
        emp_header.alignment = Alignment(horizontal='left')
        current_row += 1
        
        # Column headers
        headers = ['Day', 'Date', 'In', 'Out', 'Location', 'Zone', 'Hours/Building',
                  'Daily Total', 'Regular Hours', 'OT Hours', 'Building Address',
                  'Recorded Location', 'Distance (Mile)', 'Possible Violation']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            # White bold text on black background; no border (matching sample.xlsx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Group records by date AND location for separate rows per location
        daily_location_data = {}
        # Import parse function to match base employee ID with all variants (SP, PW, PT)
        from working_hours_calculator import parse_employee_id_for_work_type
        
        # Filter records where the BASE employee ID matches (includes 1234, 1234 SP, 1234 PW, 1234 PT)
        employee_records = []
        for r in converted_records:
            record_base_id, _ = parse_employee_id_for_work_type(str(r.employee_id))
            if record_base_id == employee_id:
                employee_records.append(r)

        for record in employee_records:
            date_key = record.check_in_date.strftime('%Y-%m-%d')
            location_key = record.location_name or 'Unknown Location'
            
            # Create nested structure: date -> location -> records
            if date_key not in daily_location_data:
                daily_location_data[date_key] = {}
            
            if location_key not in daily_location_data[date_key]:
                daily_location_data[date_key][location_key] = {
                    'records': [],
                    'location_name': location_key
                }
            
            daily_location_data[date_key][location_key]['records'].append(record)
        
        # Track weekly hours for overtime calculation
        weekly_total_hours = 0
        current_week_start = None
        grand_regular_hours = 0
        grand_ot_hours = 0
        
        # Get all dates that have records (not all weekdays)
        dates_with_records = sorted([
            date_str for date_str, day_data in emp_data['daily_hours'].items()
            if day_data.get('records_count', 0) > 0
        ])
        
        # Write daily data (ONLY DAYS WITH RECORDS)
        for date_str in dates_with_records:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            day_data = emp_data['daily_hours'][date_str]
            
            # Check for week boundary (week starts on Monday)
            week_start = date_obj - timedelta(days=date_obj.weekday())
            if current_week_start is not None and week_start != current_week_start:
                # Write weekly total row
                week_regular = min(weekly_total_hours, 40.0)
                week_overtime = max(0, weekly_total_hours - 40.0)
                
                ws.cell(row=current_row, column=7, value='Weekly Total: ').font = bold_font
                ws.cell(row=current_row, column=8, value=round(weekly_total_hours, 2)).font = bold_font
                ws.cell(row=current_row, column=9, value=round(week_regular, 2)).font = bold_font
                ws.cell(row=current_row, column=10, value=round(week_overtime, 2)).font = bold_font
                
                grand_regular_hours += week_regular
                grand_ot_hours += week_overtime
                current_row += 1
                
                weekly_total_hours = 0
            
            current_week_start = week_start
            
            # Get all locations for this date
            date_locations = daily_location_data.get(date_str, {})
            total_locations = len(date_locations)
            
            total_hours = day_data['total_hours']
            is_miss_punch = day_data.get('is_miss_punch', False)
            
            # Calculate total hours for the day (even if miss punch)
            if not is_miss_punch and total_hours > 0:
                weekly_total_hours += total_hours
            elif is_miss_punch and total_locations > 0:
                # Calculate hours even for miss punch (different locations)
                all_records_for_day = []
                for loc_data in date_locations.values():
                    all_records_for_day.extend(loc_data['records'])

                if len(all_records_for_day) >= 2:
                    sorted_all_records = sorted(all_records_for_day, key=lambda x: x.check_in_time)
                    
                    # NEW: Check if all records are IN or all OUT
                    record_types = []
                    for record in sorted_all_records:
                        action_desc = record.action_description.lower() if record.action_description else ''
                        if 'out' in action_desc or 'checkout' in action_desc:
                            record_types.append('OUT')
                        else:
                            record_types.append('IN')
                    
                    # If all same type (all IN or all OUT), hours = 0
                    if len(set(record_types)) == 1:
                        print(f"⚠️ {date_str}: All {len(sorted_all_records)} records are {record_types[0]} - 0 working hours")
                        calculated_hours = 0.0
                        total_hours = 0.0
                    else:
                        # Mixed IN/OUT - calculate from first to last
                        first_time = sorted_all_records[0].check_in_time
                        last_time = sorted_all_records[-1].check_in_time
                        
                        first_datetime = datetime.combine(date_obj, first_time)
                        last_datetime = datetime.combine(date_obj, last_time)
                        _raw_min = (last_datetime - first_datetime).total_seconds() / 60.0
                        calculated_hours = round_time_to_quarter_hour(_raw_min) / 60.0
                        
                        weekly_total_hours += calculated_hours
                        total_hours = calculated_hours
            
            # Daily total display (only shown on last location's last row)
            # Daily Total: quarter-hour rounding → decimal hours
            _dt_min = total_hours * 60.0
            daily_total_display = round_time_to_quarter_hour(_dt_min) / 60.0 if total_hours > 0 else ''
            
            # FIXED: Get all records for the day and sort by time FIRST, then group by location
            all_day_records = []
            for loc_data in date_locations.values():
                all_day_records.extend(loc_data['records'])
            
            # Sort all records by time chronologically
            all_day_records_sorted = sorted(all_day_records, key=lambda x: x.check_in_time)
            
            # Group consecutive records by location while maintaining time order
            location_groups = []
            current_location = None
            current_group = []
            
            for record in all_day_records_sorted:
                if current_location is None or record.location_name == current_location:
                    # Same location or first record
                    current_location = record.location_name
                    current_group.append(record)
                else:
                    # Different location - save current group and start new one
                    if current_group:
                        location_groups.append({
                            'location': current_location,
                            'records': current_group
                        })
                    current_location = record.location_name
                    current_group = [record]
            
            # Add the last group
            if current_group:
                location_groups.append({
                    'location': current_location,
                    'records': current_group
                })
            
            # Process each location group in chronological order
            total_groups = len(location_groups)
            for group_index, group_data in enumerate(location_groups):
                location_count = group_index + 1
                is_last_location = (location_count == total_groups)
                
                location_name = group_data['location']
                sorted_records = group_data['records']
                
                if len(sorted_records) == 1:
                    # Single record for this location
                    single_record = sorted_records[0]
                    
                    # Get the original TimeAttendance record to check action_description
                    original_record = None
                    for rec in records:
                        if (rec.employee_id == single_record.employee_id and 
                            rec.attendance_date == single_record.check_in_date and 
                            rec.attendance_time == single_record.check_in_time):
                            original_record = rec
                            break
                    
                    # Determine if this is a check-in or check-out
                    is_check_out = False
                    if original_record and original_record.action_description:
                        action_lower = original_record.action_description.lower()
                        is_check_out = 'out' in action_lower or 'checkout' in action_lower
                    
                    # Show day name and date only for first group's first record
                    day_display = date_obj.strftime('%A').upper() if location_count == 1 else ''
                    date_display = date_obj.strftime('%m/%d/%Y') if location_count == 1 else ''
                    
                    # Show daily total only if this is the last group
                    current_daily_total = daily_total_display if is_last_location else ''
                    
                    if is_check_out:
                        # Orphaned check-out
                        row_data = [
                            day_display,
                            date_display,
                            '',  # No check-in time
                            single_record.check_in_time.strftime('%I:%M:%S %p'),  # Out
                            single_record.location_name,
                            '',
                            'Missed Punch',
                            current_daily_total,
                            '',
                            '',
                            single_record.event_description or '',
                            single_record.recorded_address or '',
                            getattr(single_record, 'distance', None) or '',
                            calculate_possible_violation(getattr(single_record, 'distance', None))
                        ]
                    else:
                        # Orphaned check-in
                        row_data = [
                            day_display,
                            date_display,
                            single_record.check_in_time.strftime('%I:%M:%S %p'),  # In
                            '',  # No check-out time
                            single_record.location_name,
                            '',
                            'Missed Punch',
                            current_daily_total,
                            '',
                            '',
                            single_record.event_description or '',
                            single_record.recorded_address or '',
                            getattr(single_record, 'distance', None) or '',
                            calculate_possible_violation(getattr(single_record, 'distance', None))
                        ]
                    
                    day_border = border_day_last if is_last_location else border_day_middle
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = data_font
                        cell.border = day_border
                        # Apply orange background to Missed Punch cell (column G)
                        if col == 7:
                            cell.fill = missed_punch_fill
                    current_row += 1
                    
                else:
                    # Multiple records for this location - use action_description for pairing
                    # Determine record types based on action_description
                    record_info = []
                    for record in sorted_records:
                        action_desc = record.action_description.lower() if record.action_description else ''
                        is_out = 'out' in action_desc or 'checkout' in action_desc
                        record_info.append({
                            'record': record,
                            'is_out': is_out,
                            'used': False
                        })
                        print(f"  Record at {record.check_in_time}: action='{record.action_description}', is_out={is_out}")

                    # Create pairs based on IN -> OUT logic
                    pairs_to_write = []
                    i = 0
                    while i < len(record_info):
                        if record_info[i]['used']:
                            i += 1
                            continue
                        
                        current_is_out = record_info[i]['is_out']
                        
                        if not current_is_out:  # This is an IN
                            # Look for next OUT
                            out_found = False
                            for j in range(i + 1, len(record_info)):
                                if record_info[j]['used']:
                                    continue
                                if record_info[j]['is_out']:  # Found an OUT
                                    pairs_to_write.append({
                                        'check_in': record_info[i]['record'],
                                        'check_out': record_info[j]['record'],
                                        'is_miss_punch': (j > i + 1)  # Missed punch if not consecutive
                                    })
                                    record_info[i]['used'] = True
                                    record_info[j]['used'] = True
                                    out_found = True
                                    break
                            
                            if not out_found:  # No OUT found for this IN
                                pairs_to_write.append({
                                    'check_in': record_info[i]['record'],
                                    'check_out': None,
                                    'is_miss_punch': True
                                })
                                record_info[i]['used'] = True
                        else:  # This is an orphaned OUT
                            pairs_to_write.append({
                                'check_in': None,
                                'check_out': record_info[i]['record'],
                                'is_miss_punch': True
                            })
                            record_info[i]['used'] = True
                        
                        i += 1

                    print(f"  Created {len(pairs_to_write)} pairs")

                    # Write all pairs
                    for pair_idx, pair_data in enumerate(pairs_to_write):
                        check_in_record = pair_data['check_in']
                        check_out_record = pair_data['check_out']
                        is_miss_punch = pair_data['is_miss_punch']
                        
                        # Show day name and date only for first pair of first location
                        day_display = date_obj.strftime('%A').upper() if (location_count == 1 and pair_idx == 0) else ''
                        date_display = date_obj.strftime('%m/%d/%Y') if (location_count == 1 and pair_idx == 0) else ''
                        
                        # Calculate hours if complete pair
                        if check_in_record and check_out_record and not is_miss_punch:
                            pair_datetime_in = datetime.combine(date_obj, check_in_record.check_in_time)
                            pair_datetime_out = datetime.combine(date_obj, check_out_record.check_in_time)
                            pair_hours = (pair_datetime_out - pair_datetime_in).total_seconds() / 3600.0
                            pair_hours = round(pair_hours, 2)
                        else:
                            pair_hours = 'Missed Punch'
                        
                        # Show daily total on last pair of last location
                        is_last_pair = (pair_idx == len(pairs_to_write) - 1) and is_last_location
                        current_daily_total = daily_total_display if is_last_pair else ''
                        
                        # Build row data
                        if check_in_record and check_out_record:
                            row_data = [
                                day_display,
                                date_display,
                                check_in_record.check_in_time.strftime('%I:%M:%S %p'),  # In
                                check_out_record.check_in_time.strftime('%I:%M:%S %p'),  # Out
                                check_in_record.location_name,
                                '',
                                pair_hours,
                                current_daily_total,
                                '',
                                '',
                                check_in_record.event_description or '',
                                check_in_record.recorded_address or '',
                                getattr(check_in_record, 'distance', None) or '',
                                calculate_possible_violation(getattr(check_in_record, 'distance', None))
                            ]
                        elif check_in_record:  # IN without OUT
                            row_data = [
                                day_display,
                                date_display,
                                check_in_record.check_in_time.strftime('%I:%M:%S %p'),  # In
                                '',  # No OUT
                                check_in_record.location_name,
                                '',
                                'Missed Punch',
                                current_daily_total,
                                '',
                                '',
                                check_in_record.event_description or '',
                                check_in_record.recorded_address or '',
                                getattr(check_in_record, 'distance', None) or '',
                                calculate_possible_violation(getattr(check_in_record, 'distance', None))
                            ]
                        else:  # OUT without IN
                            row_data = [
                                day_display,
                                date_display,
                                '',  # No IN
                                check_out_record.check_in_time.strftime('%I:%M:%S %p'),  # Out
                                check_out_record.location_name,
                                '',
                                'Missed Punch',
                                current_daily_total,
                                '',
                                '',
                                check_out_record.event_description or '',
                                check_out_record.recorded_address or '',
                                getattr(check_out_record, 'distance', None) or '',
                                calculate_possible_violation(getattr(check_out_record, 'distance', None))
                            ]
                        
                        day_border = border_day_last if is_last_pair else border_day_middle
                        for col, value in enumerate(row_data, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.font = data_font
                            cell.border = day_border
                            # Apply orange background to Missed Punch cell
                            if col == 7 and value == 'Missed Punch':
                                cell.fill = missed_punch_fill
                        current_row += 1
                            
        # Write final weekly total for this employee
        if weekly_total_hours > 0:
            week_regular = min(weekly_total_hours, 40.0)
            week_overtime = max(0, weekly_total_hours - 40.0)
            
            ws.cell(row=current_row, column=7, value='Weekly Total: ').font = bold_font
            ws.cell(row=current_row, column=8, value=round(weekly_total_hours, 2)).font = bold_font
            ws.cell(row=current_row, column=9, value=round(week_regular, 2)).font = bold_font
            ws.cell(row=current_row, column=10, value=round(week_overtime, 2)).font = bold_font
            
            grand_regular_hours += week_regular
            grand_ot_hours += week_overtime
            current_row += 1
        
        # Write extra working hours rows (SP/PW/PT) if employee has any
        # Get extra hours from emp_data grand_totals
        grand_totals = emp_data.get('grand_totals', {})
        sp_hours = grand_totals.get('sp_hours', 0.0)
        pw_hours = grand_totals.get('pw_hours', 0.0)
        pt_hours = grand_totals.get('pt_hours', 0.0)
        
        # Write SP row if hours > 0
        if sp_hours > 0:
            ws.cell(row=current_row, column=7, value='Special Project (SP): ').font = Font(name='Arial', size=10, bold=True, italic=True)
            ws.cell(row=current_row, column=9, value=round(sp_hours, 2)).font = Font(name='Arial', size=10, bold=True, italic=True)
            # Log SP hours export
            logger_handler.logger.info(f"Export: Employee {employee_id} SP hours: {sp_hours:.2f}")
            current_row += 1
        
        # Write PW row if hours > 0
        if pw_hours > 0:
            ws.cell(row=current_row, column=7, value='Periodic Work (PW): ').font = Font(name='Arial', size=10, bold=True, italic=True)
            ws.cell(row=current_row, column=9, value=round(pw_hours, 2)).font = Font(name='Arial', size=10, bold=True, italic=True)
            # Log PW hours export
            logger_handler.logger.info(f"Export: Employee {employee_id} PW hours: {pw_hours:.2f}")
            current_row += 1
        
        # Write PT row if hours > 0
        if pt_hours > 0:
            ws.cell(row=current_row, column=7, value='Project Team (PT): ').font = Font(name='Arial', size=10, bold=True, italic=True)
            ws.cell(row=current_row, column=9, value=round(pt_hours, 2)).font = Font(name='Arial', size=10, bold=True, italic=True)
            # Log PT hours export
            logger_handler.logger.info(f"Export: Employee {employee_id} PT hours: {pt_hours:.2f}")
            current_row += 1
        
        # Write GRAND TOTAL row
        ws.cell(row=current_row, column=7, value='GRAND TOTAL: ').font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=current_row, column=9, value=round(grand_regular_hours, 2)).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=current_row, column=10, value=round(grand_ot_hours, 2)).font = Font(name='Arial', size=10, bold=True)
        current_row += 1
        
        # Empty row after each employee
        current_row += 1
    
    # Auto-size columns - handle merged cells properly
    for col_idx in range(1, 15):
        column_letter = get_column_letter(col_idx)
        
        # Set fixed width for Day column (column A)
        if col_idx == 1:
            ws.column_dimensions[column_letter].width = 18
            continue
        
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Filename
    if date_range_str:
        filename = f'{project_name_for_filename}time_attendance_{date_range_str}.xlsx'
    else:
        filename = f'{project_name_for_filename}time_attendance.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/time-attendance/export/excel')
@login_required
@log_user_activity('time_attendance_excel_export')
def excel_export_time_attendance():
    """Excel export with current page filters"""
    # Redirect to main export with Excel format
    return redirect(url_for('export_time_attendance', format='excel', **request.args))

@app.route('/time-attendance/export-by-building')
@login_required
@log_user_activity('time_attendance_export_by_building')
def export_time_attendance_by_building():
    """Export time attendance records grouped by building/location to Excel"""
    try:
        # Get filter parameters (same as records page)
        employee_filter = request.args.get('employee_id')
        location_filter = request.args.get('location_name')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        import_batch = request.args.get('import_batch')
        project_filter = request.args.get('project_id')
        
        # Build query with same filters as the view
        from models.time_attendance import TimeAttendance
        query = TimeAttendance.query
        
        # Apply filters
        if employee_filter:
            query = query.filter(TimeAttendance.employee_id == employee_filter)
        
        if location_filter:
            query = query.filter(TimeAttendance.location_name == location_filter)
        
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(TimeAttendance.attendance_date >= start_date_obj)
            except ValueError:
                flash('Invalid start date format.', 'error')
                return redirect(url_for('time_attendance_records'))
        
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(TimeAttendance.attendance_date <= end_date_obj)
            except ValueError:
                flash('Invalid end date format.', 'error')
                return redirect(url_for('time_attendance_records'))
        
        if import_batch:
            query = query.filter(TimeAttendance.import_batch_id == import_batch)
        
        if project_filter:
            query = query.filter(TimeAttendance.project_id == project_filter)
        
        # Order by location, date, and time
        records = query.order_by(
            TimeAttendance.location_name,
            TimeAttendance.attendance_date.desc(),
            TimeAttendance.attendance_time.desc()
        ).all()
        
        if not records:
            flash('No records found to export.', 'warning')
            return redirect(url_for('time_attendance_records'))
        
        # Get project name if project filter exists
        project_name_for_filename = ''
        if project_filter:
            try:
                from models.project import Project
                project = Project.query.get(int(project_filter))
                if project:
                    # Replace spaces and special characters with underscores
                    project_name_safe = project.name.replace(' ', '_').replace('/', '_').replace('\\', '_')
                    project_name_for_filename = f"{project_name_safe}_"
            except Exception as e:
                print(f"⚠️ Error getting project name for filename: {e}")
        
        # Log export
        logger_handler.logger.info(
            f"User {session['username']} exported {len(records)} time attendance records "
            f"by building in Excel format"
        )
        
        # Format dates for filename (MMDDYYYY format)
        date_from_formatted = ''
        date_to_formatted = ''
        if start_date:
            try:
                date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                date_from_formatted = date_obj.strftime('%m%d%Y')
            except ValueError:
                pass
        
        if end_date:
            try:
                date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                date_to_formatted = date_obj.strftime('%m%d%Y')
            except ValueError:
                pass
        
        # Build filename with date range
        date_range_str = ''
        if date_from_formatted and date_to_formatted:
            date_range_str = f"{date_from_formatted}_{date_to_formatted}"
        elif date_from_formatted:
            date_range_str = f"from_{date_from_formatted}"
        elif date_to_formatted:
            date_range_str = f"to_{date_to_formatted}"
        
        return export_time_attendance_by_building_excel(records, project_name_for_filename, date_range_str, start_date, end_date)
    
    except Exception as e:
        logger_handler.logger.error(f"Error exporting time attendance records by building: {e}")
        flash('Error generating export file. Please try again.', 'error')
        return redirect(url_for('time_attendance_records'))
    
def export_time_attendance_by_building_excel(records, project_name_for_filename, date_range_str, start_date_filter=None, end_date_filter=None):
    """Generate Excel export grouped by building/location with template format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    
    # Get date range for calculations
    if start_date_filter and end_date_filter:
        if isinstance(start_date_filter, str):
            start_date = datetime.strptime(start_date_filter, '%Y-%m-%d').date()
        else:
            start_date = start_date_filter
        
        if isinstance(end_date_filter, str):
            end_date = datetime.strptime(end_date_filter, '%Y-%m-%d').date()
        else:
            end_date = end_date_filter
    elif records:
        start_date = min(r.attendance_date for r in records)
        end_date = max(r.attendance_date for r in records)
    else:
        return None
    
    # Import parse function for work type detection
    from working_hours_calculator import parse_employee_id_for_work_type
    
    # Convert TimeAttendance records to format expected by calculator
    converted_records = []
    for record in records:
        distance_value = getattr(record, 'distance', None)
        
        record_type = 'check_in'
        if hasattr(record, 'action_description') and record.action_description:
            action_lower = record.action_description.lower()
            if 'out' in action_lower or 'checkout' in action_lower:
                record_type = 'check_out'
        
        _, work_type = parse_employee_id_for_work_type(str(record.employee_id))
        
        base_location_name = record.location_name
        if work_type and work_type in ('PT', 'SP', 'PW'):
            display_location_name = f"{base_location_name} ({work_type})"
        else:
            display_location_name = base_location_name
        
        converted_record = type('Record', (), {
            'id': record.id,
            'employee_id': str(record.employee_id),
            'employee_name': record.employee_name,
            'check_in_date': record.attendance_date,
            'check_in_time': record.attendance_time,
            'location_name': display_location_name,
            'original_location_name': base_location_name,
            'work_type': work_type,
            'latitude': None,
            'longitude': None,
            'distance': distance_value,
            'record_type': record_type,
            'action_description': record.action_description,
            'event_description': record.event_description or '',
            'recorded_address': record.recorded_address or '',
            'qr_code': type('QRCode', (), {
                'location': base_location_name,
                'location_address': record.recorded_address or '',
                'project': None
            })()
        })()
        converted_records.append(converted_record)
    
    # Group records by location (building)
    location_groups = {}
    for record in converted_records:
        loc_name = record.original_location_name or 'Unknown Location'
        if loc_name not in location_groups:
            location_groups[loc_name] = []
        location_groups[loc_name].append(record)
    
    # Sort locations alphabetically
    sorted_locations = sorted(location_groups.keys())
    
    # Log grouping info
    logger_handler.logger.info(
        f"Export by Building: Grouped {len(converted_records)} records into {len(sorted_locations)} locations"
    )
    
    # Calculate working hours using WorkingHoursCalculator for SP/PT/PW hours
    calculator = WorkingHoursCalculator()
    hours_data = calculator.calculate_all_employees_hours(
        datetime.combine(start_date, datetime.min.time()),
        datetime.combine(end_date, datetime.max.time()),
        converted_records
    )
    
    # Get employee names map
    # Look up from Employee table using the numeric base_id to get the correct name,
    # regardless of what is stored in the employee_name column (which may contain
    # work type characters such as 'Employee 3937SP' if imported with a decorated ID).
    employee_names = {}
    for record in records:
        base_id, _ = parse_employee_id_for_work_type(str(record.employee_id))
        if base_id not in employee_names:
            try:
                emp = Employee.query.filter_by(id=int(base_id)).first()
                if emp:
                    employee_names[base_id] = f"{emp.lastName}, {emp.firstName}"
                else:
                    # Fallback: use stored name if Employee table lookup fails
                    employee_names[base_id] = record.employee_name
                    logger_handler.logger.warning(f"Employee ID {base_id} not found in employee table during export (by-building); using stored name.")
            except Exception as e:
                employee_names[base_id] = record.employee_name
                logger_handler.logger.warning(f"Could not lookup employee name for ID {base_id} during export (by-building): {e}")
    
    # Setup styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    italic_bold_font = Font(name='Arial', size=10, bold=True, italic=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    missed_punch_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    # Write main headers
    current_row = 1
    
    # Row 1: Company name
    ws.merge_cells(f'A{current_row}:N{current_row}')
    title_cell = ws.cell(row=current_row, column=1, value=os.environ.get('COMPANY_NAME', 'Your Company'))
    title_cell.font = Font(name='Arial', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 2: Summary title
    ws.merge_cells(f'A{current_row}:N{current_row}')
    summary_cell = ws.cell(row=current_row, column=1, value='Summary report of Hours worked')
    summary_cell.font = Font(name='Arial', size=12, bold=True)
    summary_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 3: Project name
    project_display = project_name_for_filename.replace('_', ' ').strip() if project_name_for_filename else "[Project Name]"
    project_cell = ws.cell(row=current_row, column=1, value=project_display)
    project_cell.font = Font(name='Arial', size=11, bold=True)
    project_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 4: Date range
    date_range_text = f"Date range: {start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"
    ws.merge_cells(f'A{current_row}:N{current_row}')
    date_cell = ws.cell(row=current_row, column=1, value=date_range_text)
    date_cell.font = Font(name='Arial', size=11)
    date_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Empty rows before first building
    current_row += 2
    
    # Process each building/location
    for location_index, location_name in enumerate(sorted_locations, 1):
        location_records = location_groups[location_name]
        
        # Get zone info from QR code if available
        zone_info = ''
        try:
            qr_code = QRCode.query.filter_by(location=location_name).first()
            if qr_code:
                zone_info = getattr(qr_code, 'zone', '') or ''
        except:
            pass
        
        # Building header row
        building_header = f"{location_index}) {location_name} - Zone {zone_info}"
        ws.merge_cells(f'A{current_row}:O{current_row}')
        building_cell = ws.cell(row=current_row, column=1, value=building_header)
        building_cell.font = Font(name='Arial', size=11, bold=True)
        building_cell.alignment = Alignment(horizontal='left')
        current_row += 1
        
        # Get unique employees for this location
        employees_at_location = {}
        for record in location_records:
            base_id, _ = parse_employee_id_for_work_type(record.employee_id)
            if base_id not in employees_at_location:
                employees_at_location[base_id] = []
            employees_at_location[base_id].append(record)
        
        # Sort employees by name
        sorted_employee_ids = sorted(
            employees_at_location.keys(),
            key=lambda emp_id: employee_names.get(emp_id, f'Employee {emp_id}').lower()
        )
        
        # Process each employee at this location
        for employee_id in sorted_employee_ids:
            emp_records = employees_at_location[employee_id]
            emp_name = employee_names.get(employee_id, f'Employee {employee_id}')
            
            # Get SP/PT/PW hours from the calculator's hours_data for this employee
            emp_hours_data = hours_data.get('employees', {}).get(employee_id, {})
            grand_totals = emp_hours_data.get('grand_totals', {})
            sp_hours = grand_totals.get('sp_hours', 0.0)
            pw_hours = grand_totals.get('pw_hours', 0.0)
            pt_hours = grand_totals.get('pt_hours', 0.0)
            
            # Employee header row
            ws.merge_cells(f'A{current_row}:O{current_row}')
            emp_header = ws.cell(row=current_row, column=1, 
                                value=f'Employee ID {employee_id}: {emp_name}')
            emp_header.font = Font(name='Arial', size=11, bold=True)
            emp_header.alignment = Alignment(horizontal='left')
            current_row += 1
            
            # Column headers
            headers = ['Day', 'Date', 'In', 'Out', 'Location', 'Zone', 'Hours/Building',
                      'Daily Total', 'Regular Hours', 'OT Hours', 'Building Address',
                      'Recorded Location', 'Distance (Mile)', 'Possible Violation']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Group employee records by date
            daily_records = {}
            for record in emp_records:
                date_key = record.check_in_date.strftime('%Y-%m-%d')
                if date_key not in daily_records:
                    daily_records[date_key] = []
                daily_records[date_key].append(record)
            
            # Track weekly hours for overtime calculation
            weekly_total_hours = 0
            current_week_start = None
            grand_regular_hours = 0
            grand_ot_hours = 0
            
            # Sort dates
            sorted_dates = sorted(daily_records.keys())
            
            for date_str in sorted_dates:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                day_records = sorted(daily_records[date_str], key=lambda x: x.check_in_time)
                
                # Check for week boundary
                week_start = date_obj - timedelta(days=date_obj.weekday())
                if current_week_start is not None and week_start != current_week_start:
                    # Write weekly total row
                    week_regular = min(weekly_total_hours, 40.0)
                    week_overtime = max(0, weekly_total_hours - 40.0)
                    
                    ws.cell(row=current_row, column=7, value='Weekly Total: ').font = bold_font
                    ws.cell(row=current_row, column=8, value=round(weekly_total_hours, 2)).font = bold_font
                    ws.cell(row=current_row, column=9, value=round(week_regular, 2)).font = bold_font
                    ws.cell(row=current_row, column=10, value=round(week_overtime, 2)).font = bold_font
                    
                    grand_regular_hours += week_regular
                    grand_ot_hours += week_overtime
                    current_row += 1
                    
                    weekly_total_hours = 0
                
                current_week_start = week_start
                
                # Process day records - create IN/OUT pairs
                record_info = []
                for record in day_records:
                    action_desc = record.action_description.lower() if record.action_description else ''
                    is_out = 'out' in action_desc or 'checkout' in action_desc
                    record_info.append({
                        'record': record,
                        'is_out': is_out,
                        'used': False
                    })
                
                # Create pairs
                pairs = []
                i = 0
                while i < len(record_info):
                    if record_info[i]['used']:
                        i += 1
                        continue
                    
                    if not record_info[i]['is_out']:  # IN
                        out_found = False
                        for j in range(i + 1, len(record_info)):
                            if record_info[j]['used']:
                                continue
                            if record_info[j]['is_out']:
                                pairs.append({
                                    'check_in': record_info[i]['record'],
                                    'check_out': record_info[j]['record'],
                                    'is_miss_punch': False
                                })
                                record_info[i]['used'] = True
                                record_info[j]['used'] = True
                                out_found = True
                                break
                        
                        if not out_found:
                            pairs.append({
                                'check_in': record_info[i]['record'],
                                'check_out': None,
                                'is_miss_punch': True
                            })
                            record_info[i]['used'] = True
                    else:  # Orphaned OUT
                        pairs.append({
                            'check_in': None,
                            'check_out': record_info[i]['record'],
                            'is_miss_punch': True
                        })
                        record_info[i]['used'] = True
                    
                    i += 1
                
                # Calculate daily hours (Daily Total = quarter-hour rounding)
                _raw_day_min = 0
                for pair in pairs:
                    if pair['check_in'] and pair['check_out'] and not pair['is_miss_punch']:
                        pair_in = datetime.combine(date_obj, pair['check_in'].check_in_time)
                        pair_out = datetime.combine(date_obj, pair['check_out'].check_in_time)
                        _raw_day_min += (pair_out - pair_in).total_seconds() / 60.0
                
                daily_hours = round_time_to_quarter_hour(_raw_day_min) / 60.0
                weekly_total_hours += daily_hours
                
                # Write pairs
                for pair_idx, pair in enumerate(pairs):
                    check_in = pair['check_in']
                    check_out = pair['check_out']
                    is_miss_punch = pair['is_miss_punch']
                    
                    # Day/date only on first row
                    day_display = date_obj.strftime('%A').upper() if pair_idx == 0 else ''
                    date_display = date_obj.strftime('%m/%d/%Y') if pair_idx == 0 else ''
                    
                    # Calculate hours for this pair
                    if check_in and check_out and not is_miss_punch:
                        pair_hours = round((datetime.combine(date_obj, check_out.check_in_time) - 
                                           datetime.combine(date_obj, check_in.check_in_time)).total_seconds() / 3600.0, 2)
                    else:
                        pair_hours = 'Missed Punch'
                    
                    # Daily total only on last row of day
                    daily_total_display = daily_hours if pair_idx == len(pairs) - 1 else ''
                    
                    # Get record for address/distance info
                    ref_record = check_in or check_out
                    
                    # Build row data
                    row_data = [
                        day_display,
                        date_display,
                        check_in.check_in_time.strftime('%I:%M:%S %p') if check_in else '',
                        check_out.check_in_time.strftime('%I:%M:%S %p') if check_out else '',
                        ref_record.location_name if ref_record else '',
                        zone_info,
                        pair_hours,
                        daily_total_display if daily_total_display else '',
                        '',  # Regular Hours
                        '',  # OT Hours
                        '',  # Building Address (will be HYPERLINK)
                        '',  # Recorded Location (will be HYPERLINK)
                        getattr(ref_record, 'distance', None) or '' if ref_record else '',
                        calculate_possible_violation(getattr(ref_record, 'distance', None)) if ref_record else ''
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = data_font
                        cell.border = border
                        if col == 7 and value == 'Missed Punch':
                            cell.fill = missed_punch_fill
                    
                    # Add HYPERLINK formulas for addresses
                    if ref_record:
                        building_address = ref_record.event_description or ''
                        if building_address:
                            encoded_addr = building_address.replace(' ', '+').replace(',', '%2C')
                            hyperlink_formula = f'=HYPERLINK("https://www.google.com/maps/place/{encoded_addr}","{building_address}")'
                            ws.cell(row=current_row, column=11, value=hyperlink_formula)
                        
                        recorded_addr = ref_record.recorded_address or ''
                        if recorded_addr:
                            encoded_recorded = recorded_addr.replace(' ', '+').replace(',', '%2C')
                            recorded_hyperlink = f'=HYPERLINK("https://www.google.com/maps/place/{encoded_recorded}","{recorded_addr}")'
                            ws.cell(row=current_row, column=12, value=recorded_hyperlink)
                    
                    current_row += 1
            
            # Write final weekly total
            if weekly_total_hours > 0:
                week_regular = min(weekly_total_hours, 40.0)
                week_overtime = max(0, weekly_total_hours - 40.0)
                
                ws.cell(row=current_row, column=7, value='Weekly Total: ').font = bold_font
                ws.cell(row=current_row, column=8, value=round(weekly_total_hours, 2)).font = bold_font
                ws.cell(row=current_row, column=9, value=round(week_regular, 2)).font = bold_font
                ws.cell(row=current_row, column=10, value=round(week_overtime, 2)).font = bold_font
                
                grand_regular_hours += week_regular
                grand_ot_hours += week_overtime
                current_row += 1
            
            # ================================================================
            # Write extra working hours rows (SP/PW/PT) if employee has any
            # This matches the behavior of the regular Export to Excel
            # ================================================================
            
            # Write SP row if hours > 0
            if sp_hours > 0:
                ws.cell(row=current_row, column=7, value='Special Project (SP): ').font = italic_bold_font
                ws.cell(row=current_row, column=9, value=round(sp_hours, 2)).font = italic_bold_font
                # Log SP hours export
                logger_handler.logger.info(f"Export by Building: Employee {employee_id} SP hours: {sp_hours:.2f}")
                current_row += 1
            
            # Write PW row if hours > 0
            if pw_hours > 0:
                ws.cell(row=current_row, column=7, value='Periodic Work (PW): ').font = italic_bold_font
                ws.cell(row=current_row, column=9, value=round(pw_hours, 2)).font = italic_bold_font
                # Log PW hours export
                logger_handler.logger.info(f"Export by Building: Employee {employee_id} PW hours: {pw_hours:.2f}")
                current_row += 1
            
            # Write PT row if hours > 0
            if pt_hours > 0:
                ws.cell(row=current_row, column=7, value='Project Team (PT): ').font = italic_bold_font
                ws.cell(row=current_row, column=9, value=round(pt_hours, 2)).font = italic_bold_font
                # Log PT hours export
                logger_handler.logger.info(f"Export by Building: Employee {employee_id} PT hours: {pt_hours:.2f}")
                current_row += 1
            
            # ================================================================
            # End of extra working hours section
            # ================================================================
            
            # Write GRAND TOTAL row
            ws.cell(row=current_row, column=7, value='GRAND TOTAL: ').font = Font(name='Arial', size=10, bold=True)
            ws.cell(row=current_row, column=9, value=round(grand_regular_hours, 2)).font = Font(name='Arial', size=10, bold=True)
            ws.cell(row=current_row, column=10, value=round(grand_ot_hours, 2)).font = Font(name='Arial', size=10, bold=True)
            current_row += 1
            
            # Empty row after each employee
            current_row += 1
        
        # Empty row after each building
        current_row += 1
    
    # Auto-size columns
    for col_idx in range(1, 15):
        column_letter = get_column_letter(col_idx)
        
        if col_idx == 1:
            ws.column_dimensions[column_letter].width = 18
            continue
        
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Filename
    if date_range_str:
        filename = f'{project_name_for_filename}time_attendance_by_building_{date_range_str}.xlsx'
    else:
        filename = f'{project_name_for_filename}time_attendance_by_building.xlsx'
    
    # Log successful export
    logger_handler.logger.info(
        f"Export by Building completed: {filename} with {len(sorted_locations)} buildings"
    )
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/time-attendance/records')
@login_required
@log_user_activity('time_attendance_records_view')
def time_attendance_records():
    """Display time attendance records with filtering options"""
    try:
        # Get filter parameters
        employee_filter = request.args.get('employee_id')
        location_filter = request.args.get('location_name')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        project_filter = request.args.get('project_id')
        page = request.args.get('page', 1, type=int)
        per_page = 50  # Records per page
        
        # Build query
        query = TimeAttendance.query
        
        # Apply filters
        if employee_filter:
            query = query.filter(TimeAttendance.employee_id == employee_filter)
        
        if location_filter:
            query = query.filter(TimeAttendance.location_name == location_filter)

        if project_filter:
            query = query.filter(TimeAttendance.project_id == project_filter)
        
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(TimeAttendance.attendance_date >= start_date_obj)
            except ValueError:
                flash('Invalid start date format.', 'error')
        
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(TimeAttendance.attendance_date <= end_date_obj)
            except ValueError:
                flash('Invalid end date format.', 'error')
        
        # Order by date and time (most recent first)
        query = query.order_by(
            TimeAttendance.attendance_date.desc(),
            TimeAttendance.attendance_time.desc()
        )
        
        # Paginate results
        records = query.paginate(page=page, per_page=per_page, error_out=False)
        
        # Enhance records with QR address and location accuracy
        for record in records.items:
            # Find matching QR code by location name
            qr_code = QRCode.query.filter_by(location=record.location_name).first()
            
            if qr_code:
                record.qr_address = qr_code.location_address
                
                # Calculate location accuracy if coordinates are available
                if record.recorded_address and qr_code.location_address:
                    try:
                        # Try to calculate location accuracy
                        location_accuracy = calculate_location_accuracy_enhanced(
                            qr_address=qr_code.location_address,
                            checkin_address=record.recorded_address,
                            checkin_lat=None,  # TimeAttendance doesn't have GPS coords
                            checkin_lng=None
                        )
                        record.location_accuracy = location_accuracy
                    except Exception as e:
                        logger_handler.logger.warning(f"Could not calculate location accuracy for record {record.id}: {e}")
                        record.location_accuracy = None
                else:
                    record.location_accuracy = None
            else:
                record.qr_address = None
                record.location_accuracy = None

            # Resolve employee name by stripping work type prefix/suffix (SP, PW, PT)
            # from employee_id ONLY for the lookup. The original employee_id is kept intact.
            # e.g. '3937SP', 'SP3937', 'PW3937' -> lookup by numeric '3937'
            try:
                import re as _re
                numeric_only = _re.search(r'\d+', str(record.employee_id or ''))
                if numeric_only:
                    emp = Employee.query.filter_by(id=int(numeric_only.group(0))).first()
                    record.resolved_employee_name = f"{emp.lastName}, {emp.firstName}" if emp else record.employee_name
                else:
                    record.resolved_employee_name = record.employee_name
            except Exception as e:
                logger_handler.logger.warning(f"Could not resolve employee name for ID {record.employee_id}: {e}")
                record.resolved_employee_name = record.employee_name
        
        # Get unique employees and locations for filters
        unique_employees = TimeAttendance.get_unique_employees()
        unique_locations = TimeAttendance.get_unique_locations()
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        
        return render_template(
            'time_attendance_records.html',
            records=records,
            unique_employees=unique_employees,
            unique_locations=unique_locations,
            projects=projects
        )
        
    except Exception as e:
        logger_handler.logger.error(f"Error displaying time attendance records: {e}")
        flash('Error loading attendance records.', 'error')
        return redirect(url_for('time_attendance_dashboard'))

@app.route('/time-attendance/record/<int:record_id>')
@login_required
@log_user_activity('time_attendance_record_detail')
def time_attendance_record_detail(record_id):
    """Display detailed view of a time attendance record"""
    try:
        record = TimeAttendance.query.get_or_404(record_id)
        return render_template('time_attendance_record_detail.html', record=record)
        
    except Exception as e:
        logger_handler.logger.error(f"Error viewing time attendance record {record_id}: {e}")
        flash('Error loading record details.', 'error')
        return redirect(url_for('time_attendance_records'))

@app.route('/time-attendance/delete/<int:record_id>', methods=['POST'])
@admin_required
@log_database_operations('time_attendance_delete')
def delete_time_attendance_record(record_id):
    """Delete a time attendance record"""
    try:
        record = TimeAttendance.query.get_or_404(record_id)
        
        # Store record info for logging
        employee_info = f"{record.employee_name} (ID: {record.employee_id})"
        location_info = record.location_name
        date_info = record.attendance_date
        
        # Delete the record
        db.session.delete(record)
        db.session.commit()
        
        # Log deletion
        logger_handler.logger.info(
            f"User {session['username']} deleted time attendance record {record_id} - "
            f"Employee: {employee_info}, Location: {location_info}, Date: {date_info}"
        )
        
        flash(f'Time attendance record for {employee_info} deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('time_attendance_delete', e)
        flash('Failed to delete time attendance record.', 'error')
    
    # Get filter parameters from BOTH request.form (POST) and request.args (GET query params)
    # This handles both the records list page and the detail page
    filter_params = {}
    
    # List of possible filter parameters
    filter_keys = ['employee_id', 'location_name', 'project_id', 'start_date', 'end_date', 'page']
    
    for key in filter_keys:
        # Try to get from form data first (records list page)
        value = request.form.get(key)
        # If not in form, try query parameters (detail page)
        if not value:
            value = request.args.get(key)
        # Only include if value exists and is not empty
        if value:
            filter_params[key] = value
    
    # Redirect back with filters preserved
    return redirect(url_for('time_attendance_records', **filter_params))

@app.route('/api/time-attendance/employee/<employee_id>')
@login_required
def api_time_attendance_by_employee(employee_id):
    """API endpoint to get time attendance records for a specific employee"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        start_date_obj = None
        end_date_obj = None
        
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        records = TimeAttendance.get_by_employee_id(employee_id, start_date_obj, end_date_obj)
        
        return jsonify({
            'success': True,
            'employee_id': employee_id,
            'total_records': len(records),
            'records': [record.to_dict() for record in records]
        })
        
    except Exception as e:
        logger_handler.logger.error(f"API error getting time attendance for employee {employee_id}: {e}")
        return jsonify({
            'success': False,
            'error': 'Failed to retrieve time attendance records'
        }), 500

@app.route('/api/time-attendance/location/<location_name>')
@login_required
def api_time_attendance_by_location(location_name):
    """API endpoint to get time attendance records for a specific location"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        start_date_obj = None
        end_date_obj = None
        
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        records = TimeAttendance.get_by_location(location_name, start_date_obj, end_date_obj)
        
        return jsonify({
            'success': True,
            'location_name': location_name,
            'total_records': len(records),
            'records': [record.to_dict() for record in records]
        })
        
    except Exception as e:
        logger_handler.logger.error(f"API error getting time attendance for location {location_name}: {e}")
        return jsonify({
            'success': False,
            'error': 'Failed to retrieve time attendance records'
        }), 500
    
create_location_logging_routes(app, db, logger_handler)


# Jinja2 filters for better template functionality
@app.template_filter('days_since')
def days_since_filter(date):
    """Calculate days since a given date"""
    if not date:
        return 0
    from datetime import datetime
    now = datetime.utcnow()
    return (now - date).days

@app.template_filter('time_ago')
def time_ago_filter(date):
    """Human readable time ago"""
    if not date:
        return 'Never'
    from datetime import datetime
    now = datetime.utcnow()
    diff = now - date

    if diff.days > 365:
        years = diff.days // 365
        return f"{years} year{'s' if years != 1 else ''} ago"
    elif diff.days > 30:
        months = diff.days // 30
        return f"{months} month{'s' if months != 1 else ''} ago"
    elif diff.days > 0:
        return f"{diff.days} day{'s' if diff.days != 1 else ''} ago"
    elif diff.seconds > 3600:
        hours = diff.seconds // 3600
        return f"{hours} hour{'s' if hours != 1 else ''} ago"
    elif diff.seconds > 60:
        minutes = diff.seconds // 60
        return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
    else:
        return "Just now"

# Error handlers
@app.errorhandler(500)
def internal_error(error):
    """Handle internal server errors with user-friendly page"""
    if app.debug:
        # Let Flask handle debug errors naturally
        return None

    return '''
    <!DOCTYPE html>
    <html>
    <head><title>Server Error</title></head>
    <body style="font-family: Arial; text-align: center; margin-top: 100px;">
        <h1>🔧 Something went wrong</h1>
        <p>We're working to fix this issue. Please try again later.</p>
        <a href="/" style="color: #2563eb;">← Back to Home</a>
    </body>
    </html>
    ''', 500

@app.errorhandler(404)
def not_found(error):
    """Handle page not found errors"""
    return '''
    <!DOCTYPE html>
    <html>
    <head><title>Page Not Found</title></head>
    <body style="font-family: Arial; text-align: center; margin-top: 100px;">
        <h1>🔍 Page Not Found</h1>
        <p>The page you're looking for doesn't exist.</p>
        <a href="/" style="color: #2563eb;">← Back to Home</a>
    </body>
    </html>
    ''', 404

# Initialize database tables
@log_database_operations('database_initialization')
def create_tables():
    """Create database tables and default admin user with logging"""
    try:
        db.create_all()

        # Create default admin user if not exists
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(
                full_name='System Administrator',
                email='admin@example.com',
                username='admin',
                role='admin'
            )
            admin.set_password('admin123')  # Change this in production
            db.session.add(admin)
            db.session.commit()

            # Log admin user creation
            logger_handler.logger.info("Default admin user created during initialization")

        # Initialize logging table
        logger_handler._create_log_table()

    except Exception as e:
        logger_handler.log_database_error('database_initialization', e)
        raise

def update_existing_qr_codes():
    """Update existing QR codes with URLs and regenerate QR images with logging"""
    try:
        qr_codes = QRCode.query.filter_by(active_status=True).all()
        updated_count = 0

        for qr_code in qr_codes:
            if not qr_code.qr_url or not qr_code.qr_code_image:
                try:
                    # Generate URL if missing
                    if not qr_code.qr_url:
                        qr_code.qr_url = generate_qr_url(qr_code.name, qr_code.id)

                    # Generate QR image if missing
                    if not qr_code.qr_code_image:
                        qr_data = f"{request.url_root}qr/{qr_code.qr_url}"

                        # Use styling from database if available, otherwise defaults
                        styling = get_qr_styling(qr_code)
                        qr_code.qr_code_image = generate_qr_code(
                            data=qr_data,
                            fill_color=styling['fill_color'],
                            back_color=styling['back_color'],
                            box_size=styling['box_size'],
                            border=styling['border'],
                            error_correction=styling['error_correction']
                        )

                    updated_count += 1

                except Exception as e:
                    logger_handler.log_flask_error(
                        error_type="qr_code_update_error",
                        error_message=f"Failed to update QR code {qr_code.id}: {str(e)}"
                    )
                    continue

        if updated_count > 0:
            db.session.commit()
            logger_handler.logger.info(f"Updated {updated_count} existing QR codes with missing URLs/images")

    except Exception as e:
        logger_handler.log_database_error('update_existing_qr_codes', e)
        print(f"Error updating existing QR codes: {e}")

def add_qr_customization_columns():
    """Add QR code customization columns to existing table"""
    try:
        with app.app_context():
            # Add columns to QRCode table if they don't exist
            db.engine.execute("""
                ALTER TABLE qr_codes 
                ADD COLUMN IF NOT EXISTS fill_color VARCHAR(7) DEFAULT '#000000',
                ADD COLUMN IF NOT EXISTS back_color VARCHAR(7) DEFAULT '#FFFFFF',
                ADD COLUMN IF NOT EXISTS box_size INT DEFAULT 10,
                ADD COLUMN IF NOT EXISTS border INT DEFAULT 4,
                ADD COLUMN IF NOT EXISTS error_correction VARCHAR(1) DEFAULT 'L',
                ADD COLUMN IF NOT EXISTS style_id INT,
                ADD FOREIGN KEY (style_id) REFERENCES qr_code_styles(id)
            """)

            # Create QRCodeStyle table
            db.create_all()

            # Insert default styles
            default_styles = [
                QRCodeStyle(name="Classic Black", fill_color="#000000", back_color="#FFFFFF", is_default=True),
                QRCodeStyle(name="Blue Professional", fill_color="#2563eb", back_color="#FFFFFF"),
                QRCodeStyle(name="Green Success", fill_color="#10b981", back_color="#FFFFFF"),
                QRCodeStyle(name="Red Alert", fill_color="#ef4444", back_color="#FFFFFF"),
                QRCodeStyle(name="Dark Mode", fill_color="#FFFFFF", back_color="#1f2937"),
                QRCodeStyle(name="High Contrast", fill_color="#000000", back_color="#FFFF00"),
                QRCodeStyle(name="Corporate Blue", fill_color="#1e40af", back_color="#f8fafc"),
                QRCodeStyle(name="Minimalist Gray", fill_color="#6b7280", back_color="#FFFFFF")
            ]

            for style in default_styles:
                if not QRCodeStyle.query.filter_by(name=style.name).first():
                    db.session.add(style)

            db.session.commit()
            print("QR Code customization tables and default styles created successfully!")

    except Exception as e:
        db.session.rollback()
        print(f"Error adding QR customization columns: {e}")

def add_coordinate_columns():
    """Add coordinate columns to existing qr_codes table (MySQL compatible)"""
    try:
        # Check if columns already exist - MySQL compatible query
        result = db.session.execute(text("""
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'qr_codes' 
            AND COLUMN_NAME IN ('address_latitude', 'address_longitude', 'coordinate_accuracy', 'coordinates_updated_date')
        """))

        existing_columns = [row.COLUMN_NAME for row in result.fetchall()]

        # Add missing columns
        if 'address_latitude' not in existing_columns:
            db.session.execute(text("""
                ALTER TABLE qr_codes ADD COLUMN address_latitude FLOAT
            """))
            print("✅ Added address_latitude column")

        if 'address_longitude' not in existing_columns:
            db.session.execute(text("""
                ALTER TABLE qr_codes ADD COLUMN address_longitude FLOAT
            """))
            print("✅ Added address_longitude column")

        if 'coordinate_accuracy' not in existing_columns:
            db.session.execute(text("""
                ALTER TABLE qr_codes ADD COLUMN coordinate_accuracy VARCHAR(50) DEFAULT 'geocoded'
            """))
            print("✅ Added coordinate_accuracy column")

        if 'coordinates_updated_date' not in existing_columns:
            db.session.execute(text("""
                ALTER TABLE qr_codes ADD COLUMN coordinates_updated_date TIMESTAMP
            """))
            print("✅ Added coordinates_updated_date column")

        db.session.commit()
        print("✅ Database migration completed successfully")

    except Exception as e:
        print(f"❌ Database migration error: {e}")
        db.session.rollback()

# Application context processor for logging status
@app.context_processor
def inject_logging_status():
    """Inject logging status into all templates"""
    return {
        'logging_enabled': hasattr(app, 'logger_handler'),
        'is_admin': has_admin_privileges(session.get('role', ''))
    }

@app.context_processor
def inject_turnstile():
    """Inject Turnstile settings into all templates"""
    return {
        'turnstile_enabled': turnstile_utils.is_enabled(),
        'turnstile_site_key': turnstile_utils.get_site_key()
    }

# Before request handler for request logging
@app.before_request
def log_request_info():
    """Log request information for security monitoring"""
    # Skip logging for static files and API calls
    if (request.endpoint and
        (request.endpoint.startswith('static') or
         request.path.startswith('/api/logs'))):
        return

    # Log suspicious activity
    user_agent = request.headers.get('User-Agent', '')
    ip_address = request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)

    # Check for potential security threats
    suspicious_patterns = [
        'sqlmap', 'nikto', 'nmap', 'dirb', 'dirbuster',
        'wget', 'curl.*bot', 'scanner', 'exploit'
    ]

    if any(pattern in user_agent.lower() for pattern in suspicious_patterns):
        logger_handler.log_security_event(
            event_type="suspicious_user_agent",
            description=f"Suspicious user agent detected: {user_agent[:200]}",
            severity="HIGH",
            additional_data={'user_agent': user_agent, 'ip_address': ip_address}
        )

# After request handler for performance monitoring
@app.after_request
def log_response_info(response):
    """Log response information for performance monitoring"""
    # Skip logging for static files
    if request.endpoint and request.endpoint.startswith('static'):
        return response

    # Log slow requests (over 5 seconds)
    if hasattr(request, 'start_time'):
        duration = time.time() - request.start_time
        if duration > 5.0:
            logger_handler.logger.warning(f"Slow request: {request.path} took {duration:.2f} seconds")

    # Log error responses
    if response.status_code >= 400:
        logger_handler.logger.warning(
            f"Error response: {response.status_code} for {request.path} "
            f"by user {session.get('username', 'anonymous')}"
        )

    return response

def get_optimized_statistics(date_from=None, date_to=None, project_filter=None):
    """Optimized statistics query using new indexes"""
    try:
        # Base conditions with optimized WHERE clause ordering
        conditions = ["1=1"]  # Start with always-true condition
        params = {}
        
        # Most selective conditions first for index optimization
        if date_from and date_to:
            conditions.append("ad.check_in_date BETWEEN :date_from AND :date_to")
            params.update({'date_from': date_from, 'date_to': date_to})
        elif date_from:
            conditions.append("ad.check_in_date >= :date_from")
            params['date_from'] = date_from
        elif date_to:
            conditions.append("ad.check_in_date <= :date_to")
            params['date_to'] = date_to
            
        if project_filter:
            conditions.append("qc.project_id = :project_id")
            params['project_id'] = int(project_filter)
        
        where_clause = " AND ".join(conditions)
        
        # Optimized main statistics query
        stats_query = f"""
            SELECT 
                COUNT(*) as total_scans,
                COUNT(DISTINCT ad.employee_id) as unique_employees,
                COUNT(DISTINCT ad.qr_code_id) as unique_locations,
                COUNT(CASE WHEN ad.check_in_date = CURRENT_DATE THEN 1 END) as today_scans,
                COUNT(CASE WHEN ad.latitude IS NOT NULL THEN 1 END) as gps_enabled_scans,
                AVG(CASE WHEN ad.accuracy IS NOT NULL THEN ad.accuracy END) as avg_gps_accuracy
            FROM attendance_data ad
            LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id
            WHERE {where_clause}
        """
        
        result = db.session.execute(text(stats_query), params).fetchone()
        
        # Log successful query execution
        logger_handler.log_system_event(
            event_type="optimized_statistics_query",
            description=f"Statistics generated with {result.total_scans} total scans",
            severity="INFO"
        )
        
        return result
        
    except Exception as e:
        logger_handler.log_database_error(
            error_type="statistics_query_error",
            error_message=str(e),
            query="get_optimized_statistics"
        )
        raise

def log_slow_query_performance():
    """Monitor and log slow query performance"""
    @app.before_request
    def before_request():
        g.start_time = time.time()
        
    @app.after_request
    def after_request(response):
        if hasattr(g, 'start_time'):
            duration = time.time() - g.start_time
            
            # Log slow requests (over 2 seconds)
            if duration > 2.0:
                logger_handler.log_system_event(
                    event_type="slow_query_detected",
                    description=f"Slow request: {request.endpoint} took {duration:.2f}s",
                    severity="WARNING",
                    additional_data={
                        'duration': duration,
                        'endpoint': request.endpoint,
                        'method': request.method,
                        'user': session.get('username', 'anonymous')
                    }
                )
        
        return response
    
if __name__ == '__main__':
    with app.app_context():
        try:
            # Initialize database and logging
            create_tables()
            
            # Initialize performance optimizations
            print("🚀 Initializing performance optimizations...")
            cached_query = initialize_performance_optimizations(app, db, logger_handler)
            performance_monitor = PerformanceMonitor(app, db, logger_handler)

            if cached_query:
                print("✅ Performance optimizations completed successfully")
            else:
                print("⚠️ Performance optimizations completed with warnings")

            # Log application startup
            logger_handler.logger.info("QR Attendance Management System started successfully")

        except Exception as e:
            print(f"❌ Application startup failed: {e}")
            if hasattr(app, 'logger_handler'):
                logger_handler.log_flask_error(
                    error_type="application_startup_error",
                    error_message=str(e)
                )
            raise

    app.run(debug=os.environ.get('DEBUG'),
            host=os.environ.get('FLASK_HOST'),
            port=os.environ.get('FLASK_PORT'),
            threaded=os.environ.get ('THREADED'))