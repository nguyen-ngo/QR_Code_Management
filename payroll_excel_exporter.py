#!/usr/bin/env python3
"""
Payroll Excel Exporter
=====================

Enhanced Excel exporter for payroll reports with working hours calculations.
Based on the Java PayrollReport.java implementation.

Features:
- Employee working hours with daily/weekly breakdown
- Regular and overtime hour calculations
- Travel time inclusion options
- Professional Excel formatting
- Multiple report formats
"""

import io
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from single_checkin_calculator import SingleCheckInCalculator
from logger_handler import log_database_operations


class PayrollExcelExporter:
    """Excel exporter for payroll reports with working hours"""
    
    def __init__(self, company_name: str = "Company Name", contract_name: str = "Default Contract"):
        self.company_name = company_name
        self.contract_name = contract_name
        
        # Excel styling
        self.header1_style = None
        self.header2_style = None
        self.header3_style = None
        self.table_header_style = None
        self.data_style = None
        self.total_style = None
        
    def _setup_styles(self, workbook: Workbook):
        """Setup Excel cell styles"""
        
        # Header 1 style (Company name)
        self.header1_style = {
            'font': Font(name="Arial", size=16, bold=True, color="FFFFFF"),
            'fill': PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center")
        }
        
        # Header 2 style (Report title)
        self.header2_style = {
            'font': Font(name="Arial", size=14, bold=True, color="FFFFFF"),
            'fill': PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center")
        }
        
        # Header 3 style (Contract and date range)
        self.header3_style = {
            'font': Font(name="Arial", size=12, bold=True),
            'alignment': Alignment(horizontal="center", vertical="center")
        }
        
        # Table header style
        self.table_header_style = {
            'font': Font(name="Arial", size=11, bold=True, color="FFFFFF"),
            'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }
        
        # Data style
        self.data_style = {
            'font': Font(name="Arial", size=10),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }
        
        # Total style
        self.total_style = {
            'font': Font(name="Arial", size=11, bold=True),
            'fill': PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )
        }
    
    def _apply_style(self, cell, style_dict):
        """Apply a style dictionary to a cell"""
        for attr, value in style_dict.items():
            setattr(cell, attr, value)
    
    def _get_employee_names(self, attendance_records: List[Dict]) -> Dict[str, str]:
        """Get employee names - fallback method if names not provided"""
        print("⚠️ Excel exporter falling back to internal name lookup - this should not normally happen")
        return {}  # Return empty dict - names should be provided by the route
    
    def _get_employee_names(self, attendance_records: List[Dict]) -> Dict[str, str]:
        """Get employee names using the same CAST method as attendance report"""
        employee_names = {}
        try:
            from flask import current_app
            from sqlalchemy import text
            
            # Get unique employee IDs from attendance records
            employee_ids = []
            for record in attendance_records:
                if hasattr(record, '__dict__'):
                    emp_id = str(record.employee_id)
                else:
                    emp_id = str(record['employee_id'])
                
                if emp_id not in employee_ids:
                    employee_ids.append(emp_id)
            
            if employee_ids:
                # Use the same SQL approach as attendance report - JOIN with CAST
                placeholders = ','.join([f"'{emp_id}'" for emp_id in employee_ids])
                
                # Import db from current app context
                from app import db
                
                employee_query = db.session.execute(text(f"""
                    SELECT 
                        ad.employee_id,
                        CONCAT(e.firstName, ' ', e.lastName) as full_name 
                    FROM attendance_data ad
                    LEFT JOIN employee e ON CAST(ad.employee_id AS UNSIGNED) = e.id
                    WHERE ad.employee_id IN ({placeholders})
                    GROUP BY ad.employee_id, e.firstName, e.lastName
                """))
                
                for row in employee_query:
                    if row[1]:  # Only add if we got a name
                        employee_names[str(row[0])] = row[1]
            
            print(f"📊 Excel exporter retrieved names for {len(employee_names)} employees using CAST method")
                
        except Exception as e:
            print(f"⚠️ Excel exporter could not load employee names: {e}")
            import traceback
            print(f"⚠️ Traceback: {traceback.format_exc()}")
        
        return employee_names
    
    @log_database_operations('payroll_excel_export')
    def create_payroll_report(self, start_date: datetime, end_date: datetime, 
                             attendance_records: List[Dict], employee_names: Dict[str, str] = None,
                             include_travel_time: bool = True) -> io.BytesIO:
        """
        Create a comprehensive payroll report with working hours
        
        Args:
            start_date: Report start date
            end_date: Report end date  
            attendance_records: List of attendance records
            employee_names: Dictionary mapping employee_id to full name
            include_travel_time: Whether to include travel time in calculations
            
        Returns:
            BytesIO buffer containing the Excel file
        """
        try:
            print(f"📊 Creating payroll report from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
            
            # Get employee names using the same method as payroll routes
            if employee_names is None:
                employee_names = self._get_employee_names(attendance_records)
            
            # Calculate working hours
            calculator = SingleCheckInCalculator()
            hours_data = calculator.calculate_all_employees_hours(start_date, end_date, attendance_records)
            
            # Create workbook
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Payroll Report"
            
            # Setup styles
            self._setup_styles(workbook)
            
            # Write report headers
            current_row = self._write_report_headers(worksheet, start_date, end_date)
            
            # Write employee data
            current_row = self._write_employee_payroll_data(worksheet, hours_data, employee_names, current_row)
            
            # Write summary totals
            self._write_summary_totals(worksheet, hours_data, current_row)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(worksheet)
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            print("✅ Payroll report Excel file created successfully")
            return excel_buffer
            
        except Exception as e:
            print(f"❌ Error creating payroll report: {e}")
            raise e
    
    def _write_report_headers(self, worksheet, start_date: datetime, end_date: datetime) -> int:
        """Write report headers and return next row number"""
        current_row = 1
        
        # Company name
        cell = worksheet.cell(row=current_row, column=1, value=self.company_name)
        self._apply_style(cell, self.header1_style)
        worksheet.merge_cells(f'A{current_row}:Q{current_row}')
        current_row += 1
        
        # Report title
        cell = worksheet.cell(row=current_row, column=1, value="Payroll Report - Working Hours Summary")
        self._apply_style(cell, self.header2_style)
        worksheet.merge_cells(f'A{current_row}:Q{current_row}')
        current_row += 1
        
        # Contract name
        cell = worksheet.cell(row=current_row, column=1, value=self.contract_name)
        self._apply_style(cell, self.header3_style)
        worksheet.merge_cells(f'A{current_row}:Q{current_row}')
        current_row += 1
        
        # Date range
        date_range = f"Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        cell = worksheet.cell(row=current_row, column=1, value=date_range)
        self._apply_style(cell, self.header3_style)
        worksheet.merge_cells(f'A{current_row}:Q{current_row}')
        current_row += 2  # Add extra space
        
        return current_row
    
    def _write_employee_payroll_data(self, worksheet, hours_data: Dict, employee_names: Dict[str, str], start_row: int) -> int:
        """Write employee payroll data and return next row number"""
        current_row = start_row
        
        # Table headers
        headers = [
            "#", "Employee ID", "Employee Name", 
            "Week 1 Mon", "Week 1 Tue", "Week 1 Wed", "Week 1 Thu", "Week 1 Fri", "Week 1 Sat", "Week 1 Sun",
            "Week 1 Regular", "Week 1 OT",
            "Week 2 Mon", "Week 2 Tue", "Week 2 Wed", "Week 2 Thu", "Week 2 Fri", "Week 2 Sat", "Week 2 Sun", 
            "Week 2 Regular", "Week 2 OT",
            "Total Regular", "Total OT", "Grand Total"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col, value=header)
            self._apply_style(cell, self.table_header_style)
        current_row += 1
        
        # Employee data
        counter = 1
        for employee_id, emp_data in hours_data['employees'].items():
            # Employee basic info
            employee_name = employee_names.get(employee_id, f"Employee {employee_id}") if employee_names else f"Employee {employee_id}"
            
            row_data = [
                counter,
                employee_id,
                employee_name
            ]
            
            # Calculate daily hours for two weeks (14 days)
            daily_hours = self._calculate_daily_hours_for_payroll(emp_data, hours_data['period_start'])
            
            # Week 1 daily hours (7 days)
            week1_total = 0
            for day_idx in range(7):
                hours = daily_hours.get(day_idx, 0)
                row_data.append(round(hours, 2))
                week1_total += hours
            
            # Week 1 regular and OT
            week1_regular = min(week1_total, 40.0)
            week1_ot = max(0, week1_total - 40.0)
            row_data.extend([round(week1_regular, 2), round(week1_ot, 2)])
            
            # Week 2 daily hours (7 days)
            week2_total = 0
            for day_idx in range(7, 14):
                hours = daily_hours.get(day_idx, 0)
                row_data.append(round(hours, 2))
                week2_total += hours
            
            # Week 2 regular and OT
            week2_regular = min(week2_total, 40.0)
            week2_ot = max(0, week2_total - 40.0)
            row_data.extend([round(week2_regular, 2), round(week2_ot, 2)])
            
            # Totals
            total_regular = week1_regular + week2_regular
            total_ot = week1_ot + week2_ot
            grand_total = total_regular + total_ot
            row_data.extend([round(total_regular, 2), round(total_ot, 2), round(grand_total, 2)])
            
            # Write row data
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=current_row, column=col, value=value)
                self._apply_style(cell, self.data_style)
            
            current_row += 1
            counter += 1
        
        return current_row + 1  # Add space before summary
    
    def _calculate_daily_hours_for_payroll(self, emp_data: Dict, start_date_str: str) -> Dict[int, float]:
        """Calculate daily hours for payroll format (14 days)"""
        daily_hours = {}
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        
        for day_idx in range(14):
            current_date = start_date + timedelta(days=day_idx)
            date_key = current_date.strftime('%Y-%m-%d')
            
            if date_key in emp_data['daily_hours']:
                day_data = emp_data['daily_hours'][date_key]
                if not day_data['is_miss_punch']:
                    daily_hours[day_idx] = day_data['total_hours']
                else:
                    daily_hours[day_idx] = 0  # Miss punch = 0 hours
            else:
                daily_hours[day_idx] = 0  # No records = 0 hours
        
        return daily_hours
    
    def _write_summary_totals(self, worksheet, hours_data: Dict, start_row: int):
        """Write summary totals section"""
        current_row = start_row + 1
        
        # Summary header
        cell = worksheet.cell(row=current_row, column=1, value="PAYROLL SUMMARY")
        self._apply_style(cell, self.header2_style)
        worksheet.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2
        
        # Calculate totals across all employees
        total_employees = len(hours_data['employees'])
        total_regular_hours = 0
        total_overtime_hours = 0
        total_hours = 0
        
        for emp_data in hours_data['employees'].values():
            total_regular_hours += emp_data['grand_totals']['regular_hours']
            total_overtime_hours += emp_data['grand_totals']['overtime_hours']
            total_hours += emp_data['grand_totals']['total_hours']
        
        # Summary data
        summary_data = [
            ("Total Employees:", total_employees),
            ("Total Regular Hours:", round(total_regular_hours, 2)),
            ("Total Overtime Hours:", round(total_overtime_hours, 2)),
            ("Grand Total Hours:", round(total_hours, 2)),
            ("Travel Time Included:", "Yes" if hours_data['include_travel_time'] else "No"),
            ("Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        for label, value in summary_data:
            cell1 = worksheet.cell(row=current_row, column=1, value=label)
            self._apply_style(cell1, self.total_style)
            cell2 = worksheet.cell(row=current_row, column=2, value=value)
            self._apply_style(cell2, self.total_style)
            current_row += 1
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @log_database_operations('detailed_hours_export')
    def create_detailed_hours_report(self, start_date: datetime, end_date: datetime,
                                    attendance_records: List[Dict], employee_names: Dict[str, str] = None,
                                    include_travel_time: bool = True) -> io.BytesIO:
        """
        Create a detailed daily hours report for all employees
        
        Args:
            start_date: Report start date
            end_date: Report end date
            attendance_records: List of attendance records
            employee_names: Dictionary mapping employee_id to full name
            include_travel_time: Whether to include travel time in calculations
            
        Returns:
            BytesIO buffer containing the Excel file
        """
        try:
            print(f"📊 Creating detailed hours report from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
            
            # Get employee names using the same method as payroll routes
            if employee_names is None:
                employee_names = self._get_employee_names(attendance_records)
            
            # Calculate working hours
            calculator = SingleCheckInCalculator()
            hours_data = calculator.calculate_all_employees_hours(start_date, end_date, attendance_records)
            
            # Create workbook
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Detailed Hours Report"
            
            # Setup styles
            self._setup_styles(workbook)
            
            # Write headers
            current_row = 1
            cell = worksheet.cell(row=current_row, column=1, value=self.company_name)
            self._apply_style(cell, self.header1_style)
            worksheet.merge_cells(f'A{current_row}:J{current_row}')
            current_row += 1
            
            cell = worksheet.cell(row=current_row, column=1, value="Detailed Daily Hours Report")
            self._apply_style(cell, self.header2_style)
            worksheet.merge_cells(f'A{current_row}:J{current_row}')
            current_row += 2
            
            # Table headers
            headers = ["Employee ID", "Employee Name", "Date", "Day of Week", "Total Hours", 
                      "Status", "Records Count", "Regular Hours", "Overtime Hours", "Notes"]
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col, value=header)
                self._apply_style(cell, self.table_header_style)
            current_row += 1
            
            # Write detailed data
            for employee_id, emp_data in hours_data['employees'].items():
                employee_name = employee_names.get(employee_id, f"Employee {employee_id}") if employee_names else f"Employee {employee_id}"
                
                # Sort daily hours by date
                daily_items = sorted(emp_data['daily_hours'].items())
                
                for date_str, day_data in daily_items:
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    day_name = date_obj.strftime('%A')
                    
                    # Determine status and notes
                    if day_data['is_miss_punch']:
                        status = "Miss Punch"
                        notes = "Missing check-in or check-out"
                        hours = 0
                    elif day_data['records_count'] == 0:
                        status = "No Records"
                        notes = "No attendance records"
                        hours = 0
                    else:
                        status = "Complete"
                        notes = f"{day_data['records_count']} record(s)"
                        hours = day_data['total_hours']
                    
                    # Calculate regular/overtime for this day
                    regular_hours = min(hours, 8.0)  # Daily limit
                    overtime_hours = max(0, hours - 8.0)
                    
                    row_data = [
                        employee_id,
                        employee_name,
                        date_str,
                        day_name,
                        round(hours, 2),
                        status,
                        day_data['records_count'],
                        round(regular_hours, 2),
                        round(overtime_hours, 2),
                        notes
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = worksheet.cell(row=current_row, column=col, value=value)
                        self._apply_style(cell, self.data_style)
                        
                        # Color code miss punches
                        if status == "Miss Punch":
                            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        elif status == "No Records":
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    
                    current_row += 1
            
            # Auto-adjust columns
            self._auto_adjust_columns(worksheet)
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            print("✅ Detailed hours report Excel file created successfully")
            return excel_buffer
            
        except Exception as e:
            print(f"❌ Error creating detailed hours report: {e}")
            raise e
    
    @log_database_operations('template_hours_export')
    def create_template_format_report(self, start_date: datetime, end_date: datetime,
                                    attendance_records: List[Dict], employee_names: Dict[str, str] = None,
                                    include_travel_time: bool = True, project_name: str = None) -> io.BytesIO:
        """
        Create a template-format report matching the provided Excel template.
        This creates a single sheet with all employees' detailed reports.
        
        Args:
            start_date: Report start date
            end_date: Report end date
            attendance_records: List of attendance records
            employee_names: Dictionary mapping employee_id to full name
            include_travel_time: Whether to include travel time in calculations
            project_name: Project name for the report header
            
        Returns:
            BytesIO buffer containing the Excel file
        """
        try:
            print(f"📊 Creating single-sheet template format report from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
            
            # Get employee names using the same method as payroll routes
            if employee_names is None:
                employee_names = self._get_employee_names(attendance_records)
            
            # Calculate working hours
            calculator = SingleCheckInCalculator()
            hours_data = calculator.calculate_all_employees_hours(start_date, end_date, attendance_records)
            
            # Create workbook with single sheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Hours Report"
            
            # Setup styles
            self._setup_template_styles(workbook)
            
            # Write main report headers once at the top
            current_row = self._write_main_report_headers(worksheet, start_date, end_date, project_name)
            
            # Write all employees data in one sheet
            for employee_id, emp_data in hours_data['employees'].items():
                employee_name = employee_names.get(employee_id, f"Employee {employee_id}") if employee_names else f"Employee {employee_id}"
                
                # Write employee section
                current_row = self._write_employee_section(worksheet, employee_id, employee_name, emp_data, 
                                                        start_date, end_date, attendance_records, current_row)
                
                # Add spacing between employees
                current_row += 2
            
            # Auto-adjust column widths to match template
            self._adjust_template_columns(worksheet)
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            print("✅ Single-sheet template format report Excel file created successfully")
            return excel_buffer
            
        except Exception as e:
            print(f"❌ Error creating template format report: {e}")
            import traceback
            print(f"❌ Traceback: {traceback.format_exc()}")
            raise e
  
    def _setup_template_styles(self, workbook: Workbook):
        """Setup Excel cell styles for template format"""
        
        # Company name style (matches template)
        self.template_company_style = {
            'font': Font(name="Arial", size=14, bold=True),
            'alignment': Alignment(horizontal="left", vertical="center")
        }
        
        # Report title style
        self.template_title_style = {
            'font': Font(name="Arial", size=12, bold=True),
            'alignment': Alignment(horizontal="left", vertical="center")
        }
        
        # Project name style
        self.template_project_style = {
            'font': Font(name="Arial", size=11, bold=True),
            'alignment': Alignment(horizontal="left", vertical="center")
        }
        
        # Date range style
        self.template_date_style = {
            'font': Font(name="Arial", size=11),
            'alignment': Alignment(horizontal="left", vertical="center")
        }
        
        # Employee info style
        self.template_employee_style = {
            'font': Font(name="Arial", size=11, bold=True),
            'alignment': Alignment(horizontal="left", vertical="center")
        }
        
        # Column header style
        self.template_header_style = {
            'font': Font(name="Arial", size=10, bold=True),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }
        
        # Data cell style
        self.template_data_style = {
            'font': Font(name="Arial", size=10),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }

    def _adjust_template_columns(self, worksheet):
        """Adjust column widths to match template format"""
        # Column widths based on template analysis
        column_widths = {
            'A': 12,   # Day
            'B': 10.6,  # Date
            'C': 12.6,  # In
            'D': 12.6,  # Out
            'E': 34,  # Location
            'F': 5.6,   # Zone
            'G': 15.4,  # Hours/Building
            'H': 10.9,  # Daily Total
            'I': 14.0,  # Regular Hours
            'J': 9.6,   # OT Hours
            'K': 35,  # Building Address
            'L': 35,  # Recorded Location
            'M': 10.0,  # Distance
            'N': 15.0,  # Possible Violation
        }
            
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width    

    def _write_main_report_headers(self, worksheet, start_date: datetime, end_date: datetime, project_name: str) -> int:
        """Write main report headers at the top of the sheet"""
        current_row = 1
        
        # 1. Company Name (merged across columns A-N)
        cell = worksheet.cell(row=current_row, column=1, value=self.company_name)
        self._apply_style(cell, self.template_company_style)
        worksheet.merge_cells(f'A{current_row}:N{current_row}')
        current_row += 1
        
        # 2. Report title (merged across columns A-N)
        cell = worksheet.cell(row=current_row, column=1, value="Summary report of Hours worked")
        self._apply_style(cell, self.template_title_style)
        worksheet.merge_cells(f'A{current_row}:N{current_row}')
        current_row += 1
        
        # 3. Project name (merged across columns A-N)
        project_display = project_name if project_name else "[Project Name]"
        cell = worksheet.cell(row=current_row, column=1, value=project_display)
        self._apply_style(cell, self.template_project_style)
        worksheet.merge_cells(f'A{current_row}:N{current_row}')
        current_row += 1
        
        # 4. Date range (merged across columns A-N)
        date_range = f"Date range: {start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"
        cell = worksheet.cell(row=current_row, column=1, value=date_range)
        self._apply_style(cell, self.template_date_style)
        worksheet.merge_cells(f'A{current_row}:N{current_row}')
        current_row += 2  # Add extra space
        
        return current_row

    def _write_employee_section(self, worksheet, employee_id: str, employee_name: str, 
                        emp_data: Dict, start_date: datetime, end_date: datetime,
                        attendance_records: List[Dict], start_row: int) -> int:
        """Write individual employee section with multiple pairs per day, daily totals, weekly totals, and grand total"""
        current_row = start_row
        
        # Employee info header (merged across columns A-O)
        employee_info = f"Employee ID {employee_id}: {employee_name}"
        cell = worksheet.cell(row=current_row, column=1, value=employee_info)
        self._apply_style(cell, self.template_employee_style)
        worksheet.merge_cells(f'A{current_row}:N{current_row}')
        current_row += 1
        
        # Column headers for this employee
        headers = ["Day", "Date", "In", "Out", "Location", "Zone", "Hours/Building", 
        "Daily Total", "Regular Hours", "OT Hours", "Building Address", 
        "Recorded Location", "Distance", "Possible Violation"]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col, value=header)
            self._apply_style(cell, self.template_header_style)
        current_row += 1
        
        # Get attendance records for this employee for location info
        employee_records = [record for record in attendance_records if str(record.employee_id) == employee_id]
        
        # Group attendance records by date for location info
        daily_location_data = {}
        for record in employee_records:
            date_key = record.check_in_date.strftime('%Y-%m-%d')
            if date_key not in daily_location_data:
                daily_location_data[date_key] = {
                    'records': [],
                    'location': '',
                    'building_address': ''
                }
            daily_location_data[date_key]['records'].append(record)
        
        # Get location info for each day
        for date_str, day_info in daily_location_data.items():
            sorted_records = sorted(day_info['records'], key=lambda x: x.check_in_time)
            if sorted_records:
                # Get location info from QR code
                if hasattr(sorted_records[0], 'qr_code') and sorted_records[0].qr_code:
                    day_info['location'] = sorted_records[0].qr_code.location or ''
                    day_info['building_address'] = sorted_records[0].qr_code.location_address or ''
        
        # Initialize weekly totals tracking
        weekly_regular_hours = 0
        weekly_ot_hours = 0
        weekly_total_hours = 0
        
        # Track overall totals for grand total
        grand_regular_hours = 0
        grand_ot_hours = 0
        grand_total_hours = 0
        total_pairs_written = 0  # Track total pairs across all days
        
        # Track week boundaries (assuming payroll period starts on Monday)
        current_week_start = None
        
        # Write data rows using the calculated daily hours from emp_data
        for date_str in sorted(emp_data['daily_hours'].keys()):
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            day_hours_data = emp_data['daily_hours'][date_str]
            
            # Skip days with no hours or miss punch
            if day_hours_data.get('is_miss_punch', False) or day_hours_data.get('total_hours', 0) <= 0:
                continue
            
            # Calculate week boundaries
            week_start = date_obj - timedelta(days=date_obj.weekday())  # Monday of the week
            
            # Check if we've moved to a new week and need to write weekly total
            if current_week_start is not None and week_start != current_week_start:
                # Write weekly total row for previous week
                current_row = self._write_weekly_total_row(worksheet, current_row, 
                                                        weekly_regular_hours, weekly_ot_hours, weekly_total_hours)
                
                # Reset weekly counters
                weekly_regular_hours = 0
                weekly_ot_hours = 0
                weekly_total_hours = 0
            
            current_week_start = week_start
            
            total_hours = day_hours_data.get('total_hours', 0)
            regular_hours = min(total_hours, 8.0)  # Max 8 regular hours per day
            ot_hours = max(0, total_hours - 8.0)
            
            # Add to weekly and grand totals
            weekly_regular_hours += regular_hours
            weekly_ot_hours += ot_hours
            weekly_total_hours += total_hours
            
            grand_regular_hours += regular_hours
            grand_ot_hours += ot_hours
            grand_total_hours += total_hours
            
            # Get location info for this date
            location_info = daily_location_data.get(date_str, {})
            location = location_info.get('location', '')
            building_address = location_info.get('building_address', '')
            
            # Get all records for this day, sorted by time
            day_records = []
            if date_str in daily_location_data:
                day_records = sorted(daily_location_data[date_str]['records'], key=lambda x: x.check_in_time)
            
            # Write multiple pairs for the day (like template shows FRIDAY appearing twice)
            pairs_written = 0
            if len(day_records) >= 2:
                # Create pairs from consecutive records
                for i in range(0, len(day_records) - 1, 2):
                    if i + 1 < len(day_records):
                        start_record = day_records[i]
                        end_record = day_records[i + 1]
                        
                        current_row = self._write_record_pair_row(
                            worksheet, current_row, date_obj, start_record, end_record,
                            location, building_address, total_hours, regular_hours, ot_hours,
                            pairs_written, total_hours  # Pass total_hours for daily total on last pair
                        )
                        pairs_written += 1
                        total_pairs_written += 1  # Track total pairs
            else:
                # Single record or no records - write one row
                start_record = day_records[0] if day_records else None
                current_row = self._write_single_record_row(
                    worksheet, current_row, date_obj, start_record,
                    location, building_address, total_hours, regular_hours, ot_hours
                )
        
        # Write final weekly total if we have data
        if weekly_total_hours > 0:
            current_row = self._write_weekly_total_row(worksheet, current_row, 
                                                    weekly_regular_hours, weekly_ot_hours, weekly_total_hours)
        
        # Write grand total row
        current_row = self._write_grand_total_row(worksheet, current_row, 
                                                grand_regular_hours, grand_ot_hours)
        
        # Add space between employees
        current_row += 2
        
        # Log the export action (use print for now since logger_handler import is complex)
        print(f"✅ Template format export: Employee {employee_id} ({employee_name}) data written with {total_pairs_written} record pairs, weekly totals, and grand total")
        
        return current_row

    def _write_record_pair_row(self, worksheet, current_row: int, date_obj: datetime, 
                          start_record, end_record, location: str, building_address: str,
                          day_total_hours: float, regular_hours: float, ot_hours: float,
                          pair_index: int, daily_total_hours: float) -> int:
        """Write a single record pair row (in/out times)"""
        day_name = date_obj.strftime('%A').upper()
        date_str = date_obj.strftime('%m/%d/%Y')
        
        # Calculate hours for this specific pair
        if start_record and end_record:
            start_time = start_record.check_in_time
            end_time = end_record.check_in_time
            start_datetime = datetime.combine(start_record.check_in_date, start_time)
            end_datetime = datetime.combine(end_record.check_in_date, end_time)
            pair_duration = (end_datetime - start_datetime).total_seconds() / 3600
            pair_hours = round(pair_duration, 2)
        else:
            start_time = start_record.check_in_time if start_record else None
            end_time = end_record.check_in_time if end_record else None
            pair_hours = 0
        
        # Format times
        check_in_time = start_time.strftime('%I:%M:%S %p') if start_time else ''
        check_out_time = end_time.strftime('%I:%M:%S %p') if end_time else ''
        
        # Get location accuracy info
        location_accuracy = None
        recorded_location = ''
        distance_value = ''
        possible_violation = 'No'
        
        if start_record and hasattr(start_record, 'location_accuracy'):
            location_accuracy = getattr(start_record, 'location_accuracy', None)
            if location_accuracy is not None:
                try:
                    accuracy_value = float(location_accuracy)
                    distance_value = f"{accuracy_value:.3f}"
                    
                    if accuracy_value < 0.3:
                        recorded_location = building_address
                        possible_violation = "No"
                    else:
                        recorded_location = f"GPS Location (Accuracy: {distance_value} miles)"
                        possible_violation = "Yes" if accuracy_value > 0.5 else "Possible"
                except (ValueError, TypeError):
                    distance_value = "Unknown"
                    recorded_location = "GPS Location (Unknown Accuracy)"
                    possible_violation = "Possible"
        
        # Create building address hyperlink if available
        building_address_display = building_address
        if building_address and "," in building_address:
            # Create Google Maps hyperlink
            maps_url = f"https://www.google.com/maps/place/{building_address.replace(' ', '+')}"
            building_address_display = f'=HYPERLINK("{maps_url}","{building_address}")'
        
        # Create recorded location hyperlink if it's a GPS location
        recorded_location_display = recorded_location
        if recorded_location and "GPS Location" not in recorded_location and "," in recorded_location:
            maps_url = f"https://www.google.com/maps/place/{recorded_location.replace(' ', '+')}"
            recorded_location_display = f'=HYPERLINK("{maps_url}","{recorded_location}")'
        
        # Determine if this is the last pair of the day (for daily total)
        is_last_pair = pair_index > 0  # Show daily total on second+ pairs
        daily_total_display = daily_total_hours if is_last_pair else ""
        
        row_data = [
            day_name,                    # A - Day
            date_str,                    # B - Date  
            check_in_time,              # C - In
            check_out_time,             # D - Out
            location,                   # E - Location
            "",                         # F - Zone (empty)
            pair_hours,                 # G - Hours/Building
            daily_total_display,        # H - Daily Total (only on last pair)
            regular_hours if is_last_pair else "",  # I - Regular Hours (only on last pair)
            ot_hours if is_last_pair else "",       # J - OT Hours (only on last pair)
            building_address_display,   # K - Building Address
            recorded_location_display,  # L - Recorded Location  
            distance_value,             # M - Distance
            possible_violation          # N - Possible Violation
        ]
        
        # Write the row
        for col, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=current_row, column=col, value=value)
            self._apply_style(cell, self.template_data_style)
        
        return current_row + 1

    def _write_single_record_row(self, worksheet, current_row: int, date_obj: datetime, 
                            record, location: str, building_address: str,
                            total_hours: float, regular_hours: float, ot_hours: float) -> int:
        """Write a single record row when there's only one check-in"""
        day_name = date_obj.strftime('%A').upper()
        date_str = date_obj.strftime('%m/%d/%Y')
        
        # Single record - estimate check-out time
        if record:
            check_in_time = record.check_in_time.strftime('%I:%M:%S %p')
            # Estimate check-out based on total hours
            check_in_datetime = datetime.combine(record.check_in_date, record.check_in_time)
            estimated_checkout = check_in_datetime + timedelta(hours=total_hours)
            check_out_time = estimated_checkout.strftime('%I:%M:%S %p')
        else:
            check_in_time = ''
            check_out_time = ''
        
        # Get location accuracy info (same logic as pair row)
        location_accuracy = None
        recorded_location = ''
        distance_value = ''
        possible_violation = 'No'
        
        if record and hasattr(record, 'location_accuracy'):
            location_accuracy = getattr(record, 'location_accuracy', None)
            if location_accuracy is not None:
                try:
                    accuracy_value = float(location_accuracy)
                    distance_value = f"{accuracy_value:.3f}"
                    
                    if accuracy_value < 0.3:
                        recorded_location = building_address
                        possible_violation = "No"
                    else:
                        recorded_location = f"GPS Location (Accuracy: {distance_value} miles)"
                        possible_violation = "Yes" if accuracy_value > 0.5 else "Possible"
                except (ValueError, TypeError):
                    distance_value = "Unknown"
                    recorded_location = "GPS Location (Unknown Accuracy)"
                    possible_violation = "Possible"
        
        row_data = [
            day_name,           # A - Day
            date_str,           # B - Date
            check_in_time,      # C - In
            check_out_time,     # D - Out
            location,           # E - Location
            "",                 # F - Zone
            total_hours,        # G - Hours/Building
            total_hours,        # H - Daily Total
            regular_hours,      # I - Regular Hours
            ot_hours,          # J - OT Hours
            building_address,   # K - Building Address
            recorded_location,  # L - Recorded Location
            distance_value,     # M - Distance
            possible_violation  # N - Possible Violation
        ]
        
        # Write the row
        for col, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=current_row, column=col, value=value)
            self._apply_style(cell, self.template_data_style)
        
        return current_row + 1

    def _write_weekly_total_row(self, worksheet, current_row: int, 
                            weekly_regular: float, weekly_ot: float, weekly_total: float) -> int:
        """Write weekly total row matching template format"""
        
        # Weekly Total row (columns G-J)
        cell = worksheet.cell(row=current_row, column=7, value="Weekly Total: ")  # G
        self._apply_style(cell, self.template_data_style)
        cell.font = Font(name="Arial", size=10, bold=True)
        
        cell = worksheet.cell(row=current_row, column=8, value=round(weekly_total, 2))  # H - Daily Total
        self._apply_style(cell, self.template_data_style)
        cell.font = Font(name="Arial", size=10, bold=True)
        
        cell = worksheet.cell(row=current_row, column=9, value=round(weekly_regular, 2))  # I - Regular Hours  
        self._apply_style(cell, self.template_data_style)
        cell.font = Font(name="Arial", size=10, bold=True)
        
        cell = worksheet.cell(row=current_row, column=10, value=round(weekly_ot, 2))  # J - OT Hours
        self._apply_style(cell, self.template_data_style)  
        cell.font = Font(name="Arial", size=10, bold=True)
        
        # Log weekly total
        print(f"✅ Template export: Weekly total written - Regular: {weekly_regular:.2f}, OT: {weekly_ot:.2f}, Total: {weekly_total:.2f}")
        
        return current_row + 1

    def _write_grand_total_row(self, worksheet, current_row: int, 
                            grand_regular: float, grand_ot: float) -> int:
        """Write grand total row matching template format"""
        
        # Grand Total row (columns G, I, J)
        cell = worksheet.cell(row=current_row, column=7, value="GRAND TOTAL: ")  # G
        self._apply_style(cell, self.template_data_style)
        cell.font = Font(name="Arial", size=10, bold=True)
        
        cell = worksheet.cell(row=current_row, column=9, value=round(grand_regular, 2))  # I - Regular Hours
        self._apply_style(cell, self.template_data_style)
        cell.font = Font(name="Arial", size=10, bold=True)
        
        cell = worksheet.cell(row=current_row, column=10, value=round(grand_ot, 2))  # J - OT Hours
        self._apply_style(cell, self.template_data_style)
        cell.font = Font(name="Arial", size=10, bold=True)
        
        # Log grand total
        print(f"✅ Template export: Grand total written - Regular: {grand_regular:.2f}, OT: {grand_ot:.2f}")
        
        return current_row + 1