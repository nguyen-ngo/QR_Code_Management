"""
routes/qr_codes.py
==================
QR code management and destination handler routes.

Routes: /qr-codes/create, /qr-codes/bulk-import, /qr-codes/<id>/*,
        /qr/<string:qr_url>
"""
from flask import Blueprint, render_template, request, redirect, flash, session, jsonify, send_file, current_app, url_for
from datetime import datetime, date, timedelta, time
import io, os, base64, re, uuid, json, traceback

from extensions import db, logger_handler
from models.attendance import AttendanceData
from models.employee import Employee
from models.project import Project
from models.qrcode import QRCode, QRCodeStyle, QRCodeLocation  # ADDED: QRCodeLocation for dynamic QR
from models.user import User
from werkzeug.utils import secure_filename
from logger_handler import log_user_activity, log_database_operations
from utils.helpers import (
    admin_required,
    detect_device_info,
    generate_default_qr_code,
    generate_qr_code,
    generate_qr_url,
    get_client_ip,
    get_employee_checkin_history,
    get_qr_styling,
    login_required,
    staff_or_admin_required)
from utils.geocoding import (
    calculate_location_accuracy_enhanced,
    get_location_accuracy_level_enhanced,
    process_location_data_enhanced,
    reverse_geocode_coordinates,
    get_coordinates_from_address_enhanced)
from qr_code_import_service import QRCodeImportService
from turnstile_utils import turnstile_utils
import openpyxl

bp = Blueprint('qr_codes', __name__)


# --- ADDED: helper — returns distinct (location, location_address) pairs from
#     all standard QR codes, used to auto-populate the dynamic QR location list ---
def get_unique_qr_locations():
    """
    Query all unique (location, location_address) pairs from the qr_codes table
    (standard QR codes only).  Returns a list of dicts:
        [{'name': str, 'address': str}, ...]
    Sorted alphabetically by name, duplicates removed.
    """
    rows = (
        db.session.query(QRCode.location, QRCode.location_address)
        .filter(
            QRCode.qr_type == 'standard',
            QRCode.location.isnot(None),
            QRCode.location != ''
        )
        .distinct()
        .order_by(QRCode.location.asc())
        .all()
    )
    seen = set()
    result = []
    for loc, addr in rows:
        key = loc.strip().lower()
        if key not in seen:
            seen.add(key)
            result.append({'name': loc.strip(), 'address': (addr or '').strip()})
    return result
# --- END ADDED ---



@bp.route('/qr-codes/create', methods=['GET', 'POST'], endpoint='create_qr_code')
@login_required
@log_database_operations('qr_code_creation')
def create_qr_code():
    """Enhanced create QR code with customization options"""
    if request.method == 'POST':
        try:
            # Existing form data
            name = request.form['name']
            qr_type = request.form.get('qr_type', 'standard')  # read type first

            # For dynamic QR codes, location/address are auto-managed (not user-entered)
            if qr_type == 'dynamic':
                location = 'Dynamic'          # placeholder — selectable locations come from standard QR codes at scan time
                location_address = ''         # no single fixed address
            else:
                location = request.form.get('location', '').strip()
                location_address = request.form.get('location_address', '').strip()
                if not location:
                    flash('Location Name is required for Standard QR codes.', 'error')
                    return render_template('create_qr_code.html',
                                         projects=Project.query.filter_by(active_status=True).all(),
                                         styles=QRCodeStyle.query.all())
                if not location_address:
                    flash('Address is required for Standard QR codes.', 'error')
                    return render_template('create_qr_code.html',
                                         projects=Project.query.filter_by(active_status=True).all(),
                                         styles=QRCodeStyle.query.all())

            location_event = request.form.get('location_event', '')
            project_id = request.form.get('project_id')

            # Extract coordinates data from form
            latitude = request.form.get('latitude')
            longitude = request.form.get('longitude')
            coordinate_accuracy = request.form.get('coordinate_accuracy', 'geocoded')

            # NEW: QR Code customization data
            fill_color = request.form.get('fill_color', '#000000')
            back_color = request.form.get('back_color', '#FFFFFF')
            box_size = int(request.form.get('box_size', 10))
            border = int(request.form.get('border', 4))
            error_correction = request.form.get('error_correction', 'L')
            style_id = request.form.get('style_id')  # Pre-defined style

            # Validate colors (basic hex validation)
            if not (fill_color.startswith('#') and len(fill_color) == 7):
                fill_color = '#000000'
            if not (back_color.startswith('#') and len(back_color) == 7):
                back_color = '#FFFFFF'

            # Convert coordinates to float if they exist
            address_latitude = None
            address_longitude = None
            has_coordinates = False

            if latitude and longitude:
                try:
                    address_latitude = float(latitude)
                    address_longitude = float(longitude)
                    has_coordinates = True
                    logger_handler.logger.debug(f"Coordinates received: {address_latitude}, {address_longitude}")
                except (ValueError, TypeError) as e:
                    logger_handler.logger.warning(f"Invalid coordinates format: {e}")
                    address_latitude = None
                    address_longitude = None
                    has_coordinates = False

            # Validate project_id if provided
            project = None
            if project_id:
                try:
                    project_id = int(project_id)
                    project = db.session.get(Project, project_id)
                    if not project or not project.active_status:
                        flash('Selected project is not valid or inactive.', 'error')
                        return render_template('create_qr_code.html',
                                             projects=Project.query.filter_by(active_status=True).all(),
                                             styles=QRCodeStyle.query.all())
                except (ValueError, TypeError):
                    flash('Invalid project selection.', 'error')
                    return render_template('create_qr_code.html',
                                         projects=Project.query.filter_by(active_status=True).all(),
                                         styles=QRCodeStyle.query.all())

            # Create new QR code record first (without URL and image)
            new_qr_code = QRCode(
                name=name,
                location=location,
                location_address=location_address,
                location_event=location_event,
                qr_code_image="",  # Will be updated after URL generation
                qr_url="",  # Will be updated after ID is assigned
                created_by=session['user_id'],
                project_id=project_id,
                address_latitude=address_latitude,
                address_longitude=address_longitude,
                coordinate_accuracy=coordinate_accuracy if has_coordinates else None,
                coordinates_updated_date=datetime.utcnow() if has_coordinates else None,
                qr_type=qr_type,  # ADDED: store QR type
                # NEW: Customization fields (only if columns exist)
                **({
                    'fill_color': fill_color,
                    'back_color': back_color,
                    'box_size': box_size,
                    'border': border,
                    'error_correction': error_correction,
                    'style_id': int(style_id) if style_id and style_id.isdigit() else None
                } if hasattr(QRCode, 'fill_color') else {})
            )

            # Add to session and flush to get the ID
            db.session.add(new_qr_code)
            db.session.flush()  # This assigns the ID without committing

            # Now generate the readable URL using the ID
            qr_url = generate_qr_url(name, new_qr_code.id)

            # Generate QR code data with the destination URL and custom styling
            qr_data = f"{request.url_root}qr/{qr_url}"
            qr_image = generate_qr_code(
                data=qr_data,
                fill_color=fill_color,
                back_color=back_color,
                box_size=box_size,
                border=border,
                error_correction=error_correction
            )

            # Update the QR code with the URL and image
            new_qr_code.qr_url = qr_url
            new_qr_code.qr_code_image = qr_image

            # Now commit all changes
            db.session.commit()

            # Enhanced logging with customization information
            logger_handler.log_qr_code_created(
                qr_code_id=new_qr_code.id,
                qr_code_name=name,
                created_by_user_id=session['user_id'],
                qr_data={
                    'location': location,
                    'location_address': location_address,
                    'location_event': location_event,
                    'has_coordinates': has_coordinates,
                    'customization': {
                        'fill_color': fill_color,
                        'back_color': back_color,
                        'box_size': box_size,
                        'border': border,
                        'error_correction': error_correction
                    }
                }
            )

            # Success message with customization info
            project_info = f" in project '{project.name}'" if project else ""
            coord_info = f" with coordinates ({new_qr_code.coordinates_display})" if has_coordinates else ""
            style_info = f" with custom styling (Fill: {fill_color}, Background: {back_color})"

            flash(f'QR Code "{name}" created successfully{project_info}{coord_info}{style_info}! URL: {qr_url}', 'success')
            return redirect(url_for('dashboard.dashboard'))

        except Exception as e:
            db.session.rollback()
            logger_handler.log_database_error('qr_code_creation', e)
            flash('QR Code creation failed. Please try again.', 'error')
            logger_handler.logger.error(f"QR Code creation error: {e}", exc_info=True)

    # Get active projects and styles for dropdown
    projects = Project.query.filter_by(active_status=True).order_by(Project.name.asc()).all()
    styles = QRCodeStyle.query.order_by(QRCodeStyle.name.asc()).all()

    return render_template('create_qr_code.html', projects=projects, styles=styles)

@bp.route('/qr-codes/bulk-import', methods=['GET', 'POST'], endpoint='import_bulk_qr_codes')
@login_required
@log_database_operations('qr_code_bulk_import')
def import_bulk_qr_codes():
    """Bulk import QR codes from Excel file"""
    
    if request.method == 'GET':
        return render_template('bulk_qr_import.html')
    
    try:
        proceed_import = request.form.get('proceed_import') == 'true'
        
        if proceed_import:
            if 'pending_qr_import_file' not in session or 'pending_qr_import_filename' not in session:
                flash('Import session expired. Please upload the file again.', 'error')
                return redirect(url_for('qr_codes.import_bulk_qr_codes'))
            
            temp_path = session['pending_qr_import_file']
            filename = session['pending_qr_import_filename']
            
            if not os.path.exists(temp_path):
                flash('Temporary file not found. Please upload the file again.', 'error')
                session.pop('pending_qr_import_file', None)
                session.pop('pending_qr_import_filename', None)
                return redirect(url_for('qr_codes.import_bulk_qr_codes'))
        else:
            if 'file' not in request.files:
                flash('No file uploaded.', 'error')
                return redirect(request.url)
            
            file = request.files['file']
            if file.filename == '':
                flash('No file selected.', 'error')
                return redirect(request.url)
            
            if not file.filename.lower().endswith(('.xlsx', '.xls')):
                flash('Please upload an Excel file (.xlsx or .xls).', 'error')
                return redirect(request.url)
            
            filename = secure_filename(file.filename)
            temp_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', '/tmp'), 
                                   f"temp_qr_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
            
            os.makedirs(os.path.dirname(temp_path), exist_ok=True)
            file.save(temp_path)
            
            session['pending_qr_import_file'] = temp_path
            session['pending_qr_import_filename'] = filename
        
        validate_only = request.form.get('validate_only') == 'true' and not proceed_import
        
        import_service = QRCodeImportService(db, logger_handler)
        
        if validate_only:
            validation_result = import_service.validate_excel_file(temp_path)
            
            if validation_result['success']:
                flash(f"Validation successful! Found {validation_result['valid_rows']} valid records.", 'success')
            else:
                flash(f"Validation found errors. Please fix them before importing.", 'error')
            
            return render_template('bulk_qr_import.html', validation_result=validation_result)
        
        projects = Project.query.filter_by(active_status=True).all()
        project_lookup = {p.name: p.id for p in projects}
        
        import_result = import_service.import_from_excel(
            file_path=temp_path,
            created_by=session['user_id'],
            generate_qr_code_func=generate_qr_code,
            generate_qr_url_func=generate_qr_url,
            request_url_root=request.url_root,
            project_lookup=project_lookup,
            QRCode=QRCode,
            Project=Project,
            geocode_func=get_coordinates_from_address_enhanced
        )
        
        if import_result['success']:
            logger_handler.logger.info(
                f"User {session['username']} successfully imported {import_result['imported_records']} QR codes via bulk import "
                f"({import_result.get('geocoded_records', 0)} addresses auto-geocoded)"
            )
            
            flash(f"Import successful! Imported {import_result['imported_records']} QR codes "
                f"out of {import_result['total_rows']} total records.", 'success')
            
            # Show geocoding info
            if import_result.get('geocoded_records', 0) > 0:
                flash(f"✓ {import_result['geocoded_records']} addresses were automatically geocoded using Google Maps.", 'info')
            
            if import_result['failed_records'] > 0:
                flash(f"Note: {import_result['failed_records']} records failed to import. "
                    f"Check the error details below.", 'warning')
        else:
            flash(f"Import failed: {import_result.get('error', 'Unknown error')}", 'error')
        
        session.pop('pending_qr_import_file', None)
        session.pop('pending_qr_import_filename', None)
        
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as cleanup_error:
            logger_handler.logger.warning(f"Failed to cleanup temp file: {cleanup_error}")
        
        return render_template('bulk_qr_import.html', import_result=import_result)
    
    except Exception as e:
        logger_handler.log_database_error('qr_code_bulk_import', e)
        flash(f'Import failed: {str(e)}', 'error')
        return redirect(url_for('qr_codes.import_bulk_qr_codes'))


@bp.route('/qr-codes/bulk-import/template', endpoint='download_qr_import_template')
@login_required
def download_qr_import_template():
    """Download Excel template for bulk QR code import"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "QR Code Import Template"
        
        headers = [
            'QR Code Name',
            'QR Code Location',
            'Project',
            'Location Address',
            'Event',
            'Latitude',
            'Longitude'
        ]
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        example_data = [
            ['HQ-Entrance', 'Main Building', 'Corporate HQ', '123 Main St, Springfield, IL 62701', 'Check IN', 39.781721, -89.650148],
            ['HQ-Exit', 'Main Building', 'Corporate HQ', '123 Main St, Springfield, IL 62701', 'Check OUT', 39.781721, -89.650148],
            ['Site-A-Gate1', 'Construction Site A', 'Construction Projects', '456 Oak Ave, Chicago, IL 60601', 'Check IN', '', '']
        ]
        
        for row_num, row_data in enumerate(example_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        column_widths = [20, 20, 20, 40, 15, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger_handler.logger.info(f"User {session.get('username', 'unknown')} downloaded QR import template")
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='QR_Code_Import_Template.xlsx'
        )
    
    except Exception as e:
        logger_handler.log_flask_error('qr_import_template_download', str(e))
        flash('Error generating template. Please try again.', 'error')
        return redirect(url_for('qr_codes.import_bulk_qr_codes'))
    
@bp.route('/qr-codes/<int:qr_id>/edit', methods=['GET', 'POST'], endpoint='edit_qr_code')
@login_required
@log_database_operations('qr_code_edit')
def edit_qr_code(qr_id):
    """Enhanced edit QR code with customization support"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)

        if request.method == 'POST':
            # Track changes for logging
            old_data = {
                'name': qr_code.name,
                'location': qr_code.location,
                'location_address': qr_code.location_address,
                'location_event': qr_code.location_event,
                'project_id': qr_code.project_id,
                'qr_url': qr_code.qr_url,
                'address_latitude': qr_code.address_latitude,
                'address_longitude': qr_code.address_longitude,
                'coordinate_accuracy': qr_code.coordinate_accuracy,
                # Track old styling
                'fill_color': getattr(qr_code, 'fill_color', '#000000'),
                'back_color': getattr(qr_code, 'back_color', '#FFFFFF')
            }

            # Update QR code fields
            new_name = request.form['name']
            qr_code.name = new_name

            # --- ADDED: for dynamic QR codes, location/address are auto-managed ---
            new_qr_type = request.form.get('qr_type', 'standard')
            qr_code.qr_type = new_qr_type

            if new_qr_type == 'dynamic':
                qr_code.location = 'Dynamic'   # placeholder — selectable locations come from standard QR codes at scan time
                qr_code.location_address = ''  # no single fixed address
            else:
                qr_code.location = request.form.get('location', '').strip()
                qr_code.location_address = request.form.get('location_address', '').strip()
            # --- END ADDED ---

            qr_code.location_event = request.form.get('location_event', '')

            # Handle coordinates
            latitude = request.form.get('address_latitude')
            longitude = request.form.get('address_longitude')
            coordinate_accuracy = request.form.get('coordinate_accuracy', 'geocoded')

            if latitude and longitude:
                try:
                    qr_code.address_latitude = float(latitude)
                    qr_code.address_longitude = float(longitude)
                    qr_code.coordinate_accuracy = coordinate_accuracy
                    qr_code.coordinates_updated_date = datetime.utcnow()
                except (ValueError, TypeError):
                    pass
            elif latitude == '' and longitude == '':
                qr_code.address_latitude = None
                qr_code.address_longitude = None
                qr_code.coordinate_accuracy = None
                qr_code.coordinates_updated_date = None

            # Handle project association
            new_project_id = request.form.get('project_id')
            if new_project_id and new_project_id.strip():
                try:
                    new_project_id = int(new_project_id)
                    project = db.session.get(Project, new_project_id)
                    if project and project.active_status:
                        qr_code.project_id = new_project_id
                    else:
                        flash('Selected project is not valid or inactive.', 'error')
                        return render_template('edit_qr_code.html', qr_code=qr_code,
                                             projects=Project.query.filter_by(active_status=True).all(),
                                             styles=QRCodeStyle.query.all())
                except (ValueError, TypeError):
                    flash('Invalid project selection.', 'error')
                    return render_template('edit_qr_code.html', qr_code=qr_code,
                                         projects=Project.query.filter_by(active_status=True).all(),
                                         styles=QRCodeStyle.query.all())
            else:
                qr_code.project_id = None

            # Handle QR code customization (only if columns exist)
            fill_color = request.form.get('fill_color', '#000000')
            back_color = request.form.get('back_color', '#FFFFFF')
            box_size = int(request.form.get('box_size', 10))
            border = int(request.form.get('border', 4))
            error_correction = request.form.get('error_correction', 'L')
            style_id = request.form.get('style_id')

            # Update styling fields if they exist
            if hasattr(qr_code, 'fill_color'):
                qr_code.fill_color = fill_color
                qr_code.back_color = back_color
                qr_code.box_size = box_size
                qr_code.border = border
                qr_code.error_correction = error_correction
                qr_code.style_id = int(style_id) if style_id and style_id.isdigit() else None

            # Check if QR code needs regeneration
            name_changed = old_data['name'] != new_name
            styling_changed = (hasattr(qr_code, 'fill_color') and
                             (old_data['fill_color'] != fill_color or
                              old_data['back_color'] != back_color))

            if name_changed:
                new_qr_url = generate_qr_url(new_name, qr_code.id)
                qr_code.qr_url = new_qr_url

            # Regenerate QR code if name or styling changed
            if name_changed or styling_changed:
                qr_data = f"{request.url_root}qr/{qr_code.qr_url}"

                # Use new styling if available, otherwise use defaults
                styling = get_qr_styling(qr_code)
                qr_code.qr_code_image = generate_qr_code(
                    data=qr_data,
                    fill_color=styling['fill_color'],
                    back_color=styling['back_color'],
                    box_size=styling['box_size'],
                    border=styling['border'],
                    error_correction=styling['error_correction']
                )

            db.session.commit()

            # Success message
            flash(f'QR Code "{qr_code.name}" updated successfully!', 'success')
            return redirect(url_for('dashboard.dashboard'))

        # GET request - render edit form
        projects = Project.query.filter_by(active_status=True).order_by(Project.name.asc()).all()
        styles = QRCodeStyle.query.order_by(QRCodeStyle.name.asc()).all()
        return render_template('edit_qr_code.html', qr_code=qr_code,
                               projects=projects, styles=styles)

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('qr_code_edit', e)
        flash('QR Code update failed. Please try again.', 'error')
        return redirect(url_for('dashboard.dashboard'))

@bp.route('/qr-codes/<int:qr_id>/delete', methods=['GET', 'POST'], endpoint='delete_qr_code')
@admin_required
@log_database_operations('qr_code_deletion')
def delete_qr_code(qr_id):
    """Permanently delete QR code (Admin only) - Hard delete - PRESERVING EXACT ROUTE"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        logger_handler.logger.debug(f"Found QR Code for delete: {qr_code.name} (ID: {qr_id})")

        if request.method == 'POST':
            qr_name = qr_code.name
            qr_code_id = qr_code.id
            logger_handler.logger.info(f"User {session.get('username', 'unknown')} attempting to delete QR code: {qr_name} (ID: {qr_id})")

            # Check if QR exists before delete
            before_count = QRCode.query.count()
            logger_handler.logger.debug(f"QR count before delete: {before_count}")

            # Log QR code deletion before actual deletion
            logger_handler.log_qr_code_deleted(
                qr_code_id=qr_code_id,
                qr_code_name=qr_name,
                deleted_by_user_id=session['user_id']
            )

            # Delete the QR code
            db.session.delete(qr_code)


            db.session.commit()


            # Check count after delete
            after_count = QRCode.query.count()
            logger_handler.logger.debug(f"QR count after delete: {after_count}")
            logger_handler.logger.info(f"QR code deleted successfully: {qr_name} (ID: {qr_id}), removed {before_count - after_count} records")

            flash(f'QR code "{qr_name}" has been permanently deleted!', 'success')
            return redirect(url_for('dashboard.dashboard'))

        # GET request - show confirmation page
        logger_handler.logger.debug(f"Showing delete confirmation page for QR code ID: {qr_id}")
        return render_template('confirm_delete_qr.html', qr_code=qr_code)

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('qr_code_deletion', e)
        logger_handler.logger.error(f"Error in QR code delete route (ID: {qr_id}): {e}", exc_info=True)
        flash('Error deleting QR code. Please try again.', 'error')
        return redirect(url_for('dashboard.dashboard'))
    
@bp.route('/qr/<string:qr_url>', endpoint='qr_destination')
def qr_destination(qr_url):
    """QR code destination page where staff check in - PRESERVING EXACT ROUTE"""
    try:
        # Find QR code by URL
        qr_code = QRCode.query.filter_by(qr_url=qr_url, active_status=True).first()

        if not qr_code:
            # Log invalid QR code access attempt
            logger_handler.log_security_event(
                event_type="invalid_qr_access",
                description=f"Attempt to access invalid QR code URL: {qr_url}",
                severity="MEDIUM"
            )
            flash('QR code not found or inactive.', 'error')
            return redirect(url_for('auth.index'))

        # Log QR code access
        logger_handler.log_qr_code_accessed(
            qr_code_id=qr_code.id,
            qr_code_name=qr_code.name,
            access_method='scan'
        )

        # Load selectable locations for dynamic QR — auto-generated from all
        # active standard QR codes' unique (location, location_address) pairs.
        locations = []
        if getattr(qr_code, 'qr_type', 'standard') == 'dynamic':
            locations = (
                db.session.query(QRCode.location, QRCode.location_address)
                .filter(
                    QRCode.qr_type == 'standard',
                    QRCode.active_status == True,
                    QRCode.location.isnot(None),
                    QRCode.location != '',
                    QRCode.location != 'Dynamic'
                )
                .distinct()
                .order_by(QRCode.location.asc())
                .all()
            )

        return render_template('qr_destination.html', qr_code=qr_code, locations=locations)

    except Exception as e:
        logger_handler.log_database_error('qr_code_scan', e)
        flash('Error processing QR code scan.', 'error')
        return redirect(url_for('auth.index'))

@bp.route('/qr/<string:qr_url>/checkin', methods=['POST'], endpoint='qr_checkin')
def qr_checkin(qr_url):
    """
    Enhanced staff check-in with location accuracy calculation
    Allows multiple check-ins with minimum interval between them
    PRESERVES coordinate-to-address conversion functionality
    """
    try:
        logger_handler.logger.debug(f"Starting check-in process for QR URL: {qr_url}")

        # Find QR code by URL
        qr_code = QRCode.query.filter_by(qr_url=qr_url, active_status=True).first()

        if not qr_code:
            logger_handler.logger.warning(f"QR code not found or inactive: {qr_url}")
            return jsonify({
                'success': False,
                'message': 'QR code not found or inactive.'
            }), 404

        logger_handler.logger.debug(f"Found QR code: {qr_code.name} (ID: {qr_code.id}), location: {qr_code.location}")

        # Get and validate employee ID
        employee_id = request.form.get('employee_id', '').strip()

        # --- ADDED: dynamic QR — resolve effective location from the employee's selection ---
        selected_location_name    = request.form.get('selected_location_name', '').strip()
        selected_location_address = request.form.get('selected_location_address', '').strip()

        if getattr(qr_code, 'qr_type', 'standard') == 'dynamic' and selected_location_name:
            effective_location_name    = selected_location_name
            effective_location_address = selected_location_address or ''

            # Look up the matching standard QR code so we can inherit its
            # location_event, coordinates, and exact address — this ensures all
            # calculations (GPS accuracy, interval messages, success labels) behave
            # exactly as if the employee had scanned that standard QR directly.
            matching_qr = QRCode.query.filter_by(
                location=selected_location_name,
                qr_type='standard',
                active_status=True
            ).first()

            if matching_qr:
                # Use the standard QR's address if the selection has none
                if not effective_location_address:
                    effective_location_address = matching_qr.location_address or ''
                effective_location_event = matching_qr.location_event or qr_code.location_event or 'Check In'
                effective_address_latitude  = matching_qr.address_latitude
                effective_address_longitude = matching_qr.address_longitude
                logger_handler.logger.info(
                    f"DYNAMIC check-in: employee={employee_id}, "
                    f"selected='{selected_location_name}', "
                    f"matched standard QR #{matching_qr.id} '{matching_qr.name}'"
                )
            else:
                # No matching standard QR — use whatever the dynamic QR has
                effective_location_event    = qr_code.location_event or 'Check In'
                effective_address_latitude  = None
                effective_address_longitude = None
                logger_handler.logger.info(
                    f"DYNAMIC check-in: employee={employee_id}, "
                    f"selected='{selected_location_name}', no matching standard QR found"
                )
        else:
            effective_location_name     = qr_code.location
            effective_location_address  = qr_code.location_address
            effective_location_event    = qr_code.location_event
            effective_address_latitude  = qr_code.address_latitude
            effective_address_longitude = qr_code.address_longitude
        # --- END ADDED ---

        if not employee_id:
            return jsonify({
                'success': False,
                'message': 'Employee ID is required.'
            }), 400

        # Check for recent check-ins with 30-minute interval validation
        today = date.today()
        current_time = datetime.now()
        time_interval = current_app.config.get('TIME_INTERVAL', 30)
        the_last_checkin_time = current_time - timedelta(minutes=time_interval)

        # Find the most recent check-in for this employee at this location today
        recent_checkin = AttendanceData.query.filter_by(
            qr_code_id=qr_code.id,
            employee_id=employee_id.upper(),
            check_in_date=today
        ).order_by(AttendanceData.check_in_time.desc()).first()

        if recent_checkin:
            # Convert check_in_time (time) to datetime for comparison
            recent_checkin_datetime = datetime.combine(today, recent_checkin.check_in_time)

            # Check if 30 minutes have passed since the last check-in
            if recent_checkin_datetime > the_last_checkin_time:
                minutes_remaining = time_interval - int((current_time - recent_checkin_datetime).total_seconds() / 60)
                logger_handler.logger.info(f"Too soon for another {effective_location_event} for employee {employee_id}: {minutes_remaining} minutes remaining")

                checkin_time_str = recent_checkin.check_in_time.strftime('%H:%M')
                return jsonify({
                    'success': False,
                    'message': (
                        f'You can {effective_location_event} again in {minutes_remaining} minutes. '
                        f'Last {effective_location_event} was at {checkin_time_str}. \n'
                        f'Puedes volver a registrarte en {minutes_remaining} minutos. '
                        f'El ultimo registro fue a las {checkin_time_str}.'
                    )
                }), 400
            else:
                logger_handler.logger.debug(f"{time_interval}-minute interval satisfied for employee {employee_id}")
        else:
            logger_handler.logger.debug(f"First {effective_location_event} today for employee {employee_id}")

        # Process location data with coordinate-to-address conversion
        location_data = process_location_data_enhanced(request.form)

        # Get device and network info
        user_agent_string = request.headers.get('User-Agent', '')
        device_info = detect_device_info(user_agent_string)
        client_ip = get_client_ip()

        logger_handler.logger.debug(f"Check-in device: {device_info}, IP: {client_ip}")

        # Create attendance record
        logger_handler.logger.debug("Creating attendance record")

        # For dynamic QR: append tag to location_name so reports distinguish the source
        is_dynamic = getattr(qr_code, 'qr_type', 'standard') == 'dynamic' and bool(selected_location_name)
        record_location_name = (
            f"{effective_location_name} (Dynamic QR)" if is_dynamic else effective_location_name
        )
        # For dynamic QR: store the selected location's address as the QR-side address
        record_qr_address = effective_location_address if is_dynamic else None

        attendance = AttendanceData(
            qr_code_id=qr_code.id,
            employee_id=employee_id.upper(),
            check_in_date=today,
            check_in_time=datetime.now().time(),
            device_info=device_info,
            user_agent=user_agent_string,
            ip_address=client_ip,
            location_name=record_location_name,
            qr_address=record_qr_address,
            latitude=location_data['latitude'],
            longitude=location_data['longitude'],
            accuracy=location_data['accuracy'],
            altitude=location_data['altitude'],
            location_source=location_data['source'],
            address=location_data['address'],
            status='present',
            verification_required=False,  # Will be set below if needed
            verification_status=None
        )

        logger_handler.logger.debug("Created base attendance record")

        # Calculate location accuracy
        logger_handler.logger.debug(
            f"Location accuracy check: QR='{qr_code.name}' (ID={qr_code.id}), "
            f"lat={location_data['latitude']}, lng={location_data['longitude']}, "
            f"source={location_data['source']}"
        )

        location_accuracy = None

        try:
            # Check if we have the required data
            # CHANGED: use effective_location_address (respects dynamic QR selection)
            if not effective_location_address:
                logger_handler.logger.warning(f"QR code location_address is empty or None for QR ID: {qr_code.id}")
            elif not location_data['address'] and not (location_data['latitude'] and location_data['longitude']):
                logger_handler.logger.warning(
                    f"No check-in address or coordinates available: "
                    f"address={location_data['address']!r}, "
                    f"coords={location_data['latitude']}, {location_data['longitude']}"
                )
            else:
                logger_handler.logger.debug("Required location data available, proceeding with accuracy calculation")

                location_accuracy = calculate_location_accuracy_enhanced(
                    qr_address=effective_location_address,  # CHANGED: dynamic QR uses selected location address
                    checkin_address=location_data['address'],
                    checkin_lat=location_data['latitude'],
                    checkin_lng=location_data['longitude']
                )

                logger_handler.logger.debug(f"Location accuracy calculation result: {location_accuracy}")

                if location_accuracy is not None:
                    attendance.location_accuracy = location_accuracy
                    accuracy_level = get_location_accuracy_level_enhanced(location_accuracy)
                    logger_handler.logger.debug(f"Location accuracy set: {location_accuracy:.4f} miles ({accuracy_level})")
                else:
                    logger_handler.logger.warning("Could not calculate location accuracy — calculation returned None")

                # CHECK DISTANCE THRESHOLD FOR PHOTO VERIFICATION
                logger_handler.logger.debug(f"Photo verification enabled: {current_app.config.get('PHOTO_VERIFICATION_ENABLED', True)}")
                requires_verification = False
                verification_photo_data = None
                
                if current_app.config.get('PHOTO_VERIFICATION_ENABLED', True) and location_accuracy is not None and location_accuracy > current_app.config.get('DISTANCE_THRESHOLD_FOR_VERIFICATION', 0.3):
                    logger_handler.logger.info(f"Distance ({location_accuracy:.3f} mi) exceeds verification threshold for employee {employee_id}")
                    
                    # Check if photo was provided
                    verification_photo_data = request.form.get('verification_photo', None)
                    
                    if verification_photo_data:
                        logger_handler.logger.debug(f"Verification photo provided (size: {len(verification_photo_data)} chars)")
                        
                        # Validate photo data (basic validation)
                        if verification_photo_data.startswith('data:image/'):
                            attendance.verification_photo = verification_photo_data
                            attendance.verification_required = True
                            attendance.verification_status = 'pending'
                            attendance.verification_timestamp = datetime.now()
                            logger_handler.logger.info(f"Photo verification set to PENDING for employee {employee_id}")
                        else:
                            logger_handler.logger.warning(f"Invalid photo format provided for employee {employee_id}")
                            return jsonify({
                                'success': False,
                                'message': 'Invalid photo format. Please try again.',
                                'requires_verification': True
                            }), 400
                    else:
                        logger_handler.logger.warning(f"Photo verification required but not provided for employee {employee_id}")
                        return jsonify({
                            'success': False,
                            'message': 'Photo verification required. Distance from location is too far.',
                            'requires_verification': True,
                            'distance': round(location_accuracy, 3),
                            'threshold': current_app.config.get('DISTANCE_THRESHOLD_FOR_VERIFICATION', 0.3)
                        }), 400
                else:
                    logger_handler.logger.debug(f"Distance within threshold for employee {employee_id} — no verification needed")

        except Exception as e:
            logger_handler.logger.error(f"Error in location accuracy calculation: {e}", exc_info=True)

        # ENHANCED DEBUG: Save to database with verification
        try:
            logger_handler.logger.debug(f"Saving attendance record: employee={attendance.employee_id}, location={attendance.location_name}, accuracy={attendance.location_accuracy}")

            db.session.add(attendance)
            db.session.commit()

            # Log verification if required
            if attendance.verification_required:
                logger_handler.log_photo_verification(
                    employee_id=attendance.employee_id,
                    qr_code_id=qr_code.id,
                    distance=location_accuracy,
                    status='pending'
                )

            # VERIFICATION: Read back from database
            saved_record = db.session.get(AttendanceData, attendance.id)
            logger_handler.logger.info(f"Saved attendance record ID: {attendance.id}, db accuracy: {saved_record.location_accuracy}")

            if saved_record.location_accuracy != attendance.location_accuracy:
                logger_handler.logger.warning(f"DB accuracy mismatch: object={attendance.location_accuracy}, db={saved_record.location_accuracy}")

            # Add enhanced logging for location accuracy save
            if attendance.location_accuracy is not None:
                logger_handler.logger.info(f"Location accuracy calculated and saved: {attendance.location_accuracy:.4f} miles for employee {attendance.employee_id}")
            else:
                logger_handler.logger.warning(f"Location accuracy could not be calculated for employee {attendance.employee_id} at QR {qr_code.name}")

            # Count total check-ins for today for this employee at this location
            today_checkin_count = AttendanceData.query.filter_by(
                qr_code_id=qr_code.id,
                employee_id=employee_id.upper(),
                check_in_date=today
            ).count()

            checkin_sequence_text = f"{effective_location_event} details"

        except Exception as e:
            logger_handler.logger.error(f"Database error saving attendance record: {e}", exc_info=True)
            db.session.rollback()
            logger_handler.log_database_error('checkin_save', e)
            return jsonify({
                'success': False,
                'message': 'Database error occurred.'
            }), 500

        # Return success response with sequence information
        response_data = {
            'success': True,
            'message': f'Check-in successful! {checkin_sequence_text} for today.',
            'data': {
                'employee_id': attendance.employee_id,
                'location': effective_location_name,                    # CHANGED: use selected location name
                'location_event': effective_location_event,             # CHANGED: use resolved event
                'event': effective_location_event,                      # CHANGED: use resolved event
                'check_in_time': attendance.check_in_time.strftime('%I:%M %p'),
                'check_in_date': attendance.check_in_date.strftime('%B %d, %Y'),
                'device_info': attendance.device_info,
                'ip_address': attendance.ip_address,
                'location_accuracy': location_accuracy,
                'checkin_count_today': today_checkin_count,
                'checkin_sequence': checkin_sequence_text
            }
        }

        if location_data['address']:
            response_data['data']['address'] = location_data['address']

        if location_data['latitude'] and location_data['longitude']:
            response_data['data']['coordinates'] = f"{location_data['latitude']:.10f}, {location_data['longitude']:.10f}"

        # Enhanced logging for successful check-in with all details
        logger_handler.logger.info(
            f"Check-in completed: employee={attendance.employee_id}, "
            f"action={qr_code.location_event}, location={attendance.location_name}, "
            f"time={attendance.check_in_time.strftime('%H:%M')}, count_today={today_checkin_count}"
        )
        
        # Log to database for audit trail
        logger_handler.logger.info(f"Check-in success - Employee: {attendance.employee_id}, Location: {attendance.location_name}, Time: {attendance.check_in_time}, Action: {qr_code.location_event}")

        return jsonify(response_data), 200

    except Exception as e:
        db.session.rollback()
        logger_handler.logger.error(f"Unexpected error in check-in process (QR: {qr_url}): {e}", exc_info=True)

        return jsonify({
            'success': False,
            'message': 'An unexpected error occurred during check-in.'
        }), 500

# --- ADDED: API endpoint returning selectable locations for a dynamic QR code ---
@bp.route('/qr/<string:qr_url>/locations', methods=['GET'], endpoint='qr_get_locations')
def qr_get_locations(qr_url):
    """
    Return JSON list of active selectable locations for a dynamic QR code.
    Used by the scan page to populate the location selector.
    """
    try:
        qr_code = QRCode.query.filter_by(qr_url=qr_url, active_status=True).first()
        if not qr_code:
            return jsonify({'success': False, 'message': 'QR code not found or inactive.'}), 404

        if getattr(qr_code, 'qr_type', 'standard') != 'dynamic':
            return jsonify({'success': False, 'message': 'Not a dynamic QR code.'}), 400

        locations = (
            db.session.query(QRCode.location, QRCode.location_address)
            .filter(
                QRCode.qr_type == 'standard',
                QRCode.active_status == True,
                QRCode.location.isnot(None),
                QRCode.location != '',
                QRCode.location != 'Dynamic'
            )
            .distinct()
            .order_by(QRCode.location.asc())
            .all()
        )

        logger_handler.logger.info(
            f"qr_get_locations: QR '{qr_url}' returned {len(locations)} locations"
        )

        return jsonify({
            'success': True,
            'locations': [
                {
                    'name': loc.location,
                    'address': loc.location_address or ''
                }
                for loc in locations
            ]
        }), 200

    except Exception as e:
        logger_handler.logger.error(
            f"Error fetching locations for QR '{qr_url}': {e}", exc_info=True
        )
        return jsonify({'success': False, 'message': 'Server error fetching locations.'}), 500
# --- END ADDED ---


@bp.route('/qr-codes/<int:qr_id>/toggle-status', methods=['POST'], endpoint='toggle_qr_status')
@login_required
def toggle_qr_status(qr_id):
    """Toggle QR code active/inactive status"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)

        # Toggle the status
        qr_code.active_status = not qr_code.active_status
        db.session.commit()

        status_text = "activated" if qr_code.active_status else "deactivated"
        flash(f'QR code "{qr_code.name}" has been {status_text} successfully!', 'success')

        return jsonify({
            'success': True,
            'new_status': qr_code.active_status,
            'status_text': 'Active' if qr_code.active_status else 'Inactive',
            'message': f'QR code {status_text} successfully!'
        })

    except Exception as e:
        db.session.rollback()
        logger_handler.logger.error(f"Error toggling QR status (ID: {qr_id}): {e}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'Error updating QR code status. Please try again.'
        }), 500

@bp.route('/qr-codes/<int:qr_id>/copy-url', methods=['POST'], endpoint='copy_qr_url')
@login_required
def copy_qr_url(qr_id):
    """Log QR code URL copy action"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        
        # Log URL copy action
        logger_handler.logger.info(f"User {session.get('username', 'unknown')} copied URL for QR code {qr_code.name} (ID: {qr_id})")
        
        return jsonify({
            'success': True,
            'message': f'QR code URL copied to clipboard!',
            'url': f"{request.url_root}qr/{qr_code.qr_url}"
        })

    except Exception as e:
        logger_handler.logger.error(f"Error copying QR URL for ID {qr_id}: {e}")
        return jsonify({
            'success': False,
            'message': 'Error copying QR code URL.'
        }), 500

@bp.route('/qr-codes/<int:qr_id>/open-link', methods=['POST'], endpoint='open_qr_link')
@login_required
def open_qr_link(qr_id):
    """Log QR code link open action"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        
        # Log link open action
        logger_handler.logger.info(f"User {session.get('username', 'unknown')} opened link for QR code {qr_code.name} (ID: {qr_id})")
        
        return jsonify({
            'success': True,
            'message': f'Opening QR code link...',
            'url': f"{request.url_root}qr/{qr_code.qr_url}"
        })

    except Exception as e:
        logger_handler.logger.error(f"Error opening QR link for ID {qr_id}: {e}")
        return jsonify({
            'success': False,
            'message': 'Error opening QR code link.'
        }), 500

@bp.route('/qr-codes/<int:qr_id>/activate', methods=['POST'], endpoint='activate_qr_code')
@login_required
def activate_qr_code(qr_id):
    """Activate a QR code"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        qr_code.active_status = True
        db.session.commit()

        flash(f'QR code "{qr_code.name}" has been activated successfully!', 'success')
        return jsonify({
            'success': True,
            'new_status': True,
            'status_text': 'Active',
            'message': 'QR code activated successfully!'
        })

    except Exception as e:
        db.session.rollback()
        logger_handler.logger.error(f"Error activating QR code (ID: {qr_id}): {e}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'Error activating QR code. Please try again.'
        }), 500

@bp.route('/qr-codes/<int:qr_id>/deactivate', methods=['POST'], endpoint='deactivate_qr_code')
@login_required
def deactivate_qr_code(qr_id):
    """Deactivate a QR code"""
    try:
        qr_code = QRCode.query.get_or_404(qr_id)
        qr_code.active_status = False
        db.session.commit()

        flash(f'QR code "{qr_code.name}" has been deactivated successfully!', 'success')
        return jsonify({
            'success': True,
            'new_status': False,
            'status_text': 'Inactive',
            'message': 'QR code deactivated successfully!'
        })

    except Exception as e:
        db.session.rollback()
        logger_handler.logger.error(f"Error deactivating QR code (ID: {qr_id}): {e}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'Error deactivating QR code. Please try again.'
        }), 500
