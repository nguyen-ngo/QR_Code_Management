"""
routes/time_attendance_export.py
=================================
Excel export logic for Time Attendance — helper functions called by
routes in time_attendance.py.

Contains:
  - calculate_possible_violation()
  - _overnight_aware_sort_key()
  - _qtr()
  - export_time_attendance_excel()          (single-employee / all-employees)
  - export_time_attendance_by_building_excel()
"""
from flask import send_file, g, current_app
from datetime import datetime, date, timedelta, time
import io, os, json, re
import time as _time

from extensions import db, logger_handler
from models.employee import Employee
from models.project import Project
from models.qrcode import QRCode
from models.time_attendance import TimeAttendance
from sqlalchemy import text
from working_hours_calculator import WorkingHoursCalculator, round_time_to_quarter_hour, convert_minutes_to_base100, round_base100_hours
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import openpyxl.cell.cell

# Export helpers — no Blueprint needed, these are plain functions
# called from routes in time_attendance.py

def calculate_possible_violation(distance_value):
    """
    Calculate possible violation status based on distance
    
    Args:
        distance_value: Distance in miles (float or None)
    
    Returns:
        'Yes' if distance > 0.3, 'No' otherwise
    """
    if distance_value is None:
        return 'No'
    
    try:
        distance_float = float(distance_value)
        return 'Yes' if distance_float > 0.3 else 'No'
    except (ValueError, TypeError):
        return 'No'

def _overnight_aware_sort_key(record):
    """
    Sort key for attendance records within a single calendar-date bucket.

    Problem 1: when an overnight shift spans midnight, the check-out record's
    check_in_time (e.g. 00:01 AM) sorts numerically BEFORE the check-in time
    (e.g. 20:00 PM), producing an orphaned OUT followed by an orphaned IN.
    Fix: push early-morning check-outs (hour <= 3) past midnight by adding
    24 h worth of seconds so they sort after same-day evening check-ins.

    Problem 2: two records in the same minute (e.g. IN 06:22:04, OUT 06:22:52)
    had identical sort keys because seconds were not included, leaving the
    database-delivery order intact (DESC → OUT first).  The pairing loop then
    encountered the OUT before the IN, emitting an orphaned-OUT row followed
    by an orphaned-IN row — reversed from chronological order.
    Fix: include seconds in the key so true chronological order is preserved.
    """
    from datetime import time as _time
    t = record.check_in_time
    if isinstance(t, _time):
        # Use fractional minutes (hours*60 + minutes + seconds/60) so that
        # records sharing the same HH:MM still sort by their seconds component.
        seconds_total = t.hour * 3600 + t.minute * 60 + t.second
    else:
        seconds_total = 0
    action = (record.action_description or '').lower()
    is_out = 'out' in action or 'checkout' in action
    # Push early-morning check-outs past midnight to end of day order.
    # Use seconds-based offset (24 h = 86400 s) to remain consistent with
    # the seconds-granularity key above.
    if is_out and t.hour <= 3:
        seconds_total += 24 * 3600
    return seconds_total

def _qtr(decimal_hours: float) -> float:
    """
    Round a decimal-hours value to the nearest quarter hour (.00/.25/.50/.75).
    Pipeline: decimal hours → minutes → quarter-hour rounding → base-100 → quarter rounding.
    Examples: 4.03 → 4.0, 4.08 → 4.25, 3.87 → 4.0, 4.16 → 4.25
    Returns 0.0 for negative or zero input.
    """
    if decimal_hours <= 0:
        return 0.0
    minutes = decimal_hours * 60.0
    rounded_minutes = round_time_to_quarter_hour(minutes)
    base100 = convert_minutes_to_base100(rounded_minutes)
    return round_base100_hours(base100)

# ---------------------------------------------------------------------------
# Private export helpers — shared by both export functions below.
# ---------------------------------------------------------------------------

def _resolve_date_range(start_date_filter, end_date_filter, records, export_label='TA Excel export'):
    """
    Resolve and validate the export date range.

    Returns (start_date, end_date, filtered_records) where:
    - start_date / end_date are date objects
    - filtered_records is the input list capped to end_date + 1 day (overnight buffer)
    - end_date is capped to a maximum 14-day window
    Returns None when there are no records and no date filters.
    """
    MAX_EXPORT_DAYS = 14

    if start_date_filter and end_date_filter:
        if isinstance(start_date_filter, str):
            start_date = datetime.strptime(start_date_filter, '%Y-%m-%d').date()
        else:
            start_date = start_date_filter
        if isinstance(end_date_filter, str):
            end_date = datetime.strptime(end_date_filter, '%Y-%m-%d').date()
        else:
            end_date = end_date_filter
    elif records:
        start_date = min(r.attendance_date for r in records)
        end_date   = max(r.attendance_date for r in records)
    else:
        return None

    if (end_date - start_date).days >= MAX_EXPORT_DAYS:
        capped_end_date = start_date + timedelta(days=MAX_EXPORT_DAYS - 1)
        logger_handler.logger.info(
            f"{export_label}: date range [{start_date} \u2013 {end_date}] exceeds "
            f"{MAX_EXPORT_DAYS} days; capping end_date to {capped_end_date}."
        )
        end_date = capped_end_date

    # Preserve one extra calendar day so early-morning check-out records
    # stored on Day N+1 remain available for overnight pairing detection.
    # Display range is still controlled by dates_with_records (capped to end_date).
    filtered_records = [r for r in records if r.attendance_date <= end_date + timedelta(days=1)]

    return start_date, end_date, filtered_records


def _convert_ta_records(records):
    """
    Convert TimeAttendance ORM records to the lightweight anonymous-class
    format expected by WorkingHoursCalculator and the Excel rendering loops.

    Returns a list of converted record objects.
    """
    from working_hours_calculator import parse_employee_id_for_work_type

    converted = []
    for record in records:
        distance_value = getattr(record, 'distance', None)

        record_type = 'check_in'
        if hasattr(record, 'action_description') and record.action_description:
            action_lower = record.action_description.lower()
            if 'out' in action_lower or 'checkout' in action_lower:
                record_type = 'check_out'

        _, work_type = parse_employee_id_for_work_type(str(record.employee_id))

        base_location_name = record.location_name
        if work_type and work_type in ('PT', 'SP', 'PW'):
            display_location_name = f"{base_location_name} ({work_type})"
        else:
            display_location_name = base_location_name

        converted_record = type('Record', (), {
            'id':                   record.id,
            'employee_id':          str(record.employee_id),
            'employee_name':        getattr(record, 'employee_name', ''),
            'check_in_date':        record.attendance_date,
            'check_in_time':        record.attendance_time,
            'location_name':        display_location_name,
            'original_location_name': base_location_name,
            'work_type':            work_type,
            'latitude':             None,
            'longitude':            None,
            'distance':             distance_value,
            'record_type':          record_type,
            'action_description':   record.action_description,
            'event_description':    record.event_description or '',
            'recorded_address':     record.recorded_address or '',
            'qr_code':              type('QRCode', (), {
                'location':         base_location_name,
                'location_address': record.recorded_address or '',
                'project':          None
            })()
        })()
        converted.append(converted_record)

    return converted


def _build_employee_name_map(records):
    """
    Build a {base_employee_id: "Lastname, Firstname"} map for export headers.

    Looks up the Employee table by numeric base ID so work-type suffixes
    (e.g. '3937SP') in the stored employee_name column do not pollute labels.
    Falls back to the stored employee_name on lookup failure.
    """
    from working_hours_calculator import parse_employee_id_for_work_type

    employee_names = {}
    for record in records:
        base_id, _ = parse_employee_id_for_work_type(str(record.employee_id))
        if base_id not in employee_names:
            try:
                emp = Employee.query.filter_by(id=int(base_id)).first()
                if emp:
                    employee_names[base_id] = f"{emp.lastName}, {emp.firstName}"
                else:
                    employee_names[base_id] = getattr(record, 'employee_name', f'Employee {base_id}')
                    logger_handler.logger.warning(
                        f"Employee ID {base_id} not found in employee table during export; "
                        f"using stored name."
                    )
            except Exception as e:
                employee_names[base_id] = getattr(record, 'employee_name', f'Employee {base_id}')
                logger_handler.logger.warning(
                    f"Could not lookup employee name for ID {base_id} during export: {e}"
                )
    return employee_names


def _make_export_styles():
    """
    Return a dict of openpyxl style objects shared by both export functions.

    Keys: header_font, header_fill, data_font, bold_font, italic_bold_font,
          border, missed_punch_fill, border_day_middle, border_day_last,
          border_day_single, border_day_first, amber_fill
    """
    header_font      = Font(name='Aptos Narrow', size=11, bold=True, color='FFFFFF')
    header_fill      = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    data_font        = Font(name='Aptos Narrow', size=11)
    bold_font        = Font(name='Aptos Narrow', size=11, bold=True)
    italic_bold_font = Font(name='Aptos Narrow', size=11, bold=True, italic=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    missed_punch_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    amber_fill        = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    border_day_middle = Border()
    border_day_last   = Border(bottom=Side(style='thin'))
    border_day_single = Border(bottom=Side(style='thin'))
    border_day_first  = Border()
    return {
        'header_font':      header_font,
        'header_fill':      header_fill,
        'data_font':        data_font,
        'bold_font':        bold_font,
        'italic_bold_font': italic_bold_font,
        'border':           border,
        'missed_punch_fill': missed_punch_fill,
        'amber_fill':       amber_fill,
        'border_day_middle': border_day_middle,
        'border_day_last':   border_day_last,
        'border_day_single': border_day_single,
        'border_day_first':  border_day_first,
    }


def export_time_attendance_excel(records, project_name_for_filename, date_range_str, filter_str, start_date_filter=None, end_date_filter=None):
    """Generate Excel export with template format matching the provided template"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    
    # Resolve date range and cap to 14-day window
    result = _resolve_date_range(start_date_filter, end_date_filter, records, 'TA Excel export')
    if result is None:
        return None
    start_date, end_date, records = result

    # Import parse function at the beginning for work type detection
    from working_hours_calculator import parse_employee_id_for_work_type
    
    # Convert TimeAttendance records to format expected by calculator
    converted_records = _convert_ta_records(records)
    
    # Calculate working hours using WorkingHoursCalculator
    calculator = WorkingHoursCalculator()
    hours_data = calculator.calculate_all_employees_hours(
        datetime.combine(start_date, datetime.min.time()),
        datetime.combine(end_date, datetime.max.time()),
        converted_records
    )
    
    # Build employee name map (Lastname, Firstname keyed by base employee ID)
    employee_names = _build_employee_name_map(records)
    
    # Setup styles (shared objects)
    _styles           = _make_export_styles()
    header_font       = _styles['header_font']
    header_fill       = _styles['header_fill']
    data_font         = _styles['data_font']
    bold_font         = _styles['bold_font']
    italic_bold_font  = _styles['italic_bold_font']
    border            = _styles['border']
    missed_punch_fill = _styles['missed_punch_fill']
    amber_fill        = _styles['amber_fill']
    border_day_middle = _styles['border_day_middle']
    border_day_last   = _styles['border_day_last']
    border_day_single = _styles['border_day_single']
    border_day_first  = _styles['border_day_first']

    def get_day_border(row_position, total_rows):
        """
        Get appropriate border style based on row position within a day.
        Matches sample.xlsx: only the LAST row of each day group has a bottom border.

        Args:
            row_position: Current row number (0-indexed) within the day
            total_rows: Total number of rows for this day

        Returns:
            Border object
        """
        if total_rows == 1:
            return border_day_single
        elif row_position == total_rows - 1:
            return border_day_last
        else:
            return border_day_middle

    # Orange background for Missed Punch
    missed_punch_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    # Write main headers
    current_row = 1
    
    # Row 1: Company name
    ws.merge_cells(f'A{current_row}:N{current_row}')
    title_cell = ws.cell(row=current_row, column=1, value=current_app.config.get('COMPANY_NAME', 'QR Code Management System'))
    title_cell.font = Font(name='Aptos Narrow', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 2: Summary title
    ws.merge_cells(f'A{current_row}:N{current_row}')
    summary_cell = ws.cell(row=current_row, column=1, value='Summary report of Hours worked')
    summary_cell.font = Font(name='Aptos Narrow', size=12, bold=True)
    summary_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 3: Project name
    project_display = project_name_for_filename.replace('_', ' ').strip() if project_name_for_filename else "[Project Name]"
    project_cell = ws.cell(row=current_row, column=1, value=project_display)
    project_cell.font = Font(name='Aptos Narrow', size=11, bold=True)
    project_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 4: Date range
    date_range_text = f"Date range: {start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"
    ws.merge_cells(f'A{current_row}:N{current_row}')
    date_cell = ws.cell(row=current_row, column=1, value=date_range_text)
    date_cell.font = Font(name='Aptos Narrow', size=11)
    date_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 5: Empty row
    current_row += 1
    
    # Empty row before first employee
    current_row += 1
    
    # Sort employees by name for organized output
    sorted_employees = sorted(
        hours_data['employees'].items(),
        key=lambda x: employee_names.get(x[0], f'Employee {x[0]}').lower()
    )

    # Write data for each employee (sorted by name)
    for employee_id, emp_data in sorted_employees:
        employee_name = employee_names.get(employee_id, f'Employee {employee_id}')
        
        # Employee header row (merged A to O)
        ws.merge_cells(f'A{current_row}:O{current_row}')
        emp_header = ws.cell(row=current_row, column=1, 
                            value=f'Employee ID {employee_id}: {employee_name}')
        emp_header.font = Font(name='Aptos Narrow', size=11, bold=True)
        emp_header.alignment = Alignment(horizontal='left')
        current_row += 1
        
        # Column headers
        headers = ['Day', 'Date', 'In', 'Out', 'Location', 'Zone', 'Hours/Building',
                  'Daily Total', 'Regular Hours', 'OT Hours', 'Building Address',
                  'Recorded Location', 'Distance (Mile)', 'Possible Violation']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            # White bold text on black background; no border (matching sample.xlsx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Group records by date AND location for separate rows per location
        daily_location_data = {}
        # Import parse function to match base employee ID with all variants (SP, PW, PT)
        from working_hours_calculator import parse_employee_id_for_work_type
        
        # Filter records where the BASE employee ID matches (includes 1234, 1234 SP, 1234 PW, 1234 PT)
        employee_records = []
        for r in converted_records:
            record_base_id, _ = parse_employee_id_for_work_type(str(r.employee_id))
            if record_base_id == employee_id:
                employee_records.append(r)

        for record in employee_records:
            date_key = record.check_in_date.strftime('%Y-%m-%d')
            location_key = record.location_name or 'Unknown Location'
            
            # Create nested structure: date -> location -> records
            if date_key not in daily_location_data:
                daily_location_data[date_key] = {}
            
            if location_key not in daily_location_data[date_key]:
                daily_location_data[date_key][location_key] = {
                    'records': [],
                    'location_name': location_key
                }
            
            daily_location_data[date_key][location_key]['records'].append(record)

        # -------------------------------------------------------------------
        # OVERNIGHT SHIFT DETECTION
        # The midnight check-out record is stored in the DB with the next
        # calendar day's date (e.g. checkout at 12:18 AM on Thursday is
        # stored as check_in_date = 2026-02-26).  We need to move it into
        # Wednesday's bucket so it pairs with the 8:18 PM check-in.
        #
        # Condition to move an early-morning checkout from Day N+1 -> Day N:
        #   Day N:   has an unmatched late check-in (>= 18:00)
        #   Day N+1: has an early check-out (<= 06:00) that belongs to Day N,
        #            detected by the absence of a non-evening IN on Day N+1
        #            that could own the early OUT (or raw count imbalance).
        # -------------------------------------------------------------------
        def _is_out(r):
            a = (r.action_description or '').lower()
            return 'out' in a or 'checkout' in a

        sorted_dk = sorted(daily_location_data.keys())
        for _di, _dk in enumerate(sorted_dk):
            if _di + 1 >= len(sorted_dk):
                continue

            # Guard: _dk or _ndk may have been deleted by a prior iteration
            # when all its records were moved to the previous day's bucket.
            # Without this check, iterating the stale sorted_dk snapshot raises KeyError.
            if _dk not in daily_location_data:
                continue

            _ndk = sorted_dk[_di + 1]
            if _ndk not in daily_location_data:
                continue

            # Must be consecutive calendar days
            _dn  = datetime.strptime(_dk,  '%Y-%m-%d').date()
            _dn1 = datetime.strptime(_ndk, '%Y-%m-%d').date()
            if (_dn1 - _dn).days != 1:
                continue

            # Flatten all records for Day N and Day N+1 across locations
            _day_recs  = [r for loc in daily_location_data[_dk].values()  for r in loc['records']]
            _next_recs = [r for loc in daily_location_data[_ndk].values() for r in loc['records']]

            _day_ins  = [r for r in _day_recs  if not _is_out(r)]
            _day_outs = [r for r in _day_recs  if     _is_out(r)]
            _nxt_ins  = [r for r in _next_recs if not _is_out(r)]
            _nxt_outs = [r for r in _next_recs if     _is_out(r)]

            # Early-morning OUTs on Day N (hour <= 3) are overnight orphans from
            # Day N-1.  Counting them as regular Day N outs inflates the out-count
            # and makes the day appear balanced, which suppresses detection of an
            # unmatched late IN that needs a next-day OUT.  Exclude them.
            _day_outs_non_early = [r for r in _day_outs if r.check_in_time.hour > 3]

            # Day N must have an unmatched late check-in (more INs than non-early OUTs,
            # with at least one IN at or after 12:00 PM)
            if len(_day_ins) <= len(_day_outs_non_early):
                continue
            _late_ins = [r for r in _day_ins if r.check_in_time.hour >= 12]
            if not _late_ins:
                continue

            # Find early-morning OUTs (<=03:00) on Day N+1
            _early_outs = [r for r in _nxt_outs if r.check_in_time.hour <= 3]
            if not _early_outs:
                continue

            # Determine whether the early OUT belongs to Day N or Day N+1.
            # It belongs to Day N when Day N+1 has no morning (< 12:00) check-in
            # that could own it, OR when OUTs outnumber INs on Day N+1.
            # This handles both cases:
            #   Case A: Day N+1 has only afternoon/evening INs (all >= 12:00) -> early OUT is Day N's
            #   Case B: Day N+1 has more OUTs than INs overall -> early OUT is unmatched
            # A morning IN on Day N+1 can only own an early OUT when that IN
            # occurs STRICTLY BEFORE the early OUT's time (IN → OUT is time-ordered).
            # An IN that starts AFTER the early OUT cannot own it and must NOT block
            # the overnight move (e.g. 01:55 AM IN cannot own a 01:00 AM OUT).
            _nxt_non_evening_ins = [
                r for r in _nxt_ins
                if r.check_in_time.hour < 12
                and any(r.check_in_time < eo.check_in_time for eo in _early_outs)
            ]
            if _nxt_non_evening_ins and len(_nxt_outs) <= len(_nxt_ins):
                # Day N+1 has a morning IN that can own the early OUT, and counts
                # are balanced -> do NOT move
                continue

            # Move up to as many early OUTs as there are unmatched late INs on Day N
            _to_move = _early_outs[:len(_late_ins)]
            for _co in _to_move:
                _co_loc = _co.location_name or 'Unknown Location'
                # Add to Day N bucket
                if _co_loc not in daily_location_data[_dk]:
                    daily_location_data[_dk][_co_loc] = {'records': [], 'location_name': _co_loc}
                daily_location_data[_dk][_co_loc]['records'].append(_co)
                # Remove from Day N+1 bucket
                if _ndk in daily_location_data and _co_loc in daily_location_data[_ndk]:
                    try:
                        daily_location_data[_ndk][_co_loc]['records'].remove(_co)
                    except ValueError:
                        pass
                    if not daily_location_data[_ndk][_co_loc]['records']:
                        del daily_location_data[_ndk][_co_loc]
                if _ndk in daily_location_data and not daily_location_data[_ndk]:
                    del daily_location_data[_ndk]
                logger_handler.logger.info(
                    f"TA Export overnight shift: moved checkout {_co.check_in_time} "
                    f"from {_ndk} to {_dk} for employee {employee_id}"
                )
        # -------------------------------------------------------------------
        # END OVERNIGHT SHIFT DETECTION
        # -------------------------------------------------------------------


        # Track weekly hours for overtime calculation
        weekly_total_hours = 0
        current_week_start = None
        grand_regular_hours = 0
        grand_ot_hours = 0
        # Accumulate SP/PW/PT hours from cross-type pairs (where the calculator
        # could not detect them because it processes each work-type stream independently).
        cross_type_sp_hours = 0.0
        cross_type_pw_hours = 0.0
        cross_type_pt_hours = 0.0
        
        # Get all dates that have records (not all weekdays)
        dates_with_records = sorted([
            date_str for date_str, day_data in emp_data['daily_hours'].items()
            if day_data.get('records_count', 0) > 0
        ])
        
        # Write daily data (ONLY DAYS WITH RECORDS)
        for date_str in dates_with_records:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            day_data = emp_data['daily_hours'][date_str]
            
            # Check for week boundary anchored to start_date_filter (not calendar Monday)
            _report_start = start_date if start_date_filter else date_obj.date()
            week_start = (_report_start + timedelta(days=((date_obj.date() - _report_start).days // 7) * 7))
            if current_week_start is not None and week_start != current_week_start:
                # Write weekly total row
                week_regular = min(weekly_total_hours, 40.0)
                week_overtime = max(0, weekly_total_hours - 40.0)
                
                ws.cell(row=current_row, column=7, value='Weekly Total: ').font = bold_font
                ws.cell(row=current_row, column=8, value=_qtr(weekly_total_hours)).font = bold_font
                ws.cell(row=current_row, column=9, value=_qtr(week_regular)).font = bold_font
                ws.cell(row=current_row, column=10, value=_qtr(week_overtime)).font = bold_font
                
                grand_regular_hours += week_regular
                grand_ot_hours += week_overtime
                current_row += 1
                
                weekly_total_hours = 0
            
            current_week_start = week_start
            
            # Get all locations for this date
            date_locations = daily_location_data.get(date_str, {})
            total_locations = len(date_locations)
            
            total_hours = day_data['total_hours']
            is_miss_punch = day_data.get('is_miss_punch', False)

            # Re-evaluate is_miss_punch from actual records in daily_location_data.
            # The overnight detection may have moved a checkout into this day's bucket
            # AFTER working_hours_calculator ran, so emp_data may still say
            # is_miss_punch=True even though the records now form a valid IN/OUT pair.
            if is_miss_punch and total_locations > 0:
                _all_recs_check = [r for loc in date_locations.values() for r in loc['records']]
                _ins_c  = sum(1 for r in _all_recs_check if not _is_out(r))
                _outs_c = sum(1 for r in _all_recs_check if     _is_out(r))
                if _ins_c > 0 and _outs_c > 0 and _ins_c == _outs_c:
                    # Balanced pairs — overnight fix resolved the miss punch
                    is_miss_punch = False
                    total_hours = 0.0  # will be recalculated below

            # Calculate total hours for the day by mirroring the display pairing logic:
            # group records by base location, apply the OUT-after-IN guard within each
            # group, and sum only complete pairs.  This ensures the daily total in
            # column H matches exactly the pairs rendered in the export rows.
            _day_total_hours = 0.0
            # Track which records are consumed by same-building pairing so the
            # cross-building pass only considers true orphans.
            _same_building_used_ids = set()
            for _loc_data in date_locations.values():
                _loc_recs = sorted(_loc_data['records'], key=_overnight_aware_sort_key)
                _loc_ins  = [r for r in _loc_recs if not _is_out(r)]
                _loc_outs = [r for r in _loc_recs if     _is_out(r)]
                _out_used = [False] * len(_loc_outs)
                for _in_r in _loc_ins:
                    for _oi2, _out_r in enumerate(_loc_outs):
                        if _out_used[_oi2]:
                            continue
                        # Time-only pairing guard (mirrors Step 1/2 pairing logic).
                        _in_t_d  = _in_r.check_in_time
                        _out_t_d = _out_r.check_in_time
                        if _out_t_d.hour <= 3:
                            if _in_t_d.hour < 12:
                                continue
                            # Orphan guard: an early-morning OUT whose check_in_date
                            # matches the current day is an orphan from the PREVIOUS
                            # overnight shift — it must NOT steal an evening IN.
                            # Only OUTs moved in by overnight detection (check_in_date
                            # is later than the current day) are valid partners.
                            _out_orig_date = _out_r.check_in_date
                            if hasattr(_out_orig_date, 'date'):
                                _out_orig_date = _out_orig_date.date()
                            if _out_orig_date <= date_obj.date():
                                continue
                        elif _out_t_d <= _in_t_d:
                            continue
                        _in_ts  = datetime.combine(_in_r.check_in_date,  _in_r.check_in_time)
                        _out_ts = datetime.combine(_out_r.check_in_date, _out_r.check_in_time)
                        if _out_ts < _in_ts:
                            _out_ts += timedelta(days=1)
                        _duration = (_out_ts - _in_ts).total_seconds() / 3600.0
                        if _duration > 24:
                            continue
                        _day_total_hours += _duration
                        _out_used[_oi2] = True
                        _same_building_used_ids.add(id(_in_r))
                        _same_building_used_ids.add(id(_out_r))
                        break

            # -------------------------------------------------------------------
            # CROSS-BUILDING PAIRING
            # After same-building pairing, collect all orphaned INs and OUTs
            # across every location group for this day.  Pair them chronologically
            # (earliest available OUT that is strictly after the IN).  This handles
            # employees who check in at one building and check out at another.
            # -------------------------------------------------------------------
            _all_day_recs_flat = []
            for _loc_data in date_locations.values():
                _all_day_recs_flat.extend(_loc_data['records'])

            _orphan_ins  = sorted(
                [r for r in _all_day_recs_flat if not _is_out(r) and id(r) not in _same_building_used_ids],
                key=_overnight_aware_sort_key
            )
            _orphan_outs = sorted(
                [r for r in _all_day_recs_flat if     _is_out(r) and id(r) not in _same_building_used_ids],
                key=_overnight_aware_sort_key
            )

            # Pre-compute cross-building pairs for this day (used both for totals
            # and for row writing after the location_groups loop).
            cross_building_pairs = []   # list of {'check_in': r, 'check_out': r, 'hours': float}
            _cb_out_used = [False] * len(_orphan_outs)

            for _cb_in in _orphan_ins:
                for _cb_oi, _cb_out in enumerate(_orphan_outs):
                    if _cb_out_used[_cb_oi]:
                        continue
                    # Same time-only guard as Steps 1–3
                    _cb_in_t  = _cb_in.check_in_time
                    _cb_out_t = _cb_out.check_in_time
                    if _cb_out_t.hour <= 3:
                        if _cb_in_t.hour < 12:
                            continue
                        # Orphan guard: same-day early-morning OUT is from previous
                        # overnight shift — skip it.  Only moved OUTs (check_in_date
                        # later than current day) are valid overnight partners.
                        _cb_out_orig = _cb_out.check_in_date
                        if hasattr(_cb_out_orig, 'date'):
                            _cb_out_orig = _cb_out_orig.date()
                        if _cb_out_orig <= date_obj.date():
                            continue
                    elif _cb_out_t <= _cb_in_t:
                        continue
                    _cb_in_ts  = datetime.combine(_cb_in.check_in_date,  _cb_in.check_in_time)
                    _cb_out_ts = datetime.combine(_cb_out.check_in_date, _cb_out.check_in_time)
                    if _cb_out_ts < _cb_in_ts:
                        _cb_out_ts += timedelta(days=1)
                    _cb_dur = (_cb_out_ts - _cb_in_ts).total_seconds() / 3600.0
                    if _cb_dur > 24:
                        continue
                    _cb_out_used[_cb_oi] = True
                    cross_building_pairs.append({
                        'check_in':  _cb_in,
                        'check_out': _cb_out,
                        'hours':     _cb_dur,
                    })
                    _day_total_hours += _cb_dur
                    logger_handler.logger.info(
                        f"[TA Export] Cross-building pair for employee {employee_id} on {date_str}: "
                        f"IN {_cb_in.location_name} @ {_cb_in.check_in_time} → "
                        f"OUT {_cb_out.location_name} @ {_cb_out.check_in_time} "
                        f"({_cb_dur:.2f} h)"
                    )
                    break

            # Build a set of record ids that are part of a cross-building pair so
            # the single-record group path can suppress its Missed Punch row.
            _cross_building_record_ids = set()
            for _cbp in cross_building_pairs:
                _cross_building_record_ids.add(id(_cbp['check_in']))
                _cross_building_record_ids.add(id(_cbp['check_out']))
            # -------------------------------------------------------------------
            # END CROSS-BUILDING PAIRING PRE-COMPUTATION
            # -------------------------------------------------------------------

            total_hours = _qtr(_day_total_hours)
            weekly_total_hours += total_hours

            # Daily total display (only shown on last location's last row)
            daily_total_display = _qtr(total_hours) if total_hours > 0 else ''

            # Get all records for the day and sort by time FIRST, then group by BASE location.
            # Grouping by base location (original_location_name) ensures that records from the
            # same building but different work types (e.g. regular IN + SP OUT) land in the
            # same group so the cross-type pairing rule can resolve them.
            all_day_records = []
            for loc_data in date_locations.values():
                all_day_records.extend(loc_data['records'])

            # Sort all records by time chronologically, overnight-aware
            all_day_records_sorted = sorted(all_day_records, key=_overnight_aware_sort_key)

            # Group consecutive records by BASE location (original_location_name without work-type
            # suffix) while maintaining time order.
            def _base_loc(r):
                return getattr(r, 'original_location_name', None) or r.location_name or 'Unknown Location'

            location_groups = []
            current_base_location = None
            current_group = []

            for record in all_day_records_sorted:
                bloc = _base_loc(record)
                if current_base_location is None or bloc == current_base_location:
                    current_base_location = bloc
                    current_group.append(record)
                else:
                    if current_group:
                        location_groups.append({
                            'location': current_base_location,
                            'records': current_group
                        })
                    current_base_location = bloc
                    current_group = [record]

            # Add the last group
            if current_group:
                location_groups.append({
                    'location': current_base_location,
                    'records': current_group
                })
            
            # Process each location group in chronological order
            total_groups = len(location_groups)
            for group_index, group_data in enumerate(location_groups):
                location_count = group_index + 1
                is_last_location = (location_count == total_groups)
                
                location_name = group_data['location']
                sorted_records = group_data['records']
                
                if len(sorted_records) == 1:
                    # Single record for this location
                    single_record = sorted_records[0]

                    # If this record has been resolved by cross-building pairing,
                    # suppress the Missed Punch row here — it will be written after
                    # all location groups have been processed (Touch Point 3).
                    if id(single_record) in _cross_building_record_ids:
                        continue
                    
                    # Get the original TimeAttendance record to check action_description
                    original_record = None
                    for rec in records:
                        if (rec.employee_id == single_record.employee_id and 
                            rec.attendance_date == single_record.check_in_date and 
                            rec.attendance_time == single_record.check_in_time):
                            original_record = rec
                            break
                    
                    # Determine if this is a check-in or check-out
                    is_check_out = False
                    if original_record and original_record.action_description:
                        action_lower = original_record.action_description.lower()
                        is_check_out = 'out' in action_lower or 'checkout' in action_lower
                    
                    # Show day name and date only for first group's first record
                    day_display = date_obj.strftime('%A').upper() if location_count == 1 else ''
                    date_display = date_obj.strftime('%m/%d/%Y') if location_count == 1 else ''
                    
                    # Show daily total only if this is the last group
                    current_daily_total = daily_total_display if is_last_location else ''
                    
                    if is_check_out:
                        # Orphaned check-out
                        row_data = [
                            day_display,
                            date_display,
                            '',  # No check-in time
                            single_record.check_in_time.strftime('%I:%M:%S %p'),  # Out
                            single_record.location_name,
                            '',
                            'Missed Punch',
                            current_daily_total,
                            '',
                            '',
                            single_record.event_description or '',
                            single_record.recorded_address or '',
                            getattr(single_record, 'distance', None) or '',
                            calculate_possible_violation(getattr(single_record, 'distance', None))
                        ]
                    else:
                        # Orphaned check-in
                        row_data = [
                            day_display,
                            date_display,
                            single_record.check_in_time.strftime('%I:%M:%S %p'),  # In
                            '',  # No check-out time
                            single_record.location_name,
                            '',
                            'Missed Punch',
                            current_daily_total,
                            '',
                            '',
                            single_record.event_description or '',
                            single_record.recorded_address or '',
                            getattr(single_record, 'distance', None) or '',
                            calculate_possible_violation(getattr(single_record, 'distance', None))
                        ]
                    
                    day_border = border_day_last if is_last_location else border_day_middle
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = data_font
                        cell.border = day_border
                        # Apply orange background to Missed Punch cell (column G)
                        if col == 7:
                            cell.fill = missed_punch_fill
                    current_row += 1
                    
                else:
                    # Multiple records for this location group.
                    # Build record_info with work_type included.
                    record_info = []
                    for record in sorted_records:
                        action_desc = record.action_description.lower() if record.action_description else ''
                        is_out = 'out' in action_desc or 'checkout' in action_desc
                        wt = getattr(record, 'work_type', None)  # None = regular
                        record_info.append({
                            'record': record,
                            'is_out': is_out,
                            'work_type': wt,   # None means regular
                            'used': False
                        })
                        logger_handler.logger.debug(f"TA Export record: time={record.check_in_time}, action='{record.action_description}', is_out={is_out}, work_type={wt}")

                    ins  = [ri for ri in record_info if not ri['is_out']]
                    outs = [ri for ri in record_info if     ri['is_out']]

                    pairs_to_write = []

                    # ── STEP 1: same-type pairing ──────────────────────────────────────
                    # Pair each IN with an OUT of the same work type first.
                    # Sort INs chronologically and OUTs with overnight-aware key so that
                    # an early-morning OUT (e.g. 00:30 moved in by overnight detection)
                    # sorts AFTER same-day evening OUTs and does not steal a daytime IN.
                    ins_sorted  = sorted(ins,  key=lambda ri: _overnight_aware_sort_key(ri['record']))
                    outs_sorted = sorted(outs, key=lambda ri: _overnight_aware_sort_key(ri['record']))

                    for in_ri in ins_sorted:
                        if in_ri['used']:
                            continue
                        for out_ri in outs_sorted:
                            if out_ri['used']:
                                continue
                            # Guard: time-only pairing rule.
                            # An early-morning OUT (hour<=3) is only valid for an afternoon/evening IN (hour>=12).
                            # For all other OUTs, the OUT time must be strictly after the IN time.
                            # Using time-only (not datetime) avoids false positives from moved overnight
                            # OUT records whose check_in_date is still a later date.
                            _in_t  = in_ri['record'].check_in_time
                            _out_t = out_ri['record'].check_in_time
                            if _out_t.hour <= 3:
                                if _in_t.hour < 12:
                                    continue  # early-morning OUT cannot pair with morning IN
                                # Orphan guard: same-day early-morning OUT is from a
                                # previous overnight shift — not a valid partner for
                                # this evening IN.  Only moved OUTs (check_in_date
                                # later than current day) should pair.
                                _out_orig_d = out_ri['record'].check_in_date
                                if hasattr(_out_orig_d, 'date'):
                                    _out_orig_d = _out_orig_d.date()
                                if _out_orig_d <= date_obj.date():
                                    continue
                            elif _out_t <= _in_t:
                                continue  # same-day OUT must be strictly after IN
                            if out_ri['work_type'] == in_ri['work_type']:
                                # Matched same work type — standard pair
                                in_ri['used']  = True
                                out_ri['used'] = True
                                pairs_to_write.append({
                                    'check_in':      in_ri['record'],
                                    'check_out':     out_ri['record'],
                                    'is_miss_punch': False,
                                    'effective_work_type': in_ri['work_type']
                                })
                                break

                    # ── STEP 2: cross-type pairing (forgot the work code) ──────────────
                    # If any INs or OUTs remain unmatched after same-type pairing,
                    # attempt to pair an unmatched IN with an unmatched OUT of a
                    # *different* work type.  Hours count as the special type's hours
                    # (if either side is special, the pair is treated as special;
                    # if both are different special types, use the OUT's type as
                    # the authoritative code — it's the scan that carries the code).
                    unmatched_ins  = [ri for ri in ins_sorted  if not ri['used']]
                    unmatched_outs = [ri for ri in outs_sorted if not ri['used']]

                    for in_ri in unmatched_ins:
                        if in_ri['used']:
                            continue
                        for out_ri in unmatched_outs:
                            if out_ri['used']:
                                continue
                            # Guard: same time-only rule as Step 1.
                            _in_t2  = in_ri['record'].check_in_time
                            _out_t2 = out_ri['record'].check_in_time
                            if _out_t2.hour <= 3:
                                if _in_t2.hour < 12:
                                    continue
                                # Orphan guard: same-day early-morning OUT is from a
                                # previous overnight shift — not a valid partner for
                                # this evening IN.  Only moved OUTs (check_in_date
                                # later than current day) should pair.
                                _out_orig_d2 = out_ri['record'].check_in_date
                                if hasattr(_out_orig_d2, 'date'):
                                    _out_orig_d2 = _out_orig_d2.date()
                                if _out_orig_d2 <= date_obj.date():
                                    continue
                            elif _out_t2 <= _in_t2:
                                continue
                            # Cross-type pair: one side is regular, other is special
                            # (or both special but different codes — treat OUT's type as definitive)
                            effective_wt = out_ri['work_type'] if out_ri['work_type'] else in_ri['work_type']
                            in_ri['used']  = True
                            out_ri['used'] = True
                            pairs_to_write.append({
                                'check_in':      in_ri['record'],
                                'check_out':     out_ri['record'],
                                'is_miss_punch': False,
                                'effective_work_type': effective_wt,
                                'is_cross_type': True
                            })
                            break

                    # ── STEP 3: remaining unmatched records → Missed Punch ─────────────
                    for ri in record_info:
                        if not ri['used']:
                            ri['used'] = True
                            if ri['is_out']:
                                pairs_to_write.append({
                                    'check_in':      None,
                                    'check_out':     ri['record'],
                                    'is_miss_punch': True,
                                    'effective_work_type': ri['work_type']
                                })
                            else:
                                pairs_to_write.append({
                                    'check_in':      ri['record'],
                                    'check_out':     None,
                                    'is_miss_punch': True,
                                    'effective_work_type': ri['work_type']
                                })

                    logger_handler.logger.debug(f"TA Export: created {len(pairs_to_write)} pairs for export")

                    # Sort pairs chronologically by the anchor record's time so that
                    # orphaned records (assembled last in Steps 2-3) appear in the
                    # correct time-order position relative to complete pairs.
                    def _pair_sort_key(pd):
                        anchor = pd['check_in'] or pd['check_out']
                        return _overnight_aware_sort_key(anchor) if anchor else 0
                    pairs_to_write.sort(key=_pair_sort_key)

                    # Write all pairs
                    for pair_idx, pair_data in enumerate(pairs_to_write):
                        check_in_record = pair_data['check_in']
                        check_out_record = pair_data['check_out']
                        is_miss_punch = pair_data['is_miss_punch']
                        
                        # Show day name and date only for first pair of first location
                        day_display = date_obj.strftime('%A').upper() if (location_count == 1 and pair_idx == 0) else ''
                        date_display = date_obj.strftime('%m/%d/%Y') if (location_count == 1 and pair_idx == 0) else ''
                        
                        # Calculate hours if complete pair
                        if check_in_record and check_out_record and not is_miss_punch:
                            pair_datetime_in  = datetime.combine(check_in_record.check_in_date,  check_in_record.check_in_time)
                            pair_datetime_out = datetime.combine(check_out_record.check_in_date, check_out_record.check_in_time)
                            # If check-out time is before check-in time (overnight shift),
                            # add one day to the check-out datetime so the duration is positive and correct.
                            if pair_datetime_out < pair_datetime_in:
                                pair_datetime_out += timedelta(days=1)
                            pair_hours = (pair_datetime_out - pair_datetime_in).total_seconds() / 3600.0
                            pair_hours = round(pair_hours, 2)
                        else:
                            pair_hours = 'Missed Punch'
                        
                        # Accumulate SP/PW/PT hours for CROSS-TYPE pairs only.
                        # Same-type SP/PW/PT pairs are already captured in grand_totals
                        # by WorkingHoursCalculator; adding them again here would double-count.
                        if not is_miss_punch and isinstance(pair_hours, (int, float)) and pair_data.get('is_cross_type', False):
                            _ewt = pair_data.get('effective_work_type')
                            if _ewt == 'SP':
                                cross_type_sp_hours += pair_hours
                            elif _ewt == 'PW':
                                cross_type_pw_hours += pair_hours
                            elif _ewt == 'PT':
                                cross_type_pt_hours += pair_hours

                        # Determine whether this is an overnight pair:
                        # check-in is late evening (>= 20:00) AND check-out is early morning (<= 03:00)
                        # Both records share the same check_in_date in the DB for this scenario.
                        _is_overnight_pair = (
                            check_in_record and check_out_record and
                            check_in_record.check_in_time.hour >= 20 and
                            check_out_record.check_in_time.hour <= 3
                        )

                        # Show daily total on last pair of last location
                        is_last_pair = (pair_idx == len(pairs_to_write) - 1) and is_last_location
                        current_daily_total = daily_total_display if is_last_pair else ''
                        
                        # Build Out-time string (plain time only)
                        _out_time_str = check_out_record.check_in_time.strftime('%I:%M:%S %p') if check_out_record else ''

                        # Build Location string.
                        # For a complete pair, derive the display name from effective_work_type:
                        #   - regular pair   → base location name (no suffix)
                        #   - SP/PW/PT pair  → base location name + " (SP/PW/PT)"
                        # 'regular' is treated identically to None — no suffix is shown.
                        # For orphaned records keep their own location_name.
                        _effective_wt = pair_data.get('effective_work_type')
                        _is_special_wt = _effective_wt in ('SP', 'PW', 'PT')
                        _ref_record   = check_in_record or check_out_record
                        if check_in_record and check_out_record:
                            _base = _base_loc(check_in_record)
                            if _is_special_wt:
                                _location_str = f"{_base} ({_effective_wt})"
                            else:
                                _location_str = _base
                        else:
                            _location_str = _ref_record.location_name if _ref_record else ''

                        if _is_overnight_pair:
                            _location_str = f"{_location_str} (midnight shift)"
                        
                        # Build row data
                        if check_in_record and check_out_record:
                            row_data = [
                                day_display,
                                date_display,
                                check_in_record.check_in_time.strftime('%I:%M:%S %p'),  # In
                                _out_time_str,  # Out
                                _location_str,  # Location (effective work type + optional midnight label)
                                '',
                                pair_hours,
                                current_daily_total,
                                '',
                                '',
                                check_in_record.event_description or '',
                                check_in_record.recorded_address or '',
                                getattr(check_in_record, 'distance', None) or '',
                                calculate_possible_violation(getattr(check_in_record, 'distance', None))
                            ]
                        elif check_in_record:  # IN without OUT
                            row_data = [
                                day_display,
                                date_display,
                                check_in_record.check_in_time.strftime('%I:%M:%S %p'),  # In
                                '',  # No OUT
                                _location_str,
                                '',
                                'Missed Punch',
                                current_daily_total,
                                '',
                                '',
                                check_in_record.event_description or '',
                                check_in_record.recorded_address or '',
                                getattr(check_in_record, 'distance', None) or '',
                                calculate_possible_violation(getattr(check_in_record, 'distance', None))
                            ]
                        else:  # OUT without IN
                            row_data = [
                                day_display,
                                date_display,
                                '',  # No IN
                                check_out_record.check_in_time.strftime('%I:%M:%S %p'),  # Out
                                _location_str,
                                '',
                                'Missed Punch',
                                current_daily_total,
                                '',
                                '',
                                check_out_record.event_description or '',
                                check_out_record.recorded_address or '',
                                getattr(check_out_record, 'distance', None) or '',
                                calculate_possible_violation(getattr(check_out_record, 'distance', None))
                            ]
                        
                        day_border = border_day_last if is_last_pair else border_day_middle
                        for col, value in enumerate(row_data, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.font = data_font
                            cell.border = day_border
                            # Apply orange background to Missed Punch cell
                            if col == 7 and value == 'Missed Punch':
                                cell.fill = missed_punch_fill
                        current_row += 1

            # -------------------------------------------------------------------
            # CROSS-BUILDING PAIR ROW WRITING (Touch Point 3)
            # Write one row per cross-building pair identified during pre-computation.
            # The day-name and date columns are only shown for the very first row
            # of this day that is actually rendered; we track that with a flag.
            # -------------------------------------------------------------------
            if cross_building_pairs:
                # Determine whether any non-cross-building rows were already written
                # for this day.  We look at how many rows were consumed since the
                # start of this date's block.  The simplest proxy: check whether
                # the first location group had at least one real (non-skipped) record.
                # We use a dedicated flag instead to keep this clean.
                _cb_first_row_of_day = not any(
                    id(r) not in _cross_building_record_ids
                    for loc_data in date_locations.values()
                    for r in loc_data['records']
                )

                for _cb_idx, _cbp in enumerate(cross_building_pairs):
                    _cb_in_rec  = _cbp['check_in']
                    _cb_out_rec = _cbp['check_out']
                    _cb_hours   = _cbp['hours']
                    _cb_pair_hours = round(_cb_hours, 2)

                    _is_last_cb = (_cb_idx == len(cross_building_pairs) - 1)

                    # Show day/date only on the very first row written for this date
                    # (either this is the first row overall, or prior groups had records)
                    if _cb_idx == 0 and _cb_first_row_of_day:
                        _cb_day_display  = date_obj.strftime('%A').upper()
                        _cb_date_display = date_obj.strftime('%m/%d/%Y')
                    else:
                        _cb_day_display  = ''
                        _cb_date_display = ''

                    # Show daily total on the last cross-building row if it is
                    # also the last row written for this day.
                    _cb_daily_total = daily_total_display if _is_last_cb else ''

                    # Location label: clearly identifies both buildings
                    _cb_in_loc  = _base_loc(_cb_in_rec)
                    _cb_out_loc = _base_loc(_cb_out_rec)
                    _cb_loc_str = f"IN: {_cb_in_loc} → OUT: {_cb_out_loc}"

                    row_data = [
                        _cb_day_display,
                        _cb_date_display,
                        _cb_in_rec.check_in_time.strftime('%I:%M:%S %p'),   # In
                        _cb_out_rec.check_in_time.strftime('%I:%M:%S %p'),  # Out
                        _cb_loc_str,
                        '',
                        _cb_pair_hours,
                        _cb_daily_total,
                        '',
                        '',
                        _cb_in_rec.event_description or '',
                        _cb_in_rec.recorded_address or '',
                        getattr(_cb_in_rec, 'distance', None) or '',
                        calculate_possible_violation(getattr(_cb_in_rec, 'distance', None))
                    ]

                    _cb_border = border_day_last if _is_last_cb else border_day_middle
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = data_font
                        cell.border = _cb_border
                    current_row += 1
            # -------------------------------------------------------------------
            # END CROSS-BUILDING PAIR ROW WRITING
            # -------------------------------------------------------------------

        # Write final weekly total for this employee
        if weekly_total_hours > 0:
            week_regular = min(weekly_total_hours, 40.0)
            week_overtime = max(0, weekly_total_hours - 40.0)
            
            ws.cell(row=current_row, column=7, value='Weekly Total: ').font = bold_font
            ws.cell(row=current_row, column=8, value=_qtr(weekly_total_hours)).font = bold_font
            ws.cell(row=current_row, column=9, value=_qtr(week_regular)).font = bold_font
            ws.cell(row=current_row, column=10, value=_qtr(week_overtime)).font = bold_font
            
            grand_regular_hours += week_regular
            grand_ot_hours += week_overtime
            current_row += 1
        
        # Write extra working hours rows (SP/PW/PT) if employee has any
        # Get extra hours from emp_data grand_totals, then add any cross-type hours
        # accumulated during rendering (pairs the calculator could not detect).
        grand_totals = emp_data.get('grand_totals', {})
        sp_hours = grand_totals.get('sp_hours', 0.0) + cross_type_sp_hours
        pw_hours = grand_totals.get('pw_hours', 0.0) + cross_type_pw_hours
        pt_hours = grand_totals.get('pt_hours', 0.0) + cross_type_pt_hours
        
        # Write SP row if hours > 0
        if sp_hours > 0:
            ws.cell(row=current_row, column=7, value='Special Project (SP): ').font = Font(name='Aptos Narrow', size=11, bold=True, italic=True)
            ws.cell(row=current_row, column=9, value=round(sp_hours, 2)).font = Font(name='Aptos Narrow', size=11, bold=True, italic=True)
            # Log SP hours export
            logger_handler.logger.info(f"Export: Employee {employee_id} SP hours: {sp_hours:.2f}")
            current_row += 1
        
        # Write PW row if hours > 0
        if pw_hours > 0:
            ws.cell(row=current_row, column=7, value='Periodic Work (PW): ').font = Font(name='Aptos Narrow', size=11, bold=True, italic=True)
            ws.cell(row=current_row, column=9, value=round(pw_hours, 2)).font = Font(name='Aptos Narrow', size=11, bold=True, italic=True)
            # Log PW hours export
            logger_handler.logger.info(f"Export: Employee {employee_id} PW hours: {pw_hours:.2f}")
            current_row += 1
        
        # Write PT row if hours > 0
        if pt_hours > 0:
            ws.cell(row=current_row, column=7, value='Project Team (PT): ').font = Font(name='Aptos Narrow', size=11, bold=True, italic=True)
            ws.cell(row=current_row, column=9, value=round(pt_hours, 2)).font = Font(name='Aptos Narrow', size=11, bold=True, italic=True)
            # Log PT hours export
            logger_handler.logger.info(f"Export: Employee {employee_id} PT hours: {pt_hours:.2f}")
            current_row += 1
        
        # Write GRAND TOTAL row
        ws.cell(row=current_row, column=7, value='GRAND TOTAL: ').font = Font(name='Aptos Narrow', size=11, bold=True)
        ws.cell(row=current_row, column=9, value=_qtr(grand_regular_hours)).font = Font(name='Aptos Narrow', size=11, bold=True)
        ws.cell(row=current_row, column=10, value=_qtr(grand_ot_hours)).font = Font(name='Aptos Narrow', size=11, bold=True)
        current_row += 1
        
        # Empty row after each employee
        current_row += 1
    
    # Auto-size columns - handle merged cells properly
    for col_idx in range(1, 15):
        column_letter = get_column_letter(col_idx)
        
        # Set fixed width for Day column (column A)
        if col_idx == 1:
            ws.column_dimensions[column_letter].width = 18
            continue
        
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Filename
    if date_range_str:
        filename = f'{project_name_for_filename}time_attendance_{date_range_str}.xlsx'
    else:
        filename = f'{project_name_for_filename}time_attendance.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def export_time_attendance_by_building_excel(records, project_name_for_filename, date_range_str, start_date_filter=None, end_date_filter=None):
    """Generate Excel export grouped by building/location with template format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    
    # Resolve date range and cap to 14-day window
    result = _resolve_date_range(start_date_filter, end_date_filter, records, 'TA by-building Excel export')
    if result is None:
        return None
    start_date, end_date, records = result

    from working_hours_calculator import parse_employee_id_for_work_type
    
    # Convert TimeAttendance records to format expected by calculator
    converted_records = _convert_ta_records(records)
    
    # Group records by location (building)
    location_groups = {}
    for record in converted_records:
        loc_name = record.original_location_name or 'Unknown Location'
        if loc_name not in location_groups:
            location_groups[loc_name] = []
        location_groups[loc_name].append(record)
    
    # Sort locations alphabetically
    sorted_locations = sorted(location_groups.keys())
    
    # Log grouping info
    logger_handler.logger.info(
        f"Export by Building: Grouped {len(converted_records)} records into {len(sorted_locations)} locations"
    )
    
    # Calculate working hours using WorkingHoursCalculator for SP/PT/PW hours
    calculator = WorkingHoursCalculator()
    hours_data = calculator.calculate_all_employees_hours(
        datetime.combine(start_date, datetime.min.time()),
        datetime.combine(end_date, datetime.max.time()),
        converted_records
    )
    
    # Build employee name map (Lastname, Firstname keyed by base employee ID)
    employee_names = _build_employee_name_map(records)
    
    # Setup styles (shared objects)
    _styles          = _make_export_styles()
    header_font      = _styles['header_font']
    header_fill      = _styles['header_fill']
    data_font        = _styles['data_font']
    bold_font        = _styles['bold_font']
    italic_bold_font = _styles['italic_bold_font']
    border           = _styles['border']
    missed_punch_fill = _styles['missed_punch_fill']
    amber_fill       = _styles['amber_fill']
    border_day_middle = _styles['border_day_middle']
    border_day_last   = _styles['border_day_last']
    
    # Write main headers
    current_row = 1
    
    # Row 1: Company name
    ws.merge_cells(f'A{current_row}:N{current_row}')
    title_cell = ws.cell(row=current_row, column=1, value=current_app.config.get('COMPANY_NAME', 'QR Code Management System'))
    title_cell.font = Font(name='Aptos Narrow', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 2: Summary title
    ws.merge_cells(f'A{current_row}:N{current_row}')
    summary_cell = ws.cell(row=current_row, column=1, value='Summary report of Hours worked')
    summary_cell.font = Font(name='Aptos Narrow', size=12, bold=True)
    summary_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 3: Project name
    project_display = project_name_for_filename.replace('_', ' ').strip() if project_name_for_filename else "[Project Name]"
    project_cell = ws.cell(row=current_row, column=1, value=project_display)
    project_cell.font = Font(name='Aptos Narrow', size=11, bold=True)
    project_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Row 4: Date range
    date_range_text = f"Date range: {start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"
    ws.merge_cells(f'A{current_row}:N{current_row}')
    date_cell = ws.cell(row=current_row, column=1, value=date_range_text)
    date_cell.font = Font(name='Aptos Narrow', size=11)
    date_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Empty rows before first building
    current_row += 2
    
    # Process each building/location
    for location_index, location_name in enumerate(sorted_locations, 1):
        location_records = location_groups[location_name]
        
        # Get zone info from QR code if available
        zone_info = ''
        try:
            qr_code = QRCode.query.filter_by(location=location_name).first()
            if qr_code:
                zone_info = getattr(qr_code, 'zone', '') or ''
        except:
            pass
        
        # Building header row
        building_header = f"{location_index}) {location_name} - Zone {zone_info}"
        ws.merge_cells(f'A{current_row}:O{current_row}')
        building_cell = ws.cell(row=current_row, column=1, value=building_header)
        building_cell.font = Font(name='Aptos Narrow', size=11, bold=True)
        building_cell.alignment = Alignment(horizontal='left')
        current_row += 1
        
        # Get unique employees for this location
        employees_at_location = {}
        for record in location_records:
            base_id, _ = parse_employee_id_for_work_type(record.employee_id)
            if base_id not in employees_at_location:
                employees_at_location[base_id] = []
            employees_at_location[base_id].append(record)
        
        # Sort employees by name
        sorted_employee_ids = sorted(
            employees_at_location.keys(),
            key=lambda emp_id: employee_names.get(emp_id, f'Employee {emp_id}').lower()
        )
        
        # Process each employee at this location
        for employee_id in sorted_employee_ids:
            emp_records = employees_at_location[employee_id]
            emp_name = employee_names.get(employee_id, f'Employee {employee_id}')
            
            # Compute SP/PW/PT hours from the records already scoped to this
            # building and employee (emp_records).  Using the calculator's
            # grand_totals here would be incorrect: those totals are GLOBAL
            # (across all buildings), so an employee with SP hours at Building A
            # would incorrectly show an SP row at Building B where they have none.
            #
            # Strategy: pair same-building SP/PW/PT records the same way the
            # main loop pairs regular records, and sum the durations.
            def _building_special_hours(emp_recs, work_type_code):
                """Sum paired hours for a given work-type code at this building."""
                from datetime import datetime as _dt, timedelta as _td
                wt_recs = [r for r in emp_recs if getattr(r, 'work_type', None) == work_type_code]
                if not wt_recs:
                    return 0.0
                # Group by date
                by_date = {}
                for r in wt_recs:
                    dk = r.check_in_date.strftime('%Y-%m-%d') if hasattr(r.check_in_date, 'strftime') else str(r.check_in_date)
                    by_date.setdefault(dk, []).append(r)
                total = 0.0
                for dk, day_recs in by_date.items():
                    day_recs_s = sorted(day_recs, key=_overnight_aware_sort_key)
                    ins_r  = [r for r in day_recs_s if not ('out' in (r.action_description or '').lower() or 'checkout' in (r.action_description or '').lower())]
                    outs_r = [r for r in day_recs_s if     ('out' in (r.action_description or '').lower() or 'checkout' in (r.action_description or '').lower())]
                    used = [False] * len(outs_r)
                    d_obj = _dt.strptime(dk, '%Y-%m-%d')
                    for in_r in ins_r:
                        for oi, out_r in enumerate(outs_r):
                            if used[oi]:
                                continue
                            in_dt  = _dt.combine(d_obj, in_r.check_in_time)
                            out_dt = _dt.combine(d_obj, out_r.check_in_time)
                            if out_dt < in_dt:
                                out_dt += _td(days=1)
                            dur = (out_dt - in_dt).total_seconds() / 3600.0
                            if 0 < dur < 24:
                                total += dur
                                used[oi] = True
                                break
                return total

            sp_hours = _building_special_hours(emp_records, 'SP')
            pw_hours = _building_special_hours(emp_records, 'PW')
            pt_hours = _building_special_hours(emp_records, 'PT')
            
            # Employee header row
            ws.merge_cells(f'A{current_row}:O{current_row}')
            emp_header = ws.cell(row=current_row, column=1, 
                                value=f'Employee ID {employee_id}: {emp_name}')
            emp_header.font = Font(name='Aptos Narrow', size=11, bold=True)
            emp_header.alignment = Alignment(horizontal='left')
            current_row += 1
            
            # Column headers
            headers = ['Day', 'Date', 'In', 'Out', 'Location', 'Zone', 'Hours/Building',
                      'Daily Total', 'Regular Hours', 'OT Hours', 'Building Address',
                      'Recorded Location', 'Distance (Mile)', 'Possible Violation']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Group employee records by date
            daily_records = {}
            for record in emp_records:
                date_key = record.check_in_date.strftime('%Y-%m-%d')
                if date_key not in daily_records:
                    daily_records[date_key] = []
                daily_records[date_key].append(record)
            
            # -----------------------------------------------------------
            # OVERNIGHT SHIFT DETECTION (by-building export)
            # The midnight check-out record is stored in the DB on the
            # next calendar day's date (e.g. checkout at 01:00 AM on
            # Thursday is stored as check_in_date = Thursday).  Move it
            # into Wednesday's bucket so it pairs with the 8 PM check-in.
            #
            # Mirrors the identical logic in export_time_attendance_excel.
            # -----------------------------------------------------------
            def _bb_is_out(r):
                a = (r.action_description or '').lower()
                return 'out' in a or 'checkout' in a

            _bb_sorted_dk = sorted(daily_records.keys())
            for _bb_di, _bb_dk in enumerate(_bb_sorted_dk):
                if _bb_di + 1 >= len(_bb_sorted_dk):
                    continue
                # Guard: bucket may have been emptied by a prior iteration
                if _bb_dk not in daily_records:
                    continue
                _bb_ndk = _bb_sorted_dk[_bb_di + 1]
                if _bb_ndk not in daily_records:
                    continue
                # Must be consecutive calendar days
                _bb_dn  = datetime.strptime(_bb_dk,  '%Y-%m-%d').date()
                _bb_dn1 = datetime.strptime(_bb_ndk, '%Y-%m-%d').date()
                if (_bb_dn1 - _bb_dn).days != 1:
                    continue
                # Collect INs/OUTs for Day N and Day N+1
                _bb_day_recs  = daily_records[_bb_dk]
                _bb_next_recs = daily_records[_bb_ndk]
                _bb_day_ins   = [r for r in _bb_day_recs  if not _bb_is_out(r)]
                _bb_day_outs  = [r for r in _bb_day_recs  if     _bb_is_out(r)]
                _bb_nxt_ins   = [r for r in _bb_next_recs if not _bb_is_out(r)]
                _bb_nxt_outs  = [r for r in _bb_next_recs if     _bb_is_out(r)]
                # Exclude early-morning OUTs on Day N from the balance check:
                # they are overnight orphans from Day N-1, not Day N regulars.
                _bb_day_outs_non_early = [r for r in _bb_day_outs if r.check_in_time.hour > 3]
                # Day N must have an unmatched late check-in (>= 12:00 PM)
                if len(_bb_day_ins) <= len(_bb_day_outs_non_early):
                    continue
                _bb_late_ins = [r for r in _bb_day_ins if r.check_in_time.hour >= 12]
                if not _bb_late_ins:
                    continue
                # Find early-morning OUTs (<= 03:00) on Day N+1
                _bb_early_outs = [r for r in _bb_nxt_outs if r.check_in_time.hour <= 3]
                if not _bb_early_outs:
                    continue
                # Non-morning INs guard: do NOT move if Day N+1 has a morning
                # IN (< 12:00) that precedes the early OUT (i.e. it can own the early OUT)
                # and the counts are balanced.
                _bb_nxt_non_evening_ins = [
                    r for r in _bb_nxt_ins
                    if r.check_in_time.hour < 12
                    and any(r.check_in_time < eo.check_in_time for eo in _bb_early_outs)
                ]
                if _bb_nxt_non_evening_ins and len(_bb_nxt_outs) <= len(_bb_nxt_ins):
                    continue
                # Move up to as many early OUTs as there are unmatched late INs
                _bb_to_move = _bb_early_outs[:len(_bb_late_ins)]
                for _bb_co in _bb_to_move:
                    daily_records[_bb_dk].append(_bb_co)
                    daily_records[_bb_ndk].remove(_bb_co)
                    if not daily_records[_bb_ndk]:
                        del daily_records[_bb_ndk]
                    logger_handler.logger.info(
                        f"[TA by-building Export] Overnight: moved checkout "
                        f"{_bb_co.check_in_time} from {_bb_ndk} to {_bb_dk} "
                        f"for employee {employee_id} at {location_name}"
                    )
            # -----------------------------------------------------------
            # END OVERNIGHT SHIFT DETECTION
            # -----------------------------------------------------------

            # Track weekly hours for overtime calculation
            weekly_total_hours = 0
            current_week_start = None
            grand_regular_hours = 0
            grand_ot_hours = 0
            
            # Sort dates (re-sort after overnight detection may have removed buckets).
            # CRITICAL: cap to end_date — daily_records may contain the +1 buffer day
            # (fetched so overnight checkout records are available for pairing) but
            # that extra day must never be rendered, or it creates a spurious 3rd week.
            sorted_dates = sorted(
                dk for dk in daily_records.keys()
                if datetime.strptime(dk, '%Y-%m-%d').date() <= end_date
            )
            
            for date_str in sorted_dates:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                # Sort records overnight-aware: early-morning OUTs (<=03:00) sort after
                # evening records so they pair with the correct evening check-in.
                day_records = sorted(daily_records[date_str], key=_overnight_aware_sort_key)
                
                # Check for week boundary anchored to start_date_filter (not calendar Monday)
                _report_start = start_date if start_date_filter else date_obj.date()
                week_start = (_report_start + timedelta(days=((date_obj.date() - _report_start).days // 7) * 7))
                if current_week_start is not None and week_start != current_week_start:
                    # Write weekly total row
                    week_regular = min(weekly_total_hours, 40.0)
                    week_overtime = max(0, weekly_total_hours - 40.0)
                    
                    ws.cell(row=current_row, column=7, value='Weekly Total: ').font = bold_font
                    ws.cell(row=current_row, column=8, value=_qtr(weekly_total_hours)).font = bold_font
                    ws.cell(row=current_row, column=9, value=_qtr(week_regular)).font = bold_font
                    ws.cell(row=current_row, column=10, value=_qtr(week_overtime)).font = bold_font
                    
                    grand_regular_hours += week_regular
                    grand_ot_hours += week_overtime
                    current_row += 1
                    
                    weekly_total_hours = 0
                
                current_week_start = week_start
                
                # Re-evaluate miss-punch status after overnight detection may
                # have moved a next-day checkout into this day's bucket.
                # If INs and OUTs are now balanced, this day is no longer a
                # miss punch (mirrors logic in export_time_attendance_excel).
                _bb_all_day = day_records
                _bb_ins_count  = sum(1 for r in _bb_all_day if not _bb_is_out(r))
                _bb_outs_count = sum(1 for r in _bb_all_day if     _bb_is_out(r))
                _bb_day_is_miss_punch = (_bb_ins_count != _bb_outs_count)

                # Process day records - create IN/OUT pairs
                record_info = []
                for record in day_records:
                    action_desc = record.action_description.lower() if record.action_description else ''
                    is_out = 'out' in action_desc or 'checkout' in action_desc
                    record_info.append({
                        'record': record,
                        'is_out': is_out,
                        'used': False
                    })

                # Create pairs
                pairs = []
                ins  = [ri for ri in record_info if not ri['is_out']]
                outs = [ri for ri in record_info if     ri['is_out']]

                if len(ins) > len(outs) and len(outs) > 0:
                    # Odd-IN rule: discard all but the LATEST IN; pair it with the earliest OUT.
                    # Use overnight-aware sort so early-morning OUTs sort after evening OUTs.
                    ins_sorted  = sorted(ins,  key=lambda ri: _overnight_aware_sort_key(ri['record']))
                    outs_sorted = sorted(outs, key=lambda ri: _overnight_aware_sort_key(ri['record']))

                    latest_in   = ins_sorted[-1]
                    excess_ins  = ins_sorted[:-1]

                    # Orphan guard: when the latest IN is afternoon/evening (>=12h), skip
                    # early-morning OUTs (<=3h) whose check_in_date matches the
                    # current day — they are orphans from a previous overnight shift.
                    _oi_in_hour = latest_in['record'].check_in_time.hour
                    earliest_out = None
                    _oi_skip = []
                    for _oi_ri in outs_sorted:
                        if (earliest_out is None
                                and _oi_in_hour >= 12
                                and _oi_ri['record'].check_in_time.hour <= 3):
                            _oi_out_d = _oi_ri['record'].check_in_date
                            if hasattr(_oi_out_d, 'date'):
                                _oi_out_d = _oi_out_d.date()
                            if _oi_out_d <= date_obj.date():
                                _oi_skip.append(_oi_ri)
                                continue
                        if earliest_out is None:
                            earliest_out = _oi_ri
                            break

                    for ri in excess_ins:
                        ri['used'] = True
                        pairs.append({'check_in': ri['record'], 'check_out': None, 'is_miss_punch': True})

                    if earliest_out is not None:
                        latest_in['used']    = True
                        earliest_out['used'] = True
                        pairs.append({'check_in': latest_in['record'], 'check_out': earliest_out['record'], 'is_miss_punch': False})
                    else:
                        latest_in['used'] = True
                        pairs.append({'check_in': latest_in['record'], 'check_out': None, 'is_miss_punch': True})

                    for ri in outs_sorted:
                        if not ri['used'] and ri not in _oi_skip:
                            ri['used'] = True
                            pairs.append({'check_in': None, 'check_out': ri['record'], 'is_miss_punch': True})
                    # Orphan OUTs that were skipped
                    for ri in _oi_skip:
                        ri['used'] = True
                        pairs.append({'check_in': None, 'check_out': ri['record'], 'is_miss_punch': True})

                else:
                    # Standard pairing
                    i = 0
                    while i < len(record_info):
                        if record_info[i]['used']:
                            i += 1
                            continue

                        if not record_info[i]['is_out']:  # IN
                            out_found = False
                            for j in range(i + 1, len(record_info)):
                                if record_info[j]['used']:
                                    continue
                                if record_info[j]['is_out']:
                                    # Orphan guard: when this IN is an afternoon/evening
                                    # check-in (>=12h) and the candidate OUT is
                                    # early-morning (<=3h), the OUT is only a
                                    # valid partner if it was moved in by overnight
                                    # detection (check_in_date > current day).
                                    # Same-day early-morning OUTs are orphans from
                                    # a previous overnight shift.
                                    _in_rec  = record_info[i]['record']
                                    _out_rec = record_info[j]['record']
                                    if (_in_rec.check_in_time.hour >= 12
                                            and _out_rec.check_in_time.hour <= 3):
                                        _out_bb_date = _out_rec.check_in_date
                                        if hasattr(_out_bb_date, 'date'):
                                            _out_bb_date = _out_bb_date.date()
                                        if _out_bb_date <= date_obj.date():
                                            continue  # orphan — skip
                                    pairs.append({
                                        'check_in': record_info[i]['record'],
                                        'check_out': record_info[j]['record'],
                                        'is_miss_punch': False
                                    })
                                    record_info[i]['used'] = True
                                    record_info[j]['used'] = True
                                    out_found = True
                                    break

                            if not out_found:
                                pairs.append({
                                    'check_in': record_info[i]['record'],
                                    'check_out': None,
                                    'is_miss_punch': True
                                })
                                record_info[i]['used'] = True
                        else:  # Orphaned OUT
                            pairs.append({
                                'check_in': None,
                                'check_out': record_info[i]['record'],
                                'is_miss_punch': True
                            })
                            record_info[i]['used'] = True

                        i += 1
                
                # Calculate daily hours
                daily_hours = 0
                for pair in pairs:
                    if pair['check_in'] and pair['check_out'] and not pair['is_miss_punch']:
                        pair_in  = datetime.combine(date_obj, pair['check_in'].check_in_time)
                        pair_out = datetime.combine(date_obj, pair['check_out'].check_in_time)
                        # Overnight shift correction: if OUT is before IN on the same
                        # calendar date, the employee worked past midnight — advance
                        # pair_out by one day so the duration is always positive.
                        if pair_out < pair_in:
                            pair_out += timedelta(days=1)
                        _bb_dur = (pair_out - pair_in).total_seconds() / 3600.0
                        # 24h guard: reject implausible durations (data errors)
                        if _bb_dur <= 24:
                            daily_hours += _bb_dur
                
                daily_hours = round(daily_hours, 2)
                weekly_total_hours += daily_hours
                
                # Write pairs
                for pair_idx, pair in enumerate(pairs):
                    check_in = pair['check_in']
                    check_out = pair['check_out']
                    is_miss_punch = pair['is_miss_punch']
                    
                    # Day/date only on first row
                    day_display = date_obj.strftime('%A').upper() if pair_idx == 0 else ''
                    date_display = date_obj.strftime('%m/%d/%Y') if pair_idx == 0 else ''
                    
                    # Calculate hours for this pair
                    if check_in and check_out and not is_miss_punch:
                        _pair_in_dt  = datetime.combine(date_obj, check_in.check_in_time)
                        _pair_out_dt = datetime.combine(date_obj, check_out.check_in_time)
                        # Overnight shift correction: advance OUT by one day when it
                        # falls before IN (employee crossed midnight).
                        if _pair_out_dt < _pair_in_dt:
                            _pair_out_dt += timedelta(days=1)
                        pair_hours = round((_pair_out_dt - _pair_in_dt).total_seconds() / 3600.0, 2)
                    else:
                        pair_hours = 'Missed Punch'
                    
                    # Daily total only on last row of day
                    daily_total_display = daily_hours if pair_idx == len(pairs) - 1 else ''
                    
                    # Get record for address/distance info
                    ref_record = check_in or check_out
                    
                    # Build row data
                    row_data = [
                        day_display,
                        date_display,
                        check_in.check_in_time.strftime('%I:%M:%S %p') if check_in else '',
                        check_out.check_in_time.strftime('%I:%M:%S %p') if check_out else '',
                        ref_record.location_name if ref_record else '',
                        zone_info,
                        pair_hours,
                        daily_total_display if daily_total_display else '',
                        '',  # Regular Hours
                        '',  # OT Hours
                        '',  # Building Address (will be HYPERLINK)
                        '',  # Recorded Location (will be HYPERLINK)
                        getattr(ref_record, 'distance', None) or '' if ref_record else '',
                        calculate_possible_violation(getattr(ref_record, 'distance', None)) if ref_record else ''
                    ]
                    
                    # Use bottom-only border on the last pair row of the day;
                    # no borders on intermediate rows (matches normal TA export).
                    _bb_is_last_pair = (pair_idx == len(pairs) - 1)
                    _bb_row_border   = border_day_last if _bb_is_last_pair else border_day_middle
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = data_font
                        cell.border = _bb_row_border
                        if col == 7 and value == 'Missed Punch':
                            cell.fill = missed_punch_fill
                    
                    # Add HYPERLINK formulas for addresses
                    if ref_record:
                        building_address = ref_record.event_description or ''
                        if building_address:
                            encoded_addr = building_address.replace(' ', '+').replace(',', '%2C')
                            hyperlink_formula = f'=HYPERLINK("https://www.google.com/maps/place/{encoded_addr}","{building_address}")'
                            ws.cell(row=current_row, column=11, value=hyperlink_formula)
                        
                        recorded_addr = ref_record.recorded_address or ''
                        if recorded_addr:
                            encoded_recorded = recorded_addr.replace(' ', '+').replace(',', '%2C')
                            recorded_hyperlink = f'=HYPERLINK("https://www.google.com/maps/place/{encoded_recorded}","{recorded_addr}")'
                            ws.cell(row=current_row, column=12, value=recorded_hyperlink)
                    
                    current_row += 1
            
            # Write final weekly total
            if weekly_total_hours > 0:
                week_regular = min(weekly_total_hours, 40.0)
                week_overtime = max(0, weekly_total_hours - 40.0)
                
                ws.cell(row=current_row, column=7, value='Weekly Total: ').font = bold_font
                ws.cell(row=current_row, column=8, value=_qtr(weekly_total_hours)).font = bold_font
                ws.cell(row=current_row, column=9, value=_qtr(week_regular)).font = bold_font
                ws.cell(row=current_row, column=10, value=_qtr(week_overtime)).font = bold_font
                
                grand_regular_hours += week_regular
                grand_ot_hours += week_overtime
                current_row += 1
            
            # ================================================================
            # Write extra working hours rows (SP/PW/PT) if employee has any
            # This matches the behavior of the regular Export to Excel
            # ================================================================
            
            # Write SP row if hours > 0
            if sp_hours > 0:
                ws.cell(row=current_row, column=7, value='Special Project (SP): ').font = italic_bold_font
                ws.cell(row=current_row, column=9, value=round(sp_hours, 2)).font = italic_bold_font
                # Log SP hours export
                logger_handler.logger.info(f"Export by Building: Employee {employee_id} SP hours: {sp_hours:.2f}")
                current_row += 1
            
            # Write PW row if hours > 0
            if pw_hours > 0:
                ws.cell(row=current_row, column=7, value='Periodic Work (PW): ').font = italic_bold_font
                ws.cell(row=current_row, column=9, value=round(pw_hours, 2)).font = italic_bold_font
                # Log PW hours export
                logger_handler.logger.info(f"Export by Building: Employee {employee_id} PW hours: {pw_hours:.2f}")
                current_row += 1
            
            # Write PT row if hours > 0
            if pt_hours > 0:
                ws.cell(row=current_row, column=7, value='Project Team (PT): ').font = italic_bold_font
                ws.cell(row=current_row, column=9, value=round(pt_hours, 2)).font = italic_bold_font
                # Log PT hours export
                logger_handler.logger.info(f"Export by Building: Employee {employee_id} PT hours: {pt_hours:.2f}")
                current_row += 1
            
            # ================================================================
            # End of extra working hours section
            # ================================================================
            
            # Write GRAND TOTAL row
            ws.cell(row=current_row, column=7, value='GRAND TOTAL: ').font = Font(name='Aptos Narrow', size=11, bold=True)
            ws.cell(row=current_row, column=9, value=_qtr(grand_regular_hours)).font = Font(name='Aptos Narrow', size=11, bold=True)
            ws.cell(row=current_row, column=10, value=_qtr(grand_ot_hours)).font = Font(name='Aptos Narrow', size=11, bold=True)
            current_row += 1
            
            # Empty row after each employee
            current_row += 1
        
        # Empty row after each building
        current_row += 1
    
    # Auto-size columns
    for col_idx in range(1, 15):
        column_letter = get_column_letter(col_idx)
        
        if col_idx == 1:
            ws.column_dimensions[column_letter].width = 18
            continue
        
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Filename
    if date_range_str:
        filename = f'{project_name_for_filename}time_attendance_by_building_{date_range_str}.xlsx'
    else:
        filename = f'{project_name_for_filename}time_attendance_by_building.xlsx'
    
    # Log successful export
    logger_handler.logger.info(
        f"Export by Building completed: {filename} with {len(sorted_locations)} buildings"
    )
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )