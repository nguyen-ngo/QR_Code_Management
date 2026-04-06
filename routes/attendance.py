"""
routes/attendance.py
====================
Attendance check-in records, manual entry, verification review,
export configuration, and Excel export routes.

Routes: /attendance, /attendance/<id>/edit, /attendance/add,
        /attendance/save_manual, /api/attendance/*, /api/search_employees,
        /api/get_project_locations, /verification-review/*,
        /export-configuration, /generate-excel-export
"""
from flask import Blueprint, render_template, request, redirect, flash, session, jsonify, send_file, url_for
from datetime import datetime, date, timedelta, time
import io, os, json, re, traceback

from extensions import db, logger_handler
from models.attendance import AttendanceData
from models.employee import Employee
from models.permissions import UserLocationPermission, UserProjectPermission
from models.project import Project
from models.qrcode import QRCode
from models.user import User
from sqlalchemy import text
from logger_handler import log_user_activity, log_database_operations
from utils.helpers import (
                           admin_required,
                           get_client_ip,
                           has_admin_privileges,
                           has_staff_level_access,
                           login_required,
                           staff_or_admin_required)
from utils.geocoding import (calculate_location_accuracy_enhanced, process_location_data_enhanced,
                             check_location_accuracy_column_exists)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

bp = Blueprint('attendance', __name__)



@bp.route('/attendance', endpoint='attendance_report')
@login_required
def attendance_report():
    """Safe attendance report with backward compatibility for location_accuracy and fixed datetime handling"""
    try:
        logger_handler.logger.debug("Loading attendance report")

        # Log attendance report access
        try:
            user_role = session.get('role', 'unknown')
            logger_handler.logger.info(f"User {session.get('username', 'unknown')} accessed attendance report")
        except Exception:
            pass

        # Check if location_accuracy column exists
        has_location_accuracy = check_location_accuracy_column_exists()
        logger_handler.logger.debug(f"Location accuracy column exists: {has_location_accuracy}")

        # Get filter parameters
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        location_filter = request.args.get('location', '')
        # employee param is now a comma-separated list of IDs (multi-employee filter)
        employee_filter = request.args.get('employee', '')
        project_filter = request.args.get('project', '')

        # Build the list of selected employee IDs (strip blanks)
        employee_ids = [e.strip() for e in employee_filter.split(',') if e.strip()] if employee_filter else []

        # Build display names for each selected employee
        employee_display_names = []
        for eid in employee_ids:
            try:
                emp = Employee.query.filter_by(id=int(eid)).first()
                if emp:
                    employee_display_names.append({
                        'id': eid,
                        'name': f"{emp.lastName}, {emp.firstName}"
                    })
                else:
                    employee_display_names.append({'id': eid, 'name': f"ID: {eid}"})
            except (ValueError, TypeError):
                employee_display_names.append({'id': eid, 'name': eid})

        # Legacy single-value display name (kept for backward compat in template)
        employee_display_name = ', '.join([e['name'] for e in employee_display_names])

        # ============================================================
        # PROJECT MANAGER ACCESS CONTROL
        # ============================================================
        user_role = session.get('role')
        user_id = session.get('user_id')
        
        # Initialize permission filters
        allowed_project_ids = []
        allowed_location_names = []
        
        # Check if user is Project Manager and get their permissions
        if user_role == 'project_manager':
            logger_handler.logger.debug(f"Project Manager access control enabled for user {session.get('username')}")
            
            try:
                # Get assigned projects
                assigned_projects = UserProjectPermission.query.filter_by(user_id=user_id).all()
                allowed_project_ids = [p.project_id for p in assigned_projects]
                
                # Get assigned locations
                assigned_locations = UserLocationPermission.query.filter_by(user_id=user_id).all()
                allowed_location_names = [l.location_name for l in assigned_locations]
                
                # Log the permissions
                logger_handler.logger.info(
                    f"🔒 Project Manager {session.get('username')} restricted to: "
                    f"Projects: {allowed_project_ids}, Locations: {allowed_location_names}"
                )
                
                logger_handler.logger.debug(f"PM allowed projects: {allowed_project_ids}, locations: {allowed_location_names}")
            except Exception as perm_error:
                logger_handler.logger.warning(f"Error loading PM permissions: {perm_error}")
                logger_handler.logger.error(f"Error loading Project Manager permissions: {perm_error}")
            
            # If no permissions assigned, user cannot view anything
            if not allowed_project_ids and not allowed_location_names:
                logger_handler.logger.warning(
                    f"Project Manager {session.get('username')} has no assigned projects or locations"
                )
                flash('You do not have access to any projects or locations. Please contact an administrator.', 'warning')
                
                # Create empty stats object using named tuple style
                from collections import namedtuple
                Stats = namedtuple('Stats', ['total_checkins', 'unique_employees', 'active_locations', 
                                             'today_checkins', 'records_with_gps', 'records_with_accuracy', 
                                             'avg_location_accuracy'])
                empty_stats = Stats(0, 0, 0, 0, 0, 0, 0)
                
                # Return empty template
                return render_template('attendance_report.html',
                                    attendance_records=[],
                                    locations=[],
                                    projects=[],
                                    stats=empty_stats,
                                    date_from=date_from,
                                    date_to=date_to,
                                    location_filter=location_filter,
                                    employee_filter=employee_filter,
                                    employee_ids=employee_ids,
                                    employee_display_names=employee_display_names,
                                    employee_display_name=employee_display_name,
                                    project_filter=project_filter,
                                    today_date=datetime.now().strftime('%Y-%m-%d'),
                                    current_date_formatted=datetime.now().strftime('%B %d'),
                                    has_location_accuracy_feature=has_location_accuracy,
                                    user_role=user_role)
            
        # ============================================================
        # END: PROJECT MANAGER ACCESS CONTROL
        # ============================================================

        # Build base query - conditional based on column existence
        if has_location_accuracy:
            # New query with location accuracy
            base_query = """
                SELECT 
                    ad.id,
                    ad.employee_id,
                    ad.check_in_date,
                    ad.check_in_time,
                    ad.location_name,
                    qc.location_event,
                    COALESCE(ad.qr_address, qc.location_address) as qr_address,
                    ad.address as checked_in_address,
                    ad.latitude,
                    ad.longitude,
                    ad.location_accuracy,
                    ad.accuracy as gps_accuracy,
                    ad.device_info,
                    ad.created_timestamp,
                    ad.updated_timestamp,
                    CONCAT(e.firstName, ' ', e.lastName) as employee_name,
                    ad.verification_required,
                    ad.verification_status,
                    ad.verification_photo,
                    COALESCE(ad.is_dynamic_qr, 0) as is_dynamic_qr
                FROM attendance_data ad
                LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id
                LEFT JOIN employee e ON CAST(ad.employee_id AS UNSIGNED) = e.id
                WHERE 1=1
            """
        else:
            # Fallback query without location accuracy
            base_query = """
                SELECT 
                    ad.id,
                    ad.employee_id,
                    ad.check_in_date,
                    ad.check_in_time,
                    ad.location_name,
                    qc.location_event,
                    COALESCE(ad.qr_address, qc.location_address) as qr_address,
                    ad.address as checked_in_address,
                    ad.latitude,
                    ad.longitude,
                    NULL as location_accuracy,
                    ad.accuracy as gps_accuracy,
                    ad.device_info,
                    ad.created_timestamp,
                    ad.updated_timestamp,
                    CONCAT(e.firstName, ' ', e.lastName) as employee_name,
                    ad.verification_required,
                    ad.verification_status,
                    ad.verification_photo,
                    COALESCE(ad.is_dynamic_qr, 0) as is_dynamic_qr
                FROM attendance_data ad
                LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id
                LEFT JOIN employee e ON CAST(ad.employee_id AS UNSIGNED) = e.id
                WHERE 1=1
            """

        # Prepare filter conditions and parameters
        filter_conditions = []
        query_params = {}

        # ============================================================
        # APPLY PROJECT MANAGER FILTERS TO SQL QUERY
        # ============================================================
        if user_role == 'project_manager':
            # Filter by allowed projects
            if allowed_project_ids:
                project_placeholders = ','.join([f':project_{i}' for i in range(len(allowed_project_ids))])
                filter_conditions.append(f"qc.project_id IN ({project_placeholders})")
                for i, pid in enumerate(allowed_project_ids):
                    query_params[f'project_{i}'] = pid
            
            # Filter by allowed locations
            if allowed_location_names:
                location_placeholders = ','.join([f':location_{i}' for i in range(len(allowed_location_names))])
                filter_conditions.append(f"ad.location_name IN ({location_placeholders})")
                for i, loc in enumerate(allowed_location_names):
                    query_params[f'location_{i}'] = loc
        # ============================================================
        # END: APPLY PROJECT MANAGER FILTERS
        # ============================================================

        # Apply user-selected filters
        if date_from:
            filter_conditions.append("ad.check_in_date >= :date_from")
            query_params['date_from'] = date_from

        if date_to:
            filter_conditions.append("ad.check_in_date <= :date_to")
            query_params['date_to'] = date_to

        if location_filter:
            # Exact match — dropdown value IS the exact location_name string
            filter_conditions.append("ad.location_name = :location")
            query_params['location'] = location_filter

        if employee_ids:
            if len(employee_ids) == 1:
                filter_conditions.append("ad.employee_id = :employee_0")
                query_params['employee_0'] = employee_ids[0]
            else:
                placeholders = ', '.join([f':employee_{i}' for i in range(len(employee_ids))])
                filter_conditions.append(f"ad.employee_id IN ({placeholders})")
                for i, eid in enumerate(employee_ids):
                    query_params[f'employee_{i}'] = eid
            logger_handler.logger.info(
                f"Attendance report filtered by employee IDs: {employee_ids} "
                f"by user {session.get('username', 'unknown')}"
            )

        if project_filter:
            # For standard QR records: match by the QR code's project_id directly.
            # For dynamic QR records: the dynamic QR itself may not be in any project,
            # but the employee-selected location corresponds to a standard QR in that
            # project. Match those by checking if attendance_data.location_name
            # appears in the locations of QR codes belonging to the selected project.
            filter_conditions.append(
                "(qc.project_id = :project OR "
                "(ad.is_dynamic_qr = 1 AND ad.location_name IN ("
                "  SELECT DISTINCT qc2.location FROM qr_codes qc2 "
                "  WHERE qc2.project_id = :project AND qc2.qr_type = 'standard' "
                "  AND qc2.location IS NOT NULL AND qc2.location != ''"
                ")))"
            )
            query_params['project'] = project_filter

        # Combine query with filters
        if filter_conditions:
            base_query += " AND " + " AND ".join(filter_conditions)

        base_query += " ORDER BY ad.check_in_date DESC, ad.check_in_time DESC LIMIT 1000"

        logger_handler.logger.debug(f"Executing attendance query with filters: {list(query_params.keys())}")

        # Execute query
        result = db.session.execute(text(base_query), query_params)
        records = result.fetchall()
        logger_handler.logger.debug(f"Loaded {len(records)} attendance records")

        # Process records
        processed_records = []
        for record in records:
            try:
                record_dict = {
                    'id': record[0],
                    'employee_id': record[1],
                    'check_in_date': record[2],
                    'check_in_time': record[3],
                    'location_name': record[4],
                    'location_event': record[5],
                    'qr_address': record[6],
                    'checked_in_address': record[7],
                    'latitude': record[8],
                    'longitude': record[9],
                    'location_accuracy': record[10] if has_location_accuracy else None,
                    'gps_accuracy': record[11],
                    'device_info': record[12],
                    'created_timestamp': record[13],
                    'updated_timestamp': record[14],
                    'employee_name': record[15] or 'Unknown Employee',
                    'verification_required': record[16] if len(record) > 16 else False,
                    'verification_status': record[17] if len(record) > 17 else None,
                    'verification_photo': record[18] if len(record) > 18 else None,
                    'is_dynamic_qr': bool(record[19]) if len(record) > 19 else False
                }
                
                # Calculate accuracy_level for template display
                if record_dict['location_accuracy'] is not None:
                    accuracy_value = float(record_dict['location_accuracy'])
                    if accuracy_value <= 0.3:
                        record_dict['accuracy_level'] = 'accurate'
                    else:
                        record_dict['accuracy_level'] = 'inaccurate'
                else:
                    record_dict['accuracy_level'] = 'unknown'
                processed_records.append(record_dict)
            except Exception as rec_error:
                logger_handler.logger.warning(f"Error processing attendance record: {rec_error}")
                continue

        # Get unique locations for filter dropdown
        try:
            # ============================================================
            # FILTER LOCATIONS FOR PROJECT MANAGER
            # ============================================================
            if user_role == 'project_manager' and allowed_location_names:
                # Only show locations the PM has access to
                locations = sorted(allowed_location_names)
                logger_handler.logger.debug(f"Filtered to {len(locations)} locations for Project Manager")
            else:
                # Show all locations for Admin/Staff/Payroll
                locations_query = db.session.execute(text("""
                    SELECT DISTINCT location_name 
                    FROM attendance_data 
                    WHERE location_name IS NOT NULL 
                      AND location_name != 'Dynamic'
                      AND location_name != ''
                    ORDER BY location_name
                """))
                locations = [row[0] for row in locations_query.fetchall()]
                logger_handler.logger.debug(f"Found {len(locations)} unique locations")
            # ============================================================
            # END: FILTER LOCATIONS FOR PROJECT MANAGER
            # ============================================================
        except Exception as e:
            logger_handler.logger.warning(f"Error loading locations filter: {e}")
            locations = []

        # Get projects for filter dropdown
        try:
            # ============================================================
            # FILTER PROJECTS FOR PROJECT MANAGER
            # ============================================================
            if user_role == 'project_manager' and allowed_project_ids:
                # Only show projects the PM has access to
                project_placeholders = ','.join([str(pid) for pid in allowed_project_ids])
                projects_query = db.session.execute(text(f"""
                    SELECT p.id, p.name, COUNT(DISTINCT ad.id) as attendance_count
                    FROM projects p
                    LEFT JOIN qr_codes qc ON qc.project_id = p.id
                    LEFT JOIN attendance_data ad ON ad.qr_code_id = qc.id
                    WHERE p.active_status = true AND p.id IN ({project_placeholders})
                    GROUP BY p.id, p.name
                    ORDER BY p.name
                """))
                projects = projects_query.fetchall()
                logger_handler.logger.debug(f"Filtered to {len(projects)} projects for Project Manager")
            else:
                # Show all projects for Admin/Staff/Payroll
                projects = db.session.execute(text("""
                    SELECT p.id, p.name, COUNT(DISTINCT ad.id) as attendance_count
                    FROM projects p
                    LEFT JOIN qr_codes qc ON qc.project_id = p.id
                    LEFT JOIN attendance_data ad ON ad.qr_code_id = qc.id
                    WHERE p.active_status = true
                    GROUP BY p.id, p.name
                    HAVING COUNT(DISTINCT ad.id) > 0
                    ORDER BY p.name
                """)).fetchall()
                logger_handler.logger.debug(f"Loaded {len(projects)} projects with attendance data")
            # ============================================================
            # END: FILTER PROJECTS FOR PROJECT MANAGER
            # ============================================================
        except Exception as e:
            logger_handler.logger.warning(f"Error loading projects filter: {e}")
            projects = []

        # ============================================================
        # STATISTICS - COMPLETELY REWRITTEN FOR SAFETY
        # ============================================================
        logger_handler.logger.debug("Loading attendance statistics")
        
        # Create simple dict for stats (most compatible approach)
        stats_dict = {
            'total_checkins': 0,
            'unique_employees': 0,
            'active_locations': 0,
            'today_checkins': 0,
            'records_with_gps': 0,
            'records_with_accuracy': 0,
            'avg_location_accuracy': 0.0
        }
        
        try:
            # Build stats query
            if has_location_accuracy:
                stats_select = """
                    SELECT 
                        COALESCE(COUNT(*), 0) as total_checkins,
                        COALESCE(COUNT(DISTINCT employee_id), 0) as unique_employees,
                        COALESCE(COUNT(DISTINCT qr_code_id), 0) as active_locations,
                        COALESCE(COUNT(CASE WHEN check_in_date = CURRENT_DATE THEN 1 END), 0) as today_checkins,
                        COALESCE(COUNT(CASE WHEN latitude IS NOT NULL AND longitude IS NOT NULL THEN 1 END), 0) as records_with_gps,
                        COALESCE(COUNT(CASE WHEN location_accuracy IS NOT NULL THEN 1 END), 0) as records_with_accuracy,
                        COALESCE(AVG(location_accuracy), 0) as avg_location_accuracy
                """
            else:
                stats_select = """
                    SELECT 
                        COALESCE(COUNT(*), 0) as total_checkins,
                        COALESCE(COUNT(DISTINCT employee_id), 0) as unique_employees,
                        COALESCE(COUNT(DISTINCT qr_code_id), 0) as active_locations,
                        COALESCE(COUNT(CASE WHEN check_in_date = CURRENT_DATE THEN 1 END), 0) as today_checkins,
                        COALESCE(COUNT(CASE WHEN latitude IS NOT NULL AND longitude IS NOT NULL THEN 1 END), 0) as records_with_gps,
                        0 as records_with_accuracy,
                        0 as avg_location_accuracy
                """
            
            stats_query_text = stats_select + " FROM attendance_data ad"
            stats_params = {}
            
            # Add filters for Project Manager
            if user_role == 'project_manager':
                stats_query_text += " LEFT JOIN qr_codes qc ON ad.qr_code_id = qc.id WHERE 1=1"
                
                stats_conditions = []
                
                if allowed_project_ids:
                    project_placeholders = ','.join([f':stat_project_{i}' for i in range(len(allowed_project_ids))])
                    stats_conditions.append(f"qc.project_id IN ({project_placeholders})")
                    for i, pid in enumerate(allowed_project_ids):
                        stats_params[f'stat_project_{i}'] = pid
                
                if allowed_location_names:
                    location_placeholders = ','.join([f':stat_location_{i}' for i in range(len(allowed_location_names))])
                    stats_conditions.append(f"ad.location_name IN ({location_placeholders})")
                    for i, loc in enumerate(allowed_location_names):
                        stats_params[f'stat_location_{i}'] = loc
                
                if stats_conditions:
                    stats_query_text += " AND " + " AND ".join(stats_conditions)
            
            logger_handler.logger.debug(f"Executing stats query with params: {list(stats_params.keys())}")
            
            # Execute stats query
            stats_result = db.session.execute(text(stats_query_text), stats_params)
            stats_row = stats_result.fetchone()
            
            logger_handler.logger.debug(f"Stats row type: {type(stats_row).__name__}")
            
            # Safely extract stats from row
            if stats_row is not None and len(stats_row) >= 7:
                try:
                    stats_dict['total_checkins'] = int(stats_row[0]) if stats_row[0] is not None else 0
                    stats_dict['unique_employees'] = int(stats_row[1]) if stats_row[1] is not None else 0
                    stats_dict['active_locations'] = int(stats_row[2]) if stats_row[2] is not None else 0
                    stats_dict['today_checkins'] = int(stats_row[3]) if stats_row[3] is not None else 0
                    stats_dict['records_with_gps'] = int(stats_row[4]) if stats_row[4] is not None else 0
                    stats_dict['records_with_accuracy'] = int(stats_row[5]) if stats_row[5] is not None else 0
                    stats_dict['avg_location_accuracy'] = float(stats_row[6]) if stats_row[6] is not None else 0.0
                    logger_handler.logger.debug(f"Loaded statistics: {stats_dict['total_checkins']} total check-ins")
                except (IndexError, TypeError, ValueError) as extract_error:
                    logger_handler.logger.warning(f"Error extracting stats values: {extract_error}")
                    # stats_dict already has default values
            else:
                logger_handler.logger.warning("Stats query returned None or insufficient columns, using default stats")
                
        except Exception as stats_error:
            logger_handler.logger.error(f"Error loading statistics: {stats_error}", exc_info=True)
            # stats_dict already has default values
        
        # Convert dict to object-like for template compatibility
        class StatsObject:
            def __init__(self, stats_dict):
                for key, value in stats_dict.items():
                    setattr(self, key, value)
        
        stats = StatsObject(stats_dict)
        logger_handler.logger.debug(f"Stats object created: total_checkins={stats.total_checkins}")
        
        # ============================================================
        # END: STATISTICS
        # ============================================================
        
        # Add today's date for template
        today_date = datetime.now().strftime('%Y-%m-%d')
        current_date_formatted = datetime.now().strftime('%B %d')

        logger_handler.logger.debug("Rendering attendance report template")

        return render_template('attendance_report.html',
                     attendance_records=processed_records,
                     locations=locations,
                     projects=projects,
                     stats=stats,
                     date_from=date_from,
                     date_to=date_to,
                     location_filter=location_filter,
                     employee_filter=employee_filter,
                     employee_ids=employee_ids,
                     employee_display_names=employee_display_names,
                     employee_display_name=employee_display_name,
                     project_filter=project_filter,
                     today_date=datetime.now().strftime('%Y-%m-%d'),
                     current_date_formatted=datetime.now().strftime('%B %d'),
                     has_location_accuracy_feature=has_location_accuracy,
                     user_role=user_role)

    except Exception as e:
        logger_handler.logger.error(f"Error loading attendance report: {e}", exc_info=True)
        
        import traceback
        error_traceback = traceback.format_exc()


        # Log the error
        try:
            logger_handler.log_database_error('attendance_report', e)
        except Exception as log_error:
            logger_handler.logger.warning(f"Additional logging error: {log_error}")

        flash('Error loading attendance report. Please check the server logs for details.', 'error')
        return redirect(url_for('dashboard.dashboard'))

@bp.route('/attendance/<int:record_id>/edit', methods=['GET', 'POST'], endpoint='edit_attendance')
@login_required
@log_database_operations('attendance_update')
def edit_attendance(record_id):
    """Edit attendance record (Admin and Payroll only)"""
    # Check if user has permission to edit attendance records
    if session.get('role') not in ['admin', 'payroll', 'accounting']:
        flash('Access denied. Only administrators and accounting staff can edit attendance records.', 'error')
        return redirect(url_for('attendance.attendance_report'))

    try:
        attendance_record = AttendanceData.query.get_or_404(record_id)

        if request.method == 'POST':
            # Get the audit note from form - REQUIRED
            edit_note = request.form.get('edit_note', '').strip()
            if not edit_note:
                flash('Edit reason is required for audit purposes.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('edit_attendance.html',
                                     attendance_record=attendance_record,
                                     projects=projects,
                                     qr_codes=QRCode.query.filter_by(active_status=True).all())

            # Track changes for logging
            changes = {}
            old_values = {
                'employee_id': attendance_record.employee_id,
                'check_in_date': attendance_record.check_in_date,
                'check_in_time': attendance_record.check_in_time,
                'location_name': attendance_record.location_name,
                'qr_code_id': attendance_record.qr_code_id,
                'location_event': attendance_record.qr_code.location_event if attendance_record.qr_code else None
            }

            # Update attendance record fields
            new_employee_id = request.form['employee_id'].strip().upper()
            new_check_in_date = datetime.strptime(request.form['check_in_date'], '%Y-%m-%d').date()
            new_check_in_time = datetime.strptime(request.form['check_in_time'], '%H:%M').time()
            new_location_name = request.form['location_name'].strip()
            
            # Get the new QR code ID from the form (this determines the location event)
            new_qr_code_id = request.form.get('qr_code_id', '').strip()
            if not new_qr_code_id:
                flash('Location event selection is required.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('edit_attendance.html',
                                     attendance_record=attendance_record,
                                     projects=projects,
                                     qr_codes=QRCode.query.filter_by(active_status=True).all())
            
            # Validate the QR code exists
            new_qr_code = db.session.get(QRCode, int(new_qr_code_id))
            if not new_qr_code:
                flash('Selected location event not found.', 'error')
                projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
                return render_template('edit_attendance.html',
                                     attendance_record=attendance_record,
                                     projects=projects,
                                     qr_codes=QRCode.query.filter_by(active_status=True).all())

            # Track what changed
            if attendance_record.employee_id != new_employee_id:
                changes['employee_id'] = f"{attendance_record.employee_id} → {new_employee_id}"
            if attendance_record.check_in_date != new_check_in_date:
                changes['check_in_date'] = f"{attendance_record.check_in_date} → {new_check_in_date}"
            if attendance_record.check_in_time != new_check_in_time:
                changes['check_in_time'] = f"{attendance_record.check_in_time} → {new_check_in_time}"
            if attendance_record.location_name != new_location_name:
                changes['location_name'] = f"{attendance_record.location_name} → {new_location_name}"
            if attendance_record.qr_code_id != int(new_qr_code_id):
                old_event = attendance_record.qr_code.location_event if attendance_record.qr_code else 'Unknown'
                new_event = new_qr_code.location_event
                changes['location_event'] = f"{old_event} → {new_event}"
                changes['qr_code_id'] = f"{attendance_record.qr_code_id} → {new_qr_code_id}"

            # Apply changes
            attendance_record.employee_id = new_employee_id
            attendance_record.check_in_date = new_check_in_date
            attendance_record.check_in_time = new_check_in_time
            attendance_record.location_name = new_location_name
            attendance_record.qr_code_id = int(new_qr_code_id)
            attendance_record.updated_timestamp = datetime.utcnow()
            
            # Store the audit note with timestamp and user info
            timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
            username = session.get('username', 'Unknown')
            role = session.get('role', 'unknown')
            
            new_note_entry = f"[{timestamp}] {role.title()} '{username}': {edit_note}"
            
            if attendance_record.edit_note:
                # Append to existing notes
                attendance_record.edit_note = f"{attendance_record.edit_note}\n\n{new_note_entry}"
            else:
                # First edit note
                attendance_record.edit_note = new_note_entry

            db.session.commit()

            # Enhanced logging with audit note
            if changes:
                logger_handler.log_security_event(
                    event_type="attendance_record_update",
                    description=f"{session.get('role', 'unknown').title()} {session.get('username')} updated attendance record {record_id}",
                    severity="MEDIUM",
                    additional_data={
                        'record_id': record_id, 
                        'changes': changes, 
                        'user_role': session.get('role'),
                        'edit_reason': edit_note,
                        'editor_username': session.get('username')
                    }
                )
                logger_handler.logger.info(
                    f"User {session.get('username')} ({session.get('role', 'unknown')}) "
                    f"updated attendance record {record_id}: {changes}, reason: {edit_note}"
                )
            else:
                # Log even if no changes were made (for audit purposes)
                logger_handler.log_security_event(
                    event_type="attendance_record_edit_no_changes",
                    description=f"{session.get('role', 'unknown').title()} {session.get('username')} accessed edit form for record {record_id} but made no changes",
                    severity="LOW",
                    additional_data={
                        'record_id': record_id,
                        'user_role': session.get('role'),
                        'edit_reason': edit_note,
                        'editor_username': session.get('username')
                    }
                )
                logger_handler.logger.info(
                    f"User {session.get('username')} ({session.get('role', 'unknown')}) "
                    f"edited attendance record {record_id} with no changes, reason: {edit_note}"
                )

            flash(f'Attendance record for {new_employee_id} updated successfully! Edit reason logged for audit.', 'success')
            return redirect(url_for('attendance.attendance_report'))

        # GET request - show edit form
        # Get available projects for the dropdown
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        
        # Get available QR codes for location dropdown (for backward compatibility)
        qr_codes = QRCode.query.filter_by(active_status=True).all()

        return render_template('edit_attendance.html',
                             attendance_record=attendance_record,
                             projects=projects,
                             qr_codes=qr_codes)

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('attendance_update', e)
        logger_handler.logger.error(f"Error updating attendance record {record_id}: {e}", exc_info=True)
        flash('Error updating attendance record. Please try again.', 'error')
        return redirect(url_for('attendance.attendance_report'))

@bp.route('/attendance/add', methods=['GET'], endpoint='add_manual_attendance')
@login_required
@log_user_activity('manual_attendance_access')
def add_manual_attendance():
    """
    Display form to manually add attendance record
    Only accessible by admin and accounting roles
    """
    try:
        user_role = session.get('role')
        
        # Check authorization
        if user_role not in ['admin', 'accounting']:
            flash('You do not have permission to manually add attendance records.', 'error')
            return redirect(url_for('attendance.attendance_report'))
        
        # Get all active projects
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        
        # Get today's date for form
        today_date = datetime.now().strftime('%Y-%m-%d')
        
        logger_handler.logger.info(
            f"User {session.get('username')} ({user_role}) accessed manual attendance entry form"
        )
        
        return render_template('add_manual_attendance.html',
                             projects=projects,
                             today_date=today_date)
    
    except Exception as e:
        logger_handler.logger.error(f"Error loading manual attendance form: {e}")
        flash('Error loading form. Please try again.', 'error')
        return redirect(url_for('attendance.attendance_report'))


@bp.route('/attendance/save_manual', methods=['POST'], endpoint='save_manual_attendance')
@login_required
@log_user_activity('manual_attendance_creation')
@log_database_operations('manual_attendance_insert')
def save_manual_attendance():
    """
    Save manually created attendance record
    Only accessible by admin and accounting roles
    """
    try:
        user_role = session.get('role')
        
        # Check authorization
        if user_role not in ['admin', 'accounting']:
            return jsonify({
                'success': False,
                'message': 'You do not have permission to manually add attendance records.'
            }), 403
        
        # Get form data
        employee_id = request.form.get('employee_id', '').strip()
        location_id = request.form.get('location_id', '').strip()
        check_date = request.form.get('check_date', '').strip()
        check_time = request.form.get('check_time', '').strip()
        
        # Validate required fields
        if not all([employee_id, location_id, check_date, check_time]):
            flash('All fields are required.', 'error')
            return redirect(url_for('attendance.add_manual_attendance'))
        
        # Validate employee exists
        employee = Employee.query.filter_by(id=int(employee_id)).first()
        if not employee:
            flash(f'Employee with ID {employee_id} not found.', 'error')
            return redirect(url_for('attendance.add_manual_attendance'))
        
        # Get QR code (location)
        qr_code = db.session.get(QRCode, int(location_id))
        if not qr_code:
            flash('Selected location not found.', 'error')
            return redirect(url_for('attendance.add_manual_attendance'))
        
        # Parse date and time
        try:
            check_date_obj = datetime.strptime(check_date, '%Y-%m-%d').date()
            check_time_obj = datetime.strptime(check_time, '%H:%M').time()
        except ValueError as e:
            flash('Invalid date or time format.', 'error')
            logger_handler.logger.error(f"Date/time parsing error: {e}")
            return redirect(url_for('attendance.add_manual_attendance'))
        
        # Check if record already exists for this employee, location, date, and time
        existing_record = AttendanceData.query.filter_by(
            employee_id=str(employee_id),
            qr_code_id=qr_code.id,
            check_in_date=check_date_obj,
            check_in_time=check_time_obj
        ).first()
        
        if existing_record:
            flash('An attendance record already exists for this employee at this location, date, and time.', 'warning')
            return redirect(url_for('attendance.add_manual_attendance'))
        
        # Create new attendance record
        # Use QR code's location address for both QR address and check-in address
        # Set fixed distance of 0.010 miles
        new_attendance = AttendanceData(
            qr_code_id=qr_code.id,
            employee_id=str(employee_id),
            check_in_date=check_date_obj,
            check_in_time=check_time_obj,
            location_name=qr_code.location,
            # Use QR code's coordinates
            latitude=qr_code.address_latitude,
            longitude=qr_code.address_longitude,
            # Use QR code's address for both
            address=qr_code.location_address,
            # Set fixed distance
            location_accuracy=0.010,
            accuracy=0.010,
            # Mark as manual entry
            location_source='manual_entry',
            device_info='Manual Entry by Admin/Accounting',
            user_agent=f'Manual Entry - User: {session.get("username")}',
            ip_address=get_client_ip(),
            status='present',
            verification_required=False,
            verification_status='approved',
            created_timestamp=datetime.utcnow(),
            updated_timestamp=datetime.utcnow()
        )
        
        db.session.add(new_attendance)
        db.session.commit()
        
        # Log the manual entry
        logger_handler.logger.info(
            f"Manual attendance record created by {session.get('username')} ({user_role}): "
            f"Employee {employee.firstName} {employee.lastName} (ID: {employee_id}), "
            f"Location: {qr_code.location}, Event: {qr_code.location_event}, "
            f"Date: {check_date}, Time: {check_time}"
        )
        
        flash(f'Attendance record successfully created for {employee.firstName} {employee.lastName}.', 'success')
        return redirect(url_for('attendance.attendance_report'))
    
    except Exception as e:
        db.session.rollback()
        logger_handler.logger.error(f"Error saving manual attendance record: {e}")
        logger_handler.logger.error(f"Traceback: {traceback.format_exc()}")
        flash('Error saving attendance record. Please try again.', 'error')
        return redirect(url_for('attendance.add_manual_attendance'))


@bp.route('/api/time-attendance/locations', endpoint='time_attendance_locations_api')
@login_required
def time_attendance_locations_api():
    """Return distinct location_name values from time_attendance, optionally filtered by project_id.
    Used by the time attendance records page to dynamically scope the location dropdown."""
    try:
        project_id = request.args.get('project_id', '').strip()

        if project_id:
            try:
                project_id_int = int(project_id)
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': 'Invalid project_id'}), 400

            result = db.session.execute(text("""
                SELECT DISTINCT location_name
                FROM time_attendance
                WHERE project_id = :project_id
                  AND location_name IS NOT NULL
                ORDER BY location_name
            """), {'project_id': project_id_int})
        else:
            result = db.session.execute(text("""
                SELECT DISTINCT location_name
                FROM time_attendance
                WHERE location_name IS NOT NULL
                ORDER BY location_name
            """))

        locations = [row[0] for row in result.fetchall()]
        logger_handler.logger.info(
            f"User {session.get('username', 'unknown')} fetched time attendance locations"
            + (f" for project_id={project_id}" if project_id else " (all projects)")
        )
        return jsonify({'success': True, 'locations': locations})

    except Exception as e:
        logger_handler.logger.error(f"Error in time_attendance_locations_api: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/attendance/locations', endpoint='attendance_locations_api')
@login_required
def attendance_locations_api():
    """Return distinct location_name values from attendance_data, optionally filtered by project_id.
    Used by the attendance report page to dynamically scope the location dropdown when a project is selected."""
    try:
        project_id = request.args.get('project_id', '').strip()

        if project_id:
            try:
                project_id_int = int(project_id)
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': 'Invalid project_id'}), 400

            result = db.session.execute(text("""
                SELECT DISTINCT ad.location_name
                FROM attendance_data ad
                INNER JOIN qr_codes qc ON ad.qr_code_id = qc.id
                WHERE qc.project_id = :project_id
                  AND ad.location_name IS NOT NULL
                ORDER BY ad.location_name
            """), {'project_id': project_id_int})
        else:
            result = db.session.execute(text("""
                SELECT DISTINCT location_name
                FROM attendance_data
                WHERE location_name IS NOT NULL
                ORDER BY location_name
            """))

        locations = [row[0] for row in result.fetchall()]
        logger_handler.logger.info(
            f"User {session.get('username', 'unknown')} fetched attendance locations"
            + (f" for project_id={project_id}" if project_id else " (all projects)")
        )
        return jsonify({'success': True, 'locations': locations})

    except Exception as e:
        logger_handler.logger.error(f"Error in attendance_locations_api: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/search_employees', endpoint='search_employees_api')
@login_required
def search_employees_api():
    """
    API endpoint to search employees by name or ID.
    Returns matches from the Employee table first, then appends any IDs found
    in attendance_data that have no Employee record — so unregistered IDs
    that have attendance records can still be filtered on the attendance page.
    """
    try:
        search_query = request.args.get('q', '').strip()

        if not search_query or len(search_query) < 2:
            return jsonify({'employees': []})

        search_pattern = f"%{search_query}%"

        # 1. Registered employees — search by ID or name
        employees = Employee.query.filter(
            db.or_(
                Employee.id.like(search_pattern),
                Employee.firstName.like(search_pattern),
                Employee.lastName.like(search_pattern),
                db.func.concat(Employee.firstName, ' ', Employee.lastName).like(search_pattern)
            )
        ).limit(10).all()

        employee_list = [{
            'id': emp.id,
            'firstName': emp.firstName,
            'lastName': emp.lastName,
            'full_name': f"{emp.firstName} {emp.lastName}"
        } for emp in employees]

        registered_ids = {str(emp.id) for emp in employees}

        # 2. Unregistered IDs — present in attendance_data but not in Employee table.
        #    Only add when the search term looks like (part of) a numeric ID and we
        #    still have room in the result list.
        if len(employee_list) < 10:
            remaining_slots = 10 - len(employee_list)
            try:
                unregistered_rows = db.session.execute(
                    text("""
                        SELECT DISTINCT ad.employee_id
                        FROM attendance_data ad
                        LEFT JOIN employee e ON CAST(ad.employee_id AS UNSIGNED) = e.id
                        WHERE e.id IS NULL
                          AND ad.employee_id LIKE :pattern
                        ORDER BY ad.employee_id
                        LIMIT :lim
                    """),
                    {'pattern': search_pattern, 'lim': remaining_slots}
                ).fetchall()

                for row in unregistered_rows:
                    emp_id = str(row[0])
                    if emp_id not in registered_ids:
                        employee_list.append({
                            'id': emp_id,
                            'firstName': f'ID: {emp_id}',
                            'lastName': '(no record)',
                            'full_name': f'ID: {emp_id} (no record)'
                        })
            except Exception as unreg_err:
                logger_handler.logger.warning(f"Could not search unregistered employee IDs: {unreg_err}")

        return jsonify({'employees': employee_list})

    except Exception as e:
        logger_handler.logger.error(f"Error searching employees: {e}")
        return jsonify({'employees': [], 'error': str(e)}), 500


@bp.route('/api/get_project_locations', endpoint='get_project_locations_api')
@login_required
def get_project_locations_api():
    """
    API endpoint to get locations for a specific project
    Returns JSON with location list
    """
    try:
        project_id = request.args.get('project_id', '').strip()
        
        if not project_id:
            return jsonify({'success': False, 'locations': [], 'error': 'Project ID required'})
        
        # Get active QR codes for this project
        qr_codes = QRCode.query.filter_by(
            project_id=int(project_id),
            active_status=True
        ).order_by(QRCode.location).all()
        
        # Group QR codes by location to get unique locations
        locations_dict = {}
        for qr in qr_codes:
            location_key = f"{qr.location}||{qr.location_address}"
            
            if location_key not in locations_dict:
                locations_dict[location_key] = {
                    'location': qr.location,
                    'location_address': qr.location_address,
                    'qr_codes': {}
                }
            
            # Store QR code ID for each event type
            locations_dict[location_key]['qr_codes'][qr.location_event] = qr.id

        # Convert to list format
        location_list = [{
            'location': loc_data['location'],
            'location_address': loc_data['location_address'],
            'qr_codes': loc_data['qr_codes']
        } for loc_data in locations_dict.values()]
        
        return jsonify({'success': True, 'locations': location_list})
    
    except Exception as e:
        logger_handler.logger.error(f"Error getting project locations: {e}")
        return jsonify({'success': False, 'locations': [], 'error': str(e)}), 500
    
@bp.route('/attendance/<int:record_id>/delete', methods=['POST'], endpoint='delete_attendance')
@login_required
@log_database_operations('attendance_delete')
def delete_attendance(record_id):
    """Delete attendance record (Admin and Payroll only)"""
    # Check if user has permission to delete attendance records
    if session.get('role') not in ['admin', 'payroll', 'accounting']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': 'Access denied. Only administrators and payroll staff can delete attendance records.'
            }), 403
        else:
            flash('Access denied. Only administrators and payroll staff can delete attendance records.', 'error')
            return redirect(url_for('attendance.attendance_report'))

    try:
        attendance_record = AttendanceData.query.get_or_404(record_id)

        # Store record info for logging before deletion
        employee_id = attendance_record.employee_id
        location_name = attendance_record.location_name
        check_in_date = attendance_record.check_in_date

        # Log the deletion
        logger_handler.log_security_event(
            event_type="attendance_record_deletion",
            description=f"{session.get('role', 'unknown').title()} {session.get('username')} deleted attendance record {record_id}",
            severity="HIGH",
            additional_data={
                'record_id': record_id,
                'employee_id': employee_id,
                'location_name': location_name,
                'check_in_date': str(check_in_date),
                'user_role': session.get('role')
            }
        )

        # Delete the record
        db.session.delete(attendance_record)
        db.session.commit()

        logger_handler.logger.info(
            f"User {session.get('username')} ({session.get('role', 'unknown')}) "
            f"deleted attendance record {record_id} for employee {employee_id}"
        )

        # Return JSON response for AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': f'Attendance record for {employee_id} deleted successfully!'
            })
        else:
            flash(f'Attendance record for {employee_id} deleted successfully!', 'success')
            return redirect(url_for('attendance.attendance_report'))

    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('attendance_delete', e)
        logger_handler.logger.error(f"Error deleting attendance record {record_id}: {e}", exc_info=True)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': 'Error deleting attendance record. Please try again.'
            }), 500
        else:
            flash('Error deleting attendance record. Please try again.', 'error')
            return redirect(url_for('attendance.attendance_report'))
        
@bp.route('/verification-review', endpoint='verification_review')
@login_required
def verification_review():
    """Admin page to review pending photo verifications"""
    try:
        # Only admins can access
        if session.get('role') not in ['admin', 'payroll', 'accounting']:
            flash('Unauthorized access.', 'error')
            return redirect(url_for('dashboard.dashboard'))
        
        # Get filter parameters
        status_filter = request.args.get('status', 'pending')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        project_filter = request.args.get('project', '')
        location_filter = request.args.get('location', '')
        employee_filter = request.args.get('employee', '')
        
        # Build query - join with QRCode to access project_id
        query = AttendanceData.query.join(QRCode).filter(
            AttendanceData.verification_required == True
        )
        
        if status_filter and status_filter != 'all':
            query = query.filter(AttendanceData.verification_status == status_filter)
        
        if date_from:
            query = query.filter(AttendanceData.check_in_date >= date_from)
        
        if date_to:
            query = query.filter(AttendanceData.check_in_date <= date_to)
        
        # Apply project filter
        if project_filter:
            try:
                query = query.filter(QRCode.project_id == int(project_filter))
            except (ValueError, TypeError):
                pass
        
        # Apply location filter
        if location_filter:
            query = query.filter(AttendanceData.location_name.ilike(f'%{location_filter}%'))
        
        # Apply employee ID filter
        if employee_filter:
            query = query.filter(AttendanceData.employee_id.ilike(f'%{employee_filter}%'))
        
        # Get records with QR code information
        verifications = query.order_by(
            AttendanceData.verification_timestamp.desc()
        ).all()
        
        # Build a dictionary for employee names lookup
        employee_names = {}
        for record in verifications:
            if record.employee_id and record.employee_id not in employee_names:
                try:
                    employee = Employee.query.filter_by(id=int(record.employee_id)).first()
                    if employee:
                        employee_names[record.employee_id] = f"{employee.lastName}, {employee.firstName}"
                    else:
                        employee_names[record.employee_id] = None
                except (ValueError, TypeError):
                    employee_names[record.employee_id] = None
        
        # Build a dictionary for project names lookup
        project_names = {}
        for record in verifications:
            if record.qr_code and record.qr_code.project_id:
                project_id = record.qr_code.project_id
                if project_id not in project_names:
                    try:
                        project = db.session.get(Project, project_id)
                        if project:
                            project_names[project_id] = project.name
                        else:
                            project_names[project_id] = None
                    except Exception:
                        project_names[project_id] = None
        
        # Get counts for status badges
        pending_count = AttendanceData.query.filter(
            AttendanceData.verification_status == 'pending'
        ).count()
        
        approved_count = AttendanceData.query.filter(
            AttendanceData.verification_status == 'approved'
        ).count()
        
        rejected_count = AttendanceData.query.filter(
            AttendanceData.verification_status == 'rejected'
        ).count()
        
        # Get all projects for filter dropdown
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        
        # Get unique locations for filter dropdown
        locations = db.session.query(AttendanceData.location_name).filter(
            AttendanceData.verification_required == True
        ).distinct().order_by(AttendanceData.location_name).all()
        location_list = [loc[0] for loc in locations if loc[0]]
        
        # Log access
        logger_handler.logger.info(
            f"User {session.get('username')} ({session.get('role')}) accessed verification review page"
        )
        
        return render_template('verification_review.html',
                             verifications=verifications,
                             pending_count=pending_count,
                             approved_count=approved_count,
                             rejected_count=rejected_count,
                             status_filter=status_filter,
                             date_from=date_from,
                             date_to=date_to,
                             project_filter=project_filter,
                             location_filter=location_filter,
                             employee_filter=employee_filter,
                             projects=projects,
                             locations=location_list,
                             employee_names=employee_names,
                             project_names=project_names)
    
    except Exception as e:
        logger_handler.logger.error(f"Error in verification review: {e}")
        flash('Error loading verification review.', 'error')
        return redirect(url_for('dashboard.dashboard'))

@bp.route('/verification-review/<int:record_id>/update', methods=['POST'], endpoint='update_verification_status')
@login_required
@log_database_operations('verification_update')
def update_verification_status(record_id):
    """Update verification status (approve/reject)"""
    try:
        # Only admins can update
        if session.get('role') not in ['admin', 'payroll', 'accounting']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403
        
        record = AttendanceData.query.get_or_404(record_id)
        
        new_status = request.json.get('status')
        admin_note = request.json.get('note', '')
        
        if new_status not in ['approved', 'rejected']:
            return jsonify({
                'success': False,
                'message': 'Invalid status'
            }), 400
        
        # Update record
        record.verification_status = new_status
        record.edit_note = f"Verification {new_status} by {session.get('username')}. {admin_note}"
        
        db.session.commit()
        
        # Log the action
        logger_handler.log_photo_verification(
            employee_id=record.employee_id,
            qr_code_id=record.qr_code_id,
            distance=record.location_accuracy or 0,
            status=new_status
        )
        
        return jsonify({
            'success': True,
            'message': f'Verification {new_status} successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        logger_handler.logger.error(f"Error updating verification: {e}")
        return jsonify({
            'success': False,
            'message': 'Error updating verification status'
        }), 500
    
@bp.route('/api/attendance/<int:record_id>/verification-details', endpoint='get_verification_details')
@login_required
def get_verification_details(record_id):
    """API endpoint to get verification details for a specific record"""
    try:
        # Get the attendance record with verification data
        record = AttendanceData.query.get_or_404(record_id)
        
        # DEBUG: Log record details
        logger_handler.logger.debug(
            f"Verification details: record={record.id}, employee={record.employee_id}, "
            f"date={record.check_in_date}, time={record.check_in_time}, "
            f"has_photo={record.verification_photo is not None}, status={record.verification_status}"
        )
        
        # Check if user has permission to view
        # Allow admin and payroll staff to view verification details
        if session.get('role') not in ['admin', 'payroll', 'accounting']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403
        
        # Log the access for security audit
        logger_handler.logger.info(f"User {session.get('username')} ({session.get('role')}) accessed verification details for record {record_id}")
        
        # Safely format dates/times with error handling
        try:
            check_in_date_str = record.check_in_date.strftime('%Y-%m-%d') if record.check_in_date else 'N/A'
        except Exception as e:
            logger_handler.logger.warning(f"Error formatting check_in_date for record {record_id}: {e}")
            check_in_date_str = str(record.check_in_date) if record.check_in_date else 'N/A'
        
        try:
            check_in_time_str = record.check_in_time.strftime('%I:%M %p') if record.check_in_time else 'N/A'
        except Exception as e:
            logger_handler.logger.warning(f"Error formatting check_in_time for record {record_id}: {e}")
            check_in_time_str = str(record.check_in_time) if record.check_in_time else 'N/A'
        
        # Prepare record data with safe formatting
        try:
            check_in_date_str = record.check_in_date.strftime('%Y-%m-%d') if record.check_in_date else 'N/A'
        except:
            check_in_date_str = str(record.check_in_date) if record.check_in_date else 'N/A'
        
        try:
            check_in_time_str = record.check_in_time.strftime('%I:%M %p') if record.check_in_time else 'N/A'
        except:
            check_in_time_str = str(record.check_in_time) if record.check_in_time else 'N/A'
        
        record_data = {
            'id': record.id,
            'employee_id': record.employee_id,
            'location_name': record.location_name or 'Unknown',
            'check_in_date': check_in_date_str,
            'check_in_time': check_in_time_str,
            'location_accuracy': float(record.location_accuracy) if record.location_accuracy else None,
            'checked_in_address': record.address or 'No address',
            'verification_photo': record.verification_photo,
            'verification_status': record.verification_status,
            'verification_required': record.verification_required,
            'device_info': record.device_info or 'Unknown'
        }
        
        return jsonify({
            'success': True,
            'record': record_data
        })
    
    except Exception as e:
        logger_handler.logger.error(f"Error in get_verification_details for record {record_id}: {e}", exc_info=True)

        return jsonify({
            'success': False,
            'message': 'Error loading verification details'
        }), 500

@bp.route('/verification-review/<int:record_id>', endpoint='verification_review_detail')
@login_required
def verification_review_detail(record_id):
    """Review a single verification photo on a dedicated page"""
    try:
        # Check permissions
        if session.get('role') not in ['admin', 'payroll', 'accounting']:
            flash('Access denied. Only administrators, payroll, and accounting staff can review verification photos.', 'error')
            return redirect(url_for('attendance.attendance_report'))
        
        # Get the attendance record
        record = AttendanceData.query.get_or_404(record_id)
        
        # Check if this record has verification
        if not record.verification_required:
            flash('This record does not require verification.', 'warning')
            return redirect(url_for('attendance.attendance_report'))
        
        # Get the QR code information for additional context
        qr_code = db.session.get(QRCode, record.qr_code_id) if record.qr_code_id else None
        
        # Get employee name from Employee table
        employee_name = None
        try:
            if record.employee_id:
                employee = Employee.query.filter_by(id=int(record.employee_id)).first()
                if employee:
                    employee_name = f"{employee.lastName}, {employee.firstName}"
                else:
                    employee_name = f"Unknown (ID: {record.employee_id})"
        except (ValueError, TypeError) as e:
            logger_handler.logger.warning(f"Could not lookup employee name for ID {record.employee_id}: {e}")
            employee_name = f"Unknown (ID: {record.employee_id})"
        
        # Get event type from QR code (Check In/Check Out)
        location_event = qr_code.location_event if qr_code and qr_code.location_event else 'N/A'
        
        # Log the access for audit trail
        logger_handler.logger.info(
            f"User {session.get('username')} ({session.get('role')}) "
            f"accessed verification review for record {record_id}"
        )
        
        # Format date and time for display
        try:
            check_in_date = record.check_in_date.strftime('%m/%d/%Y') if record.check_in_date else 'N/A'
        except:
            check_in_date = str(record.check_in_date) if record.check_in_date else 'N/A'
        
        try:
            check_in_time = record.check_in_time.strftime('%I:%M %p') if record.check_in_time else 'N/A'
        except:
            check_in_time = str(record.check_in_time) if record.check_in_time else 'N/A'
        
        return render_template('verification_review_detail.html',
                             record=record,
                             qr_code=qr_code,
                             check_in_date=check_in_date,
                             check_in_time=check_in_time,
                             employee_name=employee_name,
                             location_event=location_event)
    
    except Exception as e:
        logger_handler.logger.error(f"Error loading verification review detail: {e}")
        flash('Error loading verification details.', 'error')
        return redirect(url_for('attendance.attendance_report'))

@bp.route('/api/attendance/stats', endpoint='attendance_stats_api')
@admin_required
def attendance_stats_api():
    """API endpoint for attendance statistics"""
    try:
        # Daily stats for the last 7 days
        daily_stats = db.session.execute(text("""
            SELECT 
                check_in_date,
                COUNT(*) as checkins,
                COUNT(DISTINCT employee_id) as unique_employees
            FROM attendance_data 
            WHERE check_in_date >= CURRENT_DATE - INTERVAL '7 days'
            GROUP BY check_in_date
            ORDER BY check_in_date DESC
        """)).fetchall()

        # Location stats
        location_stats = db.session.execute(text("""
            SELECT 
                location_name,
                COUNT(*) as total_checkins,
                COUNT(DISTINCT employee_id) as unique_employees
            FROM attendance_data
            GROUP BY location_name
            ORDER BY total_checkins DESC
            LIMIT 10
        """)).fetchall()

        # Peak hours
        hourly_stats = db.session.execute(text("""
            SELECT 
                EXTRACT(hour FROM check_in_time) as hour,
                COUNT(*) as checkins
            FROM attendance_data
            WHERE check_in_date >= CURRENT_DATE - INTERVAL '30 days'
            GROUP BY EXTRACT(hour FROM check_in_time)
            ORDER BY hour
        """)).fetchall()

        return jsonify({
            'daily_stats': [{'date': str(row[0]), 'checkins': row[1], 'employees': row[2]} for row in daily_stats],
            'location_stats': [{'location': row[0], 'checkins': row[1], 'employees': row[2]} for row in location_stats],
            'hourly_stats': [{'hour': int(row[0]), 'checkins': row[1]} for row in hourly_stats]
        })

    except Exception as e:
        logger_handler.logger.error(f"Error fetching attendance stats: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch attendance statistics'}), 500

@bp.route('/export-configuration', endpoint='export_configuration')
@login_required
def export_configuration():
    """Display export configuration page for customizing Excel exports"""
    try:
        user_role = session.get('role')
        if user_role not in ['admin', 'payroll', 'accounting']:
            logger_handler.logger.warning(f"User {session.get('username', 'unknown')} (role: {user_role}) attempted unauthorized access to export configuration")
            flash('Access denied. Only administrators and payroll staff can access export configuration.', 'error')
            return redirect(url_for('attendance.attendance_report'))

        # Log export configuration access using your existing logger
        try:
            logger_handler.logger.info(f"User {session.get('username', 'unknown')} (role: {user_role}) accessed export configuration")
            logger_handler.logger.info(f"User {session.get('username', 'unknown')} accessed export configuration page")
        except Exception:
            pass

        # Get current filters from session or request args
        filters = {
            'date_from': request.args.get('date_from', ''),
            'date_to': request.args.get('date_to', ''),
            'location_filter': request.args.get('location', ''),
            'employee_filter': request.args.get('employee', ''),
            'project_filter': request.args.get('project', '')
        }

        logger_handler.logger.debug(f"Export config filters: {filters}")
        
        # Get project name if project filter is applied
        project_name = None
        if filters.get('project_filter'):
            try:
                project = db.session.get(Project, int(filters['project_filter']))
                if project:
                    project_name = project.name
                    logger_handler.logger.debug(f"Project filter: ID={filters['project_filter']}, Name={project_name}")
            except Exception as e:
                logger_handler.logger.warning(f"Error fetching project name for filter: {e}")

        # Check if location accuracy feature exists
        try:
            has_location_accuracy = check_location_accuracy_column_exists()
        except Exception as e:
            logger_handler.logger.warning(f"Error checking location accuracy column: {e}")
            has_location_accuracy = False

        # Define all available columns with their default settings
        available_columns = [
            {'key': 'employee_id', 'label': 'Employee ID', 'default_name': 'ID', 'enabled': True},
            {'key': 'employee_name', 'label': 'Employee Name', 'default_name': 'Employee Name', 'enabled': False},
            {'key': 'location_name', 'label': 'Location', 'default_name': 'Location Name', 'enabled': True},
            {'key': 'status', 'label': 'Event', 'default_name': 'Action Description', 'enabled': True},
            {'key': 'check_in_date', 'label': 'Date', 'default_name': 'Date', 'enabled': True},
            {'key': 'check_in_time', 'label': 'Time', 'default_name': 'Time', 'enabled': True},
            {'key': 'qr_address', 'label': 'QR Address', 'default_name': 'Event Description', 'enabled': True},
            {'key': 'address', 'label': 'Check-in Address', 'default_name': 'Recorded Address', 'enabled': True},
            {'key': 'device_info', 'label': 'Device', 'default_name': 'Platform', 'enabled': True},
            {'key': 'ip_address', 'label': 'IP Address', 'default_name': 'IP Address', 'enabled': False},
            {'key': 'user_agent', 'label': 'User Agent', 'default_name': 'Browser/User Agent', 'enabled': False},
            {'key': 'latitude', 'label': 'Latitude', 'default_name': 'GPS Latitude', 'enabled': False},
            {'key': 'longitude', 'label': 'Longitude', 'default_name': 'GPS Longitude', 'enabled': False},
            {'key': 'accuracy', 'label': 'GPS Accuracy', 'default_name': 'GPS Accuracy (meters)', 'enabled': False},
        ]

        # Add location accuracy column if feature exists
        if has_location_accuracy:
            available_columns.append({
                'key': 'location_accuracy',
                'label': 'Location Accuracy',
                'default_name': 'Distance',
                'enabled': True  # Changed from False to True
            })

        logger_handler.logger.debug(f"Rendering export configuration with {len(available_columns)} columns")

        return render_template('export_configuration.html',
                             available_columns=available_columns,
                             filters=filters,
                             project_name=project_name,
                             has_location_accuracy_feature=has_location_accuracy)

    except Exception as e:
        logger_handler.logger.error(f"Error in export_configuration route: {e}", exc_info=True)

        # Use your existing logger error method with correct parameters
        try:
            logger_handler.log_flask_error(
                'export_configuration_error',
                str(e),
                stack_trace=traceback.format_exc()
            )
        except Exception as log_error:
            logger_handler.logger.warning(f"Could not log error: {log_error}")

        flash('Error loading export configuration page.', 'error')
        return redirect(url_for('attendance.attendance_report'))

@bp.route('/generate-excel-export', methods=['POST'], endpoint='generate_excel_export')
@login_required
def generate_excel_export():
    """Generate and download Excel file with selected columns in specified order"""
    try:
        user_role = session.get('role')
        if user_role not in ['admin', 'payroll', 'accounting']:
            logger_handler.logger.warning(f"User {session.get('username', 'unknown')} (role: {user_role}) attempted unauthorized Excel export")
            flash('Access denied. Only administrators and payroll staff can export data.', 'error')
            return redirect(url_for('attendance.attendance_report'))

        logger_handler.logger.info(f"Excel export started by user {session.get('username', 'unknown')}")

        # Log export action using your existing logger
        try:
            logger_handler.logger.info(f"User {session.get('username', 'unknown')} generated Excel export")
        except Exception:
            pass

        # Get selected columns and custom names from form
        selected_columns_raw = request.form.getlist('selected_columns')
        logger_handler.logger.debug(f"Selected columns (raw): {selected_columns_raw}")

        # Get column order from form
        column_order_json = request.form.get('column_order', '[]')
        try:
            column_order = json.loads(column_order_json) if column_order_json else []
        except (json.JSONDecodeError, TypeError):
            column_order = []

        logger_handler.logger.debug(f"Column order from form: {column_order}")

        # Determine final column order
        if column_order:
            # Use the specified order, but only include actually selected columns
            selected_columns = [col for col in column_order if col in selected_columns_raw]
            # Add any selected columns that weren't in the order (shouldn't happen, but safety check)
            for col in selected_columns_raw:
                if col not in selected_columns:
                    selected_columns.append(col)
        else:
            # Fallback to raw selection order
            selected_columns = selected_columns_raw

        logger_handler.logger.debug(f"Final column order: {selected_columns}")

        if not selected_columns:
            flash('Please select at least one column to export.', 'error')
            return redirect(url_for('attendance.export_configuration'))

        column_names = {}
        for column in selected_columns:
            column_names[column] = request.form.get(f'name_{column}', column)

        # Get filters
        filters = {
            'date_from': request.form.get('date_from'),
            'date_to': request.form.get('date_to'),
            'location_filter': request.form.get('location_filter'),
            'employee_filter': request.form.get('employee_filter'),
            'project_filter': request.form.get('project_filter')
        }

        logger_handler.logger.debug(f"Export filters: {filters}")

        # Save user preferences in session for next time
        session['export_preferences'] = {
            'selected_columns': selected_columns,
            'column_names': column_names,
            'column_order': selected_columns  # This is now the ordered list
        }

        # Generate Excel file with ordered columns
        excel_file = create_excel_export_ordered(selected_columns, column_names, filters)

        if excel_file:
            # Get project name if project filter exists
            project_name_for_filename = ''
            if filters.get('project_filter'):
                try:
                    project = db.session.get(Project, int(filters['project_filter']))
                    if project:
                        # Replace spaces and special characters with underscores
                        project_name_safe = project.name.replace(' ', '_').replace('/', '_').replace('\\', '_')
                        project_name_for_filename = f"{project_name_safe}_"
                except Exception as e:
                    logger_handler.logger.warning(f"Error getting project name for filename: {e}")
            
            # Format dates for filename (MMDDYYYY format)
            date_from_formatted = ''
            date_to_formatted = ''
            if filters.get('date_from'):
                try:
                    date_obj = datetime.strptime(filters['date_from'], '%Y-%m-%d')
                    date_from_formatted = date_obj.strftime('%m%d%Y')
                except ValueError:
                    pass
            
            if filters.get('date_to'):
                try:
                    date_obj = datetime.strptime(filters['date_to'], '%Y-%m-%d')
                    date_to_formatted = date_obj.strftime('%m%d%Y')
                except ValueError:
                    pass
            
            # Build filename components
            # Format: [project_name_]attendance_report_[fromdate_todate].xlsx
            date_range_str = ''
            if date_from_formatted and date_to_formatted:
                date_range_str = f"{date_from_formatted}_{date_to_formatted}"
            elif date_from_formatted:
                date_range_str = f"{date_from_formatted}"
            elif date_to_formatted:
                date_range_str = f"{date_to_formatted}"
            
            filename = f'{project_name_for_filename}attendance_report_{date_range_str}.xlsx'

            logger_handler.logger.info(f"Excel export generated successfully: {filename}")

            # Log successful export using your existing logger
            try:
                logger_handler.logger.info(f"Excel export generated successfully with {len(selected_columns)} columns in custom order by user {session.get('username', 'unknown')}: {filename}")
            except Exception:
                pass

            return send_file(
                excel_file,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('Error generating Excel file.', 'error')
            return redirect(url_for('attendance.export_configuration'))

    except Exception as e:
        logger_handler.logger.error(f"Error in generate_excel_export route: {e}", exc_info=True)

        # Use your existing logger error method with correct parameters
        try:
            logger_handler.log_flask_error(
                'excel_export_error',
                str(e),
                stack_trace=traceback.format_exc()
            )
        except Exception as log_error:
            logger_handler.logger.warning(f"Could not log error: {log_error}")

        flash('Error generating Excel export.', 'error')
        return redirect(url_for('attendance.export_configuration'))

def create_excel_export(selected_columns, column_names, filters):
    """Create Excel file with selected attendance data - Updated to include employee names"""
    try:
        logger_handler.logger.info(f"Creating Excel export with {len(selected_columns)} columns")

        # Import openpyxl modules
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            logger_handler.logger.error(f"openpyxl import error: {e}. Run: pip install openpyxl")
            return None

        # Build query based on filters - JOIN with QRCode to get location_event and location_address
        # Now also JOIN with Employee table to get employee names
        query = db.session.query(AttendanceData, QRCode, Employee).join(
            QRCode, AttendanceData.qr_code_id == QRCode.id
        ).outerjoin(
            Employee, text("CAST(attendance_data.employee_id AS UNSIGNED) = employee.id")
        )

        # Apply date filters
        if filters.get('date_from'):
            try:
                date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
                query = query.filter(AttendanceData.check_in_date >= date_from)
                logger_handler.logger.debug(f"Applied date_from filter: {date_from}")
            except ValueError as e:
                logger_handler.logger.warning(f"Invalid date_from format: {e}")

        if filters.get('date_to'):
            try:
                date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()
                query = query.filter(AttendanceData.check_in_date <= date_to)
                logger_handler.logger.debug(f"Applied date_to filter: {date_to}")
            except ValueError as e:
                logger_handler.logger.warning(f"Invalid date_to format: {e}")

        # Apply location filter
        if filters.get('location_filter'):
            query = query.filter(AttendanceData.location_name.like(f"%{filters['location_filter']}%"))
            logger_handler.logger.debug(f"Applied location filter: {filters['location_filter']}")

        # Apply employee filter — supports comma-separated multi-employee values
        if filters.get('employee_filter'):
            emp_ids = [e.strip() for e in filters['employee_filter'].split(',') if e.strip()]
            if len(emp_ids) == 1:
                query = query.filter(AttendanceData.employee_id == emp_ids[0])
            elif len(emp_ids) > 1:
                query = query.filter(AttendanceData.employee_id.in_(emp_ids))
            logger_handler.logger.debug(f"Applied employee filter: {emp_ids}")

        # Apply project filter
        if filters.get('project_filter'):
            try:
                project_id = int(filters['project_filter'])
                query = query.filter(QRCode.project_id == project_id)
                logger_handler.logger.debug(f"Applied project filter: {project_id}")
            except (ValueError, TypeError) as e:
                logger_handler.logger.warning(f"Invalid project filter: {e}")

        # Order by date and time
        query = query.order_by(AttendanceData.check_in_date.desc(), AttendanceData.check_in_time.desc())

        # Execute query
        results = query.all()
        logger_handler.logger.debug(f"Query returned {len(results)} records for export")

        if not results:
            logger_handler.logger.warning("No records found for export")
            return None

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Set headers based on selected columns
        headers = []
        for column_key in selected_columns:
            header_name = column_names.get(column_key, column_key)
            headers.append(header_name)

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, (attendance_record, qr_record, employee_record) in enumerate(results, 2):
            for col_idx, column_key in enumerate(selected_columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                try:
                    # Handle each column type
                    if column_key == 'employee_id':
                        cell.value = format_employee_id_for_excel(attendance_record.employee_id)
                    elif column_key == 'employee_name':
                        # NEW: Handle employee name from joined Employee table
                        if employee_record:
                            cell.value = f"{employee_record.lastName}, {employee_record.firstName}"
                        else:
                            cell.value = f"Unknown (ID: {attendance_record.employee_id})"
                    elif column_key == 'location_name':
                        cell.value = attendance_record.location_name or ''
                    elif column_key == 'status':
                        cell.value = qr_record.location_event if qr_record.location_event else 'Check In'
                    elif column_key == 'check_in_date':
                        cell.value = attendance_record.check_in_date.strftime('%Y-%m-%d') if attendance_record.check_in_date else ''
                    elif column_key == 'check_in_time':
                        cell.value = attendance_record.check_in_time.strftime('%H:%M:%S') if attendance_record.check_in_time else ''
                    elif column_key == 'qr_address':
                        # Use attendance-level qr_address first (set for dynamic QR check-ins),
                        # fall back to the QR code's location_address for standard QR.
                        cell.value = (
                            getattr(attendance_record, 'qr_address', None)
                            or (qr_record.location_address if qr_record else '')
                            or ''
                        )
                    elif column_key == 'address':
                        # Check-in address logic based on location accuracy WITH HYPERLINKS
                        # If location accuracy < 0.3 miles, use QR address; otherwise use actual check-in address
                        if hasattr(attendance_record, 'location_accuracy') and attendance_record.location_accuracy is not None:
                            try:
                                accuracy_value = float(attendance_record.location_accuracy)
                                if accuracy_value < 0.3:
                                    # High accuracy - use QR code ADDRESS (not location) with hyperlink
                                    address_text = (
                                        getattr(attendance_record, 'qr_address', None)
                                        or (qr_record.location_address if qr_record and qr_record.location_address else '')
                                        or ''
                                    )
                                    if address_text and hasattr(qr_record, 'address_latitude') and hasattr(qr_record, 'address_longitude') and qr_record.address_latitude and qr_record.address_longitude:
                                        # Format coordinates with 10 decimal places
                                        lat_formatted = f"{float(qr_record.address_latitude):.10f}"
                                        lng_formatted = f"{float(qr_record.address_longitude):.10f}"
                                        hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                        cell.value = hyperlink_formula
                                        logger_handler.logger.debug(f"Added QR address hyperlink for employee {attendance_record.employee_id}")
                                    else:
                                        cell.value = address_text
                                    logger_handler.logger.debug(f"Using QR address for employee {attendance_record.employee_id} (accuracy: {accuracy_value:.3f} miles)")
                                else:
                                    # Lower accuracy - use actual check-in address with hyperlink
                                    address_text = attendance_record.address or ''
                                    if address_text and attendance_record.latitude and attendance_record.longitude:
                                        # Format coordinates with 10 decimal places
                                        lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                        lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                        hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                        cell.value = hyperlink_formula
                                        logger_handler.logger.debug(f"Added check-in address hyperlink for employee {attendance_record.employee_id}")
                                    else:
                                        cell.value = address_text
                                    logger_handler.logger.debug(f"Using check-in address for employee {attendance_record.employee_id} (accuracy: {accuracy_value:.3f} miles)")
                            except (ValueError, TypeError):
                                # If accuracy can't be converted to float, use check-in address with hyperlink
                                address_text = attendance_record.address or ''
                                if address_text and attendance_record.latitude and attendance_record.longitude:
                                    # Format coordinates with 10 decimal places
                                    lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                    lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                    hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                    cell.value = hyperlink_formula
                                    logger_handler.logger.debug(f"Added check-in address hyperlink (fallback) for employee {attendance_record.employee_id}")
                                else:
                                    cell.value = address_text
                        else:
                            # No location accuracy data - use actual check-in address with hyperlink
                            address_text = attendance_record.address or ''
                            if address_text and attendance_record.latitude and attendance_record.longitude:
                                # Format coordinates with 10 decimal places
                                lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                cell.value = hyperlink_formula
                                logger_handler.logger.debug(f"Added check-in address hyperlink (no accuracy data) for employee {attendance_record.employee_id}")
                            else:
                                cell.value = address_text
                    elif column_key == 'device_info':
                        cell.value = attendance_record.device_info or ''
                    elif column_key == 'ip_address':
                        cell.value = attendance_record.ip_address or ''
                    elif column_key == 'user_agent':
                        cell.value = attendance_record.user_agent or ''
                    elif column_key == 'latitude':
                       cell.value = attendance_record.latitude or ''
                    elif column_key == 'longitude':
                        cell.value = attendance_record.longitude or ''
                    elif column_key == 'accuracy':
                        cell.value = attendance_record.accuracy or ''
                    elif column_key == 'location_accuracy':
                        cell.value = attendance_record.location_accuracy or ''
                    else:
                        cell.value = ''
                except Exception as cell_error:
                    logger_handler.logger.warning(f"Error setting cell value for {column_key}: {cell_error}")
                    cell.value = ''

        # Auto-adjust column widths based on content and header
        for col_idx, column_key in enumerate(selected_columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Get header name length
            header_name = column_names.get(column_key, column_key)
            max_length = len(str(header_name))
            
            # Check content in all rows (sample first 100 rows for performance)
            for row_idx in range(2, min(102, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    cell_value = str(cell.value) if cell.value else ''
                    # For HYPERLINK formulas, extract the display text
                    if cell_value.startswith('=HYPERLINK'):
                        # Extract text between last quotes: HYPERLINK("url","display_text")
                        import re
                        match = re.search(r',"([^"]+)"\)$', cell_value)
                        if match:
                            cell_value = match.group(1)
                    
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            # Set width based on column type with reasonable limits
            # Define optimal widths for specific column types
            column_width_rules = {
                'employee_id': {'min': 8, 'max': 15},
                'employee_name': {'min': 20, 'max': 30},
                'location_name': {'min': 15, 'max': 35},
                'status': {'min': 12, 'max': 20},
                'check_in_date': {'min': 12, 'max': 15},
                'check_in_time': {'min': 10, 'max': 12},
                'qr_address': {'min': 20, 'max': 40},
                'address': {'min': 20, 'max': 45},
                'device_info': {'min': 12, 'max': 20},
                'ip_address': {'min': 14, 'max': 18},
                'user_agent': {'min': 15, 'max': 30},
                'latitude': {'min': 12, 'max': 15},
                'longitude': {'min': 12, 'max': 15},
                'accuracy': {'min': 10, 'max': 15},
                'location_accuracy': {'min': 10, 'max': 15}
            }
            
            # Get rules for this column or use defaults
            rules = column_width_rules.get(column_key, {'min': 10, 'max': 40})
            
            # Calculate adjusted width: add 2 for padding, respect min/max
            adjusted_width = max_length + 2
            adjusted_width = max(rules['min'], min(adjusted_width, rules['max']))
            
            ws.column_dimensions[column_letter].width = adjusted_width
            
            logger_handler.logger.debug(f"Column {column_letter} ({column_key}): width={adjusted_width} (max_content={max_length})")

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        logger_handler.logger.info("Excel file created successfully with employee names")
        
        # Log export action with employee name column
        try:
            logger_handler.logger.info(f"Excel export with employee names generated by user {session.get('username', 'unknown')}")
        except Exception:
            pass
            
        return excel_buffer

    except Exception as e:
        logger_handler.logger.error(f"Error creating Excel export: {e}", exc_info=True)
        
        # Log error
        try:
            logger_handler.log_flask_error(
                'excel_export_error',
                str(e),
                stack_trace=traceback.format_exc()
            )
        except Exception as log_error:
            logger_handler.logger.warning(f"Could not log error: {log_error}")
            
        return None

def format_employee_id_for_excel(employee_id):
    if not employee_id:
        return ''
    emp_id_str = str(employee_id).strip()
    if emp_id_str.isdigit():
        return int(emp_id_str)
    else:
        return emp_id_str
    
def create_excel_export_ordered(selected_columns, column_names, filters):
    """Create Excel file with selected attendance data in specified column order"""
    try:
        logger_handler.logger.info(f"Creating Excel export with {len(selected_columns)} columns")

        # Import openpyxl modules
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            logger_handler.logger.error(f"openpyxl import error: {e}. Run: pip install openpyxl")
            return None

        # Build query based on filters - JOIN with QRCode to get location_event and location_address
        # Now also JOIN with Employee table to get employee names
        query = db.session.query(AttendanceData, QRCode, Employee).join(
            QRCode, AttendanceData.qr_code_id == QRCode.id
        ).outerjoin(
            Employee, text("CAST(attendance_data.employee_id AS UNSIGNED) = employee.id")
        )

        # Apply date filters
        if filters.get('date_from'):
            try:
                date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
                query = query.filter(AttendanceData.check_in_date >= date_from)
                logger_handler.logger.debug(f"Applied date_from filter: {date_from}")
            except ValueError as e:
                logger_handler.logger.warning(f"Invalid date_from format: {e}")

        if filters.get('date_to'):
            try:
                date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()
                query = query.filter(AttendanceData.check_in_date <= date_to)
                logger_handler.logger.debug(f"Applied date_to filter: {date_to}")
            except ValueError as e:
                logger_handler.logger.warning(f"Invalid date_to format: {e}")

        # Apply location filter
        if filters.get('location_filter'):
            query = query.filter(AttendanceData.location_name.like(f"%{filters['location_filter']}%"))
            logger_handler.logger.debug(f"Applied location filter: {filters['location_filter']}")

        # Apply employee filter — supports comma-separated multi-employee values
        if filters.get('employee_filter'):
            emp_ids = [e.strip() for e in filters['employee_filter'].split(',') if e.strip()]
            if len(emp_ids) == 1:
                query = query.filter(AttendanceData.employee_id == emp_ids[0])
            elif len(emp_ids) > 1:
                query = query.filter(AttendanceData.employee_id.in_(emp_ids))
            logger_handler.logger.debug(f"Applied employee filter: {emp_ids}")

        # Apply project filter
        if filters.get('project_filter'):
            try:
                project_id = int(filters['project_filter'])
                query = query.filter(QRCode.project_id == project_id)
                logger_handler.logger.debug(f"Applied project filter: {project_id}")
            except (ValueError, TypeError) as e:
                logger_handler.logger.warning(f"Invalid project filter: {e}")

        # Order by date and time
        query = query.order_by(AttendanceData.check_in_date.desc(), AttendanceData.check_in_time.desc())

        # Execute query
        results = query.all()
        logger_handler.logger.debug(f"Query returned {len(results)} records for export")

        if not results:
            logger_handler.logger.warning("No records found for export")
            return None

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Verification status color fills for location_accuracy column
        # Yellow for pending, Green for approved, Red for rejected
        verification_fill_pending = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        verification_fill_approved = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
        verification_fill_rejected = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Light Red

        # Set headers based on selected columns in the specified order
        headers = []
        for column_key in selected_columns:
            header_name = column_names.get(column_key, column_key)
            headers.append(header_name)

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, (attendance_record, qr_record, employee_record) in enumerate(results, 2):
            for col_idx, column_key in enumerate(selected_columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                try:
                    # Handle each column type
                    if column_key == 'employee_id':
                        cell.value = format_employee_id_for_excel(attendance_record.employee_id)
                    elif column_key == 'employee_name':
                        # NEW: Handle employee name from joined Employee table
                        if employee_record:
                            cell.value = f"{employee_record.lastName}, {employee_record.firstName}"
                        else:
                            cell.value = f"Unknown (ID: {attendance_record.employee_id})"
                    elif column_key == 'location_name':
                        cell.value = attendance_record.location_name or ''
                    elif column_key == 'status':
                        cell.value = qr_record.location_event if qr_record.location_event else 'Check In'
                    elif column_key == 'check_in_date':
                        cell.value = attendance_record.check_in_date.strftime('%Y-%m-%d') if attendance_record.check_in_date else ''
                    elif column_key == 'check_in_time':
                        cell.value = attendance_record.check_in_time.strftime('%H:%M:%S') if attendance_record.check_in_time else ''
                    elif column_key == 'qr_address':
                        # Use attendance-level qr_address first (set for dynamic QR check-ins),
                        # fall back to the QR code's location_address for standard QR.
                        cell.value = (
                            getattr(attendance_record, 'qr_address', None)
                            or (qr_record.location_address if qr_record else '')
                            or ''
                        )
                    elif column_key == 'address':
                        # Check-in address logic based on location accuracy WITH HYPERLINKS
                        # If location accuracy < 0.3 miles, use QR address; otherwise use actual check-in address
                        if hasattr(attendance_record, 'location_accuracy') and attendance_record.location_accuracy is not None:
                            try:
                                accuracy_value = float(attendance_record.location_accuracy)
                                if accuracy_value < 0.3:
                                    # High accuracy - use QR code ADDRESS (not location) with hyperlink
                                    address_text = (
                                        getattr(attendance_record, 'qr_address', None)
                                        or (qr_record.location_address if qr_record and qr_record.location_address else '')
                                        or ''
                                    )
                                    if address_text and hasattr(qr_record, 'address_latitude') and hasattr(qr_record, 'address_longitude') and qr_record.address_latitude and qr_record.address_longitude:
                                        # Format coordinates with 10 decimal places
                                        lat_formatted = f"{float(qr_record.address_latitude):.10f}"
                                        lng_formatted = f"{float(qr_record.address_longitude):.10f}"
                                        hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                        cell.value = hyperlink_formula
                                        logger_handler.logger.debug(f"Added QR address hyperlink for employee {attendance_record.employee_id}")
                                    else:
                                        cell.value = address_text
                                    logger_handler.logger.debug(f"Using QR address for employee {attendance_record.employee_id} (accuracy: {accuracy_value:.3f} miles)")
                                else:
                                    # Lower accuracy - use actual check-in address with hyperlink
                                    address_text = attendance_record.address or ''
                                    if address_text and attendance_record.latitude and attendance_record.longitude:
                                        # Format coordinates with 10 decimal places
                                        lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                        lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                        hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                        cell.value = hyperlink_formula
                                        logger_handler.logger.debug(f"Added check-in address hyperlink for employee {attendance_record.employee_id}")
                                    else:
                                        cell.value = address_text
                                    logger_handler.logger.debug(f"Using check-in address for employee {attendance_record.employee_id} (accuracy: {accuracy_value:.3f} miles)")
                            except (ValueError, TypeError):
                                # If accuracy can't be converted to float, use check-in address with hyperlink
                                address_text = attendance_record.address or ''
                                if address_text and attendance_record.latitude and attendance_record.longitude:
                                    # Format coordinates with 10 decimal places
                                    lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                    lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                    hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                    cell.value = hyperlink_formula
                                    logger_handler.logger.debug(f"Added check-in address hyperlink (fallback) for employee {attendance_record.employee_id}")
                                else:
                                    cell.value = address_text
                        else:
                            # No location accuracy data - use actual check-in address with hyperlink
                            address_text = attendance_record.address or ''
                            if address_text and attendance_record.latitude and attendance_record.longitude:
                                # Format coordinates with 10 decimal places
                                lat_formatted = f"{float(attendance_record.latitude):.10f}"
                                lng_formatted = f"{float(attendance_record.longitude):.10f}"
                                hyperlink_formula = f'=HYPERLINK("http://maps.google.com/maps?q={lat_formatted},{lng_formatted}","{address_text.strip()}")'
                                cell.value = hyperlink_formula
                                logger_handler.logger.debug(f"Added check-in address hyperlink (no accuracy data) for employee {attendance_record.employee_id}")
                            else:
                                cell.value = address_text
                    elif column_key == 'device_info':
                        cell.value = attendance_record.device_info or ''
                    elif column_key == 'ip_address':
                        cell.value = attendance_record.ip_address or ''
                    elif column_key == 'user_agent':
                        cell.value = attendance_record.user_agent or ''
                    elif column_key == 'latitude':
                       cell.value = attendance_record.latitude or ''
                    elif column_key == 'longitude':
                        cell.value = attendance_record.longitude or ''
                    elif column_key == 'accuracy':
                        cell.value = attendance_record.accuracy or ''
                    elif column_key == 'location_accuracy':
                        cell.value = attendance_record.location_accuracy or ''
                        # Apply color fill based on verification_status
                        # Only apply color if verification_status is not NULL
                        if hasattr(attendance_record, 'verification_status') and attendance_record.verification_status:
                            if attendance_record.verification_status == 'pending':
                                cell.fill = verification_fill_pending  # Yellow
                            elif attendance_record.verification_status == 'approved':
                                cell.fill = verification_fill_approved  # Green
                            elif attendance_record.verification_status == 'rejected':
                                cell.fill = verification_fill_rejected  # Red
                    else:
                        cell.value = ''
                except Exception as cell_error:
                    logger_handler.logger.warning(f"Error setting cell value for {column_key}: {cell_error}")
                    cell.value = ''

        # Auto-adjust column widths based on content and header
        for col_idx, column_key in enumerate(selected_columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Get header name length
            header_name = column_names.get(column_key, column_key)
            max_length = len(str(header_name))
            
            # Check content in all rows (sample first 100 rows for performance)
            for row_idx in range(2, min(102, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    cell_value = str(cell.value) if cell.value else ''
                    # For HYPERLINK formulas, extract the display text
                    if cell_value.startswith('=HYPERLINK'):
                        # Extract text between last quotes: HYPERLINK("url","display_text")
                        import re
                        match = re.search(r',"([^"]+)"\)$', cell_value)
                        if match:
                            cell_value = match.group(1)
                    
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            # Set width based on column type with reasonable limits
            # Define optimal widths for specific column types
            column_width_rules = {
                'employee_id': {'min': 8, 'max': 15},
                'employee_name': {'min': 20, 'max': 30},
                'location_name': {'min': 15, 'max': 35},
                'status': {'min': 12, 'max': 20},
                'check_in_date': {'min': 12, 'max': 15},
                'check_in_time': {'min': 10, 'max': 12},
                'qr_address': {'min': 20, 'max': 40},
                'address': {'min': 20, 'max': 45},
                'device_info': {'min': 12, 'max': 20},
                'ip_address': {'min': 14, 'max': 18},
                'user_agent': {'min': 15, 'max': 30},
                'latitude': {'min': 12, 'max': 15},
                'longitude': {'min': 12, 'max': 15},
                'accuracy': {'min': 10, 'max': 15},
                'location_accuracy': {'min': 10, 'max': 15}
            }
            
            # Get rules for this column or use defaults
            rules = column_width_rules.get(column_key, {'min': 10, 'max': 40})
            
            # Calculate adjusted width: add 2 for padding, respect min/max
            adjusted_width = max_length + 2
            adjusted_width = max(rules['min'], min(adjusted_width, rules['max']))
            
            ws.column_dimensions[column_letter].width = adjusted_width
            
            logger_handler.logger.debug(f"Column {column_letter} ({column_key}): width={adjusted_width} (max_content={max_length})")

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        logger_handler.logger.info("Excel file created successfully with employee names and verification status coloring")
        
        # Log export action with employee name column and verification status coloring
        try:
            logger_handler.logger.info(f"Excel export with employee names and verification status coloring generated by user {session.get('username', 'unknown')}")
        except Exception:
            pass
            
        return excel_buffer

    except Exception as e:
        logger_handler.logger.error(f"Error creating Excel export: {e}", exc_info=True)
        
        # Log error
        try:
            logger_handler.log_flask_error(
                'excel_export_ordered_error',
                str(e),
                stack_trace=traceback.format_exc()
            )
        except Exception as log_error:
            logger_handler.logger.warning(f"Could not log error: {log_error}")
            
        return None