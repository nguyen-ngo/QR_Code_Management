"""
routes/time_attendance.py
=========================
Time attendance dashboard, import pipeline, export (Excel / by-building),
and records management routes.

Routes: /time-attendance, /time-attendance/import/*,
        /time-attendance/export*, /time-attendance/records,
        /time-attendance/record/<id>, /time-attendance/delete/<id>,
        /api/time-attendance/*
"""
from flask import Blueprint, render_template, request, redirect, flash, session, jsonify, send_file, Response, g, current_app, url_for
from datetime import datetime, date, timedelta, time
import io, os, json, re, uuid, traceback
import time as _time

from extensions import db, logger_handler
from models.employee import Employee
from models.project import Project
from models.qrcode import QRCode
from models.time_attendance import TimeAttendance
from models.user import User
from sqlalchemy import text
from werkzeug.utils import secure_filename
from logger_handler import log_user_activity, log_database_operations
from utils.helpers import (
    admin_required,
    has_admin_privileges,
    has_staff_level_access,
    login_required,
    staff_or_admin_required)
from utils.geocoding import calculate_location_accuracy_enhanced
from working_hours_calculator import WorkingHoursCalculator, round_time_to_quarter_hour, convert_minutes_to_base100, round_base100_hours
from time_attendance_import_service import TimeAttendanceImportService
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import openpyxl.cell.cell

bp = Blueprint('time_attendance', __name__)

from routes.time_attendance_export import (
    calculate_possible_violation,
    _overnight_aware_sort_key,
    _qtr,
    export_time_attendance_excel,
    export_time_attendance_by_building_excel,
)


@bp.route('/time-attendance', endpoint='time_attendance_dashboard')
@login_required
@log_user_activity('time_attendance_view')
def time_attendance_dashboard():
    """Display time attendance dashboard with table layout"""
    try:
        # Initialize default values
        total_records = 0
        unique_employees = 0
        unique_locations = 0
        recent_imports = []
        recent_records = []
        employees = []
        locations = []
        
        # Try to get data from TimeAttendance model if it exists
        try:
            
            # Get summary statistics
            total_records = TimeAttendance.query.count()
            
            if total_records > 0:
                unique_employees = db.session.query(TimeAttendance.employee_id).distinct().count()
                unique_locations = db.session.query(TimeAttendance.location_name).distinct().count()
                
                # Get recent records (last 20 records for table display)
                recent_records = TimeAttendance.query.order_by(
                    TimeAttendance.attendance_date.desc(),
                    TimeAttendance.attendance_time.desc()
                ).limit(20).all()
                
                # Get recent imports (last 10 import batches)
                recent_imports = db.session.query(
                    TimeAttendance.import_batch_id,
                    TimeAttendance.import_date,
                    TimeAttendance.import_source,
                    db.func.count(TimeAttendance.id).label('record_count')
                ).filter(
                    TimeAttendance.import_batch_id.isnot(None)
                ).group_by(
                    TimeAttendance.import_batch_id,
                    TimeAttendance.import_date,
                    TimeAttendance.import_source
                ).order_by(
                    TimeAttendance.import_date.desc()
                ).limit(10).all()
                
                # Get filter options
                employees = TimeAttendance.get_unique_employees()
                locations = TimeAttendance.get_unique_locations()
                
        except ImportError:
            # TimeAttendance model doesn't exist yet - use defaults
            pass
        except Exception as e:
            # Database table doesn't exist yet or other error - use defaults
            print(f"TimeAttendance query error: {e}")
            pass
        
        return render_template('time_attendance_dashboard.html',
                             total_records=total_records,
                             unique_employees=unique_employees,
                             unique_locations=unique_locations,
                             recent_imports=recent_imports,
                             recent_records=recent_records,
                             employees=employees,
                             locations=locations)
                             
    except Exception as e:
        logger_handler.logger.error(f"Error in time attendance dashboard: {e}")
        flash('Error loading time attendance dashboard.', 'error')
        return redirect(url_for('dashboard.dashboard'))

@bp.route('/time-attendance/import', methods=['GET', 'POST'], endpoint='import_time_attendance')
@login_required
@log_database_operations('time_attendance_import')
def import_time_attendance():
    """Enhanced import with duplicate review"""
    if request.method == 'GET':
        # Load active projects for dropdown
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()
        return render_template('time_attendance_import.html', projects=projects)

    if request.method == 'POST':
        try:
            # Check if this is coming from invalid review (file is already in session)
            coming_from_invalid_review = request.form.get('from_invalid_review', 'false').lower() == 'true'
            coming_from_duplicate_review = request.form.get('from_duplicate_review', 'false').lower() == 'true'

            print(f"\n🔍 IMPORT FLOW DEBUG:")
            print(f"   Coming from invalid review: {coming_from_invalid_review}")
            print(f"   Coming from duplicate review: {coming_from_duplicate_review}")

            if coming_from_invalid_review or coming_from_duplicate_review:
                # Retrieve file from session
                if 'pending_import_file' not in session or 'pending_import_filename' not in session:
                    flash('Session expired. Please upload the file again.', 'error')
                    return redirect(url_for('time_attendance.import_time_attendance'))
                
                temp_path = session['pending_import_file']
                filename = session['pending_import_filename']
                
                # Verify file still exists
                if not os.path.exists(temp_path):
                    flash('Temporary file not found. Please upload the file again.', 'error')
                    session.pop('pending_import_file', None)
                    session.pop('pending_import_filename', None)
                    return redirect(url_for('time_attendance.import_time_attendance'))
                
                print(f"✅ Retrieved file from session: {filename}")
                print(f"✅ Temp path exists: {os.path.exists(temp_path)}")
                
            else:
                # Normal file upload flow - now supports multiple files
                if 'files' not in request.files:
                    flash('No files uploaded.', 'error')
                    return redirect(request.url)
                
                files = request.files.getlist('files')
                if not files or len(files) == 0:
                    flash('No files selected.', 'error')
                    return redirect(request.url)
                
                # Validate all files and save them temporarily
                temp_paths = []
                filenames = []
                
                for file in files:
                    if file.filename == '':
                        continue
                    
                    # Validate file extension
                    if not file.filename.lower().endswith(('.xlsx', '.xls')):
                        flash(f'Invalid file format: {file.filename}. Please upload only Excel files (.xlsx or .xls).', 'error')
                        # Clean up already saved files
                        for saved_path in temp_paths:
                            if os.path.exists(saved_path):
                                os.remove(saved_path)
                        return redirect(request.url)
                    
                    # Save uploaded file temporarily
                    filename = secure_filename(file.filename)
                    temp_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', '/tmp'), 
                                           f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
                    
                    os.makedirs(os.path.dirname(temp_path), exist_ok=True)
                    file.save(temp_path)
                    
                    temp_paths.append(temp_path)
                    filenames.append(filename)
                    
                    print(f"✅ Uploaded file {len(temp_paths)}: {filename}")
                    print(f"✅ Saved to: {temp_path}")
                
                if len(temp_paths) == 0:
                    flash('No valid files selected.', 'error')
                    return redirect(request.url)
                
                # Store file paths in session for duplicate/invalid review
                session['pending_import_file'] = temp_paths[0] if len(temp_paths) == 1 else temp_paths
                session['pending_import_filename'] = filenames[0] if len(filenames) == 1 else filenames
                session['pending_import_files_multiple'] = len(temp_paths) > 1
                
                temp_path = temp_paths[0] if len(temp_paths) == 1 else temp_paths
                filename = filenames[0] if len(filenames) == 1 else ', '.join(filenames)
                
                print(f"✅ Total files uploaded: {len(temp_paths)}")

            # Determine if we're processing multiple files
            is_multiple_files = session.get('pending_import_files_multiple', False)
            files_to_process = []
            
            if is_multiple_files:
                # Multiple files mode
                if isinstance(temp_path, list):
                    files_to_process = list(zip(temp_path, filename.split(', ') if isinstance(filename, str) else filename))
                else:
                    files_to_process = [(temp_path, filename)]
            else:
                # Single file mode (existing behavior)
                files_to_process = [(temp_path, filename)]
            
            print(f"📁 Processing {len(files_to_process)} file(s)")
            
            try:
                import_service = TimeAttendanceImportService(db, logger_handler)
                
                # Get import options
                skip_duplicates = request.form.get('skip_duplicates', 'true').lower() == 'true'
                validate_only = request.form.get('validate_only', 'false').lower() == 'true'
                analyze_duplicates = request.form.get('analyze_duplicates', 'false').lower() == 'true'
                analyze_invalid = request.form.get('analyze_invalid', 'false').lower() == 'true'
                
                print(f"📋 Import Options:")
                print(f"   Skip duplicates: {skip_duplicates}")
                print(f"   Validate only: {validate_only}")
                print(f"   Analyze duplicates: {analyze_duplicates}")
                print(f"   Analyze invalid: {analyze_invalid}")
                print(f"   Coming from invalid review: {coming_from_invalid_review}")
                
                # Store combined results for multiple files
                all_results = {
                    'total_files': len(files_to_process),
                    'successful_files': 0,
                    'failed_files': 0,
                    'total_imported': 0,
                    'total_duplicates': 0,
                    'total_failed': 0,
                    'file_results': [],
                    'errors': [],
                    'warnings': []
                }
                
                # Process each file
                for file_index, (current_temp_path, current_filename) in enumerate(files_to_process, 1):
                    print(f"\n📄 Processing file {file_index}/{len(files_to_process)}: {current_filename}")
                    
                    import_result = None  # Initialize to prevent reference errors
                    
                    try:
                        # For multiple files, skip review screens and import directly
                        if is_multiple_files:
                            print(f"   📦 Batch mode: processing directly without review screens")
                            
                            # Validate the file first
                            validation_result = import_service.validate_excel_file(current_temp_path)
                            
                            if not validation_result['valid']:
                                raise Exception(f"Validation failed: {'; '.join(validation_result['errors'])}")
                            
                            # Get import settings
                            project_id = request.form.get('project_id')
                            project_id = int(project_id) if project_id and project_id != '' else None
                            import_source = request.form.get('import_source', f"Batch Import - {current_filename}")
                            
                            # Import the file (always skip duplicates in batch mode)
                            import_result = import_service.import_from_excel(
                                current_temp_path,
                                created_by=session['user_id'],
                                import_source=import_source,
                                skip_duplicates=True,  # Always skip duplicates in batch mode
                                force_import_hashes=set(),
                                project_id=project_id
                            )
                            
                        else:
                            # Single file - use existing review workflow logic below
                            # This continues to the existing code after the loop
                            pass
                        
                        # Accumulate results if import was performed
                        if import_result and import_result.get('success'):
                            all_results['successful_files'] += 1
                            all_results['total_imported'] += import_result.get('imported_records', 0)
                            all_results['total_duplicates'] += import_result.get('duplicate_records', 0)
                            all_results['file_results'].append({
                                'filename': current_filename,
                                'status': 'success',
                                'imported': import_result.get('imported_records', 0),
                                'batch_id': import_result.get('batch_id', '')
                            })
                            print(f"   ✅ Imported {import_result.get('imported_records', 0)} records")
                        elif import_result:
                            # Import ran but failed
                            all_results['failed_files'] += 1
                            all_results['total_failed'] += import_result.get('failed_records', 0)
                            all_results['errors'].append(f"{current_filename}: Import failed")
                            all_results['file_results'].append({
                                'filename': current_filename,
                                'status': 'failed',
                                'error': 'Import returned unsuccessful status'
                            })
                        
                    except Exception as file_error:
                        print(f"❌ Error processing file {current_filename}: {file_error}")
                        logger_handler.logger.error(f"Error processing file {current_filename}: {file_error}")
                        all_results['failed_files'] += 1
                        all_results['errors'].append(f"{current_filename}: {str(file_error)}")
                        all_results['file_results'].append({
                            'filename': current_filename,
                            'status': 'failed',
                            'error': str(file_error)
                        })
                        continue
                    
                    finally:
                        # Cleanup individual file (only for multiple file mode, single file cleanup happens later)
                        if is_multiple_files and os.path.exists(current_temp_path):
                            try:
                                os.remove(current_temp_path)
                                print(f"   🗑️ Cleaned up temp file")
                            except Exception as cleanup_error:
                                print(f"   ⚠️ Failed to cleanup temp file: {cleanup_error}")
                
                # After processing all files
                if is_multiple_files:
                    # Log the batch import activity
                    logger_handler.logger.info(
                        f"Batch Import: User {session.get('username', 'unknown')} imported time attendance data from {len(files_to_process)} files - "
                        f"Successful: {all_results['successful_files']}/{all_results['total_files']}, "
                        f"Total imported: {all_results['total_imported']}, "
                        f"Duplicates: {all_results['total_duplicates']}"
                    )
                    
                    # Show combined results
                    if all_results['successful_files'] > 0:
                        flash(f"✅ Successfully imported {all_results['total_imported']} records from {all_results['successful_files']}/{all_results['total_files']} files.", 'success')
                        
                        if all_results['total_duplicates'] > 0:
                            flash(f"ℹ️ Skipped {all_results['total_duplicates']} duplicate records across all files.", 'info')
                    
                    if all_results['failed_files'] > 0:
                        flash(f"❌ {all_results['failed_files']} file(s) failed to import.", 'error')
                        
                        # Show first few error details
                        for error in all_results['errors'][:3]:
                            flash(f"Error: {error}", 'error')
                        
                        if len(all_results['errors']) > 3:
                            flash(f"...and {len(all_results['errors']) - 3} more errors", 'error')
                    
                    # Clear session
                    session.pop('pending_import_file', None)
                    session.pop('pending_import_filename', None)
                    session.pop('pending_import_files_multiple', None)
                    
                    print(f"\n📊 Batch Import Summary:")
                    print(f"   Total files: {all_results['total_files']}")
                    print(f"   Successful: {all_results['successful_files']}")
                    print(f"   Failed: {all_results['failed_files']}")
                    print(f"   Total imported: {all_results['total_imported']}")
                    print(f"   Total duplicates: {all_results['total_duplicates']}")
                    
                    return redirect(url_for('time_attendance.time_attendance_dashboard'))

                # Check if this is coming from duplicate review
                force_import_hashes = request.form.getlist('force_import_hashes[]')
                
                # If analyzing for duplicates, show review page (but not if coming from invalid/duplicate review)
                if analyze_duplicates and not force_import_hashes and not coming_from_invalid_review and not coming_from_duplicate_review:
                    print("🔍 Analyzing for duplicates...")
                    duplicate_analysis = import_service.analyze_for_duplicates(temp_path)
                    
                    if duplicate_analysis['duplicate_records'] > 0:
                        print(f"⚠️ Found {duplicate_analysis['duplicate_records']} duplicates")
                        # Get project_id from form
                        project_id = request.form.get('project_id')
                        # Show duplicate review page
                        return render_template('time_attendance_duplicate_review.html',
                                            analysis=duplicate_analysis,
                                            filename=filename,
                                            project_id=project_id)
                    else:
                        print("✅ No duplicates found")
                        flash('No duplicates found. Proceeding with import.', 'info')
                
                # Check for invalid rows and show review if any (but not if coming from invalid review)
                if analyze_invalid and not coming_from_invalid_review:
                    print("🔍 Analyzing for invalid rows...")
                    invalid_analysis = import_service.analyze_for_invalid_rows(temp_path)
                    
                    if invalid_analysis['invalid_rows'] > 0:
                        print(f"⚠️ Found {invalid_analysis['invalid_rows']} invalid rows")
                        # Get project_id from form
                        project_id = request.form.get('project_id')
                        # Show invalid row review page
                        return render_template('time_attendance_invalid_review.html',
                                            analysis=invalid_analysis,
                                            filename=filename,
                                            project_id=project_id)
                    else:
                        print("✅ All rows are valid")
                        flash('All rows are valid. Proceeding with import.', 'info')
                
                # If coming from invalid review, skip validation (already done)
                if not coming_from_invalid_review:
                    print("🔍 Validating file...")
                    # Validate file
                    validation_result = import_service.validate_excel_file(temp_path)
                    
                    if not validation_result['valid']:
                        print(f"❌ Validation failed: {validation_result['errors']}")
                        flash(f"File validation failed: {'; '.join(validation_result['errors'])}", 'error')
                        return render_template('time_attendance_import.html', 
                                             validation_result=validation_result)
                    
                    if validation_result['warnings']:
                        for warning in validation_result['warnings']:
                            flash(warning, 'warning')
                    
                    if validate_only:
                        print(f"✅ Validation successful: {validation_result['valid_rows']} valid records")
                        flash(f"File validation successful! Found {validation_result['valid_rows']} valid records.", 'success')
                        return render_template('time_attendance_import.html', 
                                             validation_result=validation_result)
                else:
                    print("⏭️ Skipping validation (already validated)")
                
                # Proceed with import
                print("🚀 Starting import process...")
                import_source = request.form.get('import_source', f"Manual Import - {filename}")
                project_id = request.form.get('project_id')
                project_id = int(project_id) if project_id and project_id != '' else None

                import_result = import_service.import_from_excel(
                    temp_path,
                    created_by=session['user_id'],
                    import_source=import_source,
                    skip_duplicates=skip_duplicates,
                    force_import_hashes=force_import_hashes,
                    project_id=project_id
                )
                
                if import_result['success']:
                    print(f"✅ Import successful!")
                    print(f"   Batch ID: {import_result['batch_id']}")
                    print(f"   Imported: {import_result['imported_records']}/{import_result['total_records']}")
                    print(f"   Duplicates: {import_result['duplicate_records']}")
                    print(f"   Failed: {import_result['failed_records']}")
                    
                    logger_handler.logger.info(
                        f"User {session['username']} successfully imported time attendance data - "
                        f"Batch: {import_result['batch_id']}, "
                        f"Records: {import_result['imported_records']}/{import_result['total_records']}, "
                        f"Duplicates: {import_result['duplicate_records']}, "
                        f"Forced: {import_result['forced_duplicates']}, "
                        f"Failed: {import_result['failed_records']}"
                    )
                    
                    flash(f"Import successful! Imported {import_result['imported_records']} records "
                          f"out of {import_result['total_records']} total records.", 'success')
                    
                    if import_result['duplicate_records'] > 0:
                        flash(f"Skipped {import_result['duplicate_records']} duplicate records.", 'info')
                    
                    if import_result['forced_duplicates'] > 0:
                        flash(f"Imported {import_result['forced_duplicates']} duplicate records as requested.", 'info')
                    
                    if import_result['failed_records'] > 0:
                        flash(f"Note: {import_result['failed_records']} records failed to import. "
                              f"Check the error details below.", 'warning')
                    
                    # Clean up temp file after successful import
                    if os.path.exists(temp_path):
                        try:
                            os.remove(temp_path)
                            session.pop('pending_import_file', None)
                            session.pop('pending_import_filename', None)
                            print("🗑️ Cleaned up temp file")
                        except Exception as cleanup_error:
                            print(f"⚠️ Failed to cleanup temp file: {cleanup_error}")
                    
                    return render_template('time_attendance_import_result.html', 
                                         import_result=import_result)
                else:
                    print(f"❌ Import failed: {import_result['errors']}")
                    flash(f"Import failed: {'; '.join(import_result['errors'][:3])}", 'error')
                    if len(import_result['errors']) > 3:
                        flash(f"...and {len(import_result['errors']) - 3} more errors", 'warning')
                    return render_template('time_attendance_import.html', 
                                         import_result=import_result)
                
            except Exception as import_error:
                print(f"❌ Import exception: {import_error}")
                import traceback
                print(f"❌ Traceback: {traceback.format_exc()}")
                raise
                
        except Exception as e:
            logger_handler.log_database_error('time_attendance_import', e)
            print(f"❌ Top-level exception: {e}")
            import traceback
            print(f"❌ Traceback: {traceback.format_exc()}")
            flash('Import failed due to an unexpected error.', 'error')
            return render_template('time_attendance_import.html')
    
    # GET request
    return render_template('time_attendance_import.html')


@bp.route('/time-attendance/import/analyze-duplicates', methods=['POST'], endpoint='analyze_import_duplicates')
@login_required
def analyze_import_duplicates():
    """AJAX endpoint to analyze file for duplicates"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Invalid file format'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', '/tmp'), 
                               f"analyze_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
        
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)
        
        # Store in session
        session['pending_import_file'] = temp_path
        session['pending_import_filename'] = filename
        
        try:
            import_service = TimeAttendanceImportService(db, logger_handler)
            analysis = import_service.analyze_for_duplicates(temp_path)
            
            # Convert datetime objects to strings for JSON
            for duplicate in analysis.get('duplicates', []):
                if 'new_record' in duplicate:
                    if duplicate['new_record'].get('attendance_date'):
                        duplicate['new_record']['attendance_date'] = str(duplicate['new_record']['attendance_date'])
                    if duplicate['new_record'].get('attendance_time'):
                        duplicate['new_record']['attendance_time'] = str(duplicate['new_record']['attendance_time'])
                
                if 'existing_record' in duplicate:
                    if duplicate['existing_record'].get('attendance_date'):
                        duplicate['existing_record']['attendance_date'] = str(duplicate['existing_record']['attendance_date'])
                    if duplicate['existing_record'].get('attendance_time'):
                        duplicate['existing_record']['attendance_time'] = str(duplicate['existing_record']['attendance_time'])
                    if duplicate['existing_record'].get('import_date'):
                        duplicate['existing_record']['import_date'] = str(duplicate['existing_record']['import_date'])
            
            return jsonify({
                'success': True,
                'analysis': analysis
            })
        
        except Exception as e:
            # Cleanup on error
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise e
    
    except Exception as e:
        logger_handler.logger.error(f"Duplicate analysis error: {e}")
        return jsonify({
            'success': False,
            'message': f'Analysis failed: {str(e)}'
        }), 500
    
@bp.route('/time-attendance/import/analyze-invalid', methods=['POST'], endpoint='analyze_import_invalid')
@login_required
def analyze_import_invalid():
    """AJAX endpoint to analyze file for invalid rows"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Invalid file format'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', '/tmp'), 
                               f"analyze_invalid_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
        
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)
        
        # Store in session
        session['pending_import_file'] = temp_path
        session['pending_import_filename'] = filename
        
        try:
            import_service = TimeAttendanceImportService(db, logger_handler)
            analysis = import_service.analyze_for_invalid_rows(temp_path)
            
            # Convert datetime objects to strings for JSON
            for invalid in analysis.get('invalid_details', []):
                if 'row_data' in invalid:
                    if invalid['row_data'].get('attendance_date'):
                        invalid['row_data']['attendance_date'] = str(invalid['row_data']['attendance_date'])
                    if invalid['row_data'].get('attendance_time'):
                        invalid['row_data']['attendance_time'] = str(invalid['row_data']['attendance_time'])
            
            return jsonify({
                'success': True,
                'analysis': analysis
            })
        
        except Exception as e:
            logger_handler.logger.error(f"Invalid row analysis error: {e}")
            return jsonify({
                'success': False,
                'message': f'Analysis failed: {str(e)}'
            }), 500
        
    except Exception as e:
        logger_handler.logger.error(f"Invalid row analysis error: {e}")
        return jsonify({
            'success': False,
            'message': f'Analysis failed: {str(e)}'
        }), 500


# ---------------------------------------------------------------------------
# Time Attendance Import — SSE progress streaming (disk-based, multi-worker safe)
#
# Design: progress state is written to a small JSON file on disk so that any
# gunicorn worker process can read it.  No shared in-memory state is required.
# The /stream endpoint runs the import itself (synchronously inside the SSE
# generator) while writing progress to the file and yielding events to the
# browser — compatible with gunicorn gevent workers.
# ---------------------------------------------------------------------------

def _progress_file_path(job_id: str, upload_dir: str = '/tmp') -> str:
    """Return the path for the on-disk progress file for a given job_id."""
    os.makedirs(upload_dir, exist_ok=True)
    return os.path.join(upload_dir, f"import_progress_{job_id}.json")


def _write_progress(job_id: str, event: dict, upload_dir: str = '/tmp') -> None:
    """Atomically write the latest progress event to disk."""
    path = _progress_file_path(job_id, upload_dir)
    try:
        tmp = path + '.tmp'
        with open(tmp, 'w') as f:
            json.dump(event, f)
        os.replace(tmp, path)
    except Exception:
        pass  # Best-effort; import will continue regardless


@bp.route('/time-attendance/import/start', methods=['POST'], endpoint='start_import_job')
@login_required
def start_import_job():
    """
    Validates the uploaded file, saves it to disk, stores import options in a
    progress file, then returns a job_id.  The actual import runs inside the
    SSE stream endpoint so no background thread or shared memory is needed.
    """
    try:
        if 'files' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded.'}), 400

        files = request.files.getlist('files')
        if not files or files[0].filename == '':
            return jsonify({'success': False, 'error': 'No file selected.'}), 400

        file = files[0]
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Invalid file format.'}), 400

        filename = secure_filename(file.filename)
        upload_dir = current_app.config.get('UPLOAD_FOLDER', '/tmp')
        os.makedirs(upload_dir, exist_ok=True)
        job_id = str(uuid.uuid4())
        temp_path = os.path.join(upload_dir,
                                 f"stream_{job_id}_{filename}")
        file.save(temp_path)

        # Store import options alongside the file so the stream endpoint can
        # read them without depending on session or shared memory.
        job_meta = {
            'type': 'pending',
            'temp_path': temp_path,
            'filename': filename,
            'skip_duplicates': request.form.get('skip_duplicates', 'true').lower() == 'true',
            'project_id': int(request.form.get('project_id')) if request.form.get('project_id') else None,
            'import_source': request.form.get('import_source', f"Manual Import - {filename}"),
            'created_by': session['user_id'],
            'username': session.get('username', 'unknown'),
        }
        _write_progress(job_id, job_meta, upload_dir)

        logger_handler.logger.info(
            f"User {job_meta['username']} queued time attendance import job {job_id} for file {filename}"
        )
        return jsonify({'success': True, 'job_id': job_id})

    except Exception as e:
        logger_handler.logger.error(f"Error queuing import job: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/time-attendance/import/stream/<job_id>', endpoint='stream_import_progress')
@login_required
def stream_import_progress(job_id):
    """
    SSE endpoint — runs the import synchronously while streaming progress to
    the browser.  Works across multiple gunicorn workers because all state is
    stored on disk (no in-memory job store).
    """
    # Capture upload_dir HERE in the request context — current_app is NOT
    # available inside the background thread (_run) or after context teardown.
    upload_dir = current_app.config.get('UPLOAD_FOLDER', '/tmp')
    progress_path = _progress_file_path(job_id, upload_dir)
    # Capture real app object in request context — safe to use in background thread
    _real_app = current_app._get_current_object()

    def generate():
        import time as _time

        # ── Read the job metadata written by /start ────────────────────────
        deadline = _time.time() + 15  # Wait up to 15 s for the file to appear
        meta = None
        while _time.time() < deadline:
            if os.path.exists(progress_path):
                try:
                    with open(progress_path) as f:
                        meta = json.load(f)
                    break
                except Exception:
                    pass
            yield "data: " + json.dumps({'type': 'heartbeat'}) + "\n\n"
            _time.sleep(0.3)

        if not meta or meta.get('type') != 'pending':
            yield "data: " + json.dumps({
                'type': 'error',
                'message': 'Job metadata not found. Please try importing again.'
            }) + "\n\n"
            return

        temp_path    = meta['temp_path']
        skip_dupes   = meta['skip_duplicates']
        project_id   = meta['project_id']
        import_source= meta['import_source']
        created_by   = meta['created_by']
        username     = meta['username']

        if not os.path.exists(temp_path):
            yield "data: " + json.dumps({
                'type': 'error',
                'message': 'Uploaded file not found. Please try importing again.'
            }) + "\n\n"
            return

        yield "data: " + json.dumps({'type': 'status', 'message': 'Reading and validating file...'}) + "\n\n"

        # ── Run the import with a progress callback ────────────────────────
        try:
            svc = TimeAttendanceImportService(db, logger_handler)

            # progress_callback writes to disk AND yields an SSE event.
            # We collect events in a list so the generator can yield them.
            _pending_events = []

            def on_progress(current, total, message):
                pct = int(current / total * 100) if total else 0
                event = {
                    'type': 'progress',
                    'current': current,
                    'total': total,
                    'percent': pct,
                    'message': message,
                }
                _write_progress(job_id, event, upload_dir)
                _pending_events.append(event)

            # We need to interleave yielding with the synchronous import loop.
            # Strategy: run import_from_excel; the callback appends to
            # _pending_events; after every DB commit batch (50 records) we
            # flush pending events to the SSE stream.
            import threading as _threading
            result_holder = [None]
            error_holder  = [None]
            done_event    = _threading.Event()

            def _run():
                # Push an application context so the thread can access
                # Flask-SQLAlchemy, Employee.query, etc.
                with _real_app.app_context():
                    try:
                        result_holder[0] = svc.import_from_excel(
                            temp_path,
                            created_by=created_by,
                            import_source=import_source,
                            skip_duplicates=skip_dupes,
                            force_import_hashes=[],
                            project_id=project_id,
                            progress_callback=on_progress)
                    except Exception as exc:
                        error_holder[0] = exc
                    finally:
                        done_event.set()

            t = _threading.Thread(target=_run, daemon=True)
            t.start()

            # Yield progress events as they arrive while the import thread runs
            while not done_event.is_set():
                while _pending_events:
                    yield "data: " + json.dumps(_pending_events.pop(0)) + "\n\n"
                yield "data: " + json.dumps({'type': 'heartbeat'}) + "\n\n"
                _time.sleep(0.4)

            # Drain any remaining events after the thread finishes
            while _pending_events:
                yield "data: " + json.dumps(_pending_events.pop(0)) + "\n\n"

            if error_holder[0]:
                raise error_holder[0]

            result = result_holder[0]

            if result and result['success']:
                logger_handler.logger.info(
                    f"User {username} imported {result['imported_records']} time attendance records "
                    f"via stream (batch: {result['batch_id']})"
                )

            # Sanitize result dict for JSON serialization — convert any
            # datetime objects (e.g. import_date) to ISO-format strings.
            if result and isinstance(result.get('import_date'), datetime):
                result['import_date'] = result['import_date'].isoformat()
            done_event_data = {'type': 'done', 'result': result}
            _write_progress(job_id, done_event_data, upload_dir)
            yield "data: " + json.dumps(done_event_data) + "\n\n"

        except Exception as e:
            logger_handler.logger.error(f"Import stream error for job {job_id}: {e}")
            error_event = {'type': 'error', 'message': str(e)}
            _write_progress(job_id, error_event, upload_dir)
            yield "data: " + json.dumps(error_event) + "\n\n"

        finally:
            # Clean up temp files
            for path in (temp_path, progress_path):
                try:
                    if os.path.exists(path):
                        os.remove(path)
                except Exception:
                    pass

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',   # Disable nginx buffering for SSE
        }
    )


@bp.route('/time-attendance/import/cancel-pending', endpoint='cancel_pending_import')
@login_required
def cancel_pending_import():
    """Cancel pending import and cleanup temp file"""
    try:
        if 'pending_import_file' in session:
            temp_path = session['pending_import_file']
            if os.path.exists(temp_path):
                os.remove(temp_path)
            session.pop('pending_import_file')
        
        if 'pending_import_filename' in session:
            session.pop('pending_import_filename')
        
        flash('Import cancelled.', 'info')
    except Exception as e:
        logger_handler.logger.error(f"Error cancelling import: {e}")
    
    return redirect(url_for('time_attendance.import_time_attendance'))



@bp.route('/time-attendance/import/validate', methods=['POST'], endpoint='validate_import_file')
@login_required
def validate_import_file():
    """AJAX endpoint to validate Excel file before import"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Invalid file format'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', '/tmp'), 
                               f"validate_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
        
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)
        
        try:
            # Validate file
            import_service = TimeAttendanceImportService(db, logger_handler)
            validation_result = import_service.validate_excel_file(temp_path)
            
            return jsonify({
                'success': True,
                'validation': validation_result
            })
        
        finally:
            # Cleanup
            if os.path.exists(temp_path):
                os.remove(temp_path)
    
    except Exception as e:
        logger_handler.logger.error(f"Validation error: {e}")
        return jsonify({
            'success': False,
            'message': f'Validation failed: {str(e)}'
        }), 500
    
@bp.route('/time-attendance/import/batch/<batch_id>', endpoint='view_import_batch')
@login_required
@log_user_activity('view_import_batch')
def view_import_batch(batch_id):
    """View details of a specific import batch"""
    try:
        import_service = TimeAttendanceImportService(db, logger_handler)
        batch_summary = import_service.get_import_summary(batch_id)
        
        if not batch_summary:
            flash('Import batch not found.', 'error')
            return redirect(url_for('time_attendance.time_attendance_dashboard'))
        
        return render_template('time_attendance_batch_detail.html',
                             batch_summary=batch_summary)
    
    except Exception as e:
        logger_handler.logger.error(f"Error viewing batch {batch_id}: {e}")
        flash('Error loading batch details.', 'error')
        return redirect(url_for('time_attendance.time_attendance_dashboard'))


@bp.route('/time-attendance/import/batch/<batch_id>/delete', methods=['POST'], endpoint='delete_import_batch')
@admin_required
@log_database_operations('delete_import_batch')
def delete_import_batch(batch_id):
    """Delete an entire import batch"""
    try:
        import_service = TimeAttendanceImportService(db, logger_handler)
        result = import_service.delete_import_batch(batch_id, deleted_by=session['user_id'])
        
        if result['success']:
            flash(result['message'], 'success')
            logger_handler.logger.info(
                f"User {session['username']} deleted import batch {batch_id} - "
                f"{result['deleted_count']} records removed"
            )
        else:
            flash(result['message'], 'error')
        
        return redirect(url_for('time_attendance.time_attendance_dashboard'))
    
    except Exception as e:
        logger_handler.logger.error(f"Error deleting batch {batch_id}: {e}")
        flash('Error deleting import batch.', 'error')
        return redirect(url_for('time_attendance.time_attendance_dashboard'))


@bp.route('/time-attendance/import/download-template', endpoint='download_import_template')
@login_required
def download_import_template():
    """Download Excel template for time attendance import"""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Time Attendance Template"
        
        # Define headers
        headers = ['ID', 'Name', 'Platform', 'Date', 'Time', 'Location Name', 
                   'Action Description', 'Event Description', 'Recorded Address', 'Distance']
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add sample data rows
        sample_data = [
            ['12345', 'John Doe', 'iPhone - iOS', '2025-10-06', '09:00:00', 
             'HQ Suite 210', 'Check In', 'Main Office', '123 Main St', '0.125'],
            ['67890', 'Jane Smith', 'Android', '2025-10-06', '08:45:00', 
             'Branch Office', 'Check In', 'Morning Entry', '456 Oak Avenue', '0.250'],
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Time Attendance Import Template - Instructions"],
            [""],
            ["Required Columns:"],
            ["- ID: Employee ID (required)"],
            ["- Name: Employee full name (required)"],
            ["- Date: Attendance date in YYYY-MM-DD format (required)"],
            ["- Time: Attendance time in HH:MM:SS format (required)"],
            ["- Location Name: Location where attendance was recorded (required)"],
            ["- Action Description: Type of action (e.g., Check In, Check Out) (required)"],
            [""],
            ["Optional Columns:"],
            ["- Platform: Device platform (e.g., iPhone - iOS, Android)"],
            ["- Event Description: Additional event details"],
            ["- Recorded Address: Physical address where attendance was recorded"],
            ["- Distance: Distance in miles between Building and Recorded Address (optional)"],
            [""],
            ["Important Notes:"],
            ["- Do not modify the header row"],
            ["- Ensure all required fields have values"],
            ["- Date format must be YYYY-MM-DD (e.g., 2025-10-06)"],
            ["- Time format must be HH:MM:SS (e.g., 09:00:00)"],
            ["- Remove the sample data rows before importing your actual data"],
            ["- Duplicate records will be automatically detected and skipped"],
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row_num, column=1, value=instruction[0])
        
        ws_instructions.column_dimensions['A'].width = 80
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Log download
        logger_handler.logger.info(f"User {session['username']} downloaded import template")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'time_attendance_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    except Exception as e:
        logger_handler.logger.error(f"Error generating template: {e}")
        flash('Error generating template file.', 'error')
        return redirect(url_for('time_attendance.import_time_attendance'))
    
@bp.route('/time-attendance/export', endpoint='export_time_attendance')
@login_required
@log_user_activity('time_attendance_export')
def export_time_attendance():
    """Export time attendance records to CSV or Excel"""
    try:
        export_format = request.args.get('format', 'excel').lower()
        
        # Get filter parameters (same as records page)
        employee_filter = request.args.get('employee_id')
        location_filter = request.args.get('location_name')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        import_batch = request.args.get('import_batch')
        project_filter = request.args.get('project_id')
        
        # Build query with same filters as the view
        query = TimeAttendance.query
        
        # Apply filters — employee_id supports comma-separated multi-employee values
        if employee_filter:
            employee_ids_export = [e.strip() for e in employee_filter.split(',') if e.strip()]
            from working_hours_calculator import parse_employee_id_for_work_type as _parse_wt
            all_variants = []
            for eid in employee_ids_export:
                _base_emp_id, _ = _parse_wt(str(eid))
                all_variants += [
                    _base_emp_id,
                    f"{_base_emp_id} SP", f"{_base_emp_id}SP",
                    f"SP {_base_emp_id}", f"SP{_base_emp_id}",
                    f"{_base_emp_id} PW", f"{_base_emp_id}PW",
                    f"PW {_base_emp_id}", f"PW{_base_emp_id}",
                    f"{_base_emp_id} PT", f"{_base_emp_id}PT",
                    f"PT {_base_emp_id}", f"PT{_base_emp_id}",
                ]
            query = query.filter(TimeAttendance.employee_id.in_(all_variants))

        if location_filter:
            query = query.filter(TimeAttendance.location_name == location_filter)

        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(TimeAttendance.attendance_date >= start_date_obj)
            except ValueError:
                flash('Invalid start date format.', 'error')
                return redirect(url_for('time_attendance.time_attendance_records'))

        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                # Fetch one extra calendar day beyond the requested end date so that
                # early-morning check-out records stored on Day N+1 (overnight shifts
                # ending after midnight on the last report day) are available for the
                # overnight pairing detection inside export_time_attendance_excel.
                # The displayed date range is controlled by start_date_filter /
                # end_date_filter inside that function and is not affected.
                query = query.filter(TimeAttendance.attendance_date <= end_date_obj + timedelta(days=1))
            except ValueError:
                flash('Invalid end date format.', 'error')
                return redirect(url_for('time_attendance.time_attendance_records'))

        if import_batch:
            query = query.filter(TimeAttendance.import_batch_id == import_batch)

        if project_filter:
            query = query.filter(TimeAttendance.project_id == project_filter)

        # Order by date and time (most recent first)
        records = query.order_by(
            TimeAttendance.attendance_date.desc(),
            TimeAttendance.attendance_time.desc()
        ).all()

        if not records:
            flash('No records found to export.', 'warning')
            return redirect(url_for('time_attendance.time_attendance_records'))

        # Get project name if project filter exists
        project_name_for_filename = ''
        if project_filter:
            try:
                project = Project.query.get(int(project_filter))
                if project:
                    # Replace spaces and special characters with underscores
                    project_name_safe = project.name.replace(' ', '_').replace('/', '_').replace('\\', '_')
                    project_name_for_filename = f"{project_name_safe}_"
            except Exception as e:
                print(f"⚠️ Error getting project name for filename: {e}")

        # Log export
        logger_handler.logger.info(
            f"User {session['username']} exported {len(records)} time attendance records "
            f"in {export_format.upper()} format"
        )
        
        # Format dates for filename (MMDDYYYY format)
        date_from_formatted = ''
        date_to_formatted = ''
        if start_date:
            try:
                date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                date_from_formatted = date_obj.strftime('%m%d%Y')
            except ValueError:
                pass
        
        if end_date:
            try:
                date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                date_to_formatted = date_obj.strftime('%m%d%Y')
            except ValueError:
                pass
        
        # Build filename with date range
        # Format: [project_name_]time_attendance_[fromdate_todate].xlsx/csv
        date_range_str = ''
        if date_from_formatted and date_to_formatted:
            date_range_str = f"{date_from_formatted}_{date_to_formatted}"
        elif date_from_formatted:
            date_range_str = f"from_{date_from_formatted}"
        elif date_to_formatted:
            date_range_str = f"to_{date_to_formatted}"
        
        # Keep the filter_str for backward compatibility (but not in filename anymore)
        filter_desc = []
        if employee_filter:
            filter_desc.append(f"emp_{employee_filter}")
        if location_filter:
            filter_desc.append(f"loc_{location_filter[:10]}")
        
        filter_str = "_".join(filter_desc) if filter_desc else "all"
        
        return export_time_attendance_excel(records, project_name_for_filename, date_range_str, filter_str, start_date, end_date)
    
    except Exception as e:
        logger_handler.logger.error(f"Error exporting time attendance records: {e}")
        flash('Error generating export file. Please try again.', 'error')
        return redirect(url_for('time_attendance.time_attendance_records'))

@bp.route('/time-attendance/export/excel', endpoint='excel_export_time_attendance')
@login_required
@log_user_activity('time_attendance_excel_export')
def excel_export_time_attendance():
    """Excel export with current page filters"""
    # Redirect to main export with Excel format
    return redirect(url_for('time_attendance.export_time_attendance', format='excel', **request.args))

@bp.route('/time-attendance/export-by-building', endpoint='export_time_attendance_by_building')
@login_required
@log_user_activity('time_attendance_export_by_building')
def export_time_attendance_by_building():
    """Export time attendance records grouped by building/location to Excel"""
    try:
        # Get filter parameters (same as records page)
        employee_filter = request.args.get('employee_id')
        location_filter = request.args.get('location_name')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        import_batch = request.args.get('import_batch')
        project_filter = request.args.get('project_id')
        
        # Build query with same filters as the view
        query = TimeAttendance.query
        
        # Apply filters — employee_id supports comma-separated multi-employee values
        if employee_filter:
            employee_ids_export = [e.strip() for e in employee_filter.split(',') if e.strip()]
            from working_hours_calculator import parse_employee_id_for_work_type as _parse_wt
            all_variants = []
            for eid in employee_ids_export:
                _base_emp_id, _ = _parse_wt(str(eid))
                all_variants += [
                    _base_emp_id,
                    f"{_base_emp_id} SP", f"{_base_emp_id}SP",
                    f"SP {_base_emp_id}", f"SP{_base_emp_id}",
                    f"{_base_emp_id} PW", f"{_base_emp_id}PW",
                    f"PW {_base_emp_id}", f"PW{_base_emp_id}",
                    f"{_base_emp_id} PT", f"{_base_emp_id}PT",
                    f"PT {_base_emp_id}", f"PT{_base_emp_id}",
                ]
            query = query.filter(TimeAttendance.employee_id.in_(all_variants))

        if location_filter:
            query = query.filter(TimeAttendance.location_name == location_filter)

        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(TimeAttendance.attendance_date >= start_date_obj)
            except ValueError:
                flash('Invalid start date format.', 'error')
                return redirect(url_for('time_attendance.time_attendance_records'))

        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                # Fetch one extra calendar day so that early-morning check-out records
                # stored on Day N+1 (overnight shifts ending after midnight on the last
                # report day) are included for overnight pairing detection.
                # The display range remains controlled by start_date_filter/end_date_filter
                # inside export_time_attendance_by_building_excel and is not affected.
                query = query.filter(TimeAttendance.attendance_date <= end_date_obj + timedelta(days=1))
            except ValueError:
                flash('Invalid end date format.', 'error')
                return redirect(url_for('time_attendance.time_attendance_records'))

        if import_batch:
            query = query.filter(TimeAttendance.import_batch_id == import_batch)

        if project_filter:
            query = query.filter(TimeAttendance.project_id == project_filter)

        # Order by location, date, and time
        records = query.order_by(
            TimeAttendance.location_name,
            TimeAttendance.attendance_date.desc(),
            TimeAttendance.attendance_time.desc()
        ).all()

        if not records:
            flash('No records found to export.', 'warning')
            return redirect(url_for('time_attendance.time_attendance_records'))

        # Get project name if project filter exists
        project_name_for_filename = ''
        if project_filter:
            try:
                project = Project.query.get(int(project_filter))
                if project:
                    # Replace spaces and special characters with underscores
                    project_name_safe = project.name.replace(' ', '_').replace('/', '_').replace('\\', '_')
                    project_name_for_filename = f"{project_name_safe}_"
            except Exception as e:
                print(f"⚠️ Error getting project name for filename: {e}")

        # Log export
        logger_handler.logger.info(
            f"User {session['username']} exported {len(records)} time attendance records "
            f"by building in Excel format"
        )
        
        # Format dates for filename (MMDDYYYY format)
        date_from_formatted = ''
        date_to_formatted = ''
        if start_date:
            try:
                date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                date_from_formatted = date_obj.strftime('%m%d%Y')
            except ValueError:
                pass
        
        if end_date:
            try:
                date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                date_to_formatted = date_obj.strftime('%m%d%Y')
            except ValueError:
                pass
        
        # Build filename with date range
        date_range_str = ''
        if date_from_formatted and date_to_formatted:
            date_range_str = f"{date_from_formatted}_{date_to_formatted}"
        elif date_from_formatted:
            date_range_str = f"from_{date_from_formatted}"
        elif date_to_formatted:
            date_range_str = f"to_{date_to_formatted}"
        
        return export_time_attendance_by_building_excel(records, project_name_for_filename, date_range_str, start_date, end_date)
    
    except Exception as e:
        logger_handler.logger.error(f"Error exporting time attendance records by building: {e}")
        flash('Error generating export file. Please try again.', 'error')
        return redirect(url_for('time_attendance.time_attendance_records'))
    
@bp.route('/time-attendance/records', endpoint='time_attendance_records')
@login_required
@log_user_activity('time_attendance_records_view')
def time_attendance_records():
    """Display time attendance records with filtering options"""
    try:
        # Get filter parameters
        employee_filter = request.args.get('employee_id', '')
        location_filter = request.args.get('location_name')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        project_filter = request.args.get('project_id')
        page = request.args.get('page', 1, type=int)
        per_page = 50  # Records per page

        # Build list of selected employee IDs (comma-separated multi-employee support)
        employee_ids = [e.strip() for e in employee_filter.split(',') if e.strip()] if employee_filter else []

        # Build display names for each selected employee
        import re as _re
        employee_display_names = []
        for eid in employee_ids:
            try:
                numeric_only = _re.search(r'\d+', str(eid))
                if numeric_only:
                    emp = Employee.query.filter_by(id=int(numeric_only.group(0))).first()
                    if emp:
                        employee_display_names.append({'id': eid, 'name': f"{emp.lastName}, {emp.firstName}"})
                    else:
                        employee_display_names.append({'id': eid, 'name': f"ID: {eid}"})
                else:
                    employee_display_names.append({'id': eid, 'name': eid})
            except (ValueError, TypeError):
                employee_display_names.append({'id': eid, 'name': eid})

        employee_display_name = ', '.join([e['name'] for e in employee_display_names])

        # Build query
        query = TimeAttendance.query

        # Apply filters
        if employee_ids:
            # Expand each base ID to include all SP/PW/PT work-type variants so that
            # cross-type pairs are included in results and exports.
            from working_hours_calculator import parse_employee_id_for_work_type as _parse_wt
            all_variants = []
            for eid in employee_ids:
                _base_emp_id, _ = _parse_wt(str(eid))
                all_variants += [
                    _base_emp_id,
                    f"{_base_emp_id} SP", f"{_base_emp_id}SP",
                    f"SP {_base_emp_id}", f"SP{_base_emp_id}",
                    f"{_base_emp_id} PW", f"{_base_emp_id}PW",
                    f"PW {_base_emp_id}", f"PW{_base_emp_id}",
                    f"{_base_emp_id} PT", f"{_base_emp_id}PT",
                    f"PT {_base_emp_id}", f"PT{_base_emp_id}",
                ]
            query = query.filter(TimeAttendance.employee_id.in_(all_variants))
            logger_handler.logger.info(
                f"Time attendance records filtered by employee IDs: {employee_ids} "
                f"by user {session.get('username', 'unknown')}"
            )
        
        if location_filter:
            query = query.filter(TimeAttendance.location_name == location_filter)

        if project_filter:
            query = query.filter(TimeAttendance.project_id == project_filter)
        
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(TimeAttendance.attendance_date >= start_date_obj)
            except ValueError:
                flash('Invalid start date format.', 'error')
        
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(TimeAttendance.attendance_date <= end_date_obj)
            except ValueError:
                flash('Invalid end date format.', 'error')
        
        # Order by date and time (most recent first)
        query = query.order_by(
            TimeAttendance.attendance_date.desc(),
            TimeAttendance.attendance_time.desc()
        )
        
        # Paginate results
        records = query.paginate(page=page, per_page=per_page, error_out=False)
        
        # Enhance records with QR address and location accuracy
        for record in records.items:
            # Find matching QR code by location name
            qr_code = QRCode.query.filter_by(location=record.location_name).first()
            
            if qr_code:
                record.qr_address = qr_code.location_address
                
                # Calculate location accuracy if coordinates are available
                if record.recorded_address and qr_code.location_address:
                    try:
                        # Try to calculate location accuracy
                        location_accuracy = calculate_location_accuracy_enhanced(
                            qr_address=qr_code.location_address,
                            checkin_address=record.recorded_address,
                            checkin_lat=None,  # TimeAttendance doesn't have GPS coords
                            checkin_lng=None
                        )
                        record.location_accuracy = location_accuracy
                    except Exception as e:
                        logger_handler.logger.warning(f"Could not calculate location accuracy for record {record.id}: {e}")
                        record.location_accuracy = None
                else:
                    record.location_accuracy = None
            else:
                record.qr_address = None
                record.location_accuracy = None

            # Resolve employee name by stripping work type prefix/suffix (SP, PW, PT)
            # from employee_id ONLY for the lookup. The original employee_id is kept intact.
            # e.g. '3937SP', 'SP3937', 'PW3937' -> lookup by numeric '3937'
            try:
                import re as _re
                numeric_only = _re.search(r'\d+', str(record.employee_id or ''))
                if numeric_only:
                    emp = Employee.query.filter_by(id=int(numeric_only.group(0))).first()
                    record.resolved_employee_name = f"{emp.lastName}, {emp.firstName}" if emp else record.employee_name
                else:
                    record.resolved_employee_name = record.employee_name
            except Exception as e:
                logger_handler.logger.warning(f"Could not resolve employee name for ID {record.employee_id}: {e}")
                record.resolved_employee_name = record.employee_name
        
        # Get unique employees and locations for filters
        unique_employees = TimeAttendance.get_unique_employees()
        unique_locations = TimeAttendance.get_unique_locations()
        projects = Project.query.filter_by(active_status=True).order_by(Project.name).all()

        return render_template(
            'time_attendance_records.html',
            records=records,
            unique_employees=unique_employees,
            unique_locations=unique_locations,
            projects=projects,
            employee_display_name=employee_display_name,
            employee_display_names=employee_display_names,
            employee_filter=employee_filter,
            employee_ids=employee_ids
        )
        
    except Exception as e:
        logger_handler.logger.error(f"Error displaying time attendance records: {e}")
        flash('Error loading attendance records.', 'error')
        return redirect(url_for('time_attendance.time_attendance_dashboard'))

@bp.route('/time-attendance/record/<int:record_id>', endpoint='time_attendance_record_detail')
@login_required
@log_user_activity('time_attendance_record_detail')
def time_attendance_record_detail(record_id):
    """Display detailed view of a time attendance record"""
    try:
        record = TimeAttendance.query.get_or_404(record_id)
        return render_template('time_attendance_record_detail.html', record=record)
        
    except Exception as e:
        logger_handler.logger.error(f"Error viewing time attendance record {record_id}: {e}")
        flash('Error loading record details.', 'error')
        return redirect(url_for('time_attendance.time_attendance_records'))

@bp.route('/time-attendance/delete/<int:record_id>', methods=['POST'], endpoint='delete_time_attendance_record')
@admin_required
@log_database_operations('time_attendance_delete')
def delete_time_attendance_record(record_id):
    """Delete a time attendance record"""
    try:
        record = TimeAttendance.query.get_or_404(record_id)
        
        # Store record info for logging
        employee_info = f"{record.employee_name} (ID: {record.employee_id})"
        location_info = record.location_name
        date_info = record.attendance_date
        
        # Delete the record
        db.session.delete(record)
        db.session.commit()
        
        # Log deletion
        logger_handler.logger.info(
            f"User {session['username']} deleted time attendance record {record_id} - "
            f"Employee: {employee_info}, Location: {location_info}, Date: {date_info}"
        )
        
        flash(f'Time attendance record for {employee_info} deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        logger_handler.log_database_error('time_attendance_delete', e)
        flash('Failed to delete time attendance record.', 'error')
    
    # Get filter parameters from BOTH request.form (POST) and request.args (GET query params)
    # This handles both the records list page and the detail page
    filter_params = {}
    
    # List of possible filter parameters
    filter_keys = ['employee_id', 'location_name', 'project_id', 'start_date', 'end_date', 'page']
    
    for key in filter_keys:
        # Try to get from form data first (records list page)
        value = request.form.get(key)
        # If not in form, try query parameters (detail page)
        if not value:
            value = request.args.get(key)
        # Only include if value exists and is not empty
        if value:
            filter_params[key] = value
    
    # Redirect back with filters preserved
    return redirect(url_for('time_attendance.time_attendance_records', **filter_params))

@bp.route('/api/time-attendance/employee/<employee_id>', endpoint='api_time_attendance_by_employee')
@login_required
def api_time_attendance_by_employee(employee_id):
    """API endpoint to get time attendance records for a specific employee"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        start_date_obj = None
        end_date_obj = None
        
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        records = TimeAttendance.get_by_employee_id(employee_id, start_date_obj, end_date_obj)
        
        return jsonify({
            'success': True,
            'employee_id': employee_id,
            'total_records': len(records),
            'records': [record.to_dict() for record in records]
        })
        
    except Exception as e:
        logger_handler.logger.error(f"API error getting time attendance for employee {employee_id}: {e}")
        return jsonify({
            'success': False,
            'error': 'Failed to retrieve time attendance records'
        }), 500

@bp.route('/api/time-attendance/location/<location_name>', endpoint='api_time_attendance_by_location')
@login_required
def api_time_attendance_by_location(location_name):
    """API endpoint to get time attendance records for a specific location"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        start_date_obj = None
        end_date_obj = None
        
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        records = TimeAttendance.get_by_location(location_name, start_date_obj, end_date_obj)
        
        return jsonify({
            'success': True,
            'location_name': location_name,
            'total_records': len(records),
            'records': [record.to_dict() for record in records]
        })
        
    except Exception as e:
        logger_handler.logger.error(f"API error getting time attendance for location {location_name}: {e}")
        return jsonify({
            'success': False,
            'error': 'Failed to retrieve time attendance records'
        }), 500
    


# Jinja2 filters for better template functionality