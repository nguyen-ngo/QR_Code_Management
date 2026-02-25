"""
Enhanced Time Attendance Import Service with Duplicate Review
============================================================

Added functionality to detect and present duplicates for user review.
"""

import pandas as pd
import uuid
from datetime import datetime, time
from sqlalchemy.exc import SQLAlchemyError
from typing import Dict, List, Any, Optional, Tuple
import traceback
import hashlib

class TimeAttendanceImportService:
    """Enhanced service with duplicate detection and review"""
    
    def __init__(self, db, logger_handler=None):
        """Initialize the import service with database and logger"""
        self.db = db
        self.logger = logger_handler

    def _update_progress(self, current: int, total: int, status: str = "Processing"):
        """
        Update progress information for real-time tracking
        
        Args:
            current: Current record number being processed
            total: Total number of records
            status: Status message
        """
        if hasattr(self, 'progress_callback') and self.progress_callback:
            percentage = int((current / total) * 100) if total > 0 else 0
            self.progress_callback({
                'current': current,
                'total': total,
                'percentage': percentage,
                'status': status
            })

    def _read_excel_with_formulas(self, file_path: str) -> pd.DataFrame:
        """
        Read Excel file preserving HYPERLINK formulas in Recorded Address column
        Uses openpyxl to extract formulas, then creates DataFrame
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            DataFrame with formulas preserved
        """
        from openpyxl import load_workbook
        
        # Load workbook with openpyxl to get formulas (data_only=False preserves formulas)
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active
        
        # Get header row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Find the index of 'Recorded Address' column
        recorded_address_idx = None
        try:
            recorded_address_idx = headers.index('Recorded Address')
        except ValueError:
            pass  # Column doesn't exist
        
        # Read all data rows
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_data = []
            for col_idx, cell in enumerate(row):
                # For Recorded Address column, preserve the formula if it exists
                if col_idx == recorded_address_idx and cell.value:
                    # Check if cell contains a formula
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        # This is a formula, keep it as-is
                        row_data.append(cell.value)
                        if self.logger:
                            self.logger.logger.debug(f"Found formula in Recorded Address: {cell.value[:60]}...")
                    else:
                        # Regular value
                        row_data.append(cell.value)
                else:
                    # For other columns, just get the value
                    row_data.append(cell.value)
            
            # Skip completely empty rows
            if any(val is not None for val in row_data):
                data_rows.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=headers)
        
        if self.logger:
            self.logger.logger.info(f"Read Excel with formulas preserved: {len(df)} rows, {len(headers)} columns")
        
        return df

    def _parse_excel_hyperlink(self, cell_value: str) -> str:
        """
        Parse Excel HYPERLINK formula to extract the display text (address)
        
        Handles formats like:
        - =HYPERLINK("http://maps.google.com/maps?q=38.8769894000,-77.2220616000","2815 Hartland Road, Falls Church, VA 22043")
        - Regular text (no formula)
        
        Args:
            cell_value: The cell value which may contain a HYPERLINK formula
            
        Returns:
            Extracted address text or original value if not a hyperlink
        """
        if not cell_value or not isinstance(cell_value, str):
            return cell_value
        
        cell_value = cell_value.strip()
        
        # Check if it's a HYPERLINK formula
        if cell_value.startswith('=HYPERLINK('):
            try:
                import re
                
                # Match the display text (second quoted string)
                # Pattern: =HYPERLINK("url","display_text")
                pattern = r'=HYPERLINK\s*\(\s*"[^"]*"\s*,\s*"([^"]*)"\s*\)'
                match = re.search(pattern, cell_value)
                
                if match:
                    address_text = match.group(1).strip()
                    if self.logger:
                        self.logger.logger.debug(f"Parsed HYPERLINK: '{cell_value[:60]}...' -> '{address_text}'")
                    return address_text
                else:
                    # Fallback: try to extract text between last pair of quotes
                    # Find all quoted strings
                    quoted_strings = re.findall(r'"([^"]*)"', cell_value)
                    if len(quoted_strings) >= 2:
                        # The address is typically the last quoted string
                        address_text = quoted_strings[-1].strip()
                        if self.logger:
                            self.logger.logger.debug(f"Parsed HYPERLINK (fallback): '{cell_value[:60]}...' -> '{address_text}'")
                        return address_text
                    else:
                        if self.logger:
                            self.logger.logger.warning(f"Could not parse HYPERLINK formula: {cell_value[:60]}...")
                        return None
                        
            except Exception as e:
                if self.logger:
                    self.logger.logger.error(f"Failed to parse HYPERLINK formula: {cell_value[:60]}... Error: {e}")
                return None
        
        # Not a hyperlink formula, return as-is
        return cell_value if cell_value else None
    
    def _parse_distance_field(self, row) -> Optional[float]:
        """
        Parse Distance field from Excel (optional column)
        
        Args:
            row: DataFrame row containing the 'Distance' column
            
        Returns:
            Distance value as float in miles, or None if not present/invalid
        """
        # Check if Distance column exists
        if 'Distance' not in row.index:
            return None
        
        distance_value = row.get('Distance')
        
        # Check if value exists and is not NaN
        if pd.isna(distance_value):
            return None
        
        # Try to parse as float
        try:
            # Handle string values
            if isinstance(distance_value, str):
                distance_str = distance_value.strip()
                
                # Skip empty strings
                if not distance_str or distance_str.lower() in ['', 'n/a', 'na', 'none']:
                    return None
                
                # Remove any text like "miles", "mi", "m" 
                distance_str = distance_str.lower()
                distance_str = distance_str.replace('miles', '').replace('mile', '').replace('mi', '').replace('m', '').strip()
                
                # Parse the number
                distance_float = float(distance_str)
            else:
                # Already a number
                distance_float = float(distance_value)
            
            # Validate the distance is reasonable (0 to 1000 miles)
            if 0 <= distance_float <= 1000:
                return distance_float
            else:
                if self.logger:
                    self.logger.logger.warning(f"Distance value {distance_float} is out of reasonable range")
                return None
                
        except (ValueError, TypeError) as e:
            if self.logger:
                self.logger.logger.warning(f"Could not parse distance value '{distance_value}': {e}")
            return None
    
    def _process_recorded_address(self, row) -> Optional[str]:
        """
        Process Recorded Address field from Excel - handles HYPERLINK formulas
        
        Args:
            row: DataFrame row containing the 'Recorded Address' column
            
        Returns:
            Cleaned address text, or None if not present/invalid
        """
        # Check if Recorded Address column exists
        if 'Recorded Address' not in row.index:
            return None
        
        address_value = row.get('Recorded Address')
        
        # Check if value exists and is not NaN
        if pd.isna(address_value):
            return None
        
        # Parse HYPERLINK formula if present, or return raw value
        parsed_address = self._parse_excel_hyperlink(address_value)
        
        # Clean and return
        if parsed_address:
            return str(parsed_address).strip()
        else:
            return None
    
    def _generate_record_hash(self, record_data: Dict[str, Any]) -> str:
        """
        Generate unique hash for a time attendance record
        
        Args:
            record_data: Dictionary containing record data
            
        Returns:
            SHA-256 hash string
        """
        hash_string = (
            f"{record_data['employee_id']}-"
            f"{record_data['attendance_date']}-"
            f"{record_data['attendance_time']}-"
            f"{record_data['location_name']}-"
            f"{record_data['action_description']}"
        )
        
        return hashlib.sha256(hash_string.encode()).hexdigest()
    
    def _get_existing_record_hashes(self) -> set:
        """
        Get hashes of all existing time attendance records
        
        Returns:
            Set of hash strings
        """
        try:
            from models.time_attendance import TimeAttendance
            
            records = TimeAttendance.query.all()
            hashes = set()
            
            for record in records:
                record_data = {
                    'employee_id': record.employee_id,
                    'attendance_date': record.attendance_date,
                    'attendance_time': record.attendance_time,
                    'location_name': record.location_name,
                    'action_description': record.action_description
                }
                hashes.add(self._generate_record_hash(record_data))
            
            return hashes
            
        except Exception as e:
            if self.logger:
                self.logger.logger.error(f"Failed to get existing record hashes: {e}")
            return set()
    
    def _get_existing_record_hashes_with_data(self) -> Dict[str, Dict]:
        """
        Get hashes with corresponding record data for duplicate comparison
        
        Returns:
            Dictionary mapping hash to record data
        """
        try:
            from models.time_attendance import TimeAttendance
            
            records = TimeAttendance.query.all()
            hash_map = {}
            
            for record in records:
                record_data = {
                    'employee_id': record.employee_id,
                    'employee_name': record.employee_name,
                    'attendance_date': record.attendance_date,
                    'attendance_time': record.attendance_time,
                    'location_name': record.location_name,
                    'action_description': record.action_description
                }
                record_hash = self._generate_record_hash(record_data)
                hash_map[record_hash] = record_data
            
            return hash_map
            
        except Exception as e:
            if self.logger:
                self.logger.logger.error(f"Failed to get existing record hashes with data: {e}")
            return {}
    
    def analyze_for_duplicates(self, file_path: str) -> Dict[str, Any]:
        """
        Analyze file for potential duplicates WITHOUT importing
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing duplicate analysis
        """
        analysis_result = {
            'success': False,
            'total_records': 0,
            'new_records': 0,
            'duplicate_records': 0,
            'duplicates': [],
            'errors': []
        }
        
        try:
            # Read Excel file
            df = self._read_excel_with_formulas(file_path)
            
            # Validate required columns
            required_columns = ['ID', 'Date', 'Time', 'Location Name', 'Action Description']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                analysis_result['errors'].append(f"Missing columns: {', '.join(missing_columns)}")
                return analysis_result
            
            # Remove empty rows
            df = df.dropna(how='all')
            analysis_result['total_records'] = len(df)
            
            # Get existing record hashes
            existing_hashes = self._get_existing_record_hashes_with_data()
            
            # Process each row
            duplicates_list = []
            new_records_count = 0
            
            # Process each row with enhanced validation
            for index, row in df.iterrows():
                # Update progress every 10 records or on last record
                if (index + 1) % 10 == 0 or (index + 1) == len(df):
                    self._update_progress(
                        index + 1, 
                        len(df), 
                        f"Processing row {index + 2} of {len(df) + 1}"
                    )
                
                try:
                    # Skip empty rows - only ID is required
                    if pd.isna(row['ID']):
                        continue
                    
                    # Clean the employee ID first (handles float issues like '1234.0')
                    clean_id = self._clean_employee_id(row['ID'])
                    
                    # Get employee name - either from Excel or lookup from employee table
                    employee_name = None
                    if 'Name' in df.columns and pd.notna(row.get('Name')):
                        employee_name = str(row['Name']).strip()
                    else:
                        # Lookup employee name from employee table using cleaned ID
                        employee_name = self._get_employee_name(clean_id)
                    
                    # Parse date and time
                    attendance_date = pd.to_datetime(row['Date']).date()
                    attendance_time = self._parse_time_field(row['Time'])
                    
                    # Prepare record data
                    record_data = {
                        'employee_id': clean_id,
                        'employee_name': employee_name,
                        'platform': str(row.get('Platform', '')).strip() if pd.notna(row.get('Platform')) else None,
                        'attendance_date': attendance_date,
                        'attendance_time': attendance_time,
                        'location_name': str(row['Location Name']).strip(),
                        'action_description': str(row['Action Description']).strip(),
                        'event_description': str(row.get('Event Description', '')).strip() if pd.notna(row.get('Event Description')) else None,
                        'recorded_address': self._process_recorded_address(row),
                    }
                    
                    # Check for duplicates
                    record_hash = self._generate_record_hash(record_data)
                    
                    if record_hash in existing_hashes:
                        # Found duplicate - get existing record details
                        existing_record = existing_hashes[record_hash]
                        
                        duplicates_list.append({
                            'row_number': index + 2,
                            'new_record': record_data,
                            'existing_record': existing_record,
                            'hash': record_hash
                        })
                    else:
                        new_records_count += 1
                        
                except Exception as e:
                    if self.logger:
                        self.logger.logger.warning(f"Error analyzing row {index + 2}: {e}")
                    continue
            
            analysis_result['success'] = True
            analysis_result['new_records'] = new_records_count
            analysis_result['duplicate_records'] = len(duplicates_list)
            analysis_result['duplicates'] = duplicates_list
            
            if self.logger:
                self.logger.logger.info(
                    f"Duplicate analysis complete - Total: {analysis_result['total_records']}, "
                    f"New: {new_records_count}, Duplicates: {len(duplicates_list)}"
                )
        
        except Exception as e:
            analysis_result['errors'].append(f"Analysis failed: {str(e)}")
            if self.logger:
                self.logger.logger.error(f"Duplicate analysis error: {e}")
        
        return analysis_result
    
    def analyze_for_invalid_rows(self, file_path: str) -> Dict[str, Any]:
        """
        Analyze file for invalid rows with detailed error information
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing invalid row analysis
        """
        analysis_result = {
            'success': False,
            'total_rows': 0,
            'valid_rows': 0,
            'invalid_rows': 0,
            'invalid_details': [],
            'errors': []
        }
        
        try:
            # Read Excel file
            df = self._read_excel_with_formulas(file_path)
            
            # Validate required columns
            required_columns = ['ID', 'Date', 'Time', 'Location Name', 'Action Description']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                analysis_result['errors'].append(f"Missing columns: {', '.join(missing_columns)}")
                return analysis_result
            
            # Remove empty rows
            df = df.dropna(how='all')
            analysis_result['total_rows'] = len(df)
            
            # Analyze each row
            invalid_list = []
            valid_count = 0
            
            for index, row in df.iterrows():
                row_errors = []
                row_data = {
                    'row_number': index + 2,
                    'employee_id': self._clean_employee_id(row['ID']) if pd.notna(row['ID']) else None,
                }
                
                # Check ID
                if pd.isna(row['ID']):
                    row_errors.append("Missing ID")
                else:
                    row_data['employee_id'] = self._clean_employee_id(row['ID'])
                    # Get employee name
                    if 'Name' in df.columns and pd.notna(row.get('Name')):
                        row_data['employee_name'] = str(row['Name']).strip()
                    else:
                        row_data['employee_name'] = self._get_employee_name(row_data['employee_id'])
                
                # Check and parse Date
                if pd.isna(row['Date']):
                    row_errors.append("Missing Date")
                else:
                    try:
                        attendance_date = pd.to_datetime(row['Date']).date()
                        row_data['attendance_date'] = attendance_date
                    except Exception:
                        row_errors.append(f"Invalid date format: {row['Date']}")
                
                # Check and parse Time
                if pd.isna(row['Time']):
                    row_errors.append("Missing Time")
                else:
                    try:
                        attendance_time = self._parse_time_field(row['Time'])
                        row_data['attendance_time'] = attendance_time
                    except Exception as e:
                        row_errors.append(f"Invalid time format: {row['Time']}")
                
                # Check Location Name
                if pd.isna(row['Location Name']):
                    row_errors.append("Missing Location Name")
                else:
                    row_data['location_name'] = str(row['Location Name']).strip()
                
                # Check Action Description
                if pd.isna(row['Action Description']):
                    row_errors.append("Missing Action Description")
                else:
                    row_data['action_description'] = str(row['Action Description']).strip()
                
                # Optional fields
                if pd.notna(row.get('Platform')):
                    row_data['platform'] = str(row['Platform']).strip()
                
                if pd.notna(row.get('Event Description')):
                    row_data['event_description'] = str(row['Event Description']).strip()
                
                row_data['recorded_address'] = self._process_recorded_address(row)
                
                # If row has errors, add to invalid list
                if row_errors:
                    invalid_list.append({
                        'row_number': index + 2,  # +2 for header and 0-based index
                        'row_data': row_data,
                        'errors': row_errors
                    })
                else:
                    valid_count += 1
            
            analysis_result['success'] = True
            analysis_result['valid_rows'] = valid_count
            analysis_result['invalid_rows'] = len(invalid_list)
            analysis_result['invalid_details'] = invalid_list
            
            if self.logger:
                self.logger.logger.info(
                    f"Invalid row analysis complete - Total: {analysis_result['total_rows']}, "
                    f"Valid: {valid_count}, Invalid: {len(invalid_list)}"
                )
        
        except Exception as e:
            analysis_result['errors'].append(f"Analysis failed: {str(e)}")
            if self.logger:
                self.logger.logger.error(f"Invalid row analysis error: {e}")
        
        return analysis_result
    
    def import_from_excel(self, file_path: str, created_by: int = None,
                         import_source: str = None, skip_duplicates: bool = True,
                         force_import_hashes: List[str] = None, project_id: int = None,
                         progress_callback=None) -> Dict[str, Any]:
        """
        Import time attendance data from Excel file with enhanced duplicate handling.

        Args:
            file_path: Path to the Excel file
            created_by: User ID who initiated the import
            import_source: Description of import source
            skip_duplicates: Whether to skip duplicate records
            force_import_hashes: List of hashes to force import (user confirmed duplicates)
            progress_callback: Optional callable(current, total, message) for real-time progress

        Returns:
            Dictionary containing import results
        """
        batch_id = str(uuid.uuid4())
        import_results = {
            'batch_id': batch_id,
            'total_records': 0,
            'imported_records': 0,
            'failed_records': 0,
            'duplicate_records': 0,
            'skipped_records': 0,
            'forced_duplicates': 0,
            'errors': [],
            'warnings': [],
            'success': False,
            'import_date': datetime.utcnow()
        }
        
        force_import_hashes = force_import_hashes or []
        
        try:
            # Log import start
            if self.logger:
                self.logger.logger.info(
                    f"Starting enhanced time attendance import from {file_path} by user {created_by} "
                    f"(skip_duplicates={skip_duplicates}, force_import={len(force_import_hashes)})"
                )
            
            # Read Excel file
            try:
                df = self._read_excel_with_formulas(file_path)
                
                if self.logger:
                    self.logger.logger.info(f"Read Excel file with {len(df)} rows and formulas preserved")
            except Exception as e:
                error_msg = f"Failed to read Excel file: {str(e)}"
                import_results['errors'].append(error_msg)
                if self.logger:
                    self.logger.logger.error(error_msg)
                return import_results
            
            # Validate required columns
            required_columns = ['ID', 'Date', 'Time', 'Location Name', 'Action Description']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                error_msg = f"Missing required columns: {', '.join(missing_columns)}"
                import_results['errors'].append(error_msg)
                if self.logger:
                    self.logger.logger.error(error_msg)
                return import_results
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            import_results['total_records'] = len(df)
            
            if import_results['total_records'] == 0:
                error_msg = "No valid data rows found in Excel file"
                import_results['errors'].append(error_msg)
                return import_results
            
            # ── Project-location validation ──────────────────────────────────────
            # If a project_id is provided, verify that every unique Location Name
            # in the file belongs to that project.  Return immediately on the first
            # mismatch so the user can correct the file or project selection.
            if project_id:
                try:
                    from models.qrcode import QRCode
                    from models.project import Project

                    # Collect all unique, non-empty location names from the file
                    file_locations = set(
                        str(loc).strip()
                        for loc in df['Location Name'].dropna().unique()
                        if str(loc).strip()
                    )

                    # Fetch all location names that belong to the selected project
                    project_locations = set(
                        qr.location
                        for qr in QRCode.query.filter_by(project_id=project_id)
                                              .with_entities(QRCode.location).all()
                    )

                    # Find the first location in the file that is not in the project
                    unmatched = next(
                        (loc for loc in sorted(file_locations) if loc not in project_locations),
                        None
                    )

                    if unmatched:
                        project_obj = Project.query.get(project_id)
                        project_name = project_obj.name if project_obj else f'ID {project_id}'
                        error_msg = (
                            f"Location '{unmatched}' in the file does not belong to "
                            f"project '{project_name}'. "
                            f"Please verify the selected project or correct the file."
                        )
                        import_results['errors'].append(error_msg)
                        if self.logger:
                            self.logger.logger.warning(
                                f"Project-location mismatch: {error_msg}"
                            )
                        return import_results

                    if self.logger:
                        self.logger.logger.info(
                            f"Project-location validation passed: all {len(file_locations)} "
                            f"location(s) belong to project ID {project_id}."
                        )

                except Exception as e:
                    if self.logger:
                        self.logger.logger.warning(
                            f"Could not perform project-location validation: {e}"
                        )
            # ── End project-location validation ──────────────────────────────────

            # Track duplicates using hash
            duplicate_hashes = set()
            if skip_duplicates:
                duplicate_hashes = self._get_existing_record_hashes()

            # Process each row with enhanced validation
            for index, row in df.iterrows():
                try:
                    # Skip empty rows - only ID is required
                    if pd.isna(row['ID']):
                        import_results['skipped_records'] += 1
                        import_results['warnings'].append(f"Row {index + 2}: Skipped due to missing ID")
                        continue
                    
                    # Clean the employee ID first (handles float issues like '1234.0')
                    clean_id = self._clean_employee_id(row['ID'])
                    
                    # Get employee name - either from Excel or lookup from employee table
                    employee_name = None
                    if 'Name' in df.columns and pd.notna(row.get('Name')):
                        employee_name = str(row['Name']).strip()
                    else:
                        # Lookup employee name from employee table using cleaned ID
                        employee_name = self._get_employee_name(clean_id)
                    
                    # Validate and parse date
                    try:
                        attendance_date = pd.to_datetime(row['Date']).date()
                    except Exception as date_error:
                        import_results['failed_records'] += 1
                        import_results['errors'].append(f"Row {index + 2}: Invalid date format - {str(date_error)}")
                        continue
                    
                    # Validate and parse time
                    try:
                        attendance_time = self._parse_time_field(row['Time'])
                    except Exception as time_error:
                        import_results['failed_records'] += 1
                        import_results['errors'].append(f"Row {index + 2}: Invalid time format - {str(time_error)}")
                        continue
                    
                    # Prepare record data
                    record_data = {
                        'employee_id': clean_id,
                        'employee_name': employee_name,
                        'platform': str(row.get('Platform', '')).strip() if pd.notna(row.get('Platform')) else None,
                        'attendance_date': attendance_date,
                        'attendance_time': attendance_time,
                        'location_name': str(row['Location Name']).strip(),
                        'action_description': str(row['Action Description']).strip(),
                        'event_description': str(row.get('Event Description', '')).strip() if pd.notna(row.get('Event Description')) else None,
                        'recorded_address': self._process_recorded_address(row),
                        'distance': self._parse_distance_field(row),
                    }
                    
                    # Check for duplicates
                    if skip_duplicates:
                        record_hash = self._generate_record_hash(record_data)
                        
                        # If duplicate and NOT in force import list, skip it
                        if record_hash in duplicate_hashes and record_hash not in force_import_hashes:
                            import_results['duplicate_records'] += 1
                            import_results['warnings'].append(
                                f"Row {index + 2}: Duplicate record for {record_data['employee_name']} "
                                f"on {attendance_date} at {attendance_time} - Skipped"
                            )
                            continue
                        
                        # If in force import list, track it
                        if record_hash in force_import_hashes:
                            import_results['forced_duplicates'] += 1
                        
                        duplicate_hashes.add(record_hash)
                    
                    # Create TimeAttendance record
                    from models.time_attendance import TimeAttendance
                    
                    time_attendance_record = TimeAttendance(
                        **record_data,
                        import_batch_id=batch_id,
                        import_source=import_source or f"Excel Import - {file_path}",
                        created_by=created_by,
                        project_id=project_id
                    )
                    
                    self.db.session.add(time_attendance_record)
                    import_results['imported_records'] += 1

                    # Commit in batches of 50 to prevent memory buildup on large files
                    if import_results['imported_records'] % 50 == 0:
                        self.db.session.commit()

                    # Report progress via callback if provided
                    if progress_callback:
                        processed = (import_results['imported_records'] +
                                     import_results['failed_records'] +
                                     import_results['duplicate_records'] +
                                     import_results['skipped_records'])
                        progress_callback(
                            processed,
                            import_results['total_records'],
                            f"Importing record {processed} of {import_results['total_records']}..."
                        )
                    
                except Exception as e:
                    import_results['failed_records'] += 1
                    error_msg = f"Row {index + 2}: {str(e)}"
                    import_results['errors'].append(error_msg)
                    
                    if self.logger:
                        self.logger.logger.warning(f"Failed to import row {index + 2}: {e}")
                    
                    continue
            
            # Final commit
            self.db.session.commit()
            import_results['success'] = import_results['imported_records'] > 0
            
            # Log successful import
            if self.logger:
                self.logger.logger.info(
                    f"Time attendance import completed - Batch: {batch_id}, "
                    f"Total: {import_results['total_records']}, "
                    f"Imported: {import_results['imported_records']}, "
                    f"Failed: {import_results['failed_records']}, "
                    f"Duplicates: {import_results['duplicate_records']}, "
                    f"Forced: {import_results['forced_duplicates']}, "
                    f"Skipped: {import_results['skipped_records']}"
                )
            
        except SQLAlchemyError as e:
            self.db.session.rollback()
            error_msg = f"Database error during import: {str(e)}"
            import_results['errors'].append(error_msg)
            
            if self.logger:
                self.logger.log_database_error('time_attendance_import', e)
        
        except Exception as e:
            self.db.session.rollback()
            error_msg = f"Unexpected error during import: {str(e)}"
            import_results['errors'].append(error_msg)
            import_results['traceback'] = traceback.format_exc()
            
            if self.logger:
                self.logger.logger.error(f"Time attendance import failed: {e}")
                self.logger.logger.error(f"Traceback: {traceback.format_exc()}")
        
        return import_results
    
    def _parse_time_field(self, time_value) -> time:
        """
        Parse various time formats from Excel
        
        Handles:
        - datetime.time objects
        - datetime.datetime objects
        - String formats (HH:MM, HH:MM:SS, HH:MM AM/PM)
        
        Args:
            time_value: Time value from Excel
            
        Returns:
            time object
        """
        if pd.isna(time_value):
            raise ValueError("Time value is empty")
        
        # If already a time object
        if isinstance(time_value, time):
            return time_value
        
        # If datetime object, extract time
        if isinstance(time_value, datetime):
            return time_value.time()
        
        # If string, parse it
        if isinstance(time_value, str):
            time_str = time_value.strip()
            
            # Try parsing with pandas
            try:
                dt = pd.to_datetime(time_str)
                return dt.time()
            except:
                # Try manual parsing for common formats
                try:
                    # Format: HH:MM or HH:MM:SS
                    parts = time_str.split(':')
                    if len(parts) >= 2:
                        hour = int(parts[0])
                        minute = int(parts[1])
                        second = int(parts[2]) if len(parts) > 2 else 0
                        return time(hour, minute, second)
                except:
                    pass
        
        raise ValueError(f"Could not parse time value: {time_value}")
    
    def _get_employee_name(self, employee_id: str) -> str:
        """
        Get employee name from database, formatted as 'lastname, firstname'
        
        Args:
            employee_id: Employee ID to lookup
            
        Returns:
            Employee name in 'lastname, firstname' format
        """
        try:
            from models.employee import Employee
            
            # CRITICAL: Clean the employee_id to handle float values like '1234.0'
            # Remove '.0' suffix if present and convert to integer
            cleaned_id = str(employee_id).strip()
            
            # If it's a float string like '1234.0', remove the decimal part
            if '.' in cleaned_id:
                try:
                    # Convert to float first, then to int to handle '1234.0' -> 1234
                    cleaned_id = str(int(float(cleaned_id)))
                except (ValueError, TypeError):
                    pass  # Keep original if conversion fails
            
            # Now lookup the employee
            employee = Employee.get_by_employee_id(int(cleaned_id))
            if employee:
                # Format name as "lastname, firstname"
                return f"{employee.lastName}, {employee.firstName}"
            else:
                if self.logger:
                    self.logger.logger.warning(f"Employee ID {cleaned_id} not found in employee table")
                return f"Employee {cleaned_id}"
        except Exception as e:
            if self.logger:
                self.logger.logger.warning(f"Could not lookup employee name for ID {employee_id}: {e}")
            # Return cleaned ID in error message too
            try:
                cleaned_id = str(int(float(str(employee_id).strip())))
                return f"Employee {cleaned_id}"
            except:
                return f"Employee {employee_id}"
            
    def _clean_employee_id(self, employee_id) -> str:
        """
        Clean employee ID to handle various formats
        
        Args:
            employee_id: Raw employee ID value
            
        Returns:
            Cleaned employee ID string
        """
        try:
            # Convert to string first
            id_str = str(employee_id).strip()
            
            # Handle float values like 1234.0 or '1234.0'
            if '.' in id_str:
                # Convert to float, then to int, then back to string
                # This removes the decimal part: 1234.0 -> 1234
                id_str = str(int(float(id_str)))
            
            return id_str
        except (ValueError, TypeError) as e:
            # If conversion fails, return original string
            if self.logger:
                self.logger.logger.warning(f"Could not clean employee ID '{employee_id}': {e}")
            return str(employee_id).strip()
    
    def validate_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Validate Excel file structure and content before import
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing validation results
        """
        validation_results = {
            'valid': False,
            'errors': [],
            'warnings': [],
            'total_rows': 0,
            'valid_rows': 0,
            'invalid_rows': 0
        }
        
        try:
            # Try to read the Excel file
            df = pd.read_excel(file_path)
            validation_results['total_rows'] = len(df)
            
            # Check required columns
            required_columns = ['ID', 'Date', 'Time', 'Location Name', 'Action Description']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                validation_results['errors'].append(
                    f"Missing required columns: {', '.join(missing_columns)}"
                )
            
            # Check optional columns
            optional_columns = ['Name', 'Platform', 'Event Description', 'Recorded Address', 'Distance']
            present_optional = [col for col in optional_columns if col in df.columns]
            
            if present_optional:
                validation_results['warnings'].append(
                    f"Optional columns found: {', '.join(present_optional)}"
                )
            
            # Validate data in rows
            if not missing_columns:
                valid_row_count = 0
                invalid_row_details = []
                
                for index, row in df.iterrows():
                    is_valid = True
                    missing_fields = []
                    
                    for col in required_columns:
                        if col in df.columns:
                            if pd.isna(row[col]):
                                is_valid = False
                                missing_fields.append(col)
                        else:
                            # Column doesn't exist in file
                            is_valid = False
                            missing_fields.append(f"{col} (column not found)")
                    
                    if is_valid:
                        valid_row_count += 1
                    else:
                        # Track invalid row for detailed reporting
                        invalid_row_details.append({
                            'row': index + 2,  # +2 for header and 0-based index
                            'missing': missing_fields
                        })

                validation_results['valid_rows'] = valid_row_count
                validation_results['invalid_rows'] = len(df) - valid_row_count

                if validation_results['invalid_rows'] > 0:
                    # Provide detailed warning about invalid rows
                    validation_results['warnings'].append(
                        f"{validation_results['invalid_rows']} rows have missing required data"
                    )
                    
                    # Add details about first few invalid rows for debugging
                    if invalid_row_details:
                        sample_invalid = invalid_row_details[:3]  # Show first 3 invalid rows
                        details_msg = "Examples: "
                        for detail in sample_invalid:
                            details_msg += f"Row {detail['row']} (missing: {', '.join(detail['missing'])}); "
                        validation_results['warnings'].append(details_msg.rstrip('; '))
            
            if 'Date' in df.columns:
                invalid_dates = 0
                for idx, date_val in df['Date'].items():
                    if pd.notna(date_val):
                        try:
                            pd.to_datetime(date_val)
                        except:
                            invalid_dates += 1
                
                if invalid_dates > 0:
                    validation_results['warnings'].append(
                        f"{invalid_dates} rows have invalid date format"
                    )
            
            if 'Time' in df.columns:
                invalid_times = 0
                for idx, time_val in df['Time'].items():
                    if pd.notna(time_val):
                        try:
                            self._parse_time_field(time_val)
                        except:
                            invalid_times += 1
                
                if invalid_times > 0:
                    validation_results['warnings'].append(
                        f"{invalid_times} rows have invalid time format"
                    )
            
            if all(col in df.columns for col in ['ID', 'Date', 'Time', 'Location Name']):
                duplicate_check = df[['ID', 'Date', 'Time', 'Location Name']].duplicated()
                duplicate_count = duplicate_check.sum()
                
                if duplicate_count > 0:
                    validation_results['warnings'].append(
                        f"{duplicate_count} potential duplicate records detected"
                    )
            
            validation_results['valid'] = (
                len(validation_results['errors']) == 0 and 
                validation_results['valid_rows'] > 0
            )
            
        except Exception as e:
            validation_results['errors'].append(f"Failed to validate Excel file: {str(e)}")
            if self.logger:
                self.logger.logger.error(f"Validation error: {e}")
        
        return validation_results
    
    def get_import_summary(self, batch_id: str) -> Optional[Dict[str, Any]]:
        """
        Get detailed summary of imported data by batch ID
        
        Args:
            batch_id: Import batch identifier
            
        Returns:
            Dictionary containing import summary
        """
        try:
            from models.time_attendance import TimeAttendance
            
            records = TimeAttendance.get_by_import_batch(batch_id)
            
            if not records:
                return None
            
            # Calculate summary statistics
            total_records = len(records)
            unique_employees = len(set(record.employee_id for record in records))
            unique_locations = len(set(record.location_name for record in records))
            date_range = {
                'start': min(record.attendance_date for record in records),
                'end': max(record.attendance_date for record in records)
            }
            
            # Group by action description
            actions = {}
            for record in records:
                action = record.action_description
                actions[action] = actions.get(action, 0) + 1
            
            # Group by employee
            employee_summary = {}
            for record in records:
                emp_id = record.employee_id
                if emp_id not in employee_summary:
                    employee_summary[emp_id] = {
                        'name': record.employee_name,
                        'count': 0
                    }
                employee_summary[emp_id]['count'] += 1
            
            return {
                'batch_id': batch_id,
                'total_records': total_records,
                'unique_employees': unique_employees,
                'unique_locations': unique_locations,
                'date_range': date_range,
                'actions': actions,
                'employee_summary': employee_summary,
                'import_date': records[0].import_date if records else None,
                'import_source': records[0].import_source if records else None
            }
            
        except Exception as e:
            if self.logger:
                self.logger.logger.error(f"Failed to get import summary for batch {batch_id}: {e}")
            return None
    
    def delete_import_batch(self, batch_id: str, deleted_by: int = None) -> Dict[str, Any]:
        """
        Delete all records from a specific import batch
        
        Args:
            batch_id: Import batch identifier
            deleted_by: User ID who initiated the deletion
            
        Returns:
            Dictionary containing deletion results
        """
        result = {
            'success': False,
            'deleted_count': 0,
            'message': ''
        }
        
        try:
            from models.time_attendance import TimeAttendance
            
            records = TimeAttendance.query.filter_by(import_batch_id=batch_id).all()
            deleted_count = len(records)
            
            if deleted_count == 0:
                result['message'] = 'No records found for this batch'
                return result
            
            # Delete records
            for record in records:
                self.db.session.delete(record)
            
            self.db.session.commit()
            
            # Log deletion
            if self.logger:
                self.logger.logger.info(
                    f"User {deleted_by} deleted import batch {batch_id} - "
                    f"Removed {deleted_count} records"
                )
            
            result['success'] = True
            result['deleted_count'] = deleted_count
            result['message'] = f'Successfully deleted {deleted_count} records'
            
        except Exception as e:
            self.db.session.rollback()
            result['message'] = f'Error deleting batch: {str(e)}'
            
            if self.logger:
                self.logger.logger.error(f"Failed to delete batch {batch_id}: {e}")
        
        return result