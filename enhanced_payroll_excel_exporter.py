"""
Enhanced Payroll Excel Exporter with SP/PW Support
=================================================

Extends the existing PayrollExcelExporter to support Special Project (SP) 
and Periodic Work (PW) hours in Excel reports.
"""

import io
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from single_checkin_calculator import SingleCheckInCalculator
from logger_handler import log_database_operations


class EnhancedPayrollExcelExporter:
    """Enhanced Excel exporter with SP/PW support"""
    
    def __init__(self, company_name: str = "Your Company Name"):
        self.company_name = company_name
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup Excel cell styles"""
        # Header styles
        self.header_style = {
            'font': Font(name="Arial", size=11, bold=True, color="FFFFFF"),
            'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }
        
        # Data styles
        self.data_style = {
            'font': Font(name="Arial", size=10),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }
        
        # Summary styles
        self.summary_style = {
            'font': Font(name="Arial", size=11, bold=True),
            'fill': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }
    
    def _apply_style(self, cell, style_dict):
        """Apply style dictionary to a cell"""
        for attr, value in style_dict.items():
            setattr(cell, attr, value)
    
    @log_database_operations('enhanced_payroll_export')
    def create_enhanced_payroll_report(self, start_date: datetime, end_date: datetime,
                                     attendance_records: List[Dict], employee_names: Dict[str, str] = None,
                                     project_name: str = None) -> io.BytesIO:
        """
        Create enhanced payroll report with SP/PW hours
        
        Args:
            start_date: Report start date
            end_date: Report end date
            attendance_records: List of attendance records
            employee_names: Dictionary mapping employee_id to full name
            project_name: Project name for the report
                        
        Returns:
            BytesIO buffer containing the Excel file
        """
        try:
            print(f"📊 Creating enhanced payroll report with SP/PW support from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
            
            # Get employee names using existing method
            if employee_names is None:
                employee_names = self._get_employee_names(attendance_records)
            
            # Calculate working hours using enhanced calculator
            calculator = SingleCheckInCalculator()
            working_hours_data = calculator.calculate_all_employees_hours(
                start_date, end_date, attendance_records
            )
            
            # Create workbook
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Enhanced Payroll Report"
            
            # Write report header
            current_row = self._write_enhanced_header(worksheet, start_date, end_date, project_name)
            
            # Write column headers
            headers = [
                "Employee ID", "Employee Name", "Total Hours", "Regular Hours", 
                "Overtime Hours", "SP Hours", "PW Hours", "Working Days", "Status"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col, value=header)
                self._apply_style(cell, self.header_style)
            current_row += 1
            
            # Write employee data
            total_employees = 0
            totals = {
                'total_hours': 0.0,
                'regular_hours': 0.0,
                'overtime_hours': 0.0,
                'sp_hours': 0.0,
                'pw_hours': 0.0
            }
            
            for employee_id, emp_data in working_hours_data['employees'].items():
                employee_name = employee_names.get(employee_id, f'Employee {employee_id}')
                
                # Calculate working days and status
                working_days = len([d for d in emp_data['daily_hours'].values() if d['total_hours'] > 0])
                miss_punches = len([d for d in emp_data['daily_hours'].values() if d['is_miss_punch']])
                status = f"{miss_punches} miss punch(es)" if miss_punches > 0 else "Complete"
                
                # Get grand totals
                grand_totals = emp_data['grand_totals']
                
                row_data = [
                    employee_id,
                    employee_name,
                    round(grand_totals['total_hours'], 2),
                    round(grand_totals['regular_hours'], 2),
                    round(grand_totals['overtime_hours'], 2),
                    round(grand_totals['sp_hours'], 2),
                    round(grand_totals['pw_hours'], 2),
                    working_days,
                    status
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=current_row, column=col, value=value)
                    self._apply_style(cell, self.data_style)
                    
                    # Highlight miss punches
                    if miss_punches > 0 and col == 9:  # Status column
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                # Add to totals
                totals['total_hours'] += grand_totals['total_hours']
                totals['regular_hours'] += grand_totals['regular_hours']
                totals['overtime_hours'] += grand_totals['overtime_hours']
                totals['sp_hours'] += grand_totals['sp_hours']
                totals['pw_hours'] += grand_totals['pw_hours']
                total_employees += 1
                
                current_row += 1
            
            # Write summary section
            current_row = self._write_enhanced_summary(worksheet, current_row, total_employees, totals)
            
            # Auto-adjust columns
            self._auto_adjust_columns(worksheet)
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            print("✅ Enhanced payroll report Excel file created successfully")
            return excel_buffer
            
        except Exception as e:
            print(f"❌ Error creating enhanced payroll report: {e}")
            raise e
    
    def _write_enhanced_header(self, worksheet, start_date: datetime, end_date: datetime, project_name: str) -> int:
        """Write enhanced report header"""
        current_row = 1
        
        # Company name
        cell = worksheet.cell(row=current_row, column=1, value=self.company_name)
        cell.font = Font(name="Arial", size=14, bold=True)
        worksheet.merge_cells(f'A{current_row}:I{current_row}')
        current_row += 1
        
        # Report title
        cell = worksheet.cell(row=current_row, column=1, value="Enhanced Payroll Report with SP/PW Hours")
        cell.font = Font(name="Arial", size=12, bold=True)
        worksheet.merge_cells(f'A{current_row}:I{current_row}')
        current_row += 1
        
        # Project name
        if project_name:
            cell = worksheet.cell(row=current_row, column=1, value=f"Project: {project_name}")
            cell.font = Font(name="Arial", size=11)
            worksheet.merge_cells(f'A{current_row}:I{current_row}')
            current_row += 1
        
        # Date range
        date_range = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        cell = worksheet.cell(row=current_row, column=1, value=date_range)
        cell.font = Font(name="Arial", size=11)
        worksheet.merge_cells(f'A{current_row}:I{current_row}')
        current_row += 2  # Add space
        
        return current_row
    
    def _write_enhanced_summary(self, worksheet, start_row: int, total_employees: int, totals: Dict) -> int:
        """Write enhanced summary section"""
        current_row = start_row + 1
        
        # Summary header
        cell = worksheet.cell(row=current_row, column=1, value="SUMMARY")
        cell.font = Font(name="Arial", size=12, bold=True)
        worksheet.merge_cells(f'A{current_row}:I{current_row}')
        current_row += 1
        
        # Summary data
        summary_data = [
            ("Total Employees:", total_employees),
            ("Total Hours:", round(totals['total_hours'], 2)),
            ("Regular Hours:", round(totals['regular_hours'], 2)),
            ("Overtime Hours:", round(totals['overtime_hours'], 2)),
            ("Special Project (SP) Hours:", round(totals['sp_hours'], 2)),
            ("Periodic Work (PW) Hours:", round(totals['pw_hours'], 2)),
            ("Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        for label, value in summary_data:
            cell1 = worksheet.cell(row=current_row, column=1, value=label)
            self._apply_style(cell1, self.summary_style)
            cell2 = worksheet.cell(row=current_row, column=2, value=value)
            self._apply_style(cell2, self.summary_style)
            current_row += 1
        
        return current_row
    
    def _get_employee_names(self, attendance_records: List[Dict]) -> Dict[str, str]:
        """Get employee names from attendance records"""
        # This method should be implemented based on your existing database structure
        # For now, return a basic implementation
        employee_names = {}
        for record in attendance_records:
            if hasattr(record, '__dict__'):
                emp_id = str(record.employee_id)
                if emp_id not in employee_names:
                    employee_names[emp_id] = f'Employee {emp_id}'
            else:
                emp_id = str(record['employee_id'])
                if emp_id not in employee_names:
                    employee_names[emp_id] = f'Employee {emp_id}'
        return employee_names
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @log_database_operations('detailed_sp_pw_export')
    def create_detailed_sp_pw_report(self, start_date: datetime, end_date: datetime,
                                   attendance_records: List[Dict], employee_names: Dict[str, str] = None) -> io.BytesIO:
        """
        Create detailed daily report showing SP/PW breakdown by day
        """
        try:
            print(f"📊 Creating detailed SP/PW daily report")
            
            # Get employee names
            if employee_names is None:
                employee_names = self._get_employee_names(attendance_records)
            
            # Calculate working hours
            calculator = SingleCheckInCalculator()
            working_hours_data = calculator.calculate_all_employees_hours(
                start_date, end_date, attendance_records
            )
            
            # Create workbook
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Detailed SP/PW Report"
            
            # Write header
            current_row = 1
            cell = worksheet.cell(row=current_row, column=1, value="Detailed Daily Hours Report with SP/PW Breakdown")
            cell.font = Font(name="Arial", size=14, bold=True)
            worksheet.merge_cells(f'A{current_row}:K{current_row}')
            current_row += 2
            
            # Column headers
            headers = [
                "Employee ID", "Employee Name", "Date", "Day", "Regular Hours", 
                "SP Hours", "PW Hours", "Total Hours", "Records", "Status", "Notes"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col, value=header)
                self._apply_style(cell, self.header_style)
            current_row += 1
            
            # Write daily data for each employee
            for employee_id, emp_data in working_hours_data['employees'].items():
                employee_name = employee_names.get(employee_id, f'Employee {employee_id}')
                
                for date_str, day_data in emp_data['daily_hours'].items():
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    day_name = date_obj.strftime('%A')
                    
                    # Get hours by type
                    regular_hours = day_data.get('regular_hours', 0.0)
                    sp_hours = day_data.get('sp_hours', 0.0)
                    pw_hours = day_data.get('pw_hours', 0.0)
                    total_hours = day_data.get('total_hours', 0.0)
                    
                    # Status and notes
                    if day_data.get('is_miss_punch', False):
                        status = "Miss Punch"
                        notes = "Incomplete records"
                    elif total_hours == 0:
                        status = "No Work"
                        notes = "No attendance records"
                    else:
                        status = "Complete"
                        notes = f"{day_data['records_count']} record(s)"
                    
                    row_data = [
                        employee_id,
                        employee_name,
                        date_str,
                        day_name,
                        round(regular_hours, 2),
                        round(sp_hours, 2),
                        round(pw_hours, 2),
                        round(total_hours, 2),
                        day_data['records_count'],
                        status,
                        notes
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = worksheet.cell(row=current_row, column=col, value=value)
                        self._apply_style(cell, self.data_style)
                        
                        # Color coding
                        if status == "Miss Punch":
                            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        elif sp_hours > 0 and col == 6:  # SP Hours column
                            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        elif pw_hours > 0 and col == 7:  # PW Hours column
                            cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                    
                    current_row += 1
            
            # Auto-adjust columns
            self._auto_adjust_columns(worksheet)
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            print("✅ Detailed SP/PW report Excel file created successfully")
            return excel_buffer
            
        except Exception as e:
            print(f"❌ Error creating detailed SP/PW report: {e}")
            raise e